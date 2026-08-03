from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, Response
from urllib.parse import urlparse
import csv
import io
import re
import json
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_migrate import Migrate
import os

from models import db, User, Project, ProjectHistory, TaskTemplate, PromotorTask, DailyUpdate, Product, Quote, QuoteItem, Supplier, GlassType, SupplierPricing, Reminder, PurchaseInvoice
from config import config
from forms import (LoginForm, UserForm, ProjectForm, TaskTemplateForm, 
                   TaskAssignmentForm, TaskUpdateForm, DailyUpdateForm, ProductForm)
from utils.auth import admin_required, manager_or_admin_required
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from utils.task_rollover import rollover_incomplete_tasks, get_current_week_info, get_week_date_range
from utils.s3_upload import S3Uploader
from datetime import datetime, timedelta
from sqlalchemy import text

# Initialize Flask app
app = Flask(__name__)

# Load configuration
env = os.getenv('ENVIRONMENT', 'development')
app.config.from_object(config[env])

# Initialize extensions
db.init_app(app)
migrate = Migrate(app, db)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
limiter = Limiter(get_remote_address, app=app, default_limits=[], storage_uri="memory://")
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def _run_startup_migrations():
    """Run safe ADD COLUMN migrations on every cold start (idempotent — ignores duplicates)."""
    stmts = [
        "ALTER TABLE leads ADD COLUMN facebook_lead_id VARCHAR(50) NULL",
        # WhatsApp Cloud API send-log. Created on cold start so a brand-new
        # Lambda deploy works without a separate migration step.
        """
        CREATE TABLE IF NOT EXISTS whatsapp_messages (
            id              SERIAL PRIMARY KEY,
            lead_id         INTEGER REFERENCES leads(id)    ON DELETE SET NULL,
            meeting_id      INTEGER REFERENCES meetings(id) ON DELETE SET NULL,
            to_number       VARCHAR(20)  NOT NULL,
            template_name   VARCHAR(100) NOT NULL,
            language        VARCHAR(10)  NOT NULL DEFAULT 'en',
            variables_json  TEXT,
            wamid           VARCHAR(120) UNIQUE,
            status          VARCHAR(20)  NOT NULL DEFAULT 'queued',
            error_message   TEXT,
            sent_by         INTEGER NOT NULL REFERENCES users(id),
            sent_at         TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            updated_at      TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        )
        """,
        "CREATE INDEX IF NOT EXISTS idx_wa_lead_id    ON whatsapp_messages(lead_id)",
        "CREATE INDEX IF NOT EXISTS idx_wa_meeting_id ON whatsapp_messages(meeting_id)",
        "CREATE INDEX IF NOT EXISTS idx_wa_to_number  ON whatsapp_messages(to_number)",
        "CREATE INDEX IF NOT EXISTS idx_wa_status     ON whatsapp_messages(status)",
        # Vetrova BD-creator project (2026-08-01): distinguish quotes
        # ingested from vetrova.in vs those created by BD in vcore's
        # own /quotes/vetrova/new flow. 'ingest' matches every existing
        # row's provenance so back-fill is implicit via the default.
        "ALTER TABLE vetrova_quotes ADD COLUMN source VARCHAR(20) NOT NULL DEFAULT 'ingest'",
    ]
    # Each statement runs in its own transaction. Required because Postgres
    # puts the connection into an "aborted transaction" state after ANY
    # failed statement (e.g. ALTER TABLE on a column that already exists),
    # and every subsequent statement on the same transaction silently fails
    # until ROLLBACK. Without this isolation, a single duplicate-ADD-COLUMN
    # blocks the rest of the migration list from running — which is exactly
    # how the whatsapp_messages CREATE TABLE got swallowed on 2026-06-25.
    try:
        for sql in stmts:
            try:
                with db.engine.begin() as conn:
                    conn.execute(text(sql))
            except Exception:
                pass  # Idempotent — column/table/index already exists
    except Exception:
        pass  # DB not reachable yet (local dev before first request)


with app.app_context():
    _run_startup_migrations()


@app.template_filter('fromjson')
def fromjson_filter(value):
    """Parse a JSON string in Jinja2 templates."""
    try:
        return json.loads(value)
    except Exception:
        return []


@app.template_filter('to_ist')
def to_ist_filter(utc_dt):
    """Convert a UTC datetime to IST (UTC+5:30), returns datetime object."""
    if utc_dt is None:
        return None
    return utc_dt + timedelta(hours=5, minutes=30)


@app.template_filter('panel_display')
def panel_display_filter(panel, unit):
    """Render one panel dict for the in-vcore view (always inches).
    `unit` is the dimensionUnit on the parent quote's configData, or None
    for legacy quotes (renders as feet with no conversion)."""
    from utils.bathqube_dimensions import format_panel_display
    if unit is None:
        # Legacy: customer typed in feet, keep the historical "x ft" form
        # so old quotes look identical to how they did before this feature.
        return f"{panel.get('width')}x{panel.get('height')}ft"
    return format_panel_display(panel.get('width'), panel.get('height'), unit)


# Expose the lead-facing WhatsApp template catalogue to every template so
# both /leads (bulk send) and /leads/<id> (single send) can render the
# picker dropdown without each route having to pass it in explicitly.
@app.context_processor
def _inject_lead_templates():
    from utils.lead_templates import LEAD_TEMPLATES
    return {
        'lead_templates': [t.to_dropdown_dict() for t in LEAD_TEMPLATES],
    }


# Lambda-specific: Dispose connections before each request
@app.before_request
def before_request():
    """Ensure fresh database connections for each Lambda invocation"""
    try:
        db.session.execute(text('SELECT 1'))
    except Exception:
        db.session.remove()
        db.engine.dispose()


# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute; 30 per hour")
def login():
    """Login page"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))

    form = LoginForm()
    if form.validate_on_submit():
        # Try to find user by username or email
        user = User.query.filter(
            (User.username == form.username.data) | (User.email == form.username.data)
        ).first()

        if user and user.check_password(form.password.data):
            if not user.is_active:
                flash('Your account has been deactivated. Please contact an administrator.', 'danger')
                return redirect(url_for('login'))

            login_user(user, remember=form.remember_me.data)
            flash(f'Welcome back, {user.username}!', 'success')

            # Safe redirect — only allow relative paths, never external URLs
            next_page = request.args.get('next')
            if next_page:
                parsed = urlparse(next_page)
                if parsed.scheme or parsed.netloc:
                    next_page = None  # Reject absolute/external URLs
            return redirect(next_page or url_for('dashboard'))
        else:
            flash('Invalid username/email or password.', 'danger')

    return render_template('login.html', form=form)


@app.route('/logout')
@login_required
def logout():
    """Logout user"""
    logout_user()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))


# ============================================================================
# MAIN DASHBOARD
# ============================================================================

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard"""
    # Get statistics
    total_projects = Project.query.count()
    active_projects = Project.query.filter_by(status='In Progress').count()
    completed_projects = Project.query.filter_by(status='Completed').count()
    
    # Get current week info
    week_info = get_current_week_info()
    
    # Get tasks for current user
    if current_user.is_admin() or current_user.is_manager_or_admin():
        # Admins and managers see all tasks
        total_tasks = PromotorTask.query.filter_by(
            assigned_week=week_info['week'],
            assigned_year=week_info['year']
        ).count()
        pending_tasks = PromotorTask.query.filter_by(
            assigned_week=week_info['week'],
            assigned_year=week_info['year'],
            status='Pending'
        ).count()
        overdue_tasks = PromotorTask.query.filter_by(
            assigned_week=week_info['week'],
            assigned_year=week_info['year'],
            status='Overdue'
        ).count()
    else:
        # Users see only their tasks
        total_tasks = PromotorTask.query.filter_by(
            promotor_id=current_user.id,
            assigned_week=week_info['week'],
            assigned_year=week_info['year']
        ).count()
        pending_tasks = PromotorTask.query.filter_by(
            promotor_id=current_user.id,
            assigned_week=week_info['week'],
            assigned_year=week_info['year'],
            status='Pending'
        ).count()
        overdue_tasks = PromotorTask.query.filter_by(
            promotor_id=current_user.id,
            assigned_week=week_info['week'],
            assigned_year=week_info['year'],
            status='Overdue'
        ).count()
    
    # Get recent projects
    recent_projects = Project.query.order_by(Project.updated_at.desc()).limit(5).all()
    
    return render_template('dashboard.html',
                         total_projects=total_projects,
                         active_projects=active_projects,
                         completed_projects=completed_projects,
                         total_tasks=total_tasks,
                         pending_tasks=pending_tasks,
                         overdue_tasks=overdue_tasks,
                         recent_projects=recent_projects,
                         week_info=week_info)


# ============================================================================
# PROJECT ROUTES
# ============================================================================

@app.route('/projects')
@login_required
def projects_list():
    """List all projects"""
    status_filter = request.args.get('status', '')
    owner_filter = request.args.get('owner', '')

    query = Project.query

    if status_filter:
        query = query.filter_by(status=status_filter)
    if owner_filter:
        query = query.filter_by(owner_id=int(owner_filter))

    projects = query.order_by(Project.created_at.desc()).all()
    users = User.query.filter_by(is_active=True).all()

    return render_template('projects/list.html', projects=projects, users=users)


@app.route('/projects/<int:id>')
@login_required
def project_view(id):
    """View project details and history"""
    project = Project.query.get_or_404(id)
    return render_template('projects/view.html', project=project)


@app.route('/projects/new', methods=['GET', 'POST'])
@manager_or_admin_required
def project_new():
    """Create new project"""
    form = ProjectForm()

    all_users = User.query.filter_by(is_active=True).all()
    managers_admins = [(u.id, u.username) for u in all_users if u.role in ('Admin', 'Manager')]
    form.owner_id.choices = managers_admins
    form.assigned_to_id.choices = [(0, '— Unassigned —')] + [(u.id, u.username) for u in all_users]

    if form.validate_on_submit():
        assigned = form.assigned_to_id.data if form.assigned_to_id.data else None
        project = Project(
            name=form.name.data,
            owner_id=form.owner_id.data,
            assigned_to_id=assigned,
            start_date=form.start_date.data,
            expected_end_date=form.expected_end_date.data,
            actual_end_date=form.actual_end_date.data,
            status=form.status.data,
            comments=form.comments.data
        )
        db.session.add(project)
        db.session.flush()  # get project.id before commit

        history = ProjectHistory(
            project_id=project.id,
            changed_by_id=current_user.id,
            action='Created',
            changes=None
        )
        db.session.add(history)
        db.session.commit()
        flash(f'Project "{project.name}" created successfully!', 'success')
        return redirect(url_for('projects_list'))

    return render_template('projects/form.html', form=form, title='New Project')


@app.route('/projects/<int:id>/edit', methods=['GET', 'POST'])
@manager_or_admin_required
def project_edit(id):
    """Edit project"""
    project = Project.query.get_or_404(id)
    form = ProjectForm(obj=project)

    all_users = User.query.filter_by(is_active=True).all()
    managers_admins = [(u.id, u.username) for u in all_users if u.role in ('Admin', 'Manager')]
    form.owner_id.choices = managers_admins
    form.assigned_to_id.choices = [(0, '— Unassigned —')] + [(u.id, u.username) for u in all_users]

    if form.validate_on_submit():
        # Build a human-readable diff of what changed
        user_map = {u.id: u.username for u in all_users}
        change_list = []

        def _norm(val):
            """Normalise a value for comparison — treat None and '' as equal."""
            if val is None:
                return ''
            return str(val).strip()

        old_assigned = user_map.get(project.assigned_to_id, '') if project.assigned_to_id else ''
        new_assigned_id = form.assigned_to_id.data if form.assigned_to_id.data else None
        new_assigned = user_map.get(new_assigned_id, '') if new_assigned_id else ''

        checks = [
            ('Name', _norm(project.name), _norm(form.name.data)),
            ('Owner', user_map.get(project.owner_id, ''), user_map.get(form.owner_id.data, '')),
            ('Assigned To', old_assigned, new_assigned),
            ('Status', _norm(project.status), _norm(form.status.data)),
            ('Start Date', _norm(project.start_date), _norm(form.start_date.data)),
            ('Expected End', _norm(project.expected_end_date), _norm(form.expected_end_date.data)),
            ('Actual End', _norm(project.actual_end_date), _norm(form.actual_end_date.data)),
            ('Comments', _norm(project.comments), _norm(form.comments.data)),
        ]
        for field_label, old_val, new_val in checks:
            if old_val != new_val:
                change_list.append({'field': field_label, 'old': old_val or '—', 'new': new_val or '—'})

        project.name = form.name.data
        project.owner_id = form.owner_id.data
        project.assigned_to_id = form.assigned_to_id.data if form.assigned_to_id.data else None
        project.start_date = form.start_date.data
        project.expected_end_date = form.expected_end_date.data
        project.actual_end_date = form.actual_end_date.data
        project.status = form.status.data
        project.comments = form.comments.data
        project.updated_at = datetime.utcnow()

        if change_list:
            history = ProjectHistory(
                project_id=project.id,
                changed_by_id=current_user.id,
                action='Updated',
                changes=json.dumps(change_list)
            )
            db.session.add(history)

        db.session.commit()
        flash(f'Project "{project.name}" updated successfully!', 'success')
        return redirect(url_for('project_view', id=project.id))

    # Pre-select current assigned_to in form
    if project.assigned_to_id:
        form.assigned_to_id.data = project.assigned_to_id
    else:
        form.assigned_to_id.data = 0

    return render_template('projects/form.html', form=form, title='Edit Project', project=project)


@app.route('/projects/<int:id>/comment', methods=['POST'])
@login_required
def project_add_comment(id):
    """Add a comment to a project (saved as a history entry)"""
    project = Project.query.get_or_404(id)
    text = request.form.get('comment', '').strip()
    if text:
        history = ProjectHistory(
            project_id=project.id,
            changed_by_id=current_user.id,
            action='Comment',
            changes=json.dumps([{'field': 'Comments', 'new': text}])
        )
        db.session.add(history)
        db.session.commit()
        flash('Comment added.', 'success')
    return redirect(url_for('project_view', id=id))


@app.route('/projects/<int:id>/delete', methods=['POST'])
@admin_required
def project_delete(id):
    """Delete project"""
    project = Project.query.get_or_404(id)
    project_name = project.name
    db.session.delete(project)
    db.session.commit()
    flash(f'Project "{project_name}" deleted successfully!', 'success')
    return redirect(url_for('projects_list'))


# ============================================================================
# TASK TEMPLATE ROUTES (Admin Only)
# ============================================================================

@app.route('/task-templates')
@admin_required
def task_templates_list():
    """List all task templates"""
    templates = TaskTemplate.query.order_by(TaskTemplate.created_at.desc()).all()
    return render_template('tasks/templates.html', templates=templates)


@app.route('/task-templates/new', methods=['GET', 'POST'])
@admin_required
def task_template_new():
    """Create new task template"""
    form = TaskTemplateForm()
    
    if form.validate_on_submit():
        template = TaskTemplate(
            name=form.name.data,
            description=form.description.data,
            category=form.category.data,
            priority=form.priority.data,
            is_active=form.is_active.data,
            created_by=current_user.id
        )
        db.session.add(template)
        db.session.commit()
        flash(f'Task template "{template.name}" created successfully!', 'success')
        return redirect(url_for('task_templates_list'))
    
    return render_template('tasks/template_form.html', form=form, title='New Task Template')


@app.route('/task-templates/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def task_template_edit(id):
    """Edit task template"""
    template = TaskTemplate.query.get_or_404(id)
    form = TaskTemplateForm(obj=template)
    
    if form.validate_on_submit():
        template.name = form.name.data
        template.description = form.description.data
        template.category = form.category.data
        template.priority = form.priority.data
        template.is_active = form.is_active.data
        template.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash(f'Task template "{template.name}" updated successfully!', 'success')
        return redirect(url_for('task_templates_list'))
    
    return render_template('tasks/template_form.html', form=form, title='Edit Task Template', template=template)


@app.route('/task-templates/<int:id>/delete', methods=['POST'])
@admin_required
def task_template_delete(id):
    """Delete task template"""
    template = TaskTemplate.query.get_or_404(id)
    template_name = template.name
    db.session.delete(template)
    db.session.commit()
    flash(f'Task template "{template_name}" deleted successfully!', 'success')
    return redirect(url_for('task_templates_list'))


# ============================================================================
# PROMOTOR TASK ROUTES
# ============================================================================

@app.route('/tasks/weekly')
@login_required
def tasks_weekly():
    """Weekly task board"""
    # Get week parameter or use current week
    week = request.args.get('week', type=int)
    year = request.args.get('year', type=int)
    
    if not week or not year:
        week_info = get_current_week_info()
        week = week_info['week']
        year = week_info['year']
    
    # Get date range for the week
    date_range = get_week_date_range(week, year)
    
    # Get tasks for the week
    if current_user.is_admin() or current_user.is_manager_or_admin():
        # Admins and managers see all tasks
        tasks = PromotorTask.query.filter_by(
            assigned_week=week,
            assigned_year=year
        ).order_by(PromotorTask.status, PromotorTask.priority).all()
    else:
        # Users see only their tasks
        tasks = PromotorTask.query.filter_by(
            promotor_id=current_user.id,
            assigned_week=week,
            assigned_year=year
        ).order_by(PromotorTask.status, PromotorTask.priority).all()
    
    # Group tasks by status
    tasks_by_status = {
        'Pending': [],
        'In Progress': [],
        'Completed': [],
        'Overdue': []
    }
    
    for task in tasks:
        tasks_by_status[task.status].append(task)
    
    return render_template('tasks/weekly_board.html',
                         tasks_by_status=tasks_by_status,
                         week=week,
                         year=year,
                         date_range=date_range)


@app.route('/tasks/assign', methods=['GET', 'POST'])
@login_required
def task_assign():
    """Assign task to promotor"""
    form = TaskAssignmentForm()
    
    # Populate dropdowns
    form.template_id.choices = [(t.id, t.name) for t in TaskTemplate.query.filter_by(is_active=True).all()]
    
    # Restrict user dropdown based on role
    if current_user.is_manager_or_admin():
        # Managers and admins can assign to anyone
        form.promotor_id.choices = [(u.id, u.username) for u in User.query.filter_by(is_active=True).all()]
    else:
        # Regular users can only assign to themselves
        form.promotor_id.choices = [(current_user.id, current_user.username)]
    
    form.project_id.choices = [(0, '-- None --')] + [(p.id, p.name) for p in Project.query.filter(Project.status.in_(['Not Started', 'In Progress'])).all()]
    
    if form.validate_on_submit():
        # Get current week info
        week_info = get_current_week_info()
        
        # Create task
        task = PromotorTask(
            template_id=form.template_id.data,
            task_name=form.task_name.data if form.task_name.data else None,
            promotor_id=form.promotor_id.data,
            project_id=form.project_id.data if form.project_id.data != 0 else None,
            assigned_week=week_info['week'],
            assigned_year=week_info['year'],
            original_week=week_info['week'],
            original_year=week_info['year'],
            due_date=form.due_date.data,
            priority=form.priority.data,
            comments=form.comments.data,
            created_by=current_user.id
        )
        db.session.add(task)
        db.session.commit()
        flash('Task assigned successfully!', 'success')
        return redirect(url_for('tasks_weekly'))
    
    return render_template('tasks/assign.html', form=form)


@app.route('/tasks/<int:id>/update', methods=['GET', 'POST'])
@login_required
def task_update(id):
    """Update task status"""
    task = PromotorTask.query.get_or_404(id)
    
    # Check permissions
    if not (current_user.is_admin() or current_user.is_manager_or_admin() or task.promotor_id == current_user.id):
        flash('You do not have permission to update this task.', 'danger')
        return redirect(url_for('tasks_weekly'))
    
    form = TaskUpdateForm(obj=task)
    
    if form.validate_on_submit():
        task.status = form.status.data
        task.comments = form.comments.data
        
        # Set completed date if status is Completed
        if form.status.data == 'Completed' and not task.completed_date:
            task.completed_date = datetime.utcnow()
        elif form.status.data != 'Completed':
            task.completed_date = None
        
        task.updated_at = datetime.utcnow()
        db.session.commit()
        flash('Task updated successfully!', 'success')
        return redirect(url_for('tasks_weekly'))
    
    return render_template('tasks/update.html', form=form, task=task)


# ============================================================================
# USER MANAGEMENT ROUTES (Admin Only)
# ============================================================================

@app.route('/users')
@admin_required
def users_list():
    """List all users"""
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('users/list.html', users=users)


@app.route('/users/new', methods=['GET', 'POST'])
@admin_required
def user_new():
    """Create new user"""
    form = UserForm()
    
    if form.validate_on_submit():
        user = User(
            username=form.username.data,
            email=form.email.data,
            role=form.role.data,
            is_active=form.is_active.data,
            show_in_agent_log=form.show_in_agent_log.data,
        )
        user.set_password(form.password.data)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{user.username}" created successfully!', 'success')
        return redirect(url_for('users_list'))
    
    return render_template('users/form.html', form=form, title='New User')


@app.route('/users/<int:id>/edit', methods=['GET', 'POST'])
@admin_required
def user_edit(id):
    """Edit user"""
    user = User.query.get_or_404(id)
    form = UserForm(user_id=user.id, obj=user)
    
    if form.validate_on_submit():
        user.username = form.username.data
        user.email = form.email.data
        user.role = form.role.data
        user.is_active = form.is_active.data
        user.show_in_agent_log = form.show_in_agent_log.data

        # Only update password if provided
        if form.password.data:
            user.set_password(form.password.data)
        
        user.updated_at = datetime.utcnow()
        db.session.commit()
        flash(f'User "{user.username}" updated successfully!', 'success')
        return redirect(url_for('users_list'))
    
    return render_template('users/form.html', form=form, title='Edit User', user=user)


@app.route('/users/<int:id>/delete', methods=['POST'])
@admin_required
def user_delete(id):
    """Delete user"""
    user = User.query.get_or_404(id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        flash('You cannot delete your own account!', 'danger')
        return redirect(url_for('users_list'))
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    flash(f'User "{username}" deleted successfully!', 'success')
    return redirect(url_for('users_list'))


# ============================================================================
# DAILY UPDATE ROUTES
# ============================================================================

@app.route('/daily-updates')
@login_required
def daily_updates_list():
    """List all daily updates with filtering"""
    from datetime import date, timedelta
    from sqlalchemy import and_, or_
    
    # Get filter parameters
    date_filter = request.args.get('date', '')
    user_filter = request.args.get('user', '')
    project_filter = request.args.get('project', '')
    
    # Default to today if no date specified
    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
        except ValueError:
            filter_date = date.today()
    else:
        filter_date = date.today()
    
    # Build query
    query = DailyUpdate.query.filter_by(update_date=filter_date)
    
    if user_filter:
        query = query.filter_by(user_id=int(user_filter))
    
    if project_filter:
        if project_filter == 'general':
            query = query.filter_by(is_general=True)
        else:
            query = query.filter_by(project_id=int(project_filter))
    
    # Order by created time (most recent first)
    updates = query.order_by(DailyUpdate.created_at.desc()).all()
    
    # Get all users and projects for filter dropdowns
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    projects = Project.query.filter(Project.status.in_(['Not Started', 'In Progress'])).order_by(Project.name).all()
    
    return render_template('daily_updates/list.html',
                         updates=updates,
                         users=users,
                         projects=projects,
                         filter_date=filter_date,
                         selected_user=user_filter,
                         selected_project=project_filter)


@app.route('/daily-updates/new', methods=['GET', 'POST'])
@login_required
def daily_update_new():
    """Create new daily update"""
    from datetime import date
    
    form = DailyUpdateForm()
    
    # Populate project choices - only projects user is involved with or all for managers/admins
    if current_user.is_manager_or_admin():
        # Managers and admins can update any project
        projects = Project.query.filter(Project.status.in_(['Not Started', 'In Progress'])).order_by(Project.name).all()
    else:
        # Regular users can update projects they're assigned tasks on
        user_project_ids = db.session.query(PromotorTask.project_id).filter(
            PromotorTask.promotor_id == current_user.id,
            PromotorTask.project_id.isnot(None)
        ).distinct().all()
        user_project_ids = [pid[0] for pid in user_project_ids]
        projects = Project.query.filter(Project.id.in_(user_project_ids)).order_by(Project.name).all()
    
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    
    if form.validate_on_submit():
        # Check if update already exists for this user/project/date
        today = date.today()
        existing_update = DailyUpdate.query.filter_by(
            user_id=current_user.id,
            project_id=form.project_id.data if not form.is_general.data else None,
            update_date=today,
            is_general=form.is_general.data
        ).first()
        
        if existing_update:
            flash('You have already submitted an update for this project today. Please edit the existing update instead.', 'warning')
            return redirect(url_for('daily_update_edit', id=existing_update.id))
        
        # Create new update
        update = DailyUpdate(
            user_id=current_user.id,
            project_id=form.project_id.data if not form.is_general.data else None,
            update_date=today,
            update_text=form.update_text.data,
            is_general=form.is_general.data
        )
        db.session.add(update)
        db.session.commit()
        flash('Daily update submitted successfully!', 'success')
        return redirect(url_for('daily_updates_list'))
    
    return render_template('daily_updates/form.html', form=form, title='New Daily Update')


@app.route('/daily-updates/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def daily_update_edit(id):
    """Edit existing daily update"""
    update = DailyUpdate.query.get_or_404(id)
    
    # Check permissions
    if not update.can_edit(current_user):
        flash('You can only edit your own updates from today.', 'danger')
        return redirect(url_for('daily_updates_list'))
    
    form = DailyUpdateForm(obj=update)
    
    # Populate project choices
    if current_user.is_manager_or_admin():
        projects = Project.query.filter(Project.status.in_(['Not Started', 'In Progress'])).order_by(Project.name).all()
    else:
        user_project_ids = db.session.query(PromotorTask.project_id).filter(
            PromotorTask.promotor_id == current_user.id,
            PromotorTask.project_id.isnot(None)
        ).distinct().all()
        user_project_ids = [pid[0] for pid in user_project_ids]
        projects = Project.query.filter(Project.id.in_(user_project_ids)).order_by(Project.name).all()
    
    form.project_id.choices = [(0, '-- Select Project --')] + [(p.id, p.name) for p in projects]
    
    if form.validate_on_submit():
        update.project_id = form.project_id.data if not form.is_general.data else None
        update.update_text = form.update_text.data
        update.is_general = form.is_general.data
        update.updated_at = datetime.utcnow()
        
        db.session.commit()
        flash('Daily update updated successfully!', 'success')
        return redirect(url_for('daily_updates_list'))
    
    return render_template('daily_updates/form.html', form=form, title='Edit Daily Update', update=update)


@app.route('/daily-updates/<int:id>/delete', methods=['POST'])
@login_required
def daily_update_delete(id):
    """Delete daily update"""
    update = DailyUpdate.query.get_or_404(id)
    
    # Check permissions
    if not update.can_delete(current_user):
        flash('You can only delete your own updates.', 'danger')
        return redirect(url_for('daily_updates_list'))
    
    db.session.delete(update)
    db.session.commit()
    flash('Daily update deleted successfully!', 'success')
    return redirect(url_for('daily_updates_list'))


# ============================================================================
# PRODUCT CATALOG ROUTES
# ============================================================================

@app.route('/catalog')
def catalog_list():
    """Product catalog list with filtering"""
    # Get filter parameters
    category_filter = request.args.get('category', '')
    search_query = request.args.get('search', '')
    
    # Build query
    query = Product.query.filter_by(is_active=True)
    
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    if search_query:
        query = query.filter(Product.product_name.ilike(f'%{search_query}%'))
    
    # Get products
    products = query.order_by(Product.category, Product.product_name).all()
    
    # Get filter options
    categories = [cat[0] for cat in Product.get_categories()]
    
    return render_template('catalog/list.html',
                         products=products,
                         categories=categories,
                         selected_category=category_filter,
                         search_query=search_query)


@app.route('/catalog/<int:id>')
def catalog_detail(id):
    """Product detail view with previous/next navigation"""
    product = Product.query.get_or_404(id)
    
    # Get all active products ordered by category and name
    all_products = Product.query.filter_by(is_active=True).order_by(
        Product.category, Product.product_name
    ).all()
    
    # Find current product index
    current_index = next((i for i, p in enumerate(all_products) if p.id == id), None)
    
    # Get previous and next products
    prev_product = all_products[current_index - 1] if current_index and current_index > 0 else None
    next_product = all_products[current_index + 1] if current_index is not None and current_index < len(all_products) - 1 else None
    
    return render_template('catalog/detail.html', 
                         product=product,
                         prev_product=prev_product,
                         next_product=next_product)


@app.route('/catalog/new', methods=['GET', 'POST'])
@manager_or_admin_required
def catalog_new():
    """Create new product"""
    form = ProductForm()
    
    if form.validate_on_submit():
        # Initialize S3 uploader
        s3_uploader = S3Uploader()
        
        product = Product(
            category=form.category.data,
            product_name=form.product_name.data,
            product_url=form.product_url.data,
            price=form.price.data,
            image_1_url=form.image_1_url.data,
            image_2_url=form.image_2_url.data,
            image_3_url=form.image_3_url.data,
            image_4_url=form.image_4_url.data,
            availability=form.availability.data,
            description=form.description.data,
            material=form.material.data,
            brand=form.brand.data,
            usage_application=form.usage_application.data,
            thickness=form.thickness.data,
            shape=form.shape.data,
            pattern=form.pattern.data,
            is_active=form.is_active.data
        )
        
        # Handle file uploads to S3
        for i in range(1, 5):
            file_field = getattr(form, f'image_{i}_file')
            if file_field.data:
                url = s3_uploader.upload_product_image(
                    file_field.data,
                    form.category.data,
                    form.product_name.data,
                    i
                )
                if url:
                    setattr(product, f'image_{i}_url', url)
                    flash(f'Image {i} uploaded successfully!', 'success')
                else:
                    flash(f'Failed to upload image {i}', 'warning')
        
        db.session.add(product)
        db.session.commit()
        flash(f'Product "{product.product_name}" created successfully!', 'success')
        return redirect(url_for('catalog_list'))
    
    return render_template('catalog/form.html', form=form, title='New Product')


@app.route('/catalog/<int:id>/edit', methods=['GET', 'POST'])
@manager_or_admin_required
def catalog_edit(id):
    """Edit product"""
    product = Product.query.get_or_404(id)
    form = ProductForm(obj=product)
    
    if form.validate_on_submit():
        # Initialize S3 uploader
        s3_uploader = S3Uploader()
        
        product.category = form.category.data
        product.product_name = form.product_name.data
        product.product_url = form.product_url.data
        product.price = form.price.data
        product.image_1_url = form.image_1_url.data
        product.image_2_url = form.image_2_url.data
        product.image_3_url = form.image_3_url.data
        product.image_4_url = form.image_4_url.data
        product.availability = form.availability.data
        product.description = form.description.data
        product.material = form.material.data
        product.brand = form.brand.data
        product.usage_application = form.usage_application.data
        product.thickness = form.thickness.data
        product.shape = form.shape.data
        product.pattern = form.pattern.data
        product.is_active = form.is_active.data
        product.updated_at = datetime.utcnow()
        
        # Handle file uploads to S3 (replace existing images)
        for i in range(1, 5):
            file_field = getattr(form, f'image_{i}_file')
            if file_field.data:
                url = s3_uploader.upload_product_image(
                    file_field.data,
                    form.category.data,
                    form.product_name.data,
                    i
                )
                if url:
                    setattr(product, f'image_{i}_url', url)
                    flash(f'Image {i} uploaded and replaced!', 'success')
                else:
                    flash(f'Failed to upload image {i}', 'warning')
        
        db.session.commit()
        flash(f'Product "{product.product_name}" updated successfully!', 'success')
        return redirect(url_for('catalog_detail', id=product.id))
    
    return render_template('catalog/form.html', form=form, title='Edit Product', product=product)


@app.route('/catalog/<int:id>/delete', methods=['POST'])
@admin_required
def catalog_delete(id):
    """Delete product"""
    product = Product.query.get_or_404(id)
    product_name = product.product_name
    db.session.delete(product)
    db.session.commit()
    flash(f'Product "{product_name}" deleted successfully!', 'success')
    return redirect(url_for('catalog_list'))


@app.route('/catalog/export-csv')
@manager_or_admin_required
def catalog_export_csv():
    """Export all active products as a UTF-8 CSV download"""
    from datetime import date

    products = Product.query.filter_by(is_active=True).order_by(
        Product.category, Product.product_name
    ).all()

    headers = [
        'product_name', 'category', 'description', 'price', 'price_unit',
        'min_order_qty', 'unit', 'size', 'thickness', 'color', 'finish',
        'design', 'product_type',
        'photo_url_1', 'photo_url_2', 'photo_url_3', 'photo_url_4',
        'tags'
    ]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()

    for p in products:
        specs = p.get_specifications()  # dict from JSON specifications column

        # Split stored price string (e.g. "160/sqft" or "24,999/Unit") into
        # numeric price and price_unit.
        price_val = ''
        price_unit_val = ''
        if p.price:
            m = re.match(r'^([\d,\.]+)\s*[/]?\s*(.*)', p.price.strip())
            if m:
                price_val = m.group(1).replace(',', '')
                price_unit_val = m.group(2).strip()

        def spec(*keys):
            """Return first non-empty value from specs dict for given key variants."""
            for k in keys:
                v = specs.get(k)
                if v:
                    return str(v)
            return ''

        writer.writerow({
            'product_name':  p.product_name or '',
            'category':      p.category or '',
            'description':   p.description or '',
            'price':         price_val,
            'price_unit':    spec('price_unit', 'Price Unit', 'unit_label') or price_unit_val,
            'min_order_qty': spec('min_order_qty', 'Min Order Qty', 'minimum_order_quantity', 'MOQ'),
            'unit':          spec('unit', 'Unit', 'unit_of_measurement'),
            'size':          spec('size', 'Size', 'Dimensions', 'dimensions'),
            'thickness':     p.thickness or spec('thickness', 'Thickness'),
            'color':         spec('color', 'Color', 'Colour', 'colour', 'shade', 'Shade'),
            'finish':        spec('finish', 'Finish', 'surface_finish', 'Surface Finish'),
            'design':        p.pattern or spec('design', 'Design', 'pattern', 'Pattern'),
            'product_type':  spec('product_type', 'Product Type', 'type', 'Type', 'sub_type'),
            'photo_url_1':   p.image_1_url or '',
            'photo_url_2':   p.image_2_url or '',
            'photo_url_3':   p.image_3_url or '',
            'photo_url_4':   p.image_4_url or '',
            'tags':          spec('tags', 'Tags', 'keywords', 'Keywords'),
        })

    filename = f"catalog_export_{date.today().isoformat()}.csv"
    output.seek(0)
    # utf-8-sig adds the BOM so Excel opens the file without garbled characters
    return Response(
        output.getvalue().encode('utf-8-sig'),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'}
    )


@app.route('/catalog/<int:id>/upload-image/<int:image_num>', methods=['POST'])
@manager_or_admin_required
def catalog_upload_image(id, image_num):
    """Upload/replace a single product image via AJAX"""
    try:
        product = Product.query.get_or_404(id)
        
        # Validate image number
        if image_num < 1 or image_num > 4:
            return jsonify({'success': False, 'error': 'Invalid image number'}), 400
        
        # Check if file was uploaded
        if 'image' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['image']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        # Initialize S3 uploader
        s3_uploader = S3Uploader()
        
        # Upload to S3
        url = s3_uploader.upload_product_image(
            file,
            product.category,
            product.product_name,
            image_num
        )
        
        if url:
            # Update product image URL
            setattr(product, f'image_{image_num}_url', url)
            product.updated_at = datetime.utcnow()
            db.session.commit()
            
            return jsonify({
                'success': True,
                'url': url,
                'message': f'Image {image_num} uploaded successfully'
            })
        else:
            return jsonify({'success': False, 'error': 'Failed to upload to S3'}), 500
            
    except Exception as e:
        print(f"Upload error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500



# ============================================================================
# BOM CALCULATOR ROUTES
# ============================================================================

@app.route('/bom')
@login_required
def bom_list():
    """BOM calculator list with product search"""
    search_query = request.args.get('search', '')
    
    # For now, show DGU/Insulated glass products
    products = Product.query.filter(
        Product.is_active == True
    ).filter(
        (Product.category.ilike('%insulated%')) | 
        (Product.category.ilike('%dgu%')) |
        (Product.product_name.ilike('%double glazing%'))
    ).order_by(Product.product_name).all()
    
    if search_query:
        products = [p for p in products if search_query.lower() in p.product_name.lower()]
    
    return render_template('bom/list.html', products=products, search_query=search_query)


@app.route('/bom/dgu-calculator')
@login_required
def bom_dgu_calculator():
    """DGU Glass BOM Calculator"""
    return render_template('bom/dgu_calculator.html')


# ============================================================================
# ADMIN ROLLOVER ROUTE
# ============================================================================

@app.route('/admin/rollover-tasks', methods=['POST'])
@admin_required
def admin_rollover_tasks():
    """Manually trigger task rollover"""
    try:
        result = rollover_incomplete_tasks()
        flash(f'Successfully rolled over {result["count"]} tasks to week {result["current_week"]}/{result["current_year"]}.', 'success')
    except Exception as e:
        flash(f'Error during rollover: {str(e)}', 'danger')
    
    return redirect(url_for('dashboard'))


# ============================================================================
# HEALTH CHECK & INFO ROUTES
# ============================================================================

@app.route('/health')
def health_check():
    """Health check endpoint for AWS"""
    try:
        db.session.execute(text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'database': 'connected',
            'service': 'vcore-api'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'database': 'disconnected',
            'error': str(e)
        }), 500


# ============================================================================
# WORDPRESS SYNC ROUTES (Admin Only)
# ============================================================================

@app.route('/admin/wordpress-sync')
@admin_required
def wordpress_sync_page():
    """WordPress sync admin page"""
    from utils.wordpress_sync import WordPressSync
    
    wp_sync = WordPressSync()
    stats = wp_sync.get_sync_status(db.session)
    
    return render_template('admin/wordpress_sync.html', stats=stats)


@app.route('/api/wordpress/test-connection', methods=['POST'])
@admin_required
def wordpress_test_connection():
    """Test WordPress API connection"""
    from utils.wordpress_sync import WordPressSync
    
    wp_sync = WordPressSync()
    result = wp_sync.test_connection()
    
    return jsonify(result)


@app.route('/api/wordpress/create-categories', methods=['POST'])
@admin_required
def wordpress_create_categories():
    """Create category structure in WordPress"""
    from utils.wordpress_sync import WordPressSync
    
    wp_sync = WordPressSync()
    result = wp_sync.create_categories()
    
    return jsonify(result)


@app.route('/api/wordpress/sync-all', methods=['POST'])
@admin_required
def wordpress_sync_all():
    """Sync all products to WordPress"""
    from utils.wordpress_sync import WordPressSync
    
    # Check if incremental sync is requested (default: True)
    data = request.get_json() or {}
    incremental = data.get('incremental', True)
    
    wp_sync = WordPressSync()
    result = wp_sync.sync_all_products(db.session, incremental=incremental)
    
    return jsonify(result)


@app.route('/api/wordpress/changed-products', methods=['GET'])
@admin_required
def wordpress_changed_products():
    """Get list of products that have changed since last sync"""
    from models import Product
    
    # Get products that need syncing
    changed_products = Product.query.filter(
        Product.is_active == True,
        (Product.last_wordpress_sync == None) | 
        (Product.updated_at > Product.last_wordpress_sync)
    ).all()
    
    products_list = [{
        'id': p.id,
        'name': p.product_name,
        'category': p.category,
        'updated_at': p.updated_at.isoformat() if p.updated_at else None,
        'last_sync': p.last_wordpress_sync.isoformat() if p.last_wordpress_sync else 'Never synced'
    } for p in changed_products]
    
    return jsonify({
        'success': True,
        'count': len(products_list),
        'products': products_list
    })


@app.route('/api/wordpress/sync-product/<int:product_id>', methods=['POST'])
@admin_required
def wordpress_sync_product(product_id):
    """Sync single product to WordPress"""
    from utils.wordpress_sync import WordPressSync
    
    product = Product.query.get_or_404(product_id)
    wp_sync = WordPressSync()
    result = wp_sync.sync_single_product(product)
    
    if result.get('success'):
        db.session.commit()
    
    return jsonify(result)


@app.route('/api/wordpress/sync-batch', methods=['POST'])
@admin_required
def wordpress_sync_batch():
    """Sync a batch of products to WordPress (client-side batching)"""
    from utils.wordpress_sync import WordPressSync
    from models import Product
    
    # Get batch size from request (default: 10)
    data = request.get_json() or {}
    batch_size = data.get('batch_size', 10)
    
    # Get products that need syncing
    products = Product.query.filter(
        Product.is_active == True,
        (Product.last_wordpress_sync == None) | 
        (Product.updated_at > Product.last_wordpress_sync)
    ).limit(batch_size).all()
    
    if not products:
        return jsonify({
            'success': True,
            'message': 'All products are up to date!',
            'synced': 0,
            'remaining': 0,
            'total_pending': 0
        })
    
    # Get total pending count
    total_pending = Product.query.filter(
        Product.is_active == True,
        (Product.last_wordpress_sync == None) | 
        (Product.updated_at > Product.last_wordpress_sync)
    ).count()
    
    # Sync this batch
    wp_sync = WordPressSync()
    synced = 0
    failed = 0
    errors = []
    
    for product in products:
        result = wp_sync.sync_single_product(product)
        if result.get('success'):
            synced += 1
            db.session.commit()
        else:
            failed += 1
            errors.append({
                'product_name': product.product_name,
                'error': result.get('message', 'Unknown error')
            })
    
    remaining = total_pending - synced
    
    return jsonify({
        'success': True,
        'synced': synced,
        'failed': failed,
        'batch_size': len(products),
        'remaining': remaining,
        'total_pending': total_pending,
        'errors': errors,
        'message': f'Synced {synced} of {len(products)} products in this batch'
    })


# ============================================================================
# QUOTE MANAGEMENT ROUTES
# ============================================================================

@app.route('/quotes')
@login_required
def quotes_list():
    """List all quotes with search and filter"""
    from datetime import datetime
    
    # Get filter parameters
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    quote_type_filter = request.args.get('quote_type', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Import Quote model
    from models import Quote
    
    # Build query
    query = Quote.query
    
    if search_query:
        query = query.filter(
            (Quote.customer_name.ilike(f'%{search_query}%')) |
            (Quote.quote_number.ilike(f'%{search_query}%'))
        )
    
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    if quote_type_filter:
        query = query.filter_by(quote_type=quote_type_filter)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Quote.quote_date >= from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Quote.quote_date <= to_date)
        except ValueError:
            pass
    
    # Get quotes ordered by date then by ID (newest first)
    quotes = query.order_by(Quote.quote_date.desc(), Quote.id.desc()).all()
    
    return render_template('quotes/list.html',
                         quotes=quotes,
                         search_query=search_query,
                         status_filter=status_filter,
                         quote_type_filter=quote_type_filter,
                         date_from=date_from,
                         date_to=date_to)


@app.route('/quotes/new', methods=['GET', 'POST'])
@login_required
def quote_new():
    """Create new quote with hierarchical items"""
    from models import Quote, QuoteItem
    from datetime import date, timedelta
    
    if request.method == 'POST':
        try:
            # Get form data
            data = request.form
            
            # Generate quote number
            quote_number = Quote.generate_quote_number()
            
            # Parse dates
            quote_date = datetime.strptime(data.get('quote_date'), '%Y-%m-%d').date()
            expected_date = datetime.strptime(data.get('expected_date'), '%Y-%m-%d').date() if data.get('expected_date') else None
            
            # Create quote
            quote = Quote(
                quote_number=quote_number,
                quote_date=quote_date,
                expected_date=expected_date,
                customer_name=data.get('customer_name'),
                customer_address=data.get('customer_address'),
                customer_city=data.get('customer_city'),
                customer_state=data.get('customer_state'),
                customer_phone=data.get('customer_phone'),
                customer_email=data.get('customer_email'),
                customer_gst=data.get('customer_gst'),
                customer_pan=(data.get('customer_pan') or '').strip().upper() or None,
                invoice_to=data.get('invoice_to'),
                dispatch_to=data.get('dispatch_to'),
                self_pickup=bool(data.get('self_pickup')),
                delivery_charges=float(data.get('delivery_charges') or 0),
                installation_charges=float(data.get('installation_charges') or 0),
                freight_charges=float(data.get('freight_charges') or 0),
                transport_charges=float(data.get('transport_charges') or 0),
                cutout_charges=float(data.get('cutout_charges') or 0),
                holes_charges=float(data.get('holes_charges') or 0),
                shape_cutting_charges=float(data.get('shape_cutting_charges') or 0),
                jumbo_size_charges=float(data.get('jumbo_size_charges') or 0),
                template_charges=float(data.get('template_charges') or 0),
                handling_percentage=float(data.get('handling_percentage') or 0),
                handling_charges=float(data.get('handling_charges') or 0),
                polish_charges=float(data.get('polish_charges') or 0),
                document_charges=float(data.get('document_charges') or 0),
                frosted_charges=float(data.get('frosted_charges') or 0),
                insurance_percentage=float(data.get('insurance_percentage') or 0),
                insurance_charges=float(data.get('insurance_charges') or 0),
                gst_percentage=float(data.get('gst_percentage') or 18),
                jumbo_pct_tier1=float(data.get('jumbo_pct_tier1') or 10),
                jumbo_pct_tier2=float(data.get('jumbo_pct_tier2') or 15),
                jumbo_pct_tier3=float(data.get('jumbo_pct_tier3') or 20),
                payment_terms=data.get('payment_terms'),
                status=data.get('status', 'Draft'),
                quote_type=data.get('quote_type', 'B2B'),
                created_by=current_user.id
            )
            
            db.session.add(quote)
            db.session.flush()  # Get quote ID

            # Save/update client record for autocomplete
            _upsert_client(data)

            # Process items - they come as items[0][field], items[1][field], etc.
            items_data = {}
            for key in data.keys():
                if key.startswith('items['):
                    # Parse items[0][particular] -> index=0, field=particular
                    import re
                    match = re.match(r'items\[(\d+)\]\[(\w+)\]', key)
                    if match:
                        index = int(match.group(1))
                        field = match.group(2)
                        if index not in items_data:
                            items_data[index] = {}
                        items_data[index][field] = data.get(key)
            
            # Create items in order, tracking parent IDs
            parent_id_map = {}  # Maps form item index to database ID
            index_to_group_id = {}  # Maps form index to group identifier (e.g., "group-1")
            
            for index in sorted(items_data.keys()):
                item_data = items_data[index]
                particular = item_data.get('particular', '')
                is_group = item_data.get('is_group') == 'true'
                
                # Skip only if it's a group without a particular (groups must have a name)
                if is_group and not particular:
                    continue
                
                parent_id_str = item_data.get('parent_id')  # This will be like "group-1", "group-2"
                
                # Determine actual parent_id by looking up the group identifier
                actual_parent_id = None
                if parent_id_str:
                    # Find which index created this group
                    for idx, group_id in index_to_group_id.items():
                        if group_id == parent_id_str:
                            actual_parent_id = parent_id_map.get(idx)
                            break
                
                # Create item
                item = QuoteItem(
                    quote_id=quote.id,
                    parent_id=actual_parent_id,
                    is_group=is_group,
                    sort_order=index,
                    item_number=int(item_data.get('item_number') or index + 1),
                    particular=particular,
                    image_s3_key=item_data.get('image_s3_key') or None,
                    actual_width=float(item_data.get('actual_width')) if item_data.get('actual_width') else None,
                    actual_height=float(item_data.get('actual_height')) if item_data.get('actual_height') else None,
                    chargeable_width=float(item_data.get('chargeable_width')) if item_data.get('chargeable_width') else None,
                    chargeable_height=float(item_data.get('chargeable_height')) if item_data.get('chargeable_height') else None,
                    unit=item_data.get('unit', 'MM'),
                    chargeable_extra=int(item_data.get('chargeable_extra') or 30),
                    quantity=int(item_data.get('quantity') or 1) if not is_group else 1,
                    rate_sqper=float(item_data.get('rate_sqper') or 0),
                    total=float(item_data.get('total') or 0) if not is_group else 0,
                    hole=int(item_data.get('hole') or 0) if not is_group else 0,
                    cutout=int(item_data.get('cutout') or 0) if not is_group else 0,
                    hole_price=float(item_data.get('hole_price') or 0),
                    cutout_price=float(item_data.get('cutout_price') or 0),
                )

                # Calculate unit square if dimensions provided
                if item.chargeable_width and item.chargeable_height:
                    item.calculate_unit_square()

                # Do NOT call item.calculate_total() here — the JS frontend
                # already calculates the correct total (including hole/cutout
                # charges). Calling it server-side would overwrite the
                # submitted value with an incorrect one because self.parent
                # is not accessible before the item is added to the session.

                db.session.add(item)
                db.session.flush()  # Get item ID

                # Store mapping: form index -> database ID
                parent_id_map[index] = item.id

                # If this is a group, also store its group identifier
                if is_group:
                    index_to_group_id[index] = f"group-{item.item_number}"

            # Calculate quote totals
            quote.subtotal = float(data.get('subtotal') or 0)
            quote.gst_amount = float(data.get('gst_amount') or 0)
            quote.round_off = float(data.get('round_off') or 0)
            quote.total = float(data.get('total') or 0)
            
            db.session.commit()
            flash(f'Quote {quote_number} created successfully!', 'success')
            return redirect(url_for('quote_view', id=quote.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating quote: {str(e)}', 'danger')
            return redirect(url_for('quote_new'))
    
    # GET request - show form
    from datetime import date, timedelta
    default_payment_terms = "For Confirmation the order need to give 100% of Quotation Value"
    
    return render_template('quotes/form.html',
                         title='New Quote',
                         quote=None,
                         default_quote_date=date.today(),
                         default_expected_date=date.today() + timedelta(days=8),
                         default_payment_terms=default_payment_terms)


@app.route('/quotes/<int:id>')
@login_required
def quote_view(id):
    """View quote details"""
    from models import Quote, QuoteComment
    quote = Quote.query.get_or_404(id)
    comments = QuoteComment.query.filter_by(quote_id=id).order_by(QuoteComment.created_at.desc()).all()
    return render_template('quotes/view.html', quote=quote, comments=comments)


@app.route('/quotes/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def quote_edit(id):
    """Edit existing quote with hierarchical items"""
    from models import Quote, QuoteItem
    
    quote = Quote.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            data = request.form
            
            # Update quote fields
            quote.quote_date = datetime.strptime(data.get('quote_date'), '%Y-%m-%d').date()
            quote.expected_date = datetime.strptime(data.get('expected_date'), '%Y-%m-%d').date() if data.get('expected_date') else None
            quote.customer_name = data.get('customer_name')
            quote.customer_address = data.get('customer_address')
            quote.customer_city = data.get('customer_city')
            quote.customer_state = data.get('customer_state')
            quote.customer_phone = data.get('customer_phone')
            quote.customer_email = data.get('customer_email')
            quote.customer_gst = data.get('customer_gst')
            quote.customer_pan = (data.get('customer_pan') or '').strip().upper() or None
            quote.invoice_to = data.get('invoice_to')
            quote.dispatch_to = data.get('dispatch_to')
            quote.self_pickup = bool(data.get('self_pickup'))
            quote.delivery_charges = float(data.get('delivery_charges') or 0)
            quote.installation_charges = float(data.get('installation_charges') or 0)
            quote.freight_charges = float(data.get('freight_charges') or 0)
            quote.transport_charges = float(data.get('transport_charges') or 0)
            quote.cutout_charges = float(data.get('cutout_charges') or 0)
            quote.holes_charges = float(data.get('holes_charges') or 0)
            quote.shape_cutting_charges = float(data.get('shape_cutting_charges') or 0)
            quote.jumbo_size_charges = float(data.get('jumbo_size_charges') or 0)
            quote.template_charges = float(data.get('template_charges') or 0)
            quote.handling_percentage = float(data.get('handling_percentage') or 0)
            quote.handling_charges = float(data.get('handling_charges') or 0)
            quote.polish_charges = float(data.get('polish_charges') or 0)
            quote.document_charges = float(data.get('document_charges') or 0)
            quote.frosted_charges = float(data.get('frosted_charges') or 0)
            quote.insurance_percentage = float(data.get('insurance_percentage') or 0)
            quote.insurance_charges = float(data.get('insurance_charges') or 0)
            quote.gst_percentage = float(data.get('gst_percentage') or 18)
            quote.jumbo_pct_tier1 = float(data.get('jumbo_pct_tier1') or 10)
            quote.jumbo_pct_tier2 = float(data.get('jumbo_pct_tier2') or 15)
            quote.jumbo_pct_tier3 = float(data.get('jumbo_pct_tier3') or 20)
            quote.payment_terms = data.get('payment_terms')
            quote.status = data.get('status', 'Draft')
            quote.quote_type = data.get('quote_type', 'B2B')
            
            # Delete existing items (cascade will handle children)
            QuoteItem.query.filter_by(quote_id=quote.id).delete()
            
            # Process items - same logic as quote_new
            items_data = {}
            for key in data.keys():
                if key.startswith('items['):
                    import re
                    match = re.match(r'items\[(\d+)\]\[(\w+)\]', key)
                    if match:
                        index = int(match.group(1))
                        field = match.group(2)
                        if index not in items_data:
                            items_data[index] = {}
                        items_data[index][field] = data.get(key)
            
            # Create items in order, tracking parent IDs
            parent_id_map = {}  # Maps form item index to database ID
            index_to_group_id = {}  # Maps form index to group identifier
            
            for index in sorted(items_data.keys()):
                item_data = items_data[index]
                particular = item_data.get('particular', '')
                is_group = item_data.get('is_group') == 'true'
                
                # Skip only if it's a group without a particular (groups must have a name)
                if is_group and not particular:
                    continue
                parent_id_str = item_data.get('parent_id')
                
                # Determine actual parent_id by looking up the group identifier
                actual_parent_id = None
                if parent_id_str:
                    for idx, group_id in index_to_group_id.items():
                        if group_id == parent_id_str:
                            actual_parent_id = parent_id_map.get(idx)
                            break
                
                # Create item
                item = QuoteItem(
                    quote_id=quote.id,
                    parent_id=actual_parent_id,
                    is_group=is_group,
                    sort_order=index,
                    item_number=int(item_data.get('item_number') or index + 1),
                    particular=particular,
                    image_s3_key=item_data.get('image_s3_key') or None,
                    actual_width=float(item_data.get('actual_width')) if item_data.get('actual_width') else None,
                    actual_height=float(item_data.get('actual_height')) if item_data.get('actual_height') else None,
                    chargeable_width=float(item_data.get('chargeable_width')) if item_data.get('chargeable_width') else None,
                    chargeable_height=float(item_data.get('chargeable_height')) if item_data.get('chargeable_height') else None,
                    unit=item_data.get('unit', 'MM'),
                    chargeable_extra=int(item_data.get('chargeable_extra') or 30),
                    quantity=int(item_data.get('quantity') or 1) if not is_group else 1,
                    rate_sqper=float(item_data.get('rate_sqper') or 0),
                    total=float(item_data.get('total') or 0) if not is_group else 0,
                    hole=int(item_data.get('hole') or 0) if not is_group else 0,
                    cutout=int(item_data.get('cutout') or 0) if not is_group else 0,
                    hole_price=float(item_data.get('hole_price') or 0),
                    cutout_price=float(item_data.get('cutout_price') or 0),
                )

                # Calculate unit square if dimensions provided
                if item.chargeable_width and item.chargeable_height:
                    item.calculate_unit_square()

                # Do NOT call item.calculate_total() here — same reason as
                # quote_new: self.parent is not accessible before add/flush,
                # so hole/cutout charges would be lost.

                db.session.add(item)
                db.session.flush()

                # Store mapping: form index -> database ID
                parent_id_map[index] = item.id

                # If this is a group, also store its group identifier
                if is_group:
                    index_to_group_id[index] = f"group-{item.item_number}"

            # Update quote totals
            quote.subtotal = float(data.get('subtotal') or 0)
            quote.gst_amount = float(data.get('gst_amount') or 0)
            quote.round_off = float(data.get('round_off') or 0)
            quote.total = float(data.get('total') or 0)
            quote.updated_at = datetime.utcnow()

            # Save/update client record for autocomplete
            _upsert_client(data)

            db.session.commit()
            flash(f'Quote {quote.quote_number} updated successfully!', 'success')
            return redirect(url_for('quote_view', id=quote.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating quote: {str(e)}', 'danger')
            return redirect(url_for('quote_edit', id=id))
    
    # GET request - show form with existing data
    return render_template('quotes/form.html',
                         title='Edit Quote',
                         quote=quote)


@app.route('/quotes/<int:id>/delete', methods=['POST'])
@admin_required
def quote_delete(id):
    """Delete quote"""
    from models import Quote
    quote = Quote.query.get_or_404(id)
    quote_number = quote.quote_number
    db.session.delete(quote)
    db.session.commit()
    flash(f'Quote {quote_number} deleted successfully!', 'success')
    return redirect(url_for('quotes_list'))


# ─────────────────────────────────────────────────────────
# CLIENTS
# ─────────────────────────────────────────────────────────

def _upsert_client(data):
    """Create or update a client record from quote form data.
    Matches on name (case-insensitive). Only updates fields that are non-empty."""
    from models import Client
    name = (data.get('customer_name') or '').strip()
    if not name:
        return
    client = Client.query.filter(
        db.func.lower(Client.name) == name.lower()
    ).first()
    if not client:
        client = Client(name=name)
        db.session.add(client)
    # Update only non-empty fields so partial edits don't wipe saved data
    for src, dst in [
        ('customer_phone',   'phone'),
        ('customer_email',   'email'),
        ('customer_address', 'address'),
        ('customer_city',    'city'),
        ('customer_state',   'state'),
        ('customer_gst',     'gst_number'),
        ('dispatch_to',      'dispatch_to'),
    ]:
        val = (data.get(src) or '').strip()
        if val:
            setattr(client, dst, val)
    qt = data.get('quote_type')
    if qt in ('B2B', 'B2C'):
        client.quote_type = qt


@app.route('/api/clients/search')
@login_required
def clients_search():
    """Return up to 10 clients matching the query string (for autocomplete)."""
    from models import Client
    q = (request.args.get('q') or '').strip()
    if len(q) < 1:
        return jsonify([])
    results = Client.query.filter(
        Client.name.ilike(f'%{q}%')
    ).order_by(Client.name).limit(10).all()
    return jsonify([c.to_dict() for c in results])


@app.route('/clients')
@login_required
def clients_list():
    from models import Client
    q = request.args.get('q', '').strip()
    query = Client.query
    if q:
        query = query.filter(Client.name.ilike(f'%{q}%'))
    clients = query.order_by(Client.name).all()
    return render_template('clients/list.html', clients=clients, q=q)


@app.route('/clients/new', methods=['GET', 'POST'])
@login_required
def client_new():
    from models import Client
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash('Client name is required.', 'danger')
            return redirect(url_for('client_new'))
        client = Client(
            name        = name,
            phone       = request.form.get('phone',       '').strip() or None,
            email       = request.form.get('email',       '').strip() or None,
            address     = request.form.get('address',     '').strip() or None,
            city        = request.form.get('city',        '').strip() or None,
            state       = request.form.get('state',       '').strip() or None,
            gst_number  = request.form.get('gst_number',  '').strip() or None,
            dispatch_to = request.form.get('dispatch_to', '').strip() or None,
            quote_type  = request.form.get('quote_type')  or None,
            notes       = request.form.get('notes',       '').strip() or None,
        )
        db.session.add(client)
        db.session.commit()
        flash(f'Client "{client.name}" saved.', 'success')
        return redirect(url_for('clients_list'))
    return render_template('clients/form.html', client=None)


@app.route('/clients/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def client_edit(id):
    from models import Client
    client = Client.query.get_or_404(id)
    if request.method == 'POST':
        name = (request.form.get('name') or '').strip()
        if not name:
            flash('Client name is required.', 'danger')
            return redirect(url_for('client_edit', id=id))
        client.name        = name
        client.phone       = request.form.get('phone',       '').strip() or None
        client.email       = request.form.get('email',       '').strip() or None
        client.address     = request.form.get('address',     '').strip() or None
        client.city        = request.form.get('city',        '').strip() or None
        client.state       = request.form.get('state',       '').strip() or None
        client.gst_number  = request.form.get('gst_number',  '').strip() or None
        client.dispatch_to = request.form.get('dispatch_to', '').strip() or None
        client.quote_type  = request.form.get('quote_type')  or None
        client.notes       = request.form.get('notes',       '').strip() or None
        db.session.commit()
        flash(f'Client "{client.name}" updated.', 'success')
        return redirect(url_for('clients_list'))
    return render_template('clients/form.html', client=client)


@app.route('/clients/<int:id>/delete', methods=['POST'])
@login_required
def client_delete(id):
    from models import Client
    client = Client.query.get_or_404(id)
    name = client.name
    db.session.delete(client)
    db.session.commit()
    flash(f'Client "{name}" deleted.', 'success')
    return redirect(url_for('clients_list'))


# ─────────────────────────────────────────────────────────

@app.route('/quotes/<int:id>/duplicate', methods=['POST'])
@login_required
def quote_duplicate(id):
    """Duplicate an existing quote"""
    from models import Quote, QuoteItem
    from datetime import date
    
    original_quote = Quote.query.get_or_404(id)
    
    # Create new quote with same data
    new_quote = Quote(
        quote_number=Quote.generate_quote_number(),
        quote_date=date.today(),
        expected_date=original_quote.expected_date,
        customer_name=original_quote.customer_name,
        customer_address=original_quote.customer_address,
        customer_city=original_quote.customer_city,
        customer_state=original_quote.customer_state,
        customer_phone=original_quote.customer_phone,
        customer_email=original_quote.customer_email,
        invoice_to=original_quote.invoice_to,
        dispatch_to=original_quote.dispatch_to,
        self_pickup=original_quote.self_pickup,
        delivery_charges=original_quote.delivery_charges,
        installation_charges=original_quote.installation_charges,
        freight_charges=original_quote.freight_charges,
        transport_charges=original_quote.transport_charges,
        insurance_percentage=original_quote.insurance_percentage,
        insurance_charges=original_quote.insurance_charges,
        gst_percentage=original_quote.gst_percentage,
        payment_terms=original_quote.payment_terms,
        status='Draft',
        created_by=current_user.id
    )
    
    db.session.add(new_quote)
    db.session.flush()
    
    # Duplicate items
    for original_item in original_quote.items:
        new_item = QuoteItem(
            quote_id=new_quote.id,
            item_number=original_item.item_number,
            particular=original_item.particular,
            size_width=original_item.size_width,
            size_height=original_item.size_height,
            unit=original_item.unit,
            chargeable_size_width=original_item.chargeable_size_width,
            chargeable_size_height=original_item.chargeable_size_height,
            quantity=original_item.quantity,
            first_sqper=original_item.first_sqper,
            rate_sqper=original_item.rate_sqper,
            total=original_item.total
        )
        db.session.add(new_item)
    
    new_quote.calculate_totals()
    db.session.commit()
    
    flash(f'Quote duplicated as {new_quote.quote_number}!', 'success')
    return redirect(url_for('quote_edit', id=new_quote.id))


@app.route('/quotes/<int:id>/print')
@login_required
def quote_print(id):
    """Show print-friendly version of quote — auto-marks status as Sent"""
    from models import Quote
    quote = Quote.query.get_or_404(id)
    if quote.status == 'Draft':
        quote.status = 'Sent'
        db.session.commit()
    return render_template('quotes/print.html', quote=quote)


@app.route('/quotes/<int:id>/comment', methods=['POST'])
@login_required
def quote_add_comment(id):
    """Add a comment to a quote"""
    from models import Quote, QuoteComment
    quote = Quote.query.get_or_404(id)
    text = request.form.get('comment', '').strip()
    if text:
        comment = QuoteComment(quote_id=quote.id, user_id=current_user.id, comment=text)
        db.session.add(comment)
        db.session.commit()
        flash('Comment added.', 'success')
    return redirect(url_for('quote_view', id=id))


@app.route('/quotes/<int:id>/comment/<int:comment_id>/delete', methods=['POST'])
@login_required
def quote_delete_comment(id, comment_id):
    """Delete a comment (own comment or admin)"""
    from models import QuoteComment
    c = QuoteComment.query.get_or_404(comment_id)
    if c.user_id == current_user.id or current_user.is_admin():
        db.session.delete(c)
        db.session.commit()
    return redirect(url_for('quote_view', id=id))


# ============================================================================
# QUOTE API ROUTES
# ============================================================================

@app.route('/api/quotes/next-number')
@login_required
def api_quote_next_number():
    """Get next available quote number"""
    from models import Quote
    next_number = Quote.generate_quote_number()
    return jsonify({'quote_number': next_number})


@app.route('/api/quote-items/upload-image', methods=['POST'])
@login_required
def quote_item_upload_image():
    """AJAX: upload a quote item image to S3, return the s3_key and a short-lived presigned URL."""
    file = request.files.get('image')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    allowed = {'jpg', 'jpeg', 'png', 'webp', 'gif'}
    ext = file.filename.rsplit('.', 1)[-1].lower()
    if ext not in allowed:
        return jsonify({'success': False, 'error': 'Only image files are allowed'}), 400

    try:
        from utils.s3_upload import S3Uploader
        from werkzeug.utils import secure_filename
        from datetime import datetime

        uploader  = S3Uploader()
        timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S%f')
        filename  = f"{timestamp}_{secure_filename(file.filename)}"
        s3_key    = f"quote-items/{filename}"

        uploader.s3_client.upload_fileobj(
            file, uploader.bucket_name, s3_key,
            ExtraArgs={'ContentType': file.content_type or 'image/jpeg'}
        )

        presigned = uploader.s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': uploader.bucket_name, 'Key': s3_key},
            ExpiresIn=3600
        )
        return jsonify({'success': True, 's3_key': s3_key, 'presigned_url': presigned})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.template_filter('item_image_url')
def item_image_url_filter(s3_key):
    """Jinja2 filter: convert an S3 key to a presigned URL at render time."""
    if not s3_key:
        return ''
    try:
        from utils.s3_upload import S3Uploader
        uploader = S3Uploader()
        return uploader.s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': uploader.bucket_name, 'Key': s3_key},
            ExpiresIn=3600
        )
    except Exception:
        return ''


@app.route('/api/quote-items/image-url')
@login_required
def quote_item_image_url():
    """Return a fresh presigned URL for a quote item image (used by view page)."""
    s3_key = request.args.get('key', '').strip()
    if not s3_key:
        return jsonify({'error': 'No key'}), 400
    try:
        from utils.s3_upload import S3Uploader
        uploader = S3Uploader()
        url = uploader.s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': uploader.bucket_name, 'Key': s3_key},
            ExpiresIn=3600
        )
        return redirect(url)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/products/search')
@login_required
def api_products_search():
    """Search products for quote (autocomplete)"""
    query = request.args.get('q', '')
    
    if len(query) < 2:
        return jsonify([])
    
    products = Product.query.filter(
        Product.is_active == True,
        Product.product_name.ilike(f'%{query}%')
    ).limit(10).all()
    
    results = [{
        'id': p.id,
        'name': p.product_name,
        'category': p.category,
        'price': p.price
    } for p in products]
    
    return jsonify(results)


@app.route('/api/glass-types')
@login_required
def api_glass_types():
    """API endpoint to get glass types with pricing for quote form"""
    glass_types = GlassType.query.filter_by(is_active=True).all()
    
    result = []
    for gt in glass_types:
        suppliers_pricing = []
        for pricing in gt.pricing:
            if pricing.is_active and pricing.is_currently_valid():
                suppliers_pricing.append({
                    'supplier_id': pricing.supplier_id,
                    'supplier_name': pricing.supplier.name,
                    'rate_per_sqm': float(pricing.rate_per_sqm),
                    'hole_price': float(pricing.hole_price),
                    'cutout_price': float(pricing.cutout_price),
                    'big_hole_price': float(pricing.big_hole_price),
                    'big_cutout_price': float(pricing.big_cutout_price),
                    'frosting_charge_per_sqm': float(pricing.frosting_charge_per_sqm),
                    'lead_time_days': pricing.lead_time_days
                })
        
        result.append({
            'id': gt.id,
            'name': gt.name,
            'category': gt.category,
            'thickness_mm': float(gt.thickness_mm) if gt.thickness_mm else None,
            'is_frosted': gt.is_frosted,
            'is_tinted': gt.is_tinted,
            'suppliers': suppliers_pricing,
            'best_price': float(gt.get_best_price()) if gt.get_best_price() else None
        })
    
    return jsonify(result)


# ============================================================================
# SUPPLIER MANAGEMENT ROUTES
# ============================================================================

@app.route('/suppliers')
@login_required
def suppliers_list():
    """List all suppliers"""
    from models import Supplier
    
    search_query = request.args.get('search', '')
    
    query = Supplier.query
    
    if search_query:
        query = query.filter(
            (Supplier.name.ilike(f'%{search_query}%')) |
            (Supplier.contact_person.ilike(f'%{search_query}%')) |
            (Supplier.city.ilike(f'%{search_query}%'))
        )
    
    suppliers = query.filter_by(is_active=True).order_by(Supplier.name).all()
    
    return render_template('suppliers/list.html',
                         suppliers=suppliers,
                         search_query=search_query)


@app.route('/suppliers/new', methods=['GET', 'POST'])
@login_required
def supplier_new():
    """Create new supplier"""
    from models import Supplier
    
    if request.method == 'POST':
        try:
            data = request.form
            if not data.get('name', '').strip():
                flash('Supplier name is required.', 'danger')
                return redirect(url_for('supplier_new'))
            if not data.get('phone', '').strip():
                flash('Phone number is required.', 'danger')
                return redirect(url_for('supplier_new'))
            if not data.get('email', '').strip():
                flash('Email is required.', 'danger')
                return redirect(url_for('supplier_new'))
            if not data.get('gstin', '').strip():
                flash('GSTIN is required.', 'danger')
                return redirect(url_for('supplier_new'))

            supplier = Supplier(
                name=data.get('name'),
                contact_person=data.get('contact_person'),
                phone=data.get('phone'),
                email=data.get('email'),
                address=data.get('address'),
                city=data.get('city'),
                state=data.get('state'),
                pincode=data.get('pincode'),
                gstin=data.get('gstin'),
                pan=data.get('pan'),
                bank_name=data.get('bank_name'),
                account_number=data.get('account_number'),
                ifsc_code=data.get('ifsc_code'),
                branch=data.get('branch'),
                payment_terms=data.get('payment_terms'),
                lead_time_days=int(data.get('lead_time_days')) if data.get('lead_time_days') else None,
                min_order_value=float(data.get('min_order_value')) if data.get('min_order_value') else None,
                notes=data.get('notes'),
                is_active=True
            )
            
            db.session.add(supplier)
            db.session.commit()
            
            flash(f'Supplier "{supplier.name}" created successfully!', 'success')
            return redirect(url_for('suppliers_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating supplier: {str(e)}', 'danger')
            return redirect(url_for('supplier_new'))
    
    return render_template('suppliers/form.html', title='New Supplier', supplier=None)


@app.route('/suppliers/<int:id>')
@login_required
def supplier_view(id):
    """View supplier details"""
    from models import Supplier
    supplier = Supplier.query.get_or_404(id)
    return render_template('suppliers/view.html', supplier=supplier)


@app.route('/suppliers/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def supplier_edit(id):
    """Edit supplier"""
    from models import Supplier
    
    supplier = Supplier.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            data = request.form
            if not data.get('name', '').strip():
                flash('Supplier name is required.', 'danger')
                return redirect(url_for('supplier_edit', id=id))
            if not data.get('phone', '').strip():
                flash('Phone number is required.', 'danger')
                return redirect(url_for('supplier_edit', id=id))
            if not data.get('email', '').strip():
                flash('Email is required.', 'danger')
                return redirect(url_for('supplier_edit', id=id))
            if not data.get('gstin', '').strip():
                flash('GSTIN is required.', 'danger')
                return redirect(url_for('supplier_edit', id=id))

            supplier.name = data.get('name')
            supplier.contact_person = data.get('contact_person')
            supplier.phone = data.get('phone')
            supplier.email = data.get('email')
            supplier.address = data.get('address')
            supplier.city = data.get('city')
            supplier.state = data.get('state')
            supplier.pincode = data.get('pincode')
            supplier.gstin = data.get('gstin')
            supplier.pan = data.get('pan')
            supplier.bank_name = data.get('bank_name')
            supplier.account_number = data.get('account_number')
            supplier.ifsc_code = data.get('ifsc_code')
            supplier.branch = data.get('branch')
            supplier.payment_terms = data.get('payment_terms')
            supplier.lead_time_days = int(data.get('lead_time_days')) if data.get('lead_time_days') else None
            supplier.min_order_value = float(data.get('min_order_value')) if data.get('min_order_value') else None
            supplier.notes = data.get('notes')
            supplier.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Supplier "{supplier.name}" updated successfully!', 'success')
            return redirect(url_for('supplier_view', id=supplier.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating supplier: {str(e)}', 'danger')
            return redirect(url_for('supplier_edit', id=id))
    
    return render_template('suppliers/form.html', title='Edit Supplier', supplier=supplier)


@app.route('/suppliers/<int:id>/delete', methods=['POST'])
@admin_required
def supplier_delete(id):
    """Delete supplier (soft delete)"""
    from models import Supplier
    supplier = Supplier.query.get_or_404(id)
    supplier_name = supplier.name
    supplier.is_active = False
    db.session.commit()
    flash(f'Supplier "{supplier_name}" deleted successfully!', 'success')
    return redirect(url_for('suppliers_list'))


# ============================================================================
# GLASS CATALOG ROUTES
# ============================================================================

@app.route('/glass-catalog')
@login_required
def glass_catalog_list():
    """List all glass types with pricing from all suppliers"""
    from models import GlassType, Supplier, SupplierPricing
    
    search_query = request.args.get('search', '')
    category_filter = request.args.get('category', '')
    supplier_filter = request.args.get('supplier', '')
    
    # Get all active glass types
    query = GlassType.query.filter_by(is_active=True)
    
    if search_query:
        query = query.filter(GlassType.name.ilike(f'%{search_query}%'))
    
    if category_filter:
        query = query.filter_by(category=category_filter)
    
    glass_types = query.order_by(GlassType.category, GlassType.name).all()
    
    # Get all active suppliers for filter dropdown
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    
    # Get unique categories for filter dropdown
    categories = db.session.query(GlassType.category).filter(
        GlassType.is_active == True,
        GlassType.category.isnot(None)
    ).distinct().order_by(GlassType.category).all()
    categories = [c[0] for c in categories]
    
    # Filter pricing by supplier if selected
    if supplier_filter:
        supplier_filter = int(supplier_filter)
    
    return render_template('glass_catalog/list.html',
                         glass_types=glass_types,
                         suppliers=suppliers,
                         categories=categories,
                         search_query=search_query,
                         category_filter=category_filter,
                         supplier_filter=supplier_filter)


@app.route('/glass-catalog/new', methods=['GET', 'POST'])
@login_required
def glass_catalog_new():
    """Create new glass type"""
    from models import GlassType
    
    if request.method == 'POST':
        try:
            data = request.form
            
            glass_type = GlassType(
                name=data.get('name'),
                category=data.get('category'),
                thickness_mm=float(data.get('thickness_mm')) if data.get('thickness_mm') else None,
                description=data.get('description'),
                is_frosted=bool(data.get('is_frosted')),
                is_tinted=bool(data.get('is_tinted')),
                color=data.get('color'),
                is_active=True
            )
            
            db.session.add(glass_type)
            db.session.commit()
            
            flash(f'Glass type "{glass_type.name}" created successfully!', 'success')
            return redirect(url_for('glass_catalog_view', id=glass_type.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating glass type: {str(e)}', 'danger')
            return redirect(url_for('glass_catalog_new'))
    
    return render_template('glass_catalog/form.html', title='New Glass Type', glass_type=None)


@app.route('/glass-catalog/<int:id>')
@login_required
def glass_catalog_view(id):
    """View glass type details with all supplier pricing"""
    from models import GlassType, Supplier
    
    glass_type = GlassType.query.get_or_404(id)
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    
    return render_template('glass_catalog/view.html',
                         glass_type=glass_type,
                         suppliers=suppliers)


@app.route('/glass-catalog/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def glass_catalog_edit(id):
    """Edit glass type"""
    from models import GlassType
    
    glass_type = GlassType.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            data = request.form
            
            glass_type.name = data.get('name')
            glass_type.category = data.get('category')
            glass_type.thickness_mm = float(data.get('thickness_mm')) if data.get('thickness_mm') else None
            glass_type.description = data.get('description')
            glass_type.is_frosted = bool(data.get('is_frosted'))
            glass_type.is_tinted = bool(data.get('is_tinted'))
            glass_type.color = data.get('color')
            glass_type.updated_at = datetime.utcnow()
            
            db.session.commit()
            
            flash(f'Glass type "{glass_type.name}" updated successfully!', 'success')
            return redirect(url_for('glass_catalog_view', id=glass_type.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating glass type: {str(e)}', 'danger')
            return redirect(url_for('glass_catalog_edit', id=id))
    
    return render_template('glass_catalog/form.html', title='Edit Glass Type', glass_type=glass_type)


@app.route('/glass-catalog/<int:id>/pricing', methods=['GET', 'POST'])
@login_required
def glass_catalog_pricing(id):
    """Manage supplier pricing for a glass type"""
    from models import GlassType, Supplier, SupplierPricing
    from datetime import date
    
    glass_type = GlassType.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            data = request.form
            supplier_id = int(data.get('supplier_id'))
            
            # Check if pricing already exists
            pricing = SupplierPricing.query.filter_by(
                supplier_id=supplier_id,
                glass_type_id=glass_type.id,
                is_active=True
            ).first()
            
            if pricing:
                # Update existing pricing
                pricing.rate_per_sqm = float(data.get('rate_per_sqm'))
                pricing.hole_price = float(data.get('hole_price', 0))
                pricing.cutout_price = float(data.get('cutout_price', 0))
                pricing.big_hole_price = float(data.get('big_hole_price', 0))
                pricing.big_cutout_price = float(data.get('big_cutout_price', 0))
                pricing.frosting_charge_per_sqm = float(data.get('frosting_charge_per_sqm', 0))
                pricing.tinting_charge_per_sqm = float(data.get('tinting_charge_per_sqm', 0))
                pricing.min_order_sqm = float(data.get('min_order_sqm')) if data.get('min_order_sqm') else None
                pricing.lead_time_days = int(data.get('lead_time_days')) if data.get('lead_time_days') else None
                pricing.notes = data.get('notes')
                pricing.updated_at = datetime.utcnow()
                
                flash('Pricing updated successfully!', 'success')
            else:
                # Create new pricing
                pricing = SupplierPricing(
                    supplier_id=supplier_id,
                    glass_type_id=glass_type.id,
                    rate_per_sqm=float(data.get('rate_per_sqm')),
                    hole_price=float(data.get('hole_price', 0)),
                    cutout_price=float(data.get('cutout_price', 0)),
                    big_hole_price=float(data.get('big_hole_price', 0)),
                    big_cutout_price=float(data.get('big_cutout_price', 0)),
                    frosting_charge_per_sqm=float(data.get('frosting_charge_per_sqm', 0)),
                    tinting_charge_per_sqm=float(data.get('tinting_charge_per_sqm', 0)),
                    min_order_sqm=float(data.get('min_order_sqm')) if data.get('min_order_sqm') else None,
                    lead_time_days=int(data.get('lead_time_days')) if data.get('lead_time_days') else None,
                    effective_from=date.today(),
                    notes=data.get('notes'),
                    is_active=True
                )
                db.session.add(pricing)
                flash('Pricing added successfully!', 'success')
            
            db.session.commit()
            return redirect(url_for('glass_catalog_view', id=glass_type.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error saving pricing: {str(e)}', 'danger')
            return redirect(url_for('glass_catalog_pricing', id=id))
    
    # GET request
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    
    # Check if editing existing pricing (supplier_id in query params)
    supplier_id = request.args.get('supplier_id', type=int)
    existing_pricing = None
    
    if supplier_id:
        existing_pricing = SupplierPricing.query.filter_by(
            supplier_id=supplier_id,
            glass_type_id=glass_type.id,
            is_active=True
        ).first()
    
    return render_template('glass_catalog/pricing.html',
                         glass_type=glass_type,
                         suppliers=suppliers,
                         existing_pricing=existing_pricing,
                         selected_supplier_id=supplier_id)


# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(404)
def not_found_error(error):
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('errors/500.html'), 500


# ============================================================================
# LAMBDA HANDLER
# ============================================================================

def lambda_handler(event, context):
    """AWS Lambda handler"""
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
    
    # Use awsgi to handle Lambda event
    from awsgi import response
    return response(app, event, context)


# ============================================================================
# REMINDER ROUTES
# ============================================================================

@app.route('/reminders')
@login_required
def reminders_list():
    """List all reminders for current user"""
    if current_user.role == 'Admin':
        reminders = Reminder.query.order_by(Reminder.reminder_datetime.desc()).all()
    else:
        reminders = Reminder.query.filter_by(user_id=current_user.id).order_by(Reminder.reminder_datetime.desc()).all()
    
    return render_template('reminders/list.html', reminders=reminders)


@app.route('/reminders/new', methods=['GET', 'POST'])
@login_required
def reminder_new():
    """Create a new reminder"""
    if request.method == 'POST':
        try:
            reminder_type = request.form.get('reminder_type')
            project_id = request.form.get('project_id') if reminder_type == 'project' else None
            task_id = request.form.get('task_id') if reminder_type == 'task' else None
            user_id = request.form.get('user_id', current_user.id)
            
            reminder_date = request.form.get('reminder_date')
            reminder_time = request.form.get('reminder_time', '09:00')
            reminder_datetime = datetime.strptime(f"{reminder_date} {reminder_time}", '%Y-%m-%d %H:%M')
            
            subject = request.form.get('subject') or None
            message = request.form.get('message') or None
            is_recurring = request.form.get('is_recurring') == 'on'
            recurrence_pattern = request.form.get('recurrence_pattern') if is_recurring else None
            recurrence_end_date = request.form.get('recurrence_end_date') if is_recurring else None
            
            if recurrence_end_date:
                recurrence_end_date = datetime.strptime(recurrence_end_date, '%Y-%m-%d').date()
            
            reminder = Reminder(
                reminder_type=reminder_type,
                project_id=project_id,
                task_id=task_id,
                user_id=user_id,
                reminder_datetime=reminder_datetime,
                subject=subject,
                message=message,
                is_recurring=is_recurring,
                recurrence_pattern=recurrence_pattern,
                recurrence_end_date=recurrence_end_date,
                status='pending'
            )
            
            db.session.add(reminder)
            db.session.commit()
            
            flash('Reminder created successfully!', 'success')
            return redirect(url_for('reminders_list'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating reminder: {str(e)}', 'danger')
    
    if current_user.role == 'Admin':
        projects = Project.query.filter(Project.status.in_(['New', 'In Progress'])).all()
        tasks = PromotorTask.query.filter(PromotorTask.status.in_(['Pending', 'In Progress'])).all()
        users = User.query.filter_by(is_active=True).all()
    else:
        projects = Project.query.filter_by(owner_id=current_user.id).filter(
            Project.status.in_(['New', 'In Progress'])
        ).all()
        tasks = PromotorTask.query.filter_by(promotor_id=current_user.id).filter(
            PromotorTask.status.in_(['Pending', 'In Progress'])
        ).all()
        users = [current_user]
    
    return render_template('reminders/form.html', projects=projects, tasks=tasks, users=users, reminder=None)


@app.route('/reminders/<int:id>/delete', methods=['POST'])
@login_required
def reminder_delete(id):
    """Delete a reminder"""
    reminder = Reminder.query.get_or_404(id)
    
    if current_user.role != 'Admin' and reminder.user_id != current_user.id:
        flash('You do not have permission to delete this reminder.', 'danger')
        return redirect(url_for('reminders_list'))
    
    try:
        db.session.delete(reminder)
        db.session.commit()
        flash('Reminder deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting reminder: {str(e)}', 'danger')
    
    return redirect(url_for('reminders_list'))


@app.route('/api/reminders/check', methods=['GET'])
def reminders_check():
    """Cron endpoint to check and send pending reminders"""
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        from utils.reminder_scheduler import ReminderScheduler
        scheduler = ReminderScheduler()
        result = scheduler.check_and_send_reminders()

        return jsonify({'success': True, 'result': result}), 200

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"[reminders/check] Exception: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reminders/auto-create', methods=['GET', 'POST'])
def reminders_auto_create():
    """API endpoint to create automatic reminders for all incomplete projects"""
    # Verify secret key
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get all incomplete projects
        incomplete_projects = Project.query.filter(
            Project.status.in_(['New', 'In Progress', 'On Hold'])
        ).all()
        
        # Get all active users
        active_users = User.query.filter_by(is_active=True).all()
        
        # Set reminder time to now (so the check endpoint picks them up immediately)
        reminder_time = datetime.utcnow()
        
        created_count = 0
        skipped_count = 0
        
        for project in incomplete_projects:
            # Determine recipients
            recipients = []
            
            if project.owner:
                recipients = [project.owner]
            else:
                # Send to all admins and managers
                recipients = [u for u in active_users if u.role in ['Admin', 'Manager']]
            
            for user in recipients:
                # Check if a reminder already exists for this project/user in the next hour
                existing = Reminder.query.filter(
                    Reminder.project_id == project.id,
                    Reminder.user_id == user.id,
                    Reminder.status == 'pending',
                    Reminder.reminder_datetime > datetime.utcnow(),
                    Reminder.reminder_datetime < datetime.utcnow() + timedelta(hours=1)
                ).first()
                
                if not existing:
                    # Create reminder
                    reminder = Reminder(
                        reminder_type='project',
                        project_id=project.id,
                        user_id=user.id,
                        reminder_datetime=reminder_time,
                        subject=f"Daily Project Update: {project.name}",
                        message=f"This is an automated daily reminder for project: {project.name}",
                        status='pending'
                    )
                    db.session.add(reminder)
                    created_count += 1
                else:
                    skipped_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'result': {
                'projects_checked': len(incomplete_projects),
                'reminders_created': created_count,
                'reminders_skipped': skipped_count,
                'scheduled_for': reminder_time.strftime('%Y-%m-%d %H:%M:%S UTC')
            }
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/reminders/reset-failed', methods=['GET', 'POST'])
def reminders_reset_failed():
    """API endpoint to reset failed reminders to pending status"""
    # Verify secret key
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        # Get all failed reminders
        failed_reminders = Reminder.query.filter_by(status='failed').all()
        
        reset_count = 0
        for reminder in failed_reminders:
            # Reset to pending
            reminder.status = 'pending'
            reminder.error_message = None
            
            # Update reminder time to 5 minutes from now
            reminder.reminder_datetime = datetime.utcnow() + timedelta(minutes=5)
            reset_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'result': {
                'reminders_reset': reset_count,
                'scheduled_for': (datetime.utcnow() + timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S UTC')
            }
        }), 200
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500




# ============================================================================
# QUOTE EMAIL & FOLLOW-UP ROUTES
# ============================================================================

@app.route('/quotes/<int:id>/send-email', methods=['POST'])
@login_required
def quote_send_email(id):
    """Send quotation email to client"""
    from models import Quote
    from utils.email_service import EmailService

    quote = Quote.query.get_or_404(id)

    to_email = request.form.get('to_email', '').strip()
    custom_subject = request.form.get('subject', '').strip() or None
    custom_message = request.form.get('message', '').strip() or None

    if not to_email:
        flash('Please provide a recipient email address.', 'warning')
        return redirect(url_for('quote_view', id=id))

    email_service = EmailService()
    result = email_service.send_quote_email(
        quote,
        to_email,
        custom_subject=custom_subject,
        custom_message=custom_message
    )

    if result['success']:
        if quote.status == 'Draft':
            quote.status = 'Sent'
            db.session.commit()
        flash(f'Quotation email sent successfully to {to_email}.', 'success')
    else:
        flash(f'Failed to send email: {result.get("error", "Unknown error")}', 'danger')

    return redirect(url_for('quote_view', id=id))


@app.route('/quotes/<int:id>/schedule-followup', methods=['POST'])
@login_required
def quote_schedule_followup(id):
    """Schedule a follow-up reminder for a quote"""
    from models import Quote, Reminder
    from datetime import datetime

    quote = Quote.query.get_or_404(id)

    followup_datetime_str = request.form.get('followup_datetime', '').strip()
    custom_subject = request.form.get('subject', '').strip() or None
    custom_message = request.form.get('message', '').strip() or None

    if not followup_datetime_str:
        flash('Please select a follow-up date and time.', 'warning')
        return redirect(url_for('quote_view', id=id))

    try:
        followup_datetime = datetime.strptime(followup_datetime_str, '%Y-%m-%dT%H:%M')
    except ValueError:
        flash('Invalid date/time format.', 'warning')
        return redirect(url_for('quote_view', id=id))

    if followup_datetime <= datetime.utcnow():
        flash('Follow-up time must be in the future.', 'warning')
        return redirect(url_for('quote_view', id=id))

    reminder = Reminder(
        reminder_type='quote',
        quote_id=quote.id,
        user_id=current_user.id,
        reminder_datetime=followup_datetime,
        subject=custom_subject or f'Follow-up: Quote {quote.quote_number} — {quote.customer_name}',
        message=custom_message,
        status='pending'
    )
    db.session.add(reminder)
    db.session.commit()

    flash(
        f'Follow-up reminder scheduled for {followup_datetime.strftime("%d %b %Y at %I:%M %p")}.',
        'success'
    )
    return redirect(url_for('quote_view', id=id))


# ============================================================================
# LEADFY - LEAD MANAGEMENT ROUTES
# ============================================================================

@app.route('/api/leads/webhook', methods=['POST'])
@limiter.limit("10 per minute; 50 per hour")
def leads_webhook():
    """Receive leads from glassy.in WordPress contact form"""
    import hmac

    # Block oversized bodies before parsing (prevent DoS via huge payloads)
    MAX_BODY = 32 * 1024  # 32 KB
    content_length = request.content_length
    if content_length and content_length > MAX_BODY:
        return jsonify({'success': False, 'error': 'Payload too large'}), 413

    # Constant-time secret comparison (prevents timing attacks)
    expected_secret = os.getenv('GLASSY_WEBHOOK_SECRET', '')
    auth_header = request.headers.get('Authorization', '')
    expected_header = f'Bearer {expected_secret}'
    if not expected_secret or not hmac.compare_digest(auth_header, expected_header):
        return jsonify({'success': False, 'error': 'Unauthorized'}), 401

    data = request.get_json(silent=True)
    if not data:
        return jsonify({'success': False, 'error': 'Invalid JSON'}), 400

    from models import Lead
    try:
        name       = (data.get('name') or '').strip()[:200] or None
        email      = (data.get('email') or '').strip()[:200] or None
        phone      = (data.get('phone') or '').strip()[:30] or None
        message    = (data.get('message') or '').strip()[:2000]
        source_url = (data.get('source_url') or '').strip()[:500]

        # Deduplicate: if same email submitted within the last 60 seconds, skip
        if email:
            cutoff = datetime.utcnow() - timedelta(seconds=60)
            duplicate = Lead.query.filter_by(email=email, origin='glassy.in') \
                                  .filter(Lead.created_at >= cutoff).first()
            if duplicate:
                app.logger.info(f'[Webhook] Duplicate lead skipped: {email}')
                return jsonify({'success': True}), 200

        # Compose notes from message + source URL
        notes_parts = []
        if message:
            notes_parts.append(message)
        if source_url:
            notes_parts.append(f'Source: {source_url}')
        notes = '\n'.join(notes_parts) or None

        # Find a system admin to use as created_by (required FK)
        from models import User
        admin = User.query.filter_by(role='Admin', is_active=True).first()
        if not admin:
            return jsonify({'success': False, 'error': 'Configuration error'}), 500

        lead = Lead(
            name=name,
            email=email,
            contact=phone,
            notes=notes,
            origin='glassy.in',
            stage='New Lead',
            owner_id=None,
            assigned_to_id=None,
            created_by=admin.id,
        )
        db.session.add(lead)
        db.session.commit()
        app.logger.info(f'[Webhook] New lead from glassy.in: {name} ({email})')
        return jsonify({'success': True}), 200

    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[Webhook] Error saving lead: {e}')
        return jsonify({'success': False, 'error': 'Internal error'}), 500


def _build_leads_query(args, *, restrict_to_current_user=True):
    """Apply Leadfy list filters from a dict-like `args` (request.args or
    request.form) and return the SQLAlchemy query.

    Shared by `leads_list` (paginated render) and `leads_bulk_assign`
    (select-all-matching mode). Keep filter logic in one place so the
    bulk-assign endpoint can re-run the exact same filter the user is
    looking at without re-implementing it.

    `restrict_to_current_user` enforces the same access rule as the list
    page — non-managers can only act on leads assigned to them.
    """
    from models import Lead
    from datetime import datetime

    query = Lead.query

    if restrict_to_current_user and not current_user.is_manager_or_admin():
        query = query.filter(Lead.assigned_to_id == current_user.id)

    search_query = (args.get('search') or '').strip()
    if search_query:
        query = query.filter(
            (Lead.name.ilike(f'%{search_query}%')) |
            (Lead.contact.ilike(f'%{search_query}%')) |
            (Lead.company.ilike(f'%{search_query}%'))
        )
    if args.get('stage'):
        query = query.filter(Lead.stage == args.get('stage'))
    if args.get('state'):
        query = query.filter(Lead.state.ilike(f"%{args.get('state')}%"))
    if args.get('origin'):
        query = query.filter(Lead.origin == args.get('origin'))
    if args.get('lead_type'):
        query = query.filter(Lead.lead_type == args.get('lead_type'))

    # ── Facebook ad-hierarchy filters ────────────────────────────────
    # Three nested drill-downs: campaign → adset → ad. BD picks any of
    # them via the dropdowns on /leads. We filter on IDs (not names) so
    # the URL is stable even if a campaign gets renamed in Ads Manager
    # mid-flight. The DISTINCT name lists for the dropdowns are built
    # in `leads_list` below — keep the filter logic + the dropdown
    # data-source aligned (ID + name pairs).
    if args.get('fb_campaign_id'):
        query = query.filter(Lead.fb_campaign_id == args.get('fb_campaign_id'))
    if args.get('fb_adset_id'):
        query = query.filter(Lead.fb_adset_id == args.get('fb_adset_id'))
    if args.get('fb_ad_id'):
        query = query.filter(Lead.fb_ad_id == args.get('fb_ad_id'))
    untouched = args.get('untouched', '')
    if untouched == '1':
        query = query.filter(Lead.is_untouched == True)  # noqa: E712
    elif untouched == '0':
        query = query.filter(Lead.is_untouched == False)  # noqa: E712
    if args.get('owner'):
        try:
            query = query.filter(Lead.owner_id == int(args.get('owner')))
        except ValueError:
            pass
    # `assigned_to=X` filters to leads currently assigned to user X. This is
    # what the /leads/agent-log cells link to (2026-07-02 rewrite — the matrix
    # counts by assignee, so click-through must too, otherwise the drill-down
    # would jump from "leads on rohits's plate" to "leads rohits owns" which
    # are now different sets).
    if args.get('assigned_to'):
        try:
            query = query.filter(Lead.assigned_to_id == int(args.get('assigned_to')))
        except ValueError:
            pass
    # `unowned=1` filters to leads with no owner assigned (legacy — kept for
    # any external bookmark that still hits it).
    if args.get('unowned') == '1':
        query = query.filter(Lead.owner_id.is_(None))
    # `unassigned=1` filters to leads with no current assignee. Powers the
    # per-stage Unassigned column on /leads/agent-log.
    if args.get('unassigned') == '1':
        query = query.filter(Lead.assigned_to_id.is_(None))

    for key, op in (
        ('updated_from', lambda v: Lead.updated_at >= datetime.strptime(v, '%Y-%m-%d')),
        ('updated_to',   lambda v: Lead.updated_at <= datetime.strptime(v + ' 23:59:59', '%Y-%m-%d %H:%M:%S')),
        ('created_from', lambda v: Lead.created_at >= datetime.strptime(v, '%Y-%m-%d')),
        ('created_to',   lambda v: Lead.created_at <= datetime.strptime(v + ' 23:59:59', '%Y-%m-%d %H:%M:%S')),
    ):
        v = (args.get(key) or '').strip()
        if v:
            try:
                query = query.filter(op(v))
            except ValueError:
                pass

    return query


@app.route('/leads/agent-log')
@login_required
def leads_agent_log():
    """Owner-wise stage-activity matrix for the Leadfy admin (KAN-52).

    Renders a single page with:
      • A date-range picker (Today / Yesterday / This Week / This Month /
        Last Month / Custom).
      • A matrix table: rows = lead stages, columns = lead owners, cells =
        number of *stage-change events* recorded in `lead_history` for that
        owner's leads in the chosen range. Click a cell to drill into the
        same filtered slice on /leads.

    Why "activity" and not "snapshot": the ticket title is "stage activity",
    so a cell answers "how many leads did Ansar move INTO 'PI Shared'
    yesterday?" instead of "how many of Ansar's leads are currently in 'PI
    Shared'?". The two differ when leads later move out of the stage or
    when ownership transfers. LeadHistory rows with action='stage_change'
    are the canonical activity source.

    Implementation note: the target stage of each event is currently
    embedded in `LeadHistory.description` like
    "Stage changed from <strong>X</strong> to <strong>Y</strong>". We
    extract it with a Postgres `substring(... from '...')` regex at query
    time so this ships without a schema migration. If query latency ever
    becomes an issue, add a `target_stage` column and backfill — the
    template-side rendering wouldn't need to change.

    Manager-or-admin only — same gate as bulk-assign. Regular agents would
    leak peer activity which isn't appropriate.
    """
    if not current_user.is_manager_or_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('leads_list'))

    from models import Lead, LeadHistory, User as UserModel, LEAD_STAGES_ALL, LEAD_STAGE_BADGE_CLASSES
    from sqlalchemy import func, literal_column
    from datetime import datetime, timedelta

    preset = (request.args.get('preset') or 'today').strip()
    date_from_str = (request.args.get('date_from') or '').strip()
    date_to_str = (request.args.get('date_to') or '').strip()

    # Resolve the date range. `start` is inclusive, `end` is exclusive (so a
    # "today" range covers [00:00 today, 00:00 tomorrow)). Times are UTC
    # because LeadHistory.created_at stores naive UTC via datetime.utcnow.
    now = datetime.utcnow()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)

    if preset == 'yesterday':
        end = today_start
        start = end - timedelta(days=1)
    elif preset == 'week':
        # Monday 00:00 → tomorrow 00:00 (covers today)
        start = (today_start - timedelta(days=now.weekday()))
        end = today_start + timedelta(days=1)
    elif preset == 'month':
        start = today_start.replace(day=1)
        end = today_start + timedelta(days=1)
    elif preset == 'last_month':
        # Last day of previous month at 23:59 → 1st of current month at 00:00.
        first_of_this_month = today_start.replace(day=1)
        last_month_last_day = first_of_this_month - timedelta(days=1)
        start = last_month_last_day.replace(day=1)
        end = first_of_this_month
    elif preset == 'custom' and date_from_str and date_to_str:
        try:
            start = datetime.strptime(date_from_str, '%Y-%m-%d')
            end = datetime.strptime(date_to_str, '%Y-%m-%d') + timedelta(days=1)
        except ValueError:
            # Bad input → fall through to today
            preset = 'today'
            start = today_start
            end = today_start + timedelta(days=1)
    else:
        # default 'today'
        preset = 'today'
        start = today_start
        end = today_start + timedelta(days=1)

    # Postgres-native regex extraction of the target stage from the existing
    # HTML description format. The captured group is the new stage name.
    # literal_column (not text()) so the expression is a ColumnClause and
    # carries the .label() / GROUP-BY semantics SQLAlchemy needs.
    target_stage_expr = literal_column(
        "substring(lead_history.description from 'to <strong>([^<]+)</strong>')"
    ).label('target_stage')

    grouped = (db.session.query(
                    Lead.owner_id.label('owner_id'),
                    target_stage_expr,
                    func.count(LeadHistory.id).label('cnt'),
                )
                .join(Lead, LeadHistory.lead_id == Lead.id)
                .filter(LeadHistory.action == 'stage_change')
                .filter(LeadHistory.created_at >= start)
                .filter(LeadHistory.created_at < end)
                .group_by(Lead.owner_id, target_stage_expr)
                .all())

    # Build matrix[owner_id][stage_name] = count. Drop NULL owner rows
    # (leads with no assigned owner — would render as a "no owner" column
    # which adds noise; admins can find these via the existing list filter
    # if they care).
    matrix = {}
    for r in grouped:
        if r.owner_id is None or not r.target_stage:
            continue
        matrix.setdefault(r.owner_id, {})[r.target_stage] = int(r.cnt)

    # All active users become columns even if they have 0 events in this
    # range — this makes "who hasn't done anything this week?" obvious at a
    # glance. Sort alphabetically per the ticket.
    # `show_in_agent_log` is a cosmetic filter — hidden users keep every other
    # permission (login, edit, etc), they just don't appear as a column here.
    # Manager toggles it in the user admin form.
    owners = (UserModel.query
                .filter(UserModel.is_active == True)  # noqa: E712
                .filter(UserModel.show_in_agent_log == True)  # noqa: E712
                .order_by(UserModel.username)
                .all())

    # All defined stages become rows (LEAD_STAGES_ALL is the union of the
    # default + Facebook funnels — keeps the matrix stable as we add new
    # funnels). Render in the order they appear in LEAD_STAGES_ALL so the
    # familiar progression reads top→bottom.
    stages = list(LEAD_STAGES_ALL)

    # Totals row + column + grand total.
    stage_totals = {s: sum(matrix.get(o.id, {}).get(s, 0) for o in owners) for s in stages}
    owner_totals = {o.id: sum(matrix.get(o.id, {}).get(s, 0) for s in stages) for o in owners}
    grand_total = sum(owner_totals.values())

    # ── Snapshot view ────────────────────────────────────────────────
    # Activity events (above) vs current snapshot (below) measure
    # different things — events answers "who moved which lead here this
    # week" while snapshot answers "how many leads are in stage X right
    # now". BD was getting confused because they ran a bulk-move that
    # updated Lead.stage WITHOUT writing lead_history rows, so 70% of
    # leads have no stage_change history at all → events matrix looked
    # tiny vs /leads. Surface the snapshot directly so the agent-log
    # page reconciles with /leads at a glance without losing the
    # event-flow matrix below it.
    #
    # Snapshot is date-range-INDEPENDENT (it always reflects the current
    # database state) — the date picker only affects the events matrix.
    snapshot_rows = (db.session.query(
                        Lead.owner_id.label('owner_id'),
                        Lead.stage.label('stage'),
                        func.count(Lead.id).label('cnt'))
                     .group_by(Lead.owner_id, Lead.stage)
                     .all())
    snapshot_matrix = {}     # snapshot_matrix[owner_id|None][stage] = count
    for r in snapshot_rows:
        snapshot_matrix.setdefault(r.owner_id, {})[r.stage] = int(r.cnt)
    snapshot_stage_totals = {
        s: sum(om.get(s, 0) for om in snapshot_matrix.values()) for s in stages
    }
    snapshot_owner_totals = {
        o.id: sum(snapshot_matrix.get(o.id, {}).get(s, 0) for s in stages) for o in owners
    }
    snapshot_grand_total = sum(snapshot_stage_totals.values())
    # Leads with no owner assigned — surface as a single counter so they
    # don't silently disappear from the snapshot's grand total.
    snapshot_unowned = sum(snapshot_matrix.get(None, {}).get(s, 0) for s in stages)
    # Per-stage unowned breakdown for the dedicated "Unowned" column in the
    # snapshot matrix. Without this column, visible owner cells + Total wouldn't
    # reconcile because Total sums across ALL owners including None while the
    # visible cells iterate only real owners. Adding the column makes the math
    # (visible_cells + unowned = total) obvious at a glance.
    snapshot_unowned_by_stage = {
        s: snapshot_matrix.get(None, {}).get(s, 0) for s in stages
    }
    # Leads whose `stage` value isn't in LEAD_STAGES_ALL (legacy free-text,
    # custom funnels added by BD via direct SQL, etc). They count in the
    # grand total but won't appear in any per-stage row of the matrix,
    # so we expose the count separately too.
    snapshot_other_stages_count = sum(
        cnt for om in snapshot_matrix.values()
            for s, cnt in om.items()
            if s and s not in stages
    )

    return render_template('leads/agent_log.html',
                           owners=owners,
                           stages=stages,
                           matrix=matrix,
                           stage_totals=stage_totals,
                           owner_totals=owner_totals,
                           grand_total=grand_total,
                           # Snapshot block (date-range-independent — matches /leads)
                           snapshot_matrix=snapshot_matrix,
                           snapshot_stage_totals=snapshot_stage_totals,
                           snapshot_owner_totals=snapshot_owner_totals,
                           snapshot_grand_total=snapshot_grand_total,
                           snapshot_unowned=snapshot_unowned,
                           snapshot_unowned_by_stage=snapshot_unowned_by_stage,
                           snapshot_other_stages_count=snapshot_other_stages_count,
                           preset=preset,
                           # Display the inclusive end-of-range — what the
                           # user typed / what the preset label says — not
                           # the exclusive `end` we use internally.
                           date_from=start.strftime('%Y-%m-%d'),
                           date_to=(end - timedelta(days=1)).strftime('%Y-%m-%d'),
                           badge_class_for=LEAD_STAGE_BADGE_CLASSES)


@app.route('/leads')
@login_required
def leads_list():
    """List all leads with search and filter"""
    from models import Lead, User

    search_query = request.args.get('search', '')
    stage_filter = request.args.get('stage', '')
    state_filter = request.args.get('state', '')
    origin_filter = request.args.get('origin', '')
    owner_filter = request.args.get('owner', '')
    updated_from = request.args.get('updated_from', '')
    updated_to = request.args.get('updated_to', '')
    created_from = request.args.get('created_from', '')
    created_to = request.args.get('created_to', '')
    lead_type_filter = request.args.get('lead_type', '')
    fb_campaign_filter = request.args.get('fb_campaign_id', '')
    fb_adset_filter    = request.args.get('fb_adset_id',    '')
    fb_ad_filter       = request.args.get('fb_ad_id',       '')

    query = _build_leads_query(request.args)

    # Page size — defaults to 15 but caller can pass ?per_page=… up to 500
    # so a manager who wants to select-all a long filter result can crank
    # the page up if they don't want to use the "select all matching" banner.
    DEFAULT_PER_PAGE = 15
    PER_PAGE_MAX = 500
    try:
        per_page = int(request.args.get('per_page', DEFAULT_PER_PAGE))
    except ValueError:
        per_page = DEFAULT_PER_PAGE
    per_page = max(1, min(per_page, PER_PAGE_MAX))
    page = request.args.get('page', 1, type=int)

    pagination = query.order_by(Lead.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    leads = pagination.items
    total_leads = Lead.query.count()

    # Count untouched leads matching the CURRENT filter set (minus the
    # untouched filter itself, so the badge reads e.g. "12 untouched" even
    # when the user has filtered to untouched=No). Previous behaviour was
    # `{{ leads|selectattr('is_untouched')|list|length }}` in the template
    # — that counts only rows on the current page (15 by default), so BD
    # got misled into thinking "10 untouched" was the full total when it
    # was actually 10 out of 15 visible.
    _untouched_args = {k: v for k, v in request.args.items() if k != 'untouched'}
    untouched_total_matching = (_build_leads_query(_untouched_args)
                                .filter(Lead.is_untouched.is_(True))
                                .count())

    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    origins = db.session.query(Lead.origin).filter(Lead.origin.isnot(None)).distinct().order_by(Lead.origin).all()
    origins = [o[0] for o in origins]

    states = db.session.query(Lead.state).filter(Lead.state.isnot(None)).distinct().order_by(Lead.state).all()
    states = [s[0] for s in states]

    # ── FB ad-hierarchy dropdown options ─────────────────────────────
    # (id, name) pairs so the dropdown displays the human-readable name
    # while the form submits the stable ID. If the customer narrows by
    # campaign, the adset list narrows to that campaign's adsets — same
    # idea for ad. Lists are computed at request time; for now the
    # cardinality of campaigns/adsets is small enough (tens, not
    # thousands) that an unindexed DISTINCT scan is fine. We can add a
    # cached materialised view later if Meta starts spamming us.
    fb_campaigns = db.session.query(
        Lead.fb_campaign_id, Lead.fb_campaign_name
    ).filter(Lead.fb_campaign_id.isnot(None)).distinct().order_by(
        Lead.fb_campaign_name
    ).all()
    fb_adsets_q = db.session.query(
        Lead.fb_adset_id, Lead.fb_adset_name
    ).filter(Lead.fb_adset_id.isnot(None))
    if fb_campaign_filter:
        fb_adsets_q = fb_adsets_q.filter(Lead.fb_campaign_id == fb_campaign_filter)
    fb_adsets = fb_adsets_q.distinct().order_by(Lead.fb_adset_name).all()
    fb_ads_q = db.session.query(
        Lead.fb_ad_id, Lead.fb_ad_name
    ).filter(Lead.fb_ad_id.isnot(None))
    if fb_campaign_filter:
        fb_ads_q = fb_ads_q.filter(Lead.fb_campaign_id == fb_campaign_filter)
    if fb_adset_filter:
        fb_ads_q = fb_ads_q.filter(Lead.fb_adset_id == fb_adset_filter)
    fb_ads = fb_ads_q.distinct().order_by(Lead.fb_ad_name).all()

    from models import IndiamartToken, LEAD_STAGES_ALL, LEAD_STAGES_DEFAULT, LEAD_STAGES_FACEBOOK
    indiamart_token = IndiamartToken.query.first()
    fb_token_set = bool(_fb_page_tokens())

    return render_template('leads/list.html',
                           leads=leads,
                           pagination=pagination,
                           total_leads=total_leads,
                           untouched_total_matching=untouched_total_matching,
                           users=users,
                           origins=origins,
                           states=states,
                           search_query=search_query,
                           stage_filter=stage_filter,
                           state_filter=state_filter,
                           origin_filter=origin_filter,
                           owner_filter=owner_filter,
                           lead_type_filter=lead_type_filter,
                           updated_from=updated_from,
                           updated_to=updated_to,
                           created_from=created_from,
                           created_to=created_to,
                           # FB ad-hierarchy filter state + dropdown options
                           fb_campaign_filter=fb_campaign_filter,
                           fb_adset_filter=fb_adset_filter,
                           fb_ad_filter=fb_ad_filter,
                           fb_campaigns=fb_campaigns,
                           fb_adsets=fb_adsets,
                           fb_ads=fb_ads,
                           indiamart_token=indiamart_token,
                           fb_token_set=fb_token_set,
                           stage_options_all=LEAD_STAGES_ALL,
                           stage_options_default=LEAD_STAGES_DEFAULT,
                           stage_options_facebook=LEAD_STAGES_FACEBOOK,
                           per_page=per_page)


@app.route('/leads/new', methods=['GET', 'POST'])
@login_required
def lead_new():
    """Create a new lead"""
    from models import Lead, User, default_stage_for_origin, stages_for_origin

    if request.method == 'POST':
        name = request.form.get('name', '').strip() or None
        owner_id = request.form.get('owner_id') or None
        contact = request.form.get('contact', '').strip() or None
        city = request.form.get('city', '').strip() or None
        state = request.form.get('state', '').strip() or None
        origin = request.form.get('origin', '').strip() or None
        stage = request.form.get('stage') or default_stage_for_origin(origin)

        lead = Lead(
            name=name,
            owner_id=int(owner_id) if owner_id else None,
            contact=contact,
            city=city,
            state=state,
            stage=stage,
            origin=origin,
            created_by=current_user.id
        )
        db.session.add(lead)
        db.session.commit()
        flash('Lead created successfully.', 'success')
        return redirect(url_for('leads_list'))

    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    return render_template(
        'leads/form.html', lead=None, users=users, action='new',
        stage_options=stages_for_origin(None),
    )


@app.route('/leads/<int:id>')
@login_required
def lead_view(id):
    """Lead detail page"""
    from models import Lead
    lead = Lead.query.get_or_404(id)
    history = lead.history.order_by(__import__('models').LeadHistory.created_at.desc()).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    return render_template('leads/view.html', lead=lead, history=history, users=users)


@app.route('/leads/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def lead_edit(id):
    """Edit an existing lead"""
    from models import Lead, User, LeadHistory

    from models import stages_for_origin
    lead = Lead.query.get_or_404(id)

    if request.method == 'POST':
        changes = []

        # Track stage change
        new_stage = request.form.get('stage', lead.stage)
        if new_stage != lead.stage:
            changes.append(LeadHistory(lead_id=lead.id, user_id=current_user.id,
                action='stage_change',
                description=f'Stage changed from <strong>{lead.stage}</strong> to <strong>{new_stage}</strong>'))
            lead.stage = new_stage
            lead.is_untouched = False

        # Track owner change — ONLY if the field is actually present in the
        # form. Partial-form submitters (e.g. the MOVE TO quick-change buttons
        # on the lead detail page) used to wipe whatever they didn't include,
        # because `request.form.get(...) or None` returns None for both
        # "missing field" and "empty field". A missing field means
        # "don't touch", an empty field means "clear it". Guarding with
        # `'owner_id' in request.form` distinguishes the two cleanly.
        #
        # Owner changes are manager/admin only (2026-07-02). Promotors can
        # still update stage, notes, phone, assigned_to etc. — just not
        # transfer ownership. Silently dropping the field on a Promotor edit
        # keeps their normal edits working (form template can safely include
        # the field) without letting them use it.
        if 'owner_id' in request.form and current_user.is_manager_or_admin():
            owner_id = request.form.get('owner_id') or None
            new_owner_id = int(owner_id) if owner_id else None
            if new_owner_id != lead.owner_id:
                old_owner = User.query.get(lead.owner_id).username if lead.owner_id else 'Unassigned'
                new_owner = User.query.get(new_owner_id).username if new_owner_id else 'Unassigned'
                changes.append(LeadHistory(lead_id=lead.id, user_id=current_user.id,
                    action='field_change',
                    description=f'Owner changed from <strong>{old_owner}</strong> to <strong>{new_owner}</strong>'))
                lead.owner_id = new_owner_id

        # Track assigned_to change — same guard pattern as owner_id above.
        # The MOVE TO quick-change form in templates/leads/view.html USED to
        # omit this field, which silently unassigned the lead. Fixed
        # 2026-06-08; the guard keeps it safe even if a future caller
        # forgets it.
        if 'assigned_to_id' in request.form:
            assigned_to_raw = request.form.get('assigned_to_id') or None
            new_assigned_to_id = int(assigned_to_raw) if assigned_to_raw else None
            if new_assigned_to_id != lead.assigned_to_id:
                old_assignee = User.query.get(lead.assigned_to_id).username if lead.assigned_to_id else 'Unassigned'
                new_assignee = User.query.get(new_assigned_to_id).username if new_assigned_to_id else 'Unassigned'
                changes.append(LeadHistory(lead_id=lead.id, user_id=current_user.id,
                    action='field_change',
                    description=f'Assigned To changed from <strong>{old_assignee}</strong> to <strong>{new_assignee}</strong>'))
                lead.assigned_to_id = new_assigned_to_id

        # Track other field changes — same partial-form guard. Skip fields
        # the caller didn't send so quick-action forms that only send a
        # subset (e.g. MOVE TO) leave the rest alone.
        field_map = [
            ('name', 'name', 'Name'),
            ('contact', 'contact', 'Phone'),
            ('email', 'email', 'Email'),
            ('city', 'city', 'City'),
            ('state', 'state', 'State'),
            ('origin', 'origin', 'Origin'),
        ]
        for form_key, model_attr, label in field_map:
            if form_key not in request.form:
                continue
            new_val = request.form.get(form_key, '').strip() or None
            old_val = getattr(lead, model_attr)
            if new_val != old_val:
                changes.append(LeadHistory(lead_id=lead.id, user_id=current_user.id,
                    action='field_change',
                    description=f'{label} changed from <strong>{old_val or "—"}</strong> to <strong>{new_val or "—"}</strong>'))
                setattr(lead, model_attr, new_val)

        for ch in changes:
            db.session.add(ch)
        db.session.commit()
        return redirect(url_for('lead_view', id=lead.id))

    users = User.query.filter_by(is_active=True).order_by(User.username).all()
    return render_template(
        'leads/form.html', lead=lead, users=users, action='edit',
        stage_options=stages_for_origin(lead.origin),
    )


@app.route('/leads/<int:id>/delete', methods=['POST'])
@login_required
def lead_delete(id):
    """Delete a lead (admin only)"""
    if not current_user.is_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('leads_list'))

    from models import Lead
    lead = Lead.query.get_or_404(id)
    db.session.delete(lead)
    db.session.commit()
    flash('Lead deleted.', 'success')
    return redirect(url_for('leads_list'))


@app.route('/leads/bulk-assign', methods=['POST'])
@login_required
def leads_bulk_assign():
    """Bulk-assign multiple leads to a single user (sets both owner and assigned_to).

    Manager/admin only (2026-07-02). Sales users (Promotor) can't reassign
    leads — the blast radius of a mistake here is one incident away from
    another "58 leads walked away from Abhay" event, so it's a manager call.

    Ownership model: `owner_id` transfers on every reassignment — whoever
    the lead is bulk-assigned TO becomes the current owner. Agent-log
    columns count by `owner_id`, so this keeps the "who's carrying this
    lead right now" view honest as leads pass hands.

    Two input modes:
      • `lead_ids[]` — assign exactly those IDs (original 15-per-page flow).
      • `select_all_matching=1` + filter params (same names as the leads_list
        querystring: stage, origin, search, state, owner, lead_type, untouched,
        updated_from/to, created_from/to) — re-runs the filter SQL and assigns
        every matching lead. Lets a manager assign 100s of leads at once
        without depending on pagination.
    """
    from models import Lead, User, LeadHistory

    if not current_user.is_manager_or_admin():
        return jsonify({'success': False, 'error': 'Only managers or admins can reassign leads.'}), 403

    user_id_raw = request.form.get('user_id')
    if not user_id_raw:
        return jsonify({'success': False, 'error': 'No user selected.'}), 400

    try:
        new_user_id = int(user_id_raw)
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid user id.'}), 400

    new_user = User.query.get(new_user_id)
    if not new_user or not new_user.is_active:
        return jsonify({'success': False, 'error': 'User not found or inactive.'}), 400

    select_all_matching = request.form.get('select_all_matching') == '1'

    if select_all_matching:
        # Re-apply the same filter the user is looking at in the UI. The form
        # body carries the filter params (mirrored from the URL querystring).
        leads = _build_leads_query(request.form).all()
    else:
        lead_ids = request.form.getlist('lead_ids')
        if not lead_ids:
            return jsonify({'success': False, 'error': 'No leads selected.'}), 400
        try:
            lead_ids = [int(x) for x in lead_ids]
        except ValueError:
            return jsonify({'success': False, 'error': 'Invalid lead ids.'}), 400
        # Mirror the list-page access rule: non-managers can only bulk-assign
        # leads already assigned to them.
        q = Lead.query.filter(Lead.id.in_(lead_ids))
        if not current_user.is_manager_or_admin():
            q = q.filter(Lead.assigned_to_id == current_user.id)
        leads = q.all()
    updated = 0
    for lead in leads:
        changed = False

        if lead.owner_id != new_user_id:
            old_owner = User.query.get(lead.owner_id).username if lead.owner_id else 'Unassigned'
            db.session.add(LeadHistory(
                lead_id=lead.id, user_id=current_user.id, action='field_change',
                description=f'Owner changed from <strong>{old_owner}</strong> to <strong>{new_user.username}</strong> (bulk assign)'
            ))
            lead.owner_id = new_user_id
            changed = True

        if lead.assigned_to_id != new_user_id:
            old_assignee = User.query.get(lead.assigned_to_id).username if lead.assigned_to_id else 'Unassigned'
            db.session.add(LeadHistory(
                lead_id=lead.id, user_id=current_user.id, action='field_change',
                description=f'Assigned To changed from <strong>{old_assignee}</strong> to <strong>{new_user.username}</strong> (bulk assign)'
            ))
            lead.assigned_to_id = new_user_id
            changed = True

        if changed:
            updated += 1

    db.session.commit()
    return jsonify({'success': True, 'updated': updated, 'total': len(leads)})


# ============================================================================
# INDIAMART INTEGRATION
# ============================================================================

@app.route('/leads/indiamart/save-token', methods=['POST'])
@login_required
@admin_required
def indiamart_save_token():
    """Save IndiaMart token captured from the mobile app"""
    from models import IndiamartToken
    import base64, json
    from datetime import datetime

    ak_token = request.form.get('ak_token', '').strip()
    glid = request.form.get('glid', '').strip()
    mobile = request.form.get('mobile', '').strip()
    user_ip = request.form.get('user_ip', '').strip()

    if not ak_token:
        flash('Token is required.', 'danger')
        return redirect(url_for('leads_list'))

    payload = _decode_jwt_payload(ak_token)
    expires_at = None
    refresh_token_val = None
    try:
        if payload.get('exp'):
            expires_at = datetime.utcfromtimestamp(payload['exp'])
        # Extract refresh token if IndiaMart embedded one in the JWT payload
        for key in ('refresh_token', 'rt', 'rtoken', 'refreshToken'):
            if payload.get(key):
                refresh_token_val = payload[key]
                break
    except Exception:
        pass

    token = IndiamartToken.query.first()
    if not token:
        token = IndiamartToken()
        db.session.add(token)
    token.ak_token = ak_token
    token.glid = glid
    token.mobile = mobile
    token.user_ip = user_ip
    token.expires_at = expires_at
    if refresh_token_val:
        token.refresh_token = refresh_token_val
    db.session.commit()

    flash('IndiaMart token saved successfully.', 'success')
    return redirect(url_for('leads_list'))


def _decode_jwt_payload(jwt_str):
    """Decode a JWT payload section and return the dict (no signature verification)."""
    import base64, json
    try:
        parts = jwt_str.split('.')
        if len(parts) < 2:
            return {}
        pad = parts[1] + '=' * (4 - len(parts[1]) % 4)
        return json.loads(base64.b64decode(pad))
    except Exception:
        return {}


def _save_new_ak(token, new_ak):
    """Persist a fresh AK token (and any embedded refresh_token) to the DB."""
    from datetime import datetime
    payload = _decode_jwt_payload(new_ak)
    token.ak_token = new_ak
    if payload.get('exp'):
        token.expires_at = datetime.utcfromtimestamp(payload['exp'])
    # IndiaMart sometimes embeds a long-lived refresh token in the payload
    for key in ('refresh_token', 'rt', 'rtoken', 'refreshToken'):
        if payload.get(key):
            token.refresh_token = payload[key]
            break
    db.session.commit()


def _indiamart_refresh_token(token, headers):
    """
    Try to get a fresh AK token from IndiaMart.
    Strategy 1 – checkAuth (works when current token is still valid or just expired).
    Strategy 2 – use stored refresh_token if IndiaMart returned one previously.
    Returns the current (possibly refreshed) AK string.
    """
    import requests as req_lib
    from datetime import datetime

    import logging
    log = logging.getLogger(__name__)

    datacookie = (
        f"fn=Rohit Kumar|em=vetrovaglass@gmail.com|phcc=91|iso=IN"
        f"|mb1={token.mobile}|ctid=70532|glid={token.glid}"
        f"|cmid=1|uTyp=P|utyp=P|ev=V|uv=V"
    )
    cookie_header = (
        f"ImeshVisitor=fn=Rohit Kumar|em=vetrovaglass@gmail.com|phcc=91|iso=IN"
        f"|mb1={token.mobile}|ctid=70532|glid={token.glid}"
        f"|cmid=1|uTyp=P|utyp=P|ev=V|uv=V; "
        f"im_iss=t={token.ak_token}"
    )
    base_data = {
        'APP_ACCURACY': '',
        'APP_LATITUDE': '',
        'APP_LONGITUDE': '',
        'APP_MODID': 'IOS',
        'APP_SCREEN_NAME': 'Api Request',
        'APP_USER_ID': token.glid,
        'GEOIP_COUNTRY_ISO': 'IN',
        'USER_IP': token.user_ip or '',
        'USER_IP_COUNTRY': 'India',
        'VALIDATION_GLID': token.glid,
        'VALIDATION_USERCONTACT': token.mobile,
        'VALIDATION_USER_IP': token.user_ip or '',
        'app_version_no': '13.6.4_b_4',
        'datacookie': datacookie,
        'glusrid': token.glid,
        'modid': 'IOS',
        'user_name': token.mobile,
    }
    req_headers = {**headers, 'cookie': cookie_header}

    # --- Strategy 1: checkAuth with current AK ---
    try:
        resp = req_lib.post(
            'https://mapi.indiamart.com/wservce/users/login/',
            data={**base_data,
                  'AK': token.ak_token,
                  'checkAuth': '1',
                  'im_iss': f't={token.ak_token}'},
            headers=req_headers,
            timeout=10
        )
        data = resp.json()
        log.warning(f"[IndiaMart checkAuth] HTTP {resp.status_code} | keys={list(data.keys())} | access={data.get('access')} | MSG={data.get('message','')}")
        new_ak = (data.get('jwt_token') or data.get('AK') or data.get('ak')
                  or data.get('TOKEN') or data.get('token'))
        if new_ak and data.get('access') in ('1', '2', 1, 2):
            old_exp = token.expires_at
            _save_new_ak(token, new_ak)
            # IndiaMart returns the same JWT but the server-side session is alive.
            # Extend our stored expiry by 24h from now so keepalive keeps working.
            token.expires_at = datetime.utcnow() + timedelta(hours=24)
            db.session.commit()
            log.warning(f"[IndiaMart checkAuth] Session alive, extended expiry | old_exp={old_exp} | new_exp={token.expires_at}")
            return new_ak
        else:
            log.warning(f"[IndiaMart checkAuth] No token in response. Full response: {data}")
    except Exception as e:
        log.warning(f"[IndiaMart checkAuth] Exception: {e}")

    # --- Strategy 2: use stored refresh_token (if IndiaMart ever returned one) ---
    if token.refresh_token:
        try:
            resp = req_lib.post(
                'https://mapi.indiamart.com/wservce/users/login/',
                data={**base_data,
                      'refresh_token': token.refresh_token,
                      'grant_type': 'refresh_token'},
                headers=headers,
                timeout=10
            )
            data = resp.json()
            log.warning(f"[IndiaMart refresh_token] HTTP {resp.status_code} | keys={list(data.keys())} | CODE={data.get('CODE')} | MSG={data.get('MESSAGE') or data.get('message','')}")
            new_ak = data.get('AK') or data.get('ak') or data.get('TOKEN') or data.get('token')
            if new_ak:
                _save_new_ak(token, new_ak)
                return new_ak
            else:
                log.warning(f"[IndiaMart refresh_token] No AK in response. Full response: {data}")
        except Exception as e:
            log.warning(f"[IndiaMart refresh_token] Exception: {e}")

    log.warning(f"[IndiaMart refresh] Both strategies failed. Token expires_at={token.expires_at} is_valid={token.is_valid()}")
    return token.ak_token


def _determine_lead_type(c):
    remarks = (c.get('contact_type_remarks') or '').lower()
    labels = [l.lower() for l in (c.get('label_name') or [])]
    if 'missed' in labels or 'missed' in remarks:
        return 'Missed Call'
    if c.get('is_call'):
        return 'Call'
    if 'buylead' in remarks or 'buy lead' in remarks:
        return 'Buy Lead'
    return 'Enquiry'


def _parse_fb_created_time(s):
    """Parse Facebook Graph API created_time (e.g. '2026-05-26T11:32:14+0000') to naive UTC datetime."""
    if not s:
        return None
    from datetime import timezone
    try:
        dt = datetime.strptime(s, '%Y-%m-%dT%H:%M:%S%z')
        return dt.astimezone(timezone.utc).replace(tzinfo=None)
    except Exception:
        try:
            return datetime.strptime(s[:19], '%Y-%m-%dT%H:%M:%S')
        except Exception:
            return None


def _do_indiamart_sync(owner_id, created_by_id):
    """Core sync logic. Returns (new_count, skipped_count, error_msg)

    Ownership rule (2026-07-26): when a user is flagged as the
    IndiaMart default owner (`users.is_indiamart_default_owner = TRUE`),
    every newly-imported lead is assigned to THAT user regardless of who
    triggered the sync (button click OR hourly cron). If nobody carries
    the flag, we fall back to the `owner_id` argument the caller passed
    (which is currently always `None` from both call sites — leaves
    leads unowned for a manager to route manually).
    """
    from models import IndiamartToken, Lead, User
    import requests as req_lib

    token = IndiamartToken.query.first()
    if not token:
        return 0, 0, 'No IndiaMart token saved.'

    # Resolve the effective owner ONCE up front so we don't re-query per
    # contact (a busy sync page can carry 50 rows). Default-owner flag
    # wins over the passed argument — the whole point is "every lead
    # from this integration lands with the same person".
    default_owner = User.query.filter_by(is_indiamart_default_owner=True, is_active=True).first()
    if default_owner is not None:
        effective_owner_id = default_owner.id
    else:
        effective_owner_id = owner_id

    headers = {
        'user-agent': 'IndiaMart/13.6.4 (com.indiamart.m; build:4; iOS 16.6.0) Alamofire/5.0.0-rc.2',
        'accept': '*/*',
        'accept-language': 'en-US;q=1.0',
    }

    # Attempt refresh proactively (even if expired — checkAuth may still work shortly after expiry)
    AK = _indiamart_refresh_token(token, headers)

    # Re-fetch token after potential update, then check validity
    token = IndiamartToken.query.first()
    if not token.is_valid():
        return 0, 0, 'IndiaMart token has expired and could not be auto-refreshed. Please recapture from your phone.'
    GLID = token.glid or '255155317'
    MOBILE = token.mobile or '9341980003'
    USER_IP = token.user_ip or ''

    base_params = {
        'AK': AK,
        'APP_ACCURACY': '', 'APP_LATITUDE': '', 'APP_LONGITUDE': '',
        'APP_MODID': 'IOS', 'APP_SCREEN_NAME': 'Api Request',
        'APP_USER_ID': GLID, 'VALIDATION_GLID': GLID,
        'VALIDATION_USERCONTACT': MOBILE, 'VALIDATION_USER_IP': USER_IP,
        'app_version_no': '13.6.4_b_4', 'glusrid': GLID,
        'modid': 'IOS', 'q': '*', 'rows': '50',
        'token': 'imobile@15061981', 'version': '2',
    }

    new_count = 0
    skipped_count = 0
    page = 1

    while True:
        base_params['page'] = str(page)
        try:
            resp = req_lib.get(
                'https://mapi.indiamart.com/wservce/lms/v1/search',
                params=base_params, headers=headers, timeout=15
            )
            data = resp.json()
        except Exception as e:
            return new_count, skipped_count, f'API error: {str(e)}'

        if data.get('CODE') == 429:
            return new_count, skipped_count, 'Rate limited. Try again in a minute.'

        contacts = data.get('response', {}).get('contacts', [])
        app.logger.info(f"[IndiaMart sync] page={page} contacts_returned={len(contacts)} API_CODE={data.get('CODE')} API_MSG={data.get('MESSAGE','')}")
        if not contacts:
            break

        for c in contacts:
            im_id = str(c.get('im_contact_id', ''))
            if not im_id:
                continue

            if Lead.query.filter_by(indiamart_id=im_id).first():
                skipped_count += 1
                continue

            # Parse IndiaMart dates
            def _parse_im_date(s):
                try:
                    return datetime.strptime(s, '%Y-%m-%d %H:%M:%S') if s else None
                except Exception:
                    return None

            im_added = _parse_im_date(c.get('contacts_add_date'))
            lead = Lead(
                name=c.get('contacts_name') or None,
                contact=c.get('contacts_mobile1') or None,
                city=c.get('contact_city') or None,
                state=c.get('contact_state') or None,
                company=c.get('contacts_company') or None,
                product_interest=c.get('contact_last_product') or None,
                product_qty=c.get('last_product_qty') or None,
                product_category=', '.join(c.get('mcat_name') or []) or None,
                lead_type=_determine_lead_type(c),
                has_whatsapp=bool(c.get('is_whatsapp')),
                is_gst_registered=bool(c.get('is_gst')),
                is_starred=bool(c.get('is_starred_lead')),
                last_message=c.get('last_message') or None,
                unread_count=int(c.get('unread_message_cnt') or 0),
                indiamart_added_date=im_added,
                indiamart_last_contact=_parse_im_date(c.get('last_contact_date')),
                indiamart_notes=c.get('notes_v2') or None,
                indiamart_labels=', '.join(c.get('label_name') or []) or None,
                buyer_glid=c.get('contacts_glid') or None,
                is_untouched=bool(c.get('is_contact_untouched', True)),
                origin='IndiaMart',
                stage='New Lead',
                indiamart_id=im_id,
                owner_id=effective_owner_id,
                created_by=created_by_id,
                created_at=im_added or datetime.utcnow(),
            )
            db.session.add(lead)
            new_count += 1

        db.session.commit()

        if len(contacts) < 50:
            break
        page += 1

    return new_count, skipped_count, None


@app.route('/leads/indiamart/refresh-token', methods=['POST'])
@login_required
@admin_required
def indiamart_refresh_token_route():
    """Manually trigger a token refresh attempt — useful when token is expired."""
    from models import IndiamartToken
    token = IndiamartToken.query.first()
    if not token:
        return jsonify({'success': False, 'error': 'No token saved yet.'})

    headers = {
        'user-agent': 'IndiaMart/13.6.4 (com.indiamart.m; build:4; iOS 16.6.0) Alamofire/5.0.0-rc.2',
        'accept': '*/*',
        'accept-language': 'en-US;q=1.0',
    }
    old_ak = token.ak_token
    _indiamart_refresh_token(token, headers)
    token = IndiamartToken.query.first()

    if token.ak_token != old_ak:
        return jsonify({
            'success': True,
            'message': 'Token refreshed successfully!',
            'expires_at': token.expires_at.isoformat() if token.expires_at else None,
        })
    elif token.is_valid():
        return jsonify({
            'success': True,
            'message': 'Token is still valid — no refresh needed.',
            'expires_at': token.expires_at.isoformat() if token.expires_at else None,
        })
    else:
        return jsonify({
            'success': False,
            'error': 'Could not refresh token. IndiaMart rejected the request — please recapture via mitmproxy.',
        })


@app.route('/leads/indiamart/sync', methods=['GET', 'POST'])
@login_required
def indiamart_sync():
    # Manager+ by default — hammering this button by a sales user could
    # rate-limit the paid IndiaMart API and briefly break the hourly cron
    # ingest. Individual Promotors can be granted access via the per-user
    # `can_indiamart_sync` flag (User admin form) without also inheriting
    # the rest of the Manager toolkit.
    #
    # Route also accepts GET as a safe "swallow" (2026-07-26 UX fix). The
    # sync fetches a paginated IndiaMart API and can take 30-60s+ — when
    # the user hit refresh mid-request the browser re-issued the same URL
    # as GET, and Flask's POST-only rule surfaced a 405 "Method Not
    # Allowed" page. GET now just bounces back to the leads list without
    # kicking off a sync, so the refresh becomes idempotent.
    if request.method == 'GET':
        return redirect(url_for('leads_list'))
    if not current_user.can_run_indiamart_sync():
        flash('You don’t have permission to trigger IndiaMart sync. Ask an admin to enable it on your account.', 'danger')
        return redirect(url_for('leads_list'))
    new_count, skipped_count, err = _do_indiamart_sync(None, current_user.id)
    if err:
        flash(err, 'warning' if 'expired' in err.lower() or 'rate' in err.lower() else 'danger')
    else:
        flash(f'IndiaMart sync complete: {new_count} new leads imported, {skipped_count} already existed.', 'success')
    return redirect(url_for('leads_list'))


@app.route('/leads/<int:id>/note', methods=['POST'])
@login_required
def lead_save_note(id):
    """Quick note save for a lead"""
    from models import Lead, LeadHistory
    lead = Lead.query.get_or_404(id)
    note_text = request.form.get('notes', '')
    if note_text != (lead.notes or ''):
        lead.notes = note_text
        lead.is_untouched = False
        if note_text:
            db.session.add(LeadHistory(lead_id=lead.id, user_id=current_user.id,
                action='note', description=note_text))
    db.session.commit()
    return jsonify({'success': True})


@app.route('/leads/<int:id>/set-customer-type', methods=['POST'])
@login_required
def lead_set_customer_type(id):
    from models import Lead
    lead = Lead.query.get_or_404(id)
    customer_type = request.form.get('customer_type', '').strip()
    lead.customer_type = customer_type if customer_type in ('B2B', 'B2C') else None
    db.session.commit()
    return jsonify({'success': True, 'customer_type': lead.customer_type})


def _send_lead_template(lead, template, cached_media_id=None):
    """Send one approved WhatsApp template to one lead. Return (msg, ok).

    Handles the branching between plain-body templates and document-header
    templates (catalogue attached), the "no {{1}} placeholder" retry, and
    the `whatsapp_messages` audit row. Shared by both the single-lead
    (`lead_send_template`) and the bulk (`leads_bulk_send_template`)
    endpoints so the send logic stays identical everywhere.

    Args:
        lead: Lead ORM instance with .id, .name, .contact.
        template: LeadTemplate from utils.lead_templates.
        cached_media_id: optional media_id already uploaded to Meta this
            request — pass through in bulk sends so we don't re-upload the
            catalogue N times. When None and the template needs a document,
            we upload here.

    Returns:
        (WhatsAppMessage row, ok:bool). Row is db.session.added AND
        flushed; caller commits.
    """
    from models import WhatsAppMessage
    from utils.whatsapp import (
        send_template,
        send_template_with_document,
        upload_media,
        normalize_phone,
    )
    from utils.lead_templates import build_body_variables, MissingVariable

    try:
        variables = build_body_variables(template, lead)
    except MissingVariable as e:
        # The picked template needs a field the Lead doesn't carry
        # (e.g. glassy_onboarding_invite wants star_rating/listing_url,
        # which only BulkContacts have). Record a failed row for the
        # audit log and return without hitting Meta.
        msg = WhatsAppMessage(
            lead_id=lead.id,
            to_number=normalize_phone(lead.contact) or lead.contact,
            template_name=template.name,
            language=template.language,
            variables_json=None,
            sent_by=current_user.id,
            status='failed',
            error_message=f'template {template.name!r} needs {e.field} — not on Lead',
        )
        db.session.add(msg)
        db.session.flush()
        return msg, False

    msg = WhatsAppMessage(
        lead_id=lead.id,
        to_number=normalize_phone(lead.contact) or lead.contact,
        template_name=template.name,
        language=template.language,
        variables_json=json.dumps(variables),
        sent_by=current_user.id,
        status='queued',
    )
    db.session.add(msg)
    db.session.flush()

    # Document-header template (e.g. general_followup with catalogue).
    if template.needs_document:
        media_id = cached_media_id
        if not media_id:
            try:
                with open(template.document_path, 'rb') as f:
                    pdf_bytes = f.read()
            except OSError as e:
                msg.status = 'failed'
                msg.error_message = f'Catalogue file missing: {e}'
                return msg, False
            up = upload_media(pdf_bytes, template.document_filename, mime_type='application/pdf')
            if not up.get('success'):
                msg.status = 'failed'
                msg.error_message = f'Media upload failed: {up.get("error", "unknown")}'
                return msg, False
            media_id = up['media_id']

        result = send_template_with_document(
            to=lead.contact,
            template_name=template.name,
            media_id=media_id,
            filename=template.document_filename,
            language=template.language,
            body_variables=variables if variables else None,
        )
    else:
        result = send_template(
            to=lead.contact,
            template_name=template.name,
            language=template.language,
            variables=variables,
        )
        # Body-only templates: retry without vars if Meta complains the
        # template has no {{1}} slot. Only meaningful for the plain-body
        # branch — document templates always specify their header params
        # so the "no placeholder" case doesn't apply.
        if not result.get('success') and 'parameter' in (result.get('error') or '').lower():
            result = send_template(
                to=lead.contact,
                template_name=template.name,
                language=template.language,
                variables=None,
            )
            if result.get('success'):
                msg.variables_json = json.dumps([])

    if result.get('success'):
        msg.status = 'sent'
        msg.wamid = result.get('wamid')
        return msg, True

    msg.status = 'failed'
    msg.error_message = result.get('error', 'Unknown error')
    return msg, False


@app.route('/leads/<int:id>/send-template', methods=['POST'])
@login_required
def lead_send_template(id):
    """Send a BD-picked approved WhatsApp template to one lead.

    Input (POST form):
        template_name — one of the LEAD_TEMPLATES entries (welcome,
                        general_followup, ...). Defaults to 'welcome' so
                        existing single-button callers keep working.

    See utils/lead_templates.py for the registry + how to add a template.
    Document-header templates (e.g. `general_followup`) automatically
    upload the bundled catalogue PDF to Meta and reference it in the
    template's document slot.
    """
    from models import Lead
    from utils.lead_templates import get_lead_template

    lead = Lead.query.get_or_404(id)
    if not lead.contact:
        return jsonify({'success': False, 'error': 'Lead has no phone number'}), 400

    template_name = (request.form.get('template_name') or 'welcome').strip()
    template = get_lead_template(template_name)
    if not template:
        return jsonify({'success': False, 'error': f'Unknown template: {template_name!r}'}), 400

    msg, ok = _send_lead_template(lead, template)
    db.session.commit()

    if ok:
        return jsonify({'success': True, 'wamid': msg.wamid, 'to': msg.to_number, 'template': template.name})
    return jsonify({'success': False, 'error': msg.error_message, 'template': template.name}), 502


@app.route('/leads/<int:id>/send-welcome', methods=['POST'])
@login_required
def lead_send_welcome(id):
    """Legacy alias for `lead_send_template` with template_name='welcome'.

    Kept so existing JS callers (templates/leads/view.html's older
    sendWelcome() button, any bookmarks) don't 404 during the rollout.
    New UIs should POST directly to /leads/<id>/send-template with a
    template_name form field.
    """
    from models import Lead
    from utils.lead_templates import get_lead_template

    lead = Lead.query.get_or_404(id)
    if not lead.contact:
        return jsonify({'success': False, 'error': 'Lead has no phone number'}), 400

    template = get_lead_template('welcome')
    if not template:
        return jsonify({'success': False, 'error': 'Welcome template not configured'}), 500

    msg, ok = _send_lead_template(lead, template)
    db.session.commit()

    if ok:
        return jsonify({'success': True, 'wamid': msg.wamid, 'to': msg.to_number})
    return jsonify({'success': False, 'error': msg.error_message}), 502


# ── Bulk WhatsApp template send from /leads ─────────────────────────────────
# BD picks leads via the checkbox column on /leads (same cross-page selection
# infra as bulk-assign) and picks a template from the dropdown in the bulk
# bar. Same 25-per-batch cap as the bathqube bulk send. Document-header
# templates (general_followup) upload the catalogue ONCE per batch and
# reuse the media_id for every recipient — saves ~24 uploads to Meta.

MAX_BULK_WA_LEADS = 25


@app.route('/leads/bulk-send-template', methods=['POST'])
@login_required
def leads_bulk_send_template():
    """Bulk-send a picked WhatsApp template to selected leads.

    Input (POST form):
        lead_ids[]     — lead IDs picked via the checkbox column
        template_name  — one of the LEAD_TEMPLATES entries (required)

    Behaviour:
        - Cap at MAX_BULK_WA_LEADS (25) per request; truncation is flagged.
        - Skip leads with no phone number.
        - Upload the catalogue PDF ONCE per batch when the template has a
          document header — reuse the media_id for every recipient.
        - Each attempt logs to `whatsapp_messages` with sent_by=BD user
          so /leads/agent-log activity accounting stays honest.
    """
    from models import Lead
    from utils.lead_templates import get_lead_template
    from utils.whatsapp import upload_media

    template_name = (request.form.get('template_name') or '').strip()
    template = get_lead_template(template_name)
    if not template:
        return jsonify({'success': False, 'error': f'Unknown template: {template_name!r}'}), 400

    lead_ids = request.form.getlist('lead_ids[]') or request.form.getlist('lead_ids')
    if not lead_ids:
        return jsonify({'success': False, 'error': 'No leads selected.'}), 400
    try:
        lead_ids = [int(x) for x in lead_ids]
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid lead ids.'}), 400

    leads = (Lead.query
             .filter(Lead.id.in_(lead_ids))
             .limit(MAX_BULK_WA_LEADS + 1)
             .all())
    truncated = len(leads) > MAX_BULK_WA_LEADS
    leads = leads[:MAX_BULK_WA_LEADS]

    # Upload the catalogue once per batch (document templates only).
    cached_media_id = None
    if template.needs_document:
        try:
            with open(template.document_path, 'rb') as f:
                pdf_bytes = f.read()
        except OSError as e:
            return jsonify({'success': False, 'error': f'Catalogue file missing: {e}'}), 500
        up = upload_media(pdf_bytes, template.document_filename, mime_type='application/pdf')
        if not up.get('success'):
            return jsonify({'success': False, 'error': f'Media upload failed: {up.get("error", "unknown")}'}), 502
        cached_media_id = up['media_id']

    sent = 0
    failed = 0
    skipped = 0
    failures = []

    for lead in leads:
        if not lead.contact:
            skipped += 1
            failures.append({'lead_id': lead.id, 'name': lead.name, 'error': 'no phone number'})
            continue
        msg, ok = _send_lead_template(lead, template, cached_media_id=cached_media_id)
        if ok:
            sent += 1
        else:
            failed += 1
            failures.append({'lead_id': lead.id, 'name': lead.name, 'error': msg.error_message})

    db.session.commit()

    return jsonify({
        'success': True,
        'sent': sent,
        'failed': failed,
        'skipped': skipped,
        'attempted': len(leads),
        'truncated': truncated,
        'cap': MAX_BULK_WA_LEADS,
        'template': template.name,
        'failures': failures[:20],
    })


# ── Bulk Send (marketing) ────────────────────────────────────────────────
# Standalone marketing feature — BD downloads an Excel template, fills in
# name + phone rows for a campaign, uploads, selects recipients, picks a
# marketing WhatsApp template, and sends. Replies land via webhook and
# render on the contact detail page's chat panel.
#
# Contacts live in a dedicated `bulk_contacts` table (NOT `leads`) so BD's
# real CRM funnel isn't polluted by marketing lists. See models.BulkContact.

MAX_BULK_CONTACTS_UPLOAD = 1000
MAX_BULK_CONTACTS_SEND   = 25   # per-request cap; same as MAX_BULK_WA_LEADS


# ── Bulk-contact Excel columns (canonical + accepted header aliases) ──
# Two required (name + phone) so simple imports still work. The other
# six are business-directory fields the `glassy_onboarding_invite`
# template uses — populated when BD uploads a directory export.
_BULK_COL_ALIASES = {
    'name':         ('name', 'contact name', 'contact', 'customer', 'business', 'business name', 'business (category)'),
    'phone':        ('phone', 'mobile', 'number', 'contact number', 'whatsapp', 'whatsapp number', 'mobile number'),
    'star_rating':  ('star_rating', 'star', 'stars', 'rating', 'google rating', 'star rating'),
    'listing_url':  ('listing_url', 'listing', 'glassy listing url', 'glassy listing', 'glassy url', 'directory url'),
    'category':     ('category', 'business category', 'type'),
    'location':     ('location', 'city', 'area', 'address'),
    'reviews_count':('reviews_count', 'reviews', 'review count', 'number of reviews'),
    'website':      ('website', 'url', 'site', 'business website'),
}


def _bulk_contact_excel_template_bytes():
    """Return an .xlsx workbook with the full column set (name, phone,
    star_rating, listing_url + 4 optional business fields) and two
    sample rows. BD downloads this, fills it in, uploads back."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='0B8B3A')
    center      = Alignment(horizontal='center')

    headers = ['name', 'phone', 'star_rating', 'listing_url',
               'category', 'location', 'reviews_count', 'website']
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Sample rows modeled on the Glassy India directory shape BD uploads.
    ws.append([
        'THE MARINA MALL CHENNAI', '+91 44 4017 3017', 4.4,
        'https://glassy.in/glass-business/the-marina-mall-chennai/',
        'shop/dealer', 'Old Mahabalipuram Road, Chennai', 31746,
        'marinamallchennai.com',
    ])
    ws.append([
        'Livspace Home', '+91 80 4697 1177', 4.7,
        'https://glassy.in/glass-business/livspace-home-yousuf/',
        'interior designer', 'Marathahalli, Bengaluru', 16201,
        'www.livspace.com/in/interior-design',
    ])

    widths = [30, 22, 12, 60, 22, 30, 12, 34]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _split_name_and_category(raw_name: str):
    """Column B in the Glassy directory sheet packs 'Business — category'
    (em-dash-space or hyphen-space). If BD's upload preserves that
    formatting, split it so contact.name = business, contact.category =
    category. Returns (name, category_or_None). No-op when no separator
    is present."""
    if not raw_name:
        return raw_name, None
    for sep in (' — ', ' – ', ' - '):
        if sep in raw_name:
            left, right = raw_name.split(sep, 1)
            left = left.strip()
            right = right.strip()
            if left and right and len(right) < 80:
                return left, right
    return raw_name.strip(), None


def _match_column(headers: list[str], canonical: str):
    """Return the column index for `canonical` given a list of lowercased
    header strings. Uses the alias table; returns None on no match."""
    aliases = _BULK_COL_ALIASES.get(canonical, (canonical,))
    for i, h in enumerate(headers):
        if h in aliases:
            return i
    return None


def _parse_bulk_contacts_upload(file_storage):
    """Read an uploaded .xlsx or .csv into a list of contact-shaped
    dicts. Two required columns (name + phone); when present, the
    remaining business-directory columns (star_rating, listing_url,
    category, location, reviews_count, website) are also captured.

    Returns (rows, warnings). `warnings` are per-row strings the
    import screen surfaces to BD ("row 47 skipped — missing phone")."""
    import openpyxl
    from utils.whatsapp import normalize_phone

    filename = (file_storage.filename or '').lower()
    rows = []
    warnings = []

    def _as_float(v):
        if v is None or v == '':
            return None
        try:
            return float(str(v).strip())
        except (TypeError, ValueError):
            return None

    def _as_int(v):
        if v is None or v == '':
            return None
        try:
            return int(float(str(v).strip().replace(',', '')))
        except (TypeError, ValueError):
            return None

    def _push(idx, values):
        name, phone_raw = values.get('name'), values.get('phone')
        # Blank row — silently skip
        if not name and not phone_raw:
            return
        if not name:
            warnings.append(f'row {idx}: skipped (name missing)')
            return
        if not phone_raw:
            warnings.append(f'row {idx}: skipped (phone missing)')
            return
        normalized = normalize_phone(str(phone_raw))
        if not normalized or len(normalized) < 10:
            warnings.append(f'row {idx}: skipped (phone {phone_raw!r} invalid)')
            return
        # Auto-split "Business — category" (Glassy directory shape).
        split_name, split_cat = _split_name_and_category(str(name))
        rows.append({
            'name': split_name,
            'phone': normalized,
            'star_rating': _as_float(values.get('star_rating')),
            'listing_url': (str(values['listing_url']).strip() if values.get('listing_url') else None),
            'business_category': (str(values.get('category')).strip() if values.get('category')
                                   else split_cat),
            'location': (str(values['location']).strip() if values.get('location') else None),
            'reviews_count': _as_int(values.get('reviews_count')),
            'website': (str(values['website']).strip() if values.get('website') else None),
        })

    if filename.endswith('.csv'):
        import csv as _csv
        stream = io.TextIOWrapper(file_storage.stream, encoding='utf-8', errors='replace')
        reader = _csv.DictReader(stream)
        for idx, row in enumerate(reader, start=2):
            keys = {(k or '').strip().lower(): v for k, v in row.items() if k}
            values = {}
            for canonical in _BULK_COL_ALIASES:
                for alias in _BULK_COL_ALIASES[canonical]:
                    if alias in keys and keys[alias] not in (None, ''):
                        values[canonical] = keys[alias]
                        break
            _push(idx, values)
    else:
        try:
            wb = openpyxl.load_workbook(file_storage, data_only=True)
        except Exception as e:
            raise ValueError(f'Could not read Excel file: {e}')
        ws = wb.active
        headers = [str(cell.value or '').strip().lower() for cell in ws[1]]
        idx_map = {c: _match_column(headers, c) for c in _BULK_COL_ALIASES}
        if idx_map.get('name') is None or idx_map.get('phone') is None:
            raise ValueError(
                'Expected columns "name" and "phone" in the header row. '
                'The download template has these plus optional star_rating, '
                'listing_url, category, location, reviews_count, website.'
            )
        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            values = {}
            for canonical, col_idx in idx_map.items():
                if col_idx is not None and col_idx < len(row):
                    v = row[col_idx].value
                    if v not in (None, ''):
                        values[canonical] = v
            _push(r_idx, values)

    return rows, warnings


@app.route('/bulk-send')
@login_required
def bulk_send_list():
    """Bulk Send home — list all imported contacts + upload form +
    template picker + Send button."""
    from models import BulkContact
    search   = (request.args.get('search') or '').strip()
    campaign = (request.args.get('campaign') or '').strip()
    show_opted = request.args.get('show_opted') == '1'

    q = BulkContact.query
    if not show_opted:
        q = q.filter(BulkContact.is_opted_out == False)  # noqa: E712
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            BulkContact.name.ilike(like),
            BulkContact.phone.ilike(like),
        ))
    if campaign:
        q = q.filter(BulkContact.campaign == campaign)
    contacts = q.order_by(BulkContact.id.desc()).limit(2000).all()

    # Campaign chip list — distinct non-null values, sorted.
    campaigns = [r[0] for r in db.session.query(BulkContact.campaign)
                                          .filter(BulkContact.campaign.isnot(None))
                                          .distinct()
                                          .order_by(BulkContact.campaign)
                                          .all()]
    total = BulkContact.query.count()
    opted_out_count = BulkContact.query.filter_by(is_opted_out=True).count()

    return render_template(
        'bulk_send/list.html',
        contacts=contacts, campaigns=campaigns,
        search=search, campaign_filter=campaign, show_opted=show_opted,
        total=total, opted_out_count=opted_out_count,
    )


@app.route('/bulk-send/template.xlsx')
@login_required
def bulk_send_template_download():
    """Serve the blank Excel template BD fills in and re-uploads."""
    from flask import send_file
    data = _bulk_contact_excel_template_bytes()
    return send_file(
        io.BytesIO(data),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='bulk-send-template.xlsx',
    )


@app.route('/bulk-send/import', methods=['POST'])
@login_required
def bulk_send_import():
    """Upload an Excel/CSV of {name, phone} rows. Tag them with the
    campaign name supplied in the form; dedupe by phone (updates the
    campaign tag on collision so BD can re-import to re-tag)."""
    from models import BulkContact

    f = request.files.get('file')
    if not f or not f.filename:
        flash('Please attach an Excel or CSV file.', 'warning')
        return redirect(url_for('bulk_send_list'))

    campaign = (request.form.get('campaign') or '').strip()[:100] or None

    try:
        rows, warnings = _parse_bulk_contacts_upload(f)
    except ValueError as e:
        flash(str(e), 'danger')
        return redirect(url_for('bulk_send_list'))

    if not rows:
        flash('No valid rows found. Check the file and try again.', 'warning')
        return redirect(url_for('bulk_send_list'))

    if len(rows) > MAX_BULK_CONTACTS_UPLOAD:
        flash(
            f'File has {len(rows)} rows; only the first {MAX_BULK_CONTACTS_UPLOAD} '
            f'will be imported per upload. Split into smaller files if you need more.',
            'warning',
        )
        rows = rows[:MAX_BULK_CONTACTS_UPLOAD]

    # In-file dedup by phone.
    seen_phones = set()
    deduped = []
    for r in rows:
        if r['phone'] in seen_phones:
            continue
        seen_phones.add(r['phone'])
        deduped.append(r)

    # Match against existing contacts — update campaign tag on collision.
    existing = {c.phone: c for c in BulkContact.query.filter(
        BulkContact.phone.in_(list(seen_phones))
    ).all()}

    inserted = 0
    updated  = 0
    for r in deduped:
        if r['phone'] in existing:
            row = existing[r['phone']]
            row_dirty = False
            if campaign and row.campaign != campaign:
                row.campaign = campaign
                row_dirty = True
            # Enrich empty business-directory fields on collision so a
            # re-import can top-up ratings / listing URLs without wiping
            # anything BD may have manually edited.
            for field in ('star_rating', 'listing_url', 'business_category',
                          'location', 'reviews_count', 'website'):
                incoming = r.get(field)
                if incoming is not None and getattr(row, field) in (None, ''):
                    setattr(row, field, incoming)
                    row_dirty = True
            if row_dirty:
                updated += 1
        else:
            db.session.add(BulkContact(
                name=r['name'],
                phone=r['phone'],
                campaign=campaign,
                imported_by=current_user.id,
                star_rating=r.get('star_rating'),
                listing_url=r.get('listing_url'),
                business_category=r.get('business_category'),
                location=r.get('location'),
                reviews_count=r.get('reviews_count'),
                website=r.get('website'),
            ))
            inserted += 1
    db.session.commit()

    msg_bits = []
    if inserted: msg_bits.append(f'{inserted} imported')
    if updated:  msg_bits.append(f'{updated} campaign-tag updated')
    if warnings: msg_bits.append(f'{len(warnings)} skipped')
    flash(' · '.join(msg_bits) or 'Nothing to import.', 'success' if inserted else 'info')
    if warnings:
        # Surface the first 10 warnings so BD can spot systematic issues
        # without an infinite flash message.
        for w in warnings[:10]:
            flash(w, 'warning')

    return redirect(url_for('bulk_send_list', campaign=campaign or ''))


@app.route('/bulk-send/<int:id>')
@login_required
def bulk_send_view(id):
    """One bulk contact + chat panel (same layout as leads view)."""
    from models import BulkContact
    contact = BulkContact.query.get_or_404(id)
    return render_template('bulk_send/view.html', contact=contact)


@app.route('/bulk-send/<int:id>/toggle-opt-out', methods=['POST'])
@login_required
def bulk_send_toggle_opt_out(id):
    """Manual opt-out flip — BD can undo a false-positive auto-opt-out
    or manually opt someone out from their profile page."""
    from models import BulkContact
    contact = BulkContact.query.get_or_404(id)
    contact.is_opted_out = not contact.is_opted_out
    db.session.commit()
    flash(
        f'{contact.name} — {"opted out" if contact.is_opted_out else "opt-out cleared"}.',
        'info',
    )
    return redirect(url_for('bulk_send_view', id=id))


@app.route('/bulk-send/glassy-reminders/run', methods=['POST'])
@admin_required
def bulk_send_glassy_reminders_run():
    """Fire the 4-day-no-reply reminder to Glassy directory contacts.

    Sends `glassy_onboarding_reminder` (template on the Vtspl WABA) to
    every BulkContact that:
      - received `glassy_onboarding_invite` at least 4 days ago
      - has never replied to us
      - has NOT already received the reminder (send-once guarantee)
      - is not opted out
      - has a non-empty listing_url

    Intended to run daily from an external scheduler (EventBridge cron,
    Zapier, or a curl loop). BD can also click it manually from the
    Bulk Send screen — the sends and skips are itemised in the JSON
    response so it's obvious what happened.

    Response shape:
        {
          "success": true,
          "eligible": <int>,
          "sent": <int>,
          "failed": <int>,
          "skipped_missing_url": <int>,
          "failures": [{contact_id, name, error}]
        }
    """
    from models import BulkContact, WhatsAppMessage
    from utils.whatsapp import send_template, normalize_phone
    from datetime import datetime as _dt, timedelta as _td

    cutoff = _dt.utcnow() - _td(days=4)

    # 4+ days old outbound invites that had no reply AND no reminder yet.
    # Two NOT EXISTS clauses stay in SQL so the whole eligibility check
    # is one query — the 300-contact scale doesn't need optimisation
    # gymnastics, but the pattern generalises if the list grows.
    invite_alias = db.aliased(WhatsAppMessage)
    reply_alias  = db.aliased(WhatsAppMessage)
    reminder_alias = db.aliased(WhatsAppMessage)

    eligible = (
        db.session.query(BulkContact, invite_alias.sent_at.label('invite_sent_at'))
        .join(invite_alias, invite_alias.bulk_contact_id == BulkContact.id)
        .filter(invite_alias.template_name == 'glassy_onboarding_invite')
        .filter(invite_alias.direction == 'out')
        .filter(invite_alias.sent_at <= cutoff)
        .filter(BulkContact.is_opted_out == False)  # noqa: E712
        .filter(~db.exists().where(
            (reply_alias.bulk_contact_id == BulkContact.id) &
            (reply_alias.direction == 'in')
        ))
        .filter(~db.exists().where(
            (reminder_alias.bulk_contact_id == BulkContact.id) &
            (reminder_alias.template_name == 'glassy_onboarding_reminder') &
            (reminder_alias.direction == 'out')
        ))
        .all()
    )

    sent = 0
    failed = 0
    skipped_missing_url = 0
    failures = []

    for contact, invite_sent_at in eligible:
        listing_url = (getattr(contact, 'listing_url', '') or '').strip()
        if not listing_url:
            skipped_missing_url += 1
            continue

        to_number = normalize_phone(contact.phone) or contact.phone
        variables = [listing_url]

        msg_row = WhatsAppMessage(
            bulk_contact_id=contact.id,
            direction='out',
            to_number=to_number,
            template_name='glassy_onboarding_reminder',
            language='en',
            variables_json=json.dumps(variables),
            sent_by=current_user.id,
            status='queued',
        )
        db.session.add(msg_row)
        db.session.flush()

        result = send_template(
            to=contact.phone,
            template_name='glassy_onboarding_reminder',
            language='en',
            variables=variables,
        )
        if result.get('success'):
            msg_row.wamid = result.get('wamid')
            msg_row.status = 'sent'
            sent += 1
        else:
            msg_row.status = 'failed'
            msg_row.error_message = (result.get('error') or 'unknown')[:1000]
            failed += 1
            failures.append({
                'contact_id': contact.id,
                'name': contact.name,
                'error': result.get('error') or 'unknown',
            })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[glassy-reminders] commit failed: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

    app.logger.info(
        f'[glassy-reminders] eligible={len(eligible)} sent={sent} '
        f'failed={failed} skipped_missing_url={skipped_missing_url}'
    )
    return jsonify({
        'success': True,
        'eligible': len(eligible),
        'sent': sent,
        'failed': failed,
        'skipped_missing_url': skipped_missing_url,
        'failures': failures[:50],
    })


@app.route('/bulk-send/bulk-send-template', methods=['POST'])
@login_required
def bulk_send_bulk_template():
    """Send a picked WhatsApp template to selected bulk contacts.
    Mirrors leads_bulk_send_template but works on the bulk_contacts
    table and skips opted-out rows."""
    from models import BulkContact, WhatsAppMessage
    from utils.lead_templates import (
        get_lead_template, build_body_variables, MissingVariable,
    )
    from utils.whatsapp import send_template, send_template_with_document, upload_media, normalize_phone

    template_name = (request.form.get('template_name') or '').strip()
    template = get_lead_template(template_name)
    if not template:
        return jsonify({'success': False, 'error': f'Unknown template: {template_name!r}'}), 400

    contact_ids = request.form.getlist('contact_ids[]') or request.form.getlist('contact_ids')
    if not contact_ids:
        return jsonify({'success': False, 'error': 'No contacts selected.'}), 400
    try:
        contact_ids = [int(x) for x in contact_ids]
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid contact ids.'}), 400

    contacts = (BulkContact.query
                .filter(BulkContact.id.in_(contact_ids))
                .filter(BulkContact.is_opted_out == False)  # noqa: E712
                .limit(MAX_BULK_CONTACTS_SEND + 1)
                .all())
    truncated = len(contacts) > MAX_BULK_CONTACTS_SEND
    contacts = contacts[:MAX_BULK_CONTACTS_SEND]

    # If the template needs a document header, upload once per batch.
    cached_media_id = None
    if template.needs_document:
        try:
            with open(template.document_path, 'rb') as fh:
                pdf_bytes = fh.read()
        except OSError as e:
            return jsonify({'success': False, 'error': f'Document missing: {e}'}), 500
        up = upload_media(pdf_bytes, template.document_filename, mime_type='application/pdf')
        if not up.get('success'):
            return jsonify({'success': False, 'error': f'Media upload failed: {up.get("error", "unknown")}'}), 502
        cached_media_id = up['media_id']

    sent = 0
    failed = 0
    failures = []

    skipped = 0
    for contact in contacts:
        # Build body variables for this contact. Templates with a
        # per-template var_builder read business-directory fields
        # straight off the BulkContact — we don't need a shim, the
        # BulkContact already exposes .name / .star_rating / .listing_url.
        # MissingVariable raised by the builder → skip this row with a
        # per-row failure entry so BD sees WHY (missing field name).
        try:
            variables = build_body_variables(template, contact)
        except MissingVariable as e:
            skipped += 1
            failures.append({
                'contact_id': contact.id,
                'name': contact.name,
                'error': f'missing {e.field}',
            })
            continue

        msg_row = WhatsAppMessage(
            bulk_contact_id=contact.id,
            direction='out',
            to_number=normalize_phone(contact.phone) or contact.phone,
            template_name=template.name,
            language=template.language,
            variables_json=json.dumps(variables),
            sent_by=current_user.id,
            status='queued',
        )
        db.session.add(msg_row)
        db.session.flush()

        if template.needs_document:
            result = send_template_with_document(
                to=contact.phone, template_name=template.name,
                media_id=cached_media_id, filename=template.document_filename,
                language=template.language,
                body_variables=variables if variables else None,
            )
        else:
            result = send_template(
                to=contact.phone, template_name=template.name,
                language=template.language, variables=variables,
            )
            # Retry sans-vars when Meta rejects with a param mismatch —
            # same pattern as _send_lead_template.
            if not result.get('success') and 'parameter' in (result.get('error') or '').lower():
                result = send_template(
                    to=contact.phone, template_name=template.name,
                    language=template.language, variables=None,
                )
                if result.get('success'):
                    msg_row.variables_json = json.dumps([])

        if result.get('success'):
            msg_row.status = 'sent'
            msg_row.wamid = result.get('wamid')
            sent += 1
        else:
            msg_row.status = 'failed'
            msg_row.error_message = result.get('error', 'Unknown error')
            failed += 1
            failures.append({
                'contact_id': contact.id,
                'name': contact.name,
                'error': msg_row.error_message,
            })

    db.session.commit()

    return jsonify({
        'success': True,
        'sent': sent, 'failed': failed, 'skipped': skipped,
        'attempted': len(contacts),
        'truncated': truncated,
        'cap': MAX_BULK_CONTACTS_SEND,
        'template': template.name,
        'failures': failures[:20],
    })


# ── WhatsApp inbound webhook ──────────────────────────────────────────────
# Meta's WhatsApp Cloud API POSTs here for every event we're subscribed to
# on the WABA config screen — customer replies (messages), delivery /
# read receipts (statuses), and (rarely) template quality updates.
#
# Two request patterns land on this URL:
#
#   GET  /api/whatsapp/webhook  ?hub.mode=subscribe
#                               &hub.verify_token=<t>
#                               &hub.challenge=<c>
#     — Meta's one-time subscription challenge. Echoes hub.challenge
#       when hub.verify_token matches WHATSAPP_WEBHOOK_VERIFY_TOKEN.
#
#   POST /api/whatsapp/webhook  (JSON body signed via X-Hub-Signature-256)
#     — real events. Signature is verified against WHATSAPP_APP_SECRET
#       (falling back to FB_APP_SECRET, which is the same value when the
#       WABA + Facebook page share one Meta app — the usual setup).
#
# Meta expects a 200 within seconds regardless of internal errors,
# otherwise it retries with backoff — so parsing bugs must not raise
# out. We catch + log and still return 200.

def _whatsapp_app_secret():
    """Return the shared secret Meta signs webhook payloads with.
    Prefer the WhatsApp-specific env var, fall back to the Facebook
    app secret (usually the same value — Meta apps have one App Secret
    for both surfaces)."""
    return (os.getenv('WHATSAPP_APP_SECRET')
            or os.getenv('FB_APP_SECRET')
            or '').strip()


def _verify_meta_signature(raw_body: bytes, signature_header: str) -> bool:
    """Constant-time HMAC-SHA256 verify of X-Hub-Signature-256.
    Header shape is 'sha256=<hex>'. Returns False on any mismatch or
    missing secret."""
    import hmac
    import hashlib
    secret = _whatsapp_app_secret()
    if not secret or not signature_header:
        return False
    if not signature_header.startswith('sha256='):
        return False
    expected = hmac.new(secret.encode(), raw_body, hashlib.sha256).hexdigest()
    supplied = signature_header.split('=', 1)[1]
    return hmac.compare_digest(expected, supplied)


def _lookup_lead_by_phone(phone: str):
    """Match Meta's `from` (E.164 digits, no '+') to a Lead by contact.
    Normalises the DB side the same way `normalize_phone` does at send
    time — Indian 10-digit numbers get '91' prepended, everything else
    stays as-is. Returns the newest matching lead when multiple share
    a phone, or None."""
    from models import Lead
    from utils.whatsapp import normalize_phone
    normalized = normalize_phone(phone)
    if not normalized:
        return None
    # Query for any Lead whose normalized contact matches.
    #   contact could have been stored as raw '9876543210', '+91 9876543210',
    #   '919876543210', etc — we compare against both the raw and the
    #   normalized form.
    #
    # PostgreSQL: regexp_replace strips everything non-digit; then we
    # optionally prepend '91' when the resulting string is 10 chars long.
    # Everything happens server-side so no full-table scan in Python.
    rows = db.session.execute(
        db.text("""
            SELECT id
              FROM leads
             WHERE contact IS NOT NULL
               AND (
                     regexp_replace(contact, '[^0-9]', '', 'g') = :n
                  OR '91' || regexp_replace(contact, '[^0-9]', '', 'g') = :n
                  OR regexp_replace(contact, '[^0-9]', '', 'g') = substr(:n, 3)
                   )
             ORDER BY id DESC
             LIMIT 1
        """),
        {'n': normalized},
    ).first()
    if rows is None:
        return None
    return Lead.query.get(rows[0])


def _lookup_bulk_contact_by_phone(phone: str):
    """Match Meta's `from` to a BulkContact by phone. Runs AFTER the
    Lead lookup — if a phone lives in both tables (BD imported someone
    who's already a lead), Leads win because they carry more context
    (stage, owner, product interest). Returns the newest matching
    BulkContact or None."""
    from models import BulkContact
    from utils.whatsapp import normalize_phone
    normalized = normalize_phone(phone)
    if not normalized:
        return None
    rows = db.session.execute(
        db.text("""
            SELECT id
              FROM bulk_contacts
             WHERE phone IS NOT NULL
               AND (
                     regexp_replace(phone, '[^0-9]', '', 'g') = :n
                  OR '91' || regexp_replace(phone, '[^0-9]', '', 'g') = :n
                  OR regexp_replace(phone, '[^0-9]', '', 'g') = substr(:n, 3)
                   )
             ORDER BY id DESC
             LIMIT 1
        """),
        {'n': normalized},
    ).first()
    if rows is None:
        return None
    return BulkContact.query.get(rows[0])


# STOP-word patterns for auto opt-out on marketing replies. Case-insensitive,
# whole-word match on the first token of the body — so a friendly reply
# containing the word "stop" mid-sentence doesn't auto-opt-out somebody
# who's actually engaged. See _handle_inbound_message.
_STOP_WORDS = {'stop', 'unsubscribe', 'remove', 'optout', 'opt-out', 'end'}


def _download_meta_media(media_id: str):
    """Fetch the Meta media-download URL for a given media id, using the
    WhatsApp Cloud API token. We only STORE the URL (short-lived, ~5 min)
    for phase-1; a follow-up patch can copy the bytes into S3 for
    permanence. Returns (url, mime) or (None, None) on failure."""
    token = os.getenv('WHATSAPP_TOKEN', '').strip()
    api_ver = os.getenv('WHATSAPP_API_VERSION', 'v21.0')
    if not token or not media_id:
        return None, None
    try:
        r = requests.get(
            f'https://graph.facebook.com/{api_ver}/{media_id}',
            headers={'Authorization': f'Bearer {token}'},
            timeout=6,
        )
        if r.status_code != 200:
            return None, None
        j = r.json() or {}
        return j.get('url'), j.get('mime_type')
    except Exception as e:
        app.logger.error(f'[whatsapp-webhook] media lookup failed: {e}')
        return None, None


def _handle_inbound_message(msg: dict, meta_from: str, contact_name: str = None):
    """Persist one inbound `message` object from Meta's webhook payload.
    Idempotent on `wamid` — if Meta re-delivers we no-op."""
    from models import WhatsAppMessage
    from datetime import datetime as _dt
    wamid = msg.get('id')
    if not wamid:
        return
    # Idempotent guard
    existing = WhatsAppMessage.query.filter_by(wamid=wamid).first()
    if existing:
        return

    msg_type = msg.get('type') or 'text'
    body_text = None
    media_url = None
    media_mime = None
    media_caption = None

    if msg_type == 'text':
        body_text = (msg.get('text') or {}).get('body')
    elif msg_type in ('image', 'video', 'audio', 'document', 'sticker'):
        obj = msg.get(msg_type) or {}
        media_id = obj.get('id')
        media_caption = obj.get('caption')
        media_url, media_mime = _download_meta_media(media_id)
    elif msg_type == 'button':
        body_text = (msg.get('button') or {}).get('text')
    elif msg_type == 'interactive':
        # Interactive replies (button + list). Meta packs the reply
        # payload under `interactive.button_reply.title` etc.
        inter = msg.get('interactive') or {}
        for sub in ('button_reply', 'list_reply'):
            v = inter.get(sub) or {}
            if v.get('title'):
                body_text = v['title']
                break
    else:
        body_text = f'[{msg_type} message]'

    # Meta timestamp is unix seconds as a string.
    ts_str = msg.get('timestamp')
    try:
        recv_at = _dt.utcfromtimestamp(int(ts_str)) if ts_str else _dt.utcnow()
    except (TypeError, ValueError):
        recv_at = _dt.utcnow()

    # Route the reply to whatever context the LAST outbound to this
    # phone was in — the reply is a continuation of that conversation,
    # not an unrelated event. If the last outbound was a bulk-send, the
    # reply attaches to the BulkContact; if it was a lead-send, to the
    # Lead. Only when there's no prior outbound do we fall back to the
    # look-up-Lead-first priority.
    #
    # Prior design (Leads win unconditionally) broke Bulk Send when
    # BD imported existing customers — replies to marketing blasts
    # would silently attach to the lead's chat panel and never show
    # up on the bulk contact.
    from models import WhatsAppMessage
    from utils.whatsapp import normalize_phone
    lead = None
    bulk_contact = None
    normalized_from = normalize_phone(meta_from)
    if normalized_from:
        last_out = (WhatsAppMessage.query
                    .filter(WhatsAppMessage.direction == 'out')
                    .filter(WhatsAppMessage.to_number == normalized_from)
                    .filter(db.or_(
                        WhatsAppMessage.bulk_contact_id.isnot(None),
                        WhatsAppMessage.lead_id.isnot(None),
                    ))
                    .order_by(WhatsAppMessage.sent_at.desc())
                    .first())
        if last_out:
            if last_out.bulk_contact_id:
                from models import BulkContact
                bulk_contact = BulkContact.query.get(last_out.bulk_contact_id)
            elif last_out.lead_id:
                from models import Lead
                lead = Lead.query.get(last_out.lead_id)
    # Fallback when there's no prior outbound to this phone —
    # match by contact/phone across leads first, then bulk_contacts.
    if not lead and not bulk_contact:
        lead = _lookup_lead_by_phone(meta_from)
        bulk_contact = None if lead else _lookup_bulk_contact_by_phone(meta_from)

    # Auto opt-out: if this is a plain-text reply from a bulk marketing
    # contact and the first token matches a STOP-word, flip the
    # contact's is_opted_out flag so future bulk-sends skip them.
    # We only auto-flip for BulkContacts — Leads have real BD context
    # and shouldn't be silently muted by a stray "stop".
    if bulk_contact and body_text:
        first = (body_text.strip().split() or [''])[0].lower().strip('.,!?')
        if first in _STOP_WORDS and not bulk_contact.is_opted_out:
            bulk_contact.is_opted_out = True
            app.logger.info(
                f'[whatsapp-webhook] auto-opted-out bulk contact {bulk_contact.id} '
                f'({bulk_contact.phone}) — first word {first!r}'
            )

    row = WhatsAppMessage(
        lead_id=lead.id if lead else None,
        bulk_contact_id=bulk_contact.id if bulk_contact else None,
        direction='in',
        from_number=meta_from,
        body_text=body_text,
        media_url=media_url,
        media_mime=media_mime,
        media_caption=media_caption,
        received_at=recv_at,
        sent_at=recv_at,   # so ORDER BY sent_at gives a chronological chat
        wamid=wamid,
        status='received',
        # sent_by is nullable for inbound
    )
    db.session.add(row)

    # ── Auto-response: Glassy directory "Yes" quick-reply ─────────────
    # When an imported bulk contact taps "Yes" on the
    # glassy_onboarding_invite template, fire the follow-up
    # glassy_onboarding_yes_reply carrying their personalised website
    # snippet (listing_url as {{1}}). Fires once per inbound "Yes";
    # subsequent taps by the same contact are ignored (any prior
    # outbound of the same template_name to them counts as "already
    # responded"). Failures are logged, never raised — the primary
    # inbound persistence must not depend on the reply landing.
    try:
        _maybe_auto_reply_glassy_yes(msg, body_text, bulk_contact)
    except Exception as e:
        app.logger.error(f'[whatsapp-webhook] glassy auto-reply failed: {e}')


def _maybe_auto_reply_glassy_yes(msg: dict, body_text: str, bulk_contact) -> None:
    """Send glassy_onboarding_yes_reply if the inbound is a `Yes` button
    tap on a glassy_onboarding_invite send. No-op otherwise."""
    from models import WhatsAppMessage, BulkContact  # noqa: F401
    from utils.whatsapp import send_template, normalize_phone

    if bulk_contact is None or not body_text:
        return
    # Only fire on button-tap-style replies with title "Yes" (Meta packs
    # the tap under interactive.button_reply.title; the parent handler
    # already lifts it into body_text).
    if body_text.strip().lower() != 'yes':
        return
    # Confirm the reply is CONTEXTUAL to a previous glassy_onboarding_invite
    # send — otherwise a naked "Yes" from a contact who never got the
    # invite could trigger a phantom follow-up. Meta gives us the source
    # wamid on msg.context.id.
    context_wamid = ((msg.get('context') or {}).get('id') or '').strip()
    if not context_wamid:
        return
    parent = WhatsAppMessage.query.filter_by(wamid=context_wamid, direction='out').first()
    if not parent or parent.template_name != 'glassy_onboarding_invite':
        return

    # Guard against double-sends — if we already fired the follow-up to
    # this contact, don't fire again (they may tap Yes twice).
    already = (WhatsAppMessage.query
               .filter_by(bulk_contact_id=bulk_contact.id,
                          template_name='glassy_onboarding_yes_reply',
                          direction='out')
               .first())
    if already:
        app.logger.info(
            f'[whatsapp-webhook] glassy Yes-reply already sent to bulk_contact '
            f'{bulk_contact.id}; skipping duplicate'
        )
        return

    listing_url = (getattr(bulk_contact, 'listing_url', '') or '').strip()
    if not listing_url:
        app.logger.warning(
            f'[whatsapp-webhook] bulk_contact {bulk_contact.id} tapped Yes but '
            f'has no listing_url — cannot personalise the reply, skipping'
        )
        return

    to_number = normalize_phone(bulk_contact.phone) or bulk_contact.phone
    variables = [listing_url]

    # Log-first so the row exists even if send_template raises. The
    # outer webhook route commits after we return; a failed send just
    # leaves a 'failed' status.
    out_row = WhatsAppMessage(
        bulk_contact_id=bulk_contact.id,
        direction='out',
        to_number=to_number,
        template_name='glassy_onboarding_yes_reply',
        language='en',
        variables_json=json.dumps(variables),
        # sent_by is nullable — this is a system-triggered send, not
        # attributable to any specific BD user.
        status='queued',
    )
    db.session.add(out_row)
    db.session.flush()

    result = send_template(
        to=bulk_contact.phone,
        template_name='glassy_onboarding_yes_reply',
        language='en',
        variables=variables,
    )
    if result.get('success'):
        out_row.wamid = result.get('wamid')
        out_row.status = 'sent'
        app.logger.info(
            f'[whatsapp-webhook] auto-sent glassy_onboarding_yes_reply to '
            f'bulk_contact {bulk_contact.id} ({to_number})'
        )
    else:
        out_row.status = 'failed'
        out_row.error_message = (result.get('error') or 'unknown')[:1000]
        app.logger.error(
            f'[whatsapp-webhook] glassy_onboarding_yes_reply send failed for '
            f'bulk_contact {bulk_contact.id}: {result.get("error")}'
        )


def _handle_status_update(st: dict):
    """Update an outbound row's status from a Meta status event."""
    from models import WhatsAppMessage
    wamid = st.get('id')
    status = st.get('status')  # sent | delivered | read | failed
    if not wamid or not status:
        return
    row = WhatsAppMessage.query.filter_by(wamid=wamid).first()
    if not row:
        return  # a status event for a message we don't have — ignore
    # Don't move backwards: read > delivered > sent > queued.
    order = {'queued': 0, 'sent': 1, 'delivered': 2, 'read': 3, 'failed': 9}
    if order.get(status, -1) < order.get(row.status, -1):
        return
    row.status = status
    if status == 'failed':
        errs = st.get('errors') or []
        if errs:
            row.error_message = (errs[0].get('title') or errs[0].get('message')
                                  or 'Meta reported failure')


@app.route('/api/whatsapp/webhook', methods=['GET', 'POST'])
def whatsapp_webhook():
    """Meta WhatsApp Cloud API webhook receiver.

    GET  → subscription challenge; echo `hub.challenge` when the token
           matches WHATSAPP_WEBHOOK_VERIFY_TOKEN.
    POST → real events. Verify HMAC-SHA256 signature, parse
           entry[].changes[].value, dispatch to _handle_inbound_message /
           _handle_status_update. Always returns 200 (Meta retries on any
           non-2xx, which just multiplies our error rate)."""
    if request.method == 'GET':
        # Meta subscription challenge
        mode      = request.args.get('hub.mode')
        token     = request.args.get('hub.verify_token')
        challenge = request.args.get('hub.challenge')
        expected  = os.getenv('WHATSAPP_WEBHOOK_VERIFY_TOKEN', '').strip()
        if mode == 'subscribe' and expected and token == expected:
            # Meta wants a plain-text response with just the challenge.
            return challenge or '', 200
        return 'forbidden', 403

    # POST — event delivery
    raw = request.get_data() or b''
    sig = request.headers.get('X-Hub-Signature-256', '')

    if not _verify_meta_signature(raw, sig):
        # Don't 401 in production — Meta backs off aggressively on 4xx.
        # Log and 200. If someone's spoofing, the signature check has
        # already stopped them from writing anything.
        app.logger.warning('[whatsapp-webhook] signature check failed — dropping event')
        return '', 200

    try:
        payload = request.get_json(silent=True) or {}
    except Exception:
        payload = {}

    try:
        for entry in payload.get('entry', []) or []:
            for change in entry.get('changes', []) or []:
                value = change.get('value') or {}
                # Contact block gives us `wa_id` (phone) + optional name;
                # we only need the wa_id since the message.from carries it too.
                contacts = value.get('contacts') or []
                contact_name = None
                if contacts and isinstance(contacts[0], dict):
                    contact_name = (contacts[0].get('profile') or {}).get('name')

                # Inbound messages
                for msg in value.get('messages') or []:
                    frm = msg.get('from') or (contacts[0].get('wa_id') if contacts else None)
                    _handle_inbound_message(msg, frm, contact_name=contact_name)

                # Outbound status updates
                for st in value.get('statuses') or []:
                    _handle_status_update(st)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[whatsapp-webhook] handler crashed: {e}')

    return '', 200


def _fb_page_tokens():
    """Return {page_id: page_access_token} for every Page vcore is wired to.

    Two env shapes supported, in priority order:

      1. FB_PAGE_ACCESS_TOKENS (preferred, multi-Page)
         JSON dict mapping page_id -> page access token. Example:
         {"1163731123483467": "EAAUTIm...", "860083723853421": "EAAUTIm..."}

      2. FB_PAGE_ID + FB_PAGE_ACCESS_TOKEN (legacy single-Page)
         Kept so prod stays live during the cut-over; once
         FB_PAGE_ACCESS_TOKENS is set in Lambda env we can drop the legacy
         pair, but no rush — both being present is fine (the JSON wins).

    Empty dict means "nothing configured" — callers translate that to a
    user-facing error.
    """
    raw = os.getenv('FB_PAGE_ACCESS_TOKENS', '').strip()
    if raw:
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict) and parsed:
                # Coerce both keys + values to str so the dict is safe to
                # use directly for Graph URL composition.
                return {str(k): str(v) for k, v in parsed.items() if k and v}
        except Exception as e:
            app.logger.error(f'[FB] FB_PAGE_ACCESS_TOKENS not valid JSON: {e}')
            # Fall through to legacy single-Page env so prod doesn't go
            # dark on a typo in the new var.
    legacy_id = os.getenv('FB_PAGE_ID', '').strip()
    legacy_token = os.getenv('FB_PAGE_ACCESS_TOKEN', '').strip()
    if legacy_id and legacy_token:
        return {legacy_id: legacy_token}
    return {}


def _do_facebook_sync(created_by_id):
    """Fetch leads from all Lead Ad forms on every configured Page and save new ones.
    Returns (new_count, skipped_count, error_msg).
    """
    import requests as req_lib
    from models import Lead

    pages = _fb_page_tokens()
    if not pages:
        return 0, 0, (
            'No Facebook Pages configured. Set FB_PAGE_ACCESS_TOKENS '
            '(JSON dict of page_id -> token) or the legacy FB_PAGE_ID + '
            'FB_PAGE_ACCESS_TOKEN pair in env.'
        )

    new_count = 0
    skipped_count = 0
    per_page_errors = []

    for page_id, page_token in pages.items():
        # Step 1 — get all Lead Ad forms for this Page
        try:
            resp = req_lib.get(
                f'https://graph.facebook.com/v19.0/{page_id}/leadgen_forms',
                params={'access_token': page_token, 'fields': 'id,name', 'limit': 100},
                timeout=15
            )
            forms_data = resp.json()
        except Exception as e:
            per_page_errors.append(f'Page {page_id} forms fetch: {e}')
            continue

        if 'error' in forms_data:
            msg = forms_data['error'].get('message', 'Unknown error')
            per_page_errors.append(f'Page {page_id}: {msg}')
            continue

        forms = forms_data.get('data', [])
        if not forms:
            # Not an error — a Page can legitimately have zero lead forms.
            continue

        # Step 2 — for each form, paginate through all leads
        for form in forms:
            form_id = form.get('id')
            form_name = form.get('name', '')
            url = f'https://graph.facebook.com/v19.0/{form_id}/leads'
            params = {
                'access_token': page_token,
                # Nested field expansion pulls campaign/adset/ad names in the
                # same round trip. Same field list as facebook_webhook_receive
                # below — keep them in sync so BD's campaign filter on /leads
                # works for cron-ingested + manually-synced + webhook-ingested
                # leads identically. NOTE: `form_name` is not a leadgen-edge
                # field (Graph returns error #100); we capture form_name
                # separately via the parent form metadata above.
                'fields': (
                    'id,created_time,field_data,'
                    'ad_id,adset_id,campaign_id,'
                    'campaign_name,adset_name,ad_name'
                ),
                'limit': 100,
            }

            while url:
                try:
                    resp = req_lib.get(url, params=params, timeout=15)
                    data = resp.json()
                except Exception as e:
                    app.logger.error(f'[Facebook sync] Error fetching leads for form {form_id}: {e}')
                    break

                if 'error' in data:
                    app.logger.error(f'[Facebook sync] API error for form {form_id}: {data["error"]}')
                    break

                for lead_entry in data.get('data', []):
                    fb_lead_id = str(lead_entry.get('id', ''))
                    if not fb_lead_id:
                        continue

                    # Dedup check
                    if Lead.query.filter_by(facebook_lead_id=fb_lead_id).first():
                        skipped_count += 1
                        continue

                    # Parse field_data: [{"name": "full_name", "values": ["John"]}, ...]
                    fields = {f['name']: (f['values'][0] if f.get('values') else '')
                              for f in lead_entry.get('field_data', [])}

                    # Skip Facebook test leads (generated by Meta's "Test on Facebook" tool)
                    if any('<test lead' in str(v).lower() for v in fields.values()):
                        skipped_count += 1
                        continue

                    name = (fields.get('full_name') or fields.get('name') or
                            fields.get('first_name', '') + ' ' + fields.get('last_name', '')).strip() or None
                    phone = (fields.get('phone_number') or fields.get('phone') or
                             fields.get('mobile') or fields.get('contact')) or None
                    email = fields.get('email') or None
                    city  = fields.get('city') or None
                    state = fields.get('state') or None

                    # Build notes from form name + any extra fields
                    notes = f'Ad Form: {form_name}' if form_name else None

                    fb_created = _parse_fb_created_time(lead_entry.get('created_time'))

                    # Same ad-hierarchy capture as the webhook ingest.
                    # None-safe: missing legs (organic forms, archived ads)
                    # land as NULL columns and the lead falls into the
                    # "Unknown campaign" bucket on /leads filter dropdowns
                    # rather than blocking ingest.
                    lead = Lead(
                        name=name,
                        contact=phone,
                        email=email,
                        city=city,
                        state=state,
                        notes=notes,
                        origin='Facebook',
                        stage='New Lead',
                        lead_type='Enquiry',
                        facebook_lead_id=fb_lead_id,
                        fb_page_id       = page_id,
                        fb_campaign_id   = lead_entry.get('campaign_id')   or None,
                        fb_campaign_name = lead_entry.get('campaign_name') or None,
                        fb_adset_id      = lead_entry.get('adset_id')      or None,
                        fb_adset_name    = lead_entry.get('adset_name')    or None,
                        fb_ad_id         = lead_entry.get('ad_id')         or None,
                        fb_ad_name       = lead_entry.get('ad_name')       or None,
                        # Form name comes from the parent loop (the Step-1
                        # forms query already gave us form_name); promote it
                        # off `notes` for filter use.
                        fb_form_name     = form_name or None,
                        owner_id=None,
                        assigned_to_id=None,
                        created_by=created_by_id,
                        created_at=fb_created or datetime.utcnow(),
                    )
                    db.session.add(lead)
                    new_count += 1

                db.session.commit()

                # Follow pagination cursor
                next_page = data.get('paging', {}).get('next')
                if next_page:
                    url = next_page
                    params = {}  # next URL already contains all params
                else:
                    break

    # Surface per-Page failures but don't block — partial success is
    # better than silently skipping a Page because its sibling errored.
    err_msg = '; '.join(per_page_errors) if per_page_errors else None
    if err_msg and new_count == 0 and skipped_count == 0:
        return 0, 0, err_msg
    return new_count, skipped_count, err_msg


@app.route('/leads/facebook/sync', methods=['POST'])
@login_required
def facebook_sync():
    """Manual sync — triggered by clicking the Sync Facebook button."""
    # Manager+ only — same reasoning as IndiaMart sync above.
    if not current_user.is_manager_or_admin():
        flash('Only managers or admins can trigger Facebook sync.', 'danger')
        return redirect(url_for('leads_list'))
    new_count, skipped_count, err = _do_facebook_sync(current_user.id)
    if err:
        flash(f'Facebook sync failed: {err}', 'danger')
    else:
        flash(f'Facebook sync complete: {new_count} new leads imported, {skipped_count} already existed.', 'success')
    return redirect(url_for('leads_list'))


@app.route('/api/leads/facebook-sync', methods=['GET'])
def facebook_cron_sync():
    """Cron endpoint — auto-sync Facebook leads every hour."""
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401

    admin = User.query.filter_by(role='admin').first() or User.query.first()
    if not admin:
        return jsonify({'error': 'No users found'}), 500

    new_count, skipped_count, err = _do_facebook_sync(admin.id)
    if err:
        return jsonify({'success': False, 'error': err}), 500
    return jsonify({'success': True, 'new': new_count, 'skipped': skipped_count}), 200


# ─── Real-time Facebook Lead Ads webhook ──────────────────────────────────────
# Meta POSTs here within seconds of a Lead Ads form submission, so leads appear
# in vcore without waiting for the periodic sync. The flow:
#   1. Meta verifies the endpoint with a GET (hub.mode=subscribe + verify_token)
#   2. On every new lead, Meta POSTs a notification with the lead_id only
#   3. We verify the X-Hub-Signature-256 (HMAC of body using FB_APP_SECRET)
#   4. We fetch the full lead via Graph API using FB_PAGE_ACCESS_TOKEN
#   5. Same de-dupe + test-lead filter as the manual sync
#
# Setup: see DEPLOYMENT.md / one-time Meta dashboard wiring.

@app.route('/api/leads/facebook-webhook', methods=['GET'])
def facebook_webhook_verify():
    """Meta's verification handshake — called once when the webhook is registered.

    Meta sends: GET /api/leads/facebook-webhook?hub.mode=subscribe&hub.verify_token=<our token>&hub.challenge=<random>
    We must echo back hub.challenge ONLY IF the verify_token matches what we stored.
    """
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    expected = os.getenv('FB_WEBHOOK_VERIFY_TOKEN', '')

    if mode == 'subscribe' and expected and token == expected:
        app.logger.info('[FB webhook] verification handshake OK')
        return challenge or '', 200
    app.logger.warning('[FB webhook] verification rejected (token mismatch or missing)')
    return 'forbidden', 403


@app.route('/api/leads/facebook-webhook', methods=['POST'])
def facebook_webhook_receive():
    """Receive a real-time lead notification from Meta and import the lead."""
    import hmac as _hmac
    import hashlib as _hashlib
    import requests as req_lib
    from models import Lead, default_stage_for_origin

    raw = request.get_data(cache=True)

    # 1. HMAC signature verification (Meta signs every request with FB_APP_SECRET)
    sig_header = request.headers.get('X-Hub-Signature-256', '')
    app_secret = os.getenv('FB_APP_SECRET', '')
    if not app_secret:
        app.logger.error('[FB webhook] FB_APP_SECRET not configured — refusing to process')
        return jsonify({'error': 'server misconfigured'}), 500
    if not sig_header.startswith('sha256='):
        return jsonify({'error': 'missing signature'}), 401
    expected_sig = _hmac.new(app_secret.encode(), raw, _hashlib.sha256).hexdigest()
    if not _hmac.compare_digest(expected_sig, sig_header.split('=', 1)[1]):
        app.logger.warning('[FB webhook] signature mismatch')
        return jsonify({'error': 'invalid signature'}), 401

    # 2. Parse payload
    try:
        body = json.loads(raw.decode('utf-8'))
    except Exception:
        return jsonify({'error': 'invalid json'}), 400

    if body.get('object') != 'page':
        # Other object types (user, instagram, etc.) — ignore
        return jsonify({'ok': True, 'ignored': body.get('object')}), 200

    # Multi-Page support: each entry carries `entry.id` = the Page ID that
    # triggered the change. Look up the right Page Access Token per entry
    # so BathQube + Glassy.in webhooks can share this single endpoint.
    page_tokens = _fb_page_tokens()
    if not page_tokens:
        app.logger.error('[FB webhook] no Page tokens configured (FB_PAGE_ACCESS_TOKENS / legacy pair)')
        return jsonify({'error': 'server misconfigured'}), 500

    # 3. Iterate notifications. Each `entry` is one page change batch; each
    # `changes` entry is one field change; we only care about `leadgen`.
    admin = User.query.filter_by(role='admin').first() or User.query.first()
    creator_id = admin.id if admin else None

    new_count = 0
    skipped = 0
    errors = []

    for entry in body.get('entry', []):
        # `entry.id` is the Page ID. Route to the matching Page token.
        # If we get a webhook for a Page we don't have a token for, log
        # + skip rather than crash — Meta sometimes fans out to apps
        # subscribed via BM groups even when a specific Page wasn't
        # explicitly connected.
        entry_page_id = str(entry.get('id') or '')
        page_token = page_tokens.get(entry_page_id)
        if not page_token:
            errors.append(f'no token for page {entry_page_id}')
            continue

        for change in entry.get('changes', []):
            if change.get('field') != 'leadgen':
                continue
            value = change.get('value', {}) or {}
            fb_lead_id = str(value.get('leadgen_id') or '')
            if not fb_lead_id:
                continue

            # Dedupe before hitting Graph API
            if Lead.query.filter_by(facebook_lead_id=fb_lead_id).first():
                skipped += 1
                continue

            # Fetch the full lead from Graph API
            try:
                resp = req_lib.get(
                    f'https://graph.facebook.com/v19.0/{fb_lead_id}',
                    params={
                        'access_token': page_token,
                        # Nested-field expansion pulls campaign / adset /
                        # ad / form names in ONE round trip so we don't
                        # have to make 4 sequential Graph calls per lead.
                        # If a particular leg is missing on Meta's side
                        # (rare; usually only on archived ads) the key
                        # is simply absent from the response.
                        # NOTE: `form_name` is NOT a field on the
                        # leadgen edge — Graph returns error #100
                        # ("Tried accessing nonexisting field"). The
                        # form_id stays in our `notes` blob; if BD ever
                        # wants the human-readable form name we'd have
                        # to make a second call to /{form_id}?fields=name,
                        # not worth the round-trip yet.
                        'fields': (
                            'id,created_time,field_data,form_id,'
                            'ad_id,adset_id,campaign_id,'
                            'campaign_name,adset_name,ad_name'
                        ),
                    },
                    timeout=10,
                )
                lead_data = resp.json()
            except Exception as e:
                errors.append(f'fetch {fb_lead_id}: {e}')
                continue
            if 'error' in lead_data:
                errors.append(f'fb error for {fb_lead_id}: {lead_data["error"].get("message")}')
                continue

            fields = {f['name']: (f['values'][0] if f.get('values') else '')
                      for f in lead_data.get('field_data', [])}

            # Skip Meta's "Test on Facebook" generated test leads
            if any('<test lead' in str(v).lower() for v in fields.values()):
                skipped += 1
                continue

            name = (fields.get('full_name') or fields.get('name') or
                    (fields.get('first_name', '') + ' ' + fields.get('last_name', '')).strip() or None)
            phone = (fields.get('phone_number') or fields.get('phone')
                     or fields.get('mobile') or fields.get('contact')) or None
            email = fields.get('email') or None
            city = fields.get('city') or None
            state = fields.get('state') or None

            fb_created = _parse_fb_created_time(lead_data.get('created_time'))
            if fb_created is None:
                # Webhook also delivers a Unix timestamp on `value.created_time`
                ts = value.get('created_time')
                if isinstance(ts, (int, float)):
                    fb_created = datetime.utcfromtimestamp(ts)
            # Capture the ad-hierarchy metadata for BD's campaign filters
            # on /leads. None-safe: if Meta omitted a leg (archived ad,
            # organic lead form, etc.) the column simply stays NULL and
            # the lead falls into the "Unknown campaign" filter bucket
            # rather than blocking ingest.
            fb_campaign_id   = lead_data.get('campaign_id')   or None
            fb_campaign_name = lead_data.get('campaign_name') or None
            fb_adset_id      = lead_data.get('adset_id')      or None
            fb_adset_name    = lead_data.get('adset_name')    or None
            fb_ad_id         = lead_data.get('ad_id')         or None
            fb_ad_name       = lead_data.get('ad_name')       or None
            # form_name isn't fetchable on the leadgen edge — leave NULL.
            # Resolvable via a second Graph call if BD ever asks.
            fb_form_name     = None

            lead = Lead(
                name=name, contact=phone, email=email, city=city, state=state,
                notes=f'Webhook · form_id={lead_data.get("form_id", "")}',
                origin='Facebook',
                stage=default_stage_for_origin('Facebook'),
                lead_type='Enquiry',
                facebook_lead_id=fb_lead_id, owner_id=None, assigned_to_id=None,
                created_by=creator_id,
                created_at=fb_created or datetime.utcnow(),
                fb_page_id=entry_page_id,
                fb_campaign_id=fb_campaign_id,
                fb_campaign_name=fb_campaign_name,
                fb_adset_id=fb_adset_id,
                fb_adset_name=fb_adset_name,
                fb_ad_id=fb_ad_id,
                fb_ad_name=fb_ad_name,
                fb_form_name=fb_form_name,
            )
            db.session.add(lead)
            new_count += 1

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'db error', 'detail': str(e)}), 500

    app.logger.info(f'[FB webhook] processed: new={new_count} skipped={skipped} errors={len(errors)}')
    # Meta only checks for HTTP 200 — body is for our own logging
    return jsonify({
        'ok': True, 'new': new_count, 'skipped': skipped,
        'errors': errors if errors else None,
    }), 200


# ─── WhatsApp Cloud API delivery-status webhook ───────────────────────────────
# Meta POSTs status updates (sent → delivered → read → failed) for every
# template message we send. Without this endpoint, vcore's "Sent ✓" badge
# only proves Meta accepted the request — not that the customer received it.
# With it, the WhatsAppMessage.status column reflects the true delivery
# state and failures (template paused, recipient not on WhatsApp, marketing
# opt-out, quality-rating drop) surface in error_message for BD to act on.
#
# Setup (one-time, in Meta Business Manager → WhatsApp Manager →
# Configuration → Webhooks):
#   Callback URL:  https://vcore.glassy.in/api/whatsapp/webhook
#   Verify token:  value of WHATSAPP_WEBHOOK_VERIFY_TOKEN env var
#   Subscribe to:  `messages` field on the WABA

# Status-progression order — webhooks can arrive out of order (e.g.
# 'read' before 'delivered' if Meta batches them); guard against
# downgrades so a freshly-delivered message doesn't snap back to 'sent'.
_WA_STATUS_RANK = {
    'queued':       0,
    'sent':         1,
    'delivered':    2,
    'read':         3,
    # Failure states are terminal — give them a high rank so they
    # always win over the in-flight states.
    'failed':       9,
    'undelivered':  9,
    'warning':      9,
}


@app.route('/api/whatsapp/webhook', methods=['GET'])
def whatsapp_webhook_verify():
    """Meta's verification handshake — called once when the webhook is registered."""
    mode = request.args.get('hub.mode')
    token = request.args.get('hub.verify_token')
    challenge = request.args.get('hub.challenge')
    expected = os.getenv('WHATSAPP_WEBHOOK_VERIFY_TOKEN', '')
    if mode == 'subscribe' and expected and token == expected:
        app.logger.info('[WA webhook] verification handshake OK')
        return challenge or '', 200
    app.logger.warning('[WA webhook] verification rejected (token mismatch or missing)')
    return 'forbidden', 403


@app.route('/api/whatsapp/webhook', methods=['POST'])
def whatsapp_webhook_receive():
    """Receive WhatsApp delivery-status updates from Meta.

    Payload shape (statuses are the only field-type we care about; inbound
    customer messages are ignored — vcore is send-only today):

      {"object":"whatsapp_business_account",
       "entry":[{"id":"<WABA_ID>","changes":[{
         "field":"messages",
         "value":{
           "statuses":[{"id":"<wamid>","status":"delivered",
                        "timestamp":"...","recipient_id":"91...",
                        "errors":[{...}]}]
         }
       }]}]}
    """
    import hmac as _hmac
    import hashlib as _hashlib

    raw = request.get_data(cache=True)

    # 1. HMAC verification — same secret as FB webhook (the WABA lives
    # under the same Meta App so FB_APP_SECRET signs both).
    sig_header = request.headers.get('X-Hub-Signature-256', '')
    app_secret = os.getenv('FB_APP_SECRET', '')
    if not app_secret:
        app.logger.error('[WA webhook] FB_APP_SECRET not configured — refusing to process')
        return jsonify({'error': 'server misconfigured'}), 500
    if not sig_header.startswith('sha256='):
        return jsonify({'error': 'missing signature'}), 401
    expected_sig = _hmac.new(app_secret.encode(), raw, _hashlib.sha256).hexdigest()
    if not _hmac.compare_digest(expected_sig, sig_header.split('=', 1)[1]):
        app.logger.warning('[WA webhook] signature mismatch')
        return jsonify({'error': 'invalid signature'}), 401

    try:
        body = json.loads(raw.decode('utf-8'))
    except Exception:
        return jsonify({'error': 'invalid json'}), 400

    if body.get('object') != 'whatsapp_business_account':
        return jsonify({'ok': True, 'ignored': body.get('object')}), 200

    from models import WhatsAppMessage

    updated = 0
    not_found = 0
    failures = 0
    notes = []

    for entry in body.get('entry', []):
        for change in entry.get('changes', []):
            if change.get('field') != 'messages':
                continue
            value = change.get('value', {}) or {}
            for st in (value.get('statuses') or []):
                wamid = st.get('id')
                new_status = (st.get('status') or '').lower()
                if not wamid or not new_status:
                    continue
                msg = WhatsAppMessage.query.filter_by(wamid=wamid).first()
                if not msg:
                    # Webhook for a message we didn't send (or a status
                    # update arriving before the DB commit landed). Log
                    # and move on rather than 500 — Meta retries on 500.
                    not_found += 1
                    notes.append(f'no row for wamid={wamid[:24]}…')
                    continue

                old_rank = _WA_STATUS_RANK.get(msg.status, 0)
                new_rank = _WA_STATUS_RANK.get(new_status, 0)
                if new_rank > old_rank:
                    msg.status = new_status
                    updated += 1

                # Capture failure details so BD can see WHY a message
                # was rejected (template paused, recipient not on
                # WhatsApp, opt-out, etc.).
                if new_status in ('failed', 'undelivered'):
                    failures += 1
                    errors = st.get('errors') or []
                    if errors:
                        e = errors[0] if isinstance(errors[0], dict) else {}
                        detail = (
                            e.get('error_data', {}).get('details') if isinstance(e.get('error_data'), dict) else None
                        )
                        msg.error_message = (
                            detail
                            or e.get('message')
                            or e.get('title')
                            or f'Meta-side failure (code={e.get("code", "?")})'
                        )

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[WA webhook] db commit failed: {e}')
        return jsonify({'error': 'db error', 'detail': str(e)}), 500

    app.logger.info(
        f'[WA webhook] updated={updated} not_found={not_found} failures={failures}'
    )
    return jsonify({
        'ok': True, 'updated': updated, 'not_found': not_found,
        'failures': failures, 'notes': notes or None,
    }), 200


@app.route('/api/leads/indiamart-sync', methods=['GET'])
def indiamart_cron_sync():
    """Cron endpoint — auto-sync IndiaMart leads every hour"""
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401

    # Use first admin user as owner
    admin = User.query.filter_by(role='admin').first() or User.query.first()
    if not admin:
        return jsonify({'error': 'No users found'}), 500

    new_count, skipped_count, err = _do_indiamart_sync(None, admin.id)
    if err:
        return jsonify({'success': False, 'error': err}), 500
    return jsonify({'success': True, 'new': new_count, 'skipped': skipped_count}), 200


@app.route('/api/leads/indiamart-keepalive', methods=['GET'])
def indiamart_keepalive():
    """
    Cron endpoint — call every 12 hours to keep the IndiaMart token alive.
    Only refreshes the token; does NOT import leads.
    Must be called while the token is still valid (before the 24hr expiry).
    """
    cron_secret = request.headers.get('X-Cron-Secret') or request.args.get('secret')
    expected_secret = os.getenv('CRON_SECRET')
    if not expected_secret or cron_secret != expected_secret:
        return jsonify({'error': 'Unauthorized'}), 401

    from models import IndiamartToken
    token = IndiamartToken.query.first()
    if not token:
        return jsonify({'success': False, 'error': 'No token saved.'}), 400

    headers = {
        'user-agent': 'IndiaMart/13.6.4 (com.indiamart.m; build:4; iOS 16.6.0) Alamofire/5.0.0-rc.2',
        'accept': '*/*',
        'accept-language': 'en-US;q=1.0',
    }

    old_ak = token.ak_token
    _indiamart_refresh_token(token, headers)
    token = IndiamartToken.query.first()

    refreshed = token.ak_token != old_ak
    return jsonify({
        'success': True,
        'refreshed': refreshed,
        'token_valid': token.is_valid(),
        'expires_at': token.expires_at.isoformat() if token.expires_at else None,
    }), 200


# ============================================================================
# TAX INVOICES — GST invoice generation flow (Phase 1: constants + helpers)
# ============================================================================
# When a quote reaches `closed_won`, BD can promote it to a tax invoice
# via the Tally screen. The invoice freezes a snapshot of line items at
# generation time + carries the regulatory metadata GST requires
# (HSN, GSTIN, sequential per-FY numbering, etc.).

# Vetrova seller defaults — written into each TaxInvoice as a snapshot
# at create time, so changing these constants later doesn't rewrite
# historical invoices. Confirmed by BD on 2026-06-29.
TAX_INVOICE_SELLER_DEFAULTS = {
    'name':       'VETROVA TECH SERVICES PRIVATE LIMITED',
    'address':    ('A-72, Unit-I Hosur Road KKSIDC IND estate, '
                   'Bommasandra Industrial Area Phase 4, Bengaluru'),
    'gstin':      '29AALCV4455A1Z7',
    'state':      'Karnataka',
    'state_code': '29',
    'pan':        'AALCV4455A',
    'udyam':      'UDYAM-KR-03-0632749',
    'email':      'support@glassy.in',
    'cin':        'U62099KA2018PTC127405',
}

# Bank details printed on the invoice. Mirrors the Bathqube quote PDF
# block (single source of truth: any future bank change updates here).
TAX_INVOICE_BANK_DEFAULTS = {
    'account_name': 'Vetrova Tech Services Private Limited',
    'bank_name':    'IDFC First Bank',
    'account_no':   '10249972220',
    'ifsc':         'IDFB0080158',
    'branch':       'Bengaluru',
    'upi_id':       '8550011196@ybl',
}

# Standard GST declaration printed on every tax invoice (same text the
# reference invoice carries — required by GST law).
TAX_INVOICE_DECLARATION = (
    'We declare that this invoice shows the actual price of the goods '
    'described and that all particulars are true and correct.'
)

# Default HSN code per item-source, confirmed by BD on 2026-06-29.
# `_default_hsn_for_*` resolvers below pull from this map. BD can
# override per-line in the invoice form.
TAX_INVOICE_HSN_DEFAULTS = {
    'bathqube_glass':   '70071900',   # toughened safety glass
    'bathqube_mirror':  '70091000',   # mirrors of glass
    'upvc':             '39252000',   # UPVC builders' ware
    'hardware':         '83024110',   # mountings for doors / windows
    'transportation':   '9965',       # transport service SAC
    'installation':     '995462',     # glazing service SAC
}


def _current_financial_year():
    """India FY runs April → March. Today 2026-06-29 falls in FY 26-27,
    written as '2627' for invoice numbering.

    Returns a two-year string e.g. '2627' that's safe for the next ~70
    years (no Y3K issue worth pre-empting).
    """
    today = datetime.utcnow()
    yy = today.year % 100
    if today.month >= 4:
        return f'{yy:02d}{(yy + 1) % 100:02d}'
    return f'{(yy - 1) % 100:02d}{yy:02d}'


def _amount_in_words_inr(amount):
    """Convert ₹ amount to Indian-format English words.

    GST law requires the invoice to print the amount in words. Indian
    numbering uses lakh (100k) + crore (10M) rather than million/billion,
    so a naive en-US converter would render `₹1,82,832` as "one hundred
    eighty-two thousand…" instead of the legally-expected
    "INR One Lakh Eighty Two Thousand…".

    Handles amounts up to 99 crore (more than realistic for any single
    Vetrova invoice). Returns the empty string for invalid input — the
    caller renders a blank words-line rather than failing PDF gen.
    """
    try:
        amt = float(amount)
    except (TypeError, ValueError):
        return ''
    rupees = int(amt)
    paise = round((amt - rupees) * 100)
    if paise == 100:
        rupees += 1
        paise = 0

    UNITS = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine',
             'Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen',
             'Seventeen', 'Eighteen', 'Nineteen']
    TENS = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']

    def two_digits(n):
        if n < 20:
            return UNITS[n]
        return (TENS[n // 10] + (' ' + UNITS[n % 10] if n % 10 else '')).strip()

    def three_digits(n):
        if n == 0:
            return ''
        hundreds = n // 100
        rest = n % 100
        parts = []
        if hundreds:
            parts.append(UNITS[hundreds] + ' Hundred')
        if rest:
            parts.append(two_digits(rest))
        return ' '.join(parts)

    if rupees == 0:
        words = 'Zero'
    else:
        crore = rupees // 10000000
        lakh  = (rupees // 100000) % 100
        thou  = (rupees // 1000) % 100
        rem   = rupees % 1000
        parts = []
        if crore:
            parts.append(f'{two_digits(crore)} Crore')
        if lakh:
            parts.append(f'{two_digits(lakh)} Lakh')
        if thou:
            parts.append(f'{two_digits(thou)} Thousand')
        if rem:
            parts.append(three_digits(rem))
        words = ' '.join(parts).strip()

    result = f'INR {words}'
    if paise:
        result += f' and {two_digits(paise)} paise'
    result += ' Only'
    return result


def _seed_tax_invoice_from_bathqube(quote, invoice):
    """Pre-fill a TaxInvoice + items from a BathqubeQuote. Called once
    at invoice-create time. Items map 1:1 from BathqubeQuoteItem rows.

    HSN defaults: glass panels → 70071900, transport/install extras →
    9965/995462. BD can override per-line on the form.
    """
    from models import TaxInvoiceItem, BathqubeQuoteItem
    # Buyer (Bill-to) — taken from the quote
    invoice.buyer_name = quote.customer_name or ''
    invoice.buyer_address = '\n'.join([
        x for x in [quote.site_address, quote.pincode] if x
    ]) or quote.customer_name or ''
    invoice.buyer_state = 'Karnataka'   # Vetrova's home state; BD can edit on form
    invoice.buyer_state_code = '29'
    # B2C — only PAN can be printed on the invoice, no GSTIN.
    invoice.buyer_pan = (quote.customer_pan or '').strip() or None
    # Consignee (Ship-to) — same as buyer by default; BD overrides if different
    invoice.consignee_name = invoice.buyer_name
    invoice.consignee_address = invoice.buyer_address
    invoice.consignee_state = invoice.buyer_state
    invoice.consignee_state_code = invoice.buyer_state_code

    # Seed items 1:1 from the quote
    sort = 0
    for src in (quote.items or []):
        desc = (src.description or 'Shower glass')[:500]
        is_extra = bool(getattr(src, 'is_extra', False))
        # Pick HSN: extras typically transport/install; everything else
        # defaults to toughened-glass.
        if is_extra:
            low = desc.lower()
            if 'transport' in low or 'delivery' in low:
                hsn = TAX_INVOICE_HSN_DEFAULTS['transportation']
            elif 'install' in low or 'fitting' in low or 'labour' in low:
                hsn = TAX_INVOICE_HSN_DEFAULTS['installation']
            elif 'mirror' in low:
                hsn = TAX_INVOICE_HSN_DEFAULTS['bathqube_mirror']
            else:
                hsn = TAX_INVOICE_HSN_DEFAULTS['hardware']
        else:
            hsn = TAX_INVOICE_HSN_DEFAULTS['bathqube_glass']
        invoice.items.append(TaxInvoiceItem(
            sort_order = sort,
            description = desc,
            hsn_code   = hsn,
            quantity   = float(src.quantity or 1),
            unit       = 'nos',
            rate       = float(src.rate or 0),
            amount     = float(src.amount or 0),
            is_extra   = is_extra,
        ))
        sort += 1


def _seed_tax_invoice_from_upvc(quote, invoice):
    """Pre-fill a TaxInvoice + items from a UpvcQuote. Each UPVC item
    becomes a single invoice line, expressed in sqft × ₹/sqft because
    that's how the source quote priced it."""
    from models import TaxInvoiceItem
    invoice.buyer_name = quote.customer_name or ''
    invoice.buyer_address = '\n'.join([
        x for x in [quote.site_address, quote.pincode] if x
    ]) or quote.customer_name or ''
    invoice.buyer_state = 'Karnataka'
    invoice.buyer_state_code = '29'
    invoice.consignee_name = invoice.buyer_name
    invoice.consignee_address = invoice.buyer_address
    invoice.consignee_state = invoice.buyer_state
    invoice.consignee_state_code = invoice.buyer_state_code

    sort = 0
    for src in (quote.items or []):
        # Build a descriptive line: "Master BR — Sliding · 3-track ·
        # 2100×2400 mm · Wooden"
        type_bits = [(src.track_type or '').capitalize()]
        if src.track_system:
            type_bits.append(src.track_system)
        if src.colour:
            type_bits.append(src.colour.capitalize())
        dims = ''
        if src.width and src.height:
            dims = f' · {float(src.width):g}×{float(src.height):g} {src.unit}'
        label = (src.label or '').strip()
        desc_parts = []
        if label:
            desc_parts.append(label)
        desc_parts.append(' · '.join(type_bits))
        desc = (' — '.join(desc_parts) + dims)[:500]
        # Quantity field: total sqft (qty * sqft) so the rate stays
        # ₹/sqft; matches what the UPVC quote priced on.
        total_sqft = float(src.quantity or 1) * float(src.sqft or 0)
        invoice.items.append(TaxInvoiceItem(
            sort_order = sort,
            description = desc,
            hsn_code   = TAX_INVOICE_HSN_DEFAULTS['upvc'],
            quantity   = round(total_sqft, 4) or 1.0,
            unit       = 'sqft',
            rate       = float(src.rate or 0),
            amount     = float(src.amount or 0),
            is_extra   = False,
        ))
        sort += 1


# State-name → GST state code map for India (just the codes we actually
# encounter in BD's customer base). Used by the lead-quote seeder to set
# buyer_state_code when the source Quote has only a state name. Unknown
# states default to '29' (Karnataka — Vetrova's home base, matches the
# most common intra-state case so CGST+SGST stays the default).
_GST_STATE_CODES = {
    'andhra pradesh': '37', 'arunachal pradesh': '12', 'assam': '18',
    'bihar': '10', 'chhattisgarh': '22', 'delhi': '07', 'goa': '30',
    'gujarat': '24', 'haryana': '06', 'himachal pradesh': '02',
    'jammu and kashmir': '01', 'jharkhand': '20', 'karnataka': '29',
    'kerala': '32', 'madhya pradesh': '23', 'maharashtra': '27',
    'manipur': '14', 'meghalaya': '17', 'mizoram': '15', 'nagaland': '13',
    'odisha': '21', 'puducherry': '34', 'punjab': '03', 'rajasthan': '08',
    'sikkim': '11', 'tamil nadu': '33', 'telangana': '36', 'tripura': '16',
    'uttar pradesh': '09', 'uttarakhand': '05', 'west bengal': '19',
    'chandigarh': '04', 'andaman and nicobar islands': '35',
    'dadra and nagar haveli and daman and diu': '26', 'lakshadweep': '31',
    'ladakh': '38',
}


def _state_code_for(name):
    """Return the GST state code for a state name, or None if not recognised."""
    if not name:
        return None
    return _GST_STATE_CODES.get(name.strip().lower())


def _seed_tax_invoice_from_lead_quote(quote, invoice):
    """Pre-fill a TaxInvoice + items from a regular Quote (the /quotes
    flow — not Bathqube, not UPVC).

    Quote has a richer shape than the configurator-driven quotes:
      - Hierarchical line items (group + children); only leaf items
        become invoice lines, group headers are skipped.
      - 14 optional charge fields (delivery / installation / freight /
        transport / cut-out / holes / shape-cutting / jumbo-size /
        template / handling / polish / document / frosted / insurance).
        Each non-zero charge is appended as an `is_extra=True` line so
        the tax invoice reflects every leg of the price.
      - Optional `customer_gst` (B2B). When present, copied onto the
        buyer GSTIN so the invoice prints it correctly.
      - Separate `invoice_to` (bill) vs `dispatch_to` (ship). When both
        are set, consignee uses dispatch_to; buyer uses invoice_to.
    """
    from models import TaxInvoiceItem

    # Buyer (Bill-to) — prefer invoice_to when set, else stitch from
    # customer_address + city.
    buyer_addr = (quote.invoice_to or '').strip()
    if not buyer_addr:
        addr_lines = []
        if quote.customer_address:
            addr_lines.append(quote.customer_address.strip())
        if quote.customer_city:
            addr_lines.append(quote.customer_city.strip())
        buyer_addr = '\n'.join(addr_lines)
    invoice.buyer_name      = quote.customer_name or ''
    invoice.buyer_address   = buyer_addr or quote.customer_name or ''
    invoice.buyer_gstin     = (quote.customer_gst or '').strip() or None
    # PAN — typically only filled for B2C quotes (B2B uses GSTIN
    # instead). Either or both can coexist; the PDF prints whichever
    # is present.
    invoice.buyer_pan       = (quote.customer_pan or '').strip() or None
    invoice.buyer_state     = quote.customer_state or 'Karnataka'
    invoice.buyer_state_code = _state_code_for(quote.customer_state) or '29'

    # Consignee (Ship-to). If dispatch_to is set, that's the ship-to
    # address (and the rest of the consignee fields mirror buyer). Else
    # same-as-buyer.
    if (quote.dispatch_to or '').strip():
        invoice.consignee_name      = quote.customer_name or ''
        invoice.consignee_address   = quote.dispatch_to.strip()
        invoice.consignee_gstin     = invoice.buyer_gstin
        invoice.consignee_state     = invoice.buyer_state
        invoice.consignee_state_code = invoice.buyer_state_code
    else:
        invoice.consignee_name      = invoice.buyer_name
        invoice.consignee_address   = invoice.buyer_address
        invoice.consignee_gstin     = invoice.buyer_gstin
        invoice.consignee_state     = invoice.buyer_state
        invoice.consignee_state_code = invoice.buyer_state_code

    # GST % on the quote may differ from the default 18% — copy onto
    # the invoice so the recompute uses the right rate.
    if quote.gst_percentage:
        # Note: we don't store gst_percentage on TaxInvoice in v1 (single
        # rate assumed); the caller _recompute_tax_invoice_totals takes
        # gst_percent as a kwarg. The route uses 18% by default which
        # matches Vetrova's standard slab — if quote.gst_percentage
        # differs significantly BD will edit the line items + amounts
        # on the form.
        pass

    # Quote-level metadata — useful starting points (BD can override
    # on the invoice form)
    invoice.buyers_order_no    = quote.quote_number
    invoice.buyers_order_date  = quote.quote_date
    invoice.terms_of_delivery  = 'SELF PICKUP' if quote.self_pickup else (invoice.terms_of_delivery or 'EX OUR SITE')

    # Line items — flatten the hierarchy (skip group headers).
    #
    # Reconciliation contract: on every TaxInvoiceItem, qty × rate ≡ amount
    # (the EDIT-form re-save asserts this — if it ever diverges, BD's first
    # edit silently recomputes amount = qty × rate, collapsing the subtotal).
    #
    # For a regular QuoteItem, `quantity` is the piece-count and `rate_sqper`
    # is ₹/sqm, with `total = unit_square × rate_sqper × quantity` (see
    # QuoteItem.calculate_total). So we store TOTAL AREA as the invoice
    # quantity, mirroring _seed_tax_invoice_from_upvc:
    #
    #     invoice.quantity = unit_square × QI.quantity      (sqm)
    #     invoice.rate     = rate_sqper                     (₹/sqm)
    #     invoice.amount   = quantity × rate                (₹)
    #
    # When unit_square is missing (line typed without dims), we fall back to
    # piece-count × per-piece rate so the math still reconciles.
    sort = 0
    for src in (quote.items or []):
        if getattr(src, 'is_group', False):
            continue
        desc_parts = [(src.particular or '').strip()]
        if src.actual_width and src.actual_height:
            unit_str = (src.unit or 'MM').strip()
            desc_parts.append(
                f'({float(src.actual_width):g}×{float(src.actual_height):g} {unit_str})'
            )
        desc = ' '.join(p for p in desc_parts if p)[:500] or 'Item'

        piece_qty   = float(src.quantity or 1)
        area_sqm    = float(src.unit_square or 0)
        rate_per_sm = float(src.rate_sqper or 0)
        src_total   = float(src.total or 0)

        if area_sqm > 0 and rate_per_sm > 0:
            # Standard area-priced line — store total sqm so qty×rate = amount
            total_sqm = round(area_sqm * piece_qty, 4)
            inv_qty   = total_sqm or 1.0
            inv_unit  = 'sqm'
            inv_rate  = round(rate_per_sm, 2)
            inv_amt   = round(inv_qty * inv_rate, 2)
        else:
            # Per-piece line (no dimensions). Use the line's total as the
            # effective unit price so qty×rate stays consistent.
            inv_qty   = piece_qty or 1.0
            inv_unit  = 'nos'
            inv_rate  = round((src_total / inv_qty) if inv_qty else src_total, 2)
            inv_amt   = round(inv_qty * inv_rate, 2)

        invoice.items.append(TaxInvoiceItem(
            sort_order = sort,
            description = desc,
            hsn_code   = TAX_INVOICE_HSN_DEFAULTS['bathqube_glass'],
            quantity   = inv_qty,
            unit       = inv_unit,
            rate       = inv_rate,
            amount     = inv_amt,
            is_extra   = False,
        ))
        sort += 1

    # Extra charges → one is_extra line per non-zero field.
    extras = [
        ('Delivery charges',       'transportation', quote.delivery_charges),
        ('Installation charges',   'installation',   quote.installation_charges),
        ('Freight charges',        'transportation', quote.freight_charges),
        ('Transport charges',      'transportation', quote.transport_charges),
        ('Cut-out charges',        'hardware',       quote.cutout_charges),
        ('Holes charges',          'hardware',       quote.holes_charges),
        ('Shape cutting charges',  'hardware',       quote.shape_cutting_charges),
        ('Jumbo size charges',     'bathqube_glass', quote.jumbo_size_charges),
        ('Template charges',       'installation',   quote.template_charges),
        ('Handling charges',       'hardware',       quote.handling_charges),
        ('Polish charges',         'hardware',       quote.polish_charges),
        ('Document charges',       'hardware',       quote.document_charges),
        ('Frosted charges',        'bathqube_glass', quote.frosted_charges),
        ('Insurance charges',      'hardware',       quote.insurance_charges),
    ]
    for label, hsn_key, raw_amount in extras:
        amt = float(raw_amount or 0)
        if amt <= 0:
            continue
        invoice.items.append(TaxInvoiceItem(
            sort_order = sort,
            description = label,
            hsn_code   = TAX_INVOICE_HSN_DEFAULTS[hsn_key],
            quantity   = 1,
            unit       = 'lot',
            rate       = amt,
            amount     = amt,
            is_extra   = True,
        ))
        sort += 1


def _recompute_tax_invoice_totals(invoice, gst_percent=18.0):
    """Sum item amounts → CGST/SGST (intra-state) or IGST (inter-state)
    → round-off → total. Writes the result back onto the invoice and
    refreshes amount_in_words.

    GST rate is invoice-wide (18% by default — Vetrova's products fall
    in the 18% slab). For multi-rate quotes we'd need per-item rates +
    HSN-grouped tax computation; not in scope for v1.
    """
    subtotal = sum(float(it.amount or 0) for it in invoice.items)
    if invoice.is_inter_state:
        invoice.igst = round(subtotal * gst_percent / 100, 2)
        invoice.cgst = 0
        invoice.sgst = 0
        gross = subtotal + invoice.igst
    else:
        half = gst_percent / 2
        invoice.cgst = round(subtotal * half / 100, 2)
        invoice.sgst = round(subtotal * half / 100, 2)
        invoice.igst = 0
        gross = subtotal + float(invoice.cgst) + float(invoice.sgst)
    # Round to nearest rupee for the printed total; record the
    # round-off delta so the invoice still reconciles.
    rounded_total = round(gross)
    invoice.round_off = round(rounded_total - gross, 2)
    invoice.subtotal = round(subtotal, 2)
    invoice.total = rounded_total
    invoice.amount_in_words = _amount_in_words_inr(rounded_total)


def _next_tax_invoice_number():
    """Allocate the next sequential per-FY tax-invoice number.

    Format: VTS/<FY>/<NNNN> e.g. VTS/2627/0001. Matches the reference
    Arihant invoice's prefix-style scheme. NNNN resets to 0001 every
    April 1.

    Concurrency: we read the max existing sequence within the current
    FY and add 1. The unique index on `invoice_number` is the actual
    safety net — if two BD users hit Generate Invoice in the same
    second, one of them will get a constraint-violation INSERT and the
    caller should retry. In practice BD-driven invoice creation is so
    low-frequency this never happens, but the index keeps us honest.
    """
    from models import TaxInvoice
    fy = _current_financial_year()
    prefix = f'VTS/{fy}/'
    last = (TaxInvoice.query
                      .filter(TaxInvoice.invoice_number.like(f'{prefix}%'))
                      .order_by(TaxInvoice.id.desc())
                      .first())
    seq = 1
    if last and last.invoice_number:
        try:
            seq = int(last.invoice_number.rsplit('/', 1)[-1]) + 1
        except (ValueError, IndexError):
            # Couldn't parse the tail — fall back to a count-based seq.
            # Safer to skip a number than collide with an existing one.
            seq = TaxInvoice.query.filter_by(financial_year=fy).count() + 1
    return f'{prefix}{seq:04d}', fy


def _create_tax_invoice_draft(source_kind, source_quote, creator_id):
    """Allocate a new TaxInvoice number + seed buyer/items from the
    source quote. Returns the persisted (still 'draft') row.

    Caller must commit. Raises ValueError if a non-cancelled invoice
    already exists for this source quote (prevents double-issue).
    """
    from models import TaxInvoice
    # Guard against double-issue. Cancelled invoices don't block a new
    # one (BD might cancel + reissue).
    existing_q = TaxInvoice.query.filter(TaxInvoice.status != 'cancelled')
    if source_kind == 'bathqube':
        existing_q = existing_q.filter_by(bathqube_quote_id=source_quote.id)
    elif source_kind == 'upvc':
        existing_q = existing_q.filter_by(upvc_quote_id=source_quote.id)
    else:
        existing_q = existing_q.filter_by(lead_quote_id=source_quote.id)
    existing = existing_q.first()
    if existing:
        raise ValueError(
            f'Tax invoice {existing.invoice_number} already exists '
            f'for this quote (status: {existing.status}). Cancel it first '
            'if you need to re-issue.'
        )

    invoice_number, fy = _next_tax_invoice_number()
    seller = TAX_INVOICE_SELLER_DEFAULTS
    bank = TAX_INVOICE_BANK_DEFAULTS

    inv = TaxInvoice(
        invoice_number = invoice_number,
        financial_year = fy,
        invoice_date   = datetime.utcnow().date(),
        # Seller snapshot
        seller_name = seller['name'], seller_address = seller['address'],
        seller_gstin = seller['gstin'], seller_state = seller['state'],
        seller_state_code = seller['state_code'], seller_pan = seller['pan'],
        seller_udyam = seller['udyam'], seller_email = seller['email'],
        seller_cin = seller['cin'],
        # Bank snapshot
        bank_account_name = bank['account_name'], bank_name = bank['bank_name'],
        bank_account_no = bank['account_no'], bank_ifsc = bank['ifsc'],
        bank_branch = bank['branch'], upi_id = bank['upi_id'],
        declaration = TAX_INVOICE_DECLARATION,
        # Default invoice-meta fields
        mode_of_payment = 'PROMPT',
        dispatched_through = 'ROAD',
        terms_of_delivery = 'EX OUR SITE',
        destination = 'Bengaluru',
        status = 'draft',
        created_by = creator_id,
    )
    # Link to source quote
    if source_kind == 'bathqube':
        inv.bathqube_quote_id = source_quote.id
        _seed_tax_invoice_from_bathqube(source_quote, inv)
    elif source_kind == 'upvc':
        inv.upvc_quote_id = source_quote.id
        _seed_tax_invoice_from_upvc(source_quote, inv)
    else:
        inv.lead_quote_id = source_quote.id
        _seed_tax_invoice_from_lead_quote(source_quote, inv)

    db.session.add(inv)
    db.session.flush()
    _recompute_tax_invoice_totals(inv)
    return inv


@app.route('/quotes/bathqube/<int:id>/invoice/new', methods=['POST'])
@login_required
def bathqube_tax_invoice_new(id):
    """Generate a draft tax invoice from a closed_won Bathqube quote."""
    quote = BathqubeQuote.query.get_or_404(id)
    if quote.stage != 'closed_won':
        flash('Tax invoice can only be generated for Closed Won quotes.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))
    try:
        inv = _create_tax_invoice_draft('bathqube', quote, current_user.id)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))
    db.session.commit()
    flash(f'Draft invoice {inv.invoice_number} created — review + edit before issuing.', 'success')
    return redirect(url_for('tax_invoice_edit', id=inv.id))


@app.route('/quotes/upvc/<int:id>/invoice/new', methods=['POST'])
@login_required
def upvc_tax_invoice_new(id):
    """Generate a draft tax invoice from a closed_won UPVC quote."""
    quote = UpvcQuote.query.get_or_404(id)
    if quote.stage != 'closed_won':
        flash('Tax invoice can only be generated for Closed Won quotes.', 'warning')
        return redirect(url_for('upvc_quote_view', id=id))
    try:
        inv = _create_tax_invoice_draft('upvc', quote, current_user.id)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('upvc_quote_view', id=id))
    db.session.commit()
    flash(f'Draft invoice {inv.invoice_number} created — review + edit before issuing.', 'success')
    return redirect(url_for('tax_invoice_edit', id=inv.id))


@app.route('/quotes/<int:id>/invoice/new', methods=['POST'])
@login_required
def quote_tax_invoice_new(id):
    """Generate a draft tax invoice from an Accepted regular Quote.

    Regular /quotes status flow is Draft → Sent → Accepted → (Rejected/
    Expired). Only Accepted quotes are eligible for invoicing — matches
    the Bathqube/UPVC `closed_won` gate.
    """
    from models import Quote
    quote = Quote.query.get_or_404(id)
    if quote.status != 'Accepted':
        flash(
            f'Tax invoice can only be generated for Accepted quotes. '
            f'This one is {quote.status}.',
            'warning',
        )
        return redirect(url_for('quote_view', id=id))
    try:
        inv = _create_tax_invoice_draft('lead', quote, current_user.id)
    except ValueError as e:
        flash(str(e), 'warning')
        return redirect(url_for('quote_view', id=id))
    db.session.commit()
    flash(f'Draft invoice {inv.invoice_number} created — review + edit before issuing.', 'success')
    return redirect(url_for('tax_invoice_edit', id=inv.id))


@app.route('/tax-invoices')
@login_required
def tax_invoice_list():
    """List all tax invoices — filterable by status + FY + date range."""
    from models import TaxInvoice
    status = (request.args.get('status') or '').strip()
    fy = (request.args.get('fy') or '').strip()
    search = (request.args.get('search') or '').strip()
    q = TaxInvoice.query
    if status:
        q = q.filter_by(status=status)
    if fy:
        q = q.filter_by(financial_year=fy)
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            TaxInvoice.invoice_number.ilike(like),
            TaxInvoice.buyer_name.ilike(like),
        ))
    invoices = q.order_by(TaxInvoice.id.desc()).limit(500).all()
    return render_template('tax_invoices/list.html', invoices=invoices,
                           status_filter=status, fy_filter=fy, search=search)


@app.route('/tax-invoices/<int:id>')
@login_required
def tax_invoice_view(id):
    from models import TaxInvoice
    inv = TaxInvoice.query.get_or_404(id)
    return render_template('tax_invoices/view.html', invoice=inv)


@app.route('/tax-invoices/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def tax_invoice_edit(id):
    """Edit a draft tax invoice (full edit) or an issued one (soft fields only).

    Soft fields = invoice metadata (vehicle no, dispatched_through,
    IRN paste-in, etc.) that don't affect the printed financial amounts.
    Money + line items are locked once status='issued'.
    """
    from models import TaxInvoice, TaxInvoiceItem
    inv = TaxInvoice.query.get_or_404(id)
    if inv.status == 'cancelled':
        flash('Cancelled invoices cannot be edited.', 'warning')
        return redirect(url_for('tax_invoice_view', id=id))

    if request.method == 'GET':
        return render_template('tax_invoices/form.html', invoice=inv,
                               hsn_defaults=TAX_INVOICE_HSN_DEFAULTS)

    form = request.form
    is_draft = (inv.status == 'draft')

    # Soft fields — always editable
    def _setf(field, key=None, transform=None):
        key = key or field
        v = (form.get(key) or '').strip()
        if transform and v:
            try:
                v = transform(v)
            except Exception:
                return
        setattr(inv, field, v or None)

    def _set_date(field, key=None):
        key = key or field
        v = (form.get(key) or '').strip()
        if not v:
            setattr(inv, field, None)
            return
        try:
            setattr(inv, field, datetime.strptime(v, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Customer / consignee text fields — locked after issue but soft enough
    # that we still allow address typo fixes on issued invoices.
    for f in ('buyer_name', 'buyer_address', 'buyer_gstin', 'buyer_pan',
              'buyer_state', 'buyer_state_code',
              'consignee_name', 'consignee_address', 'consignee_gstin',
              'consignee_state', 'consignee_state_code'):
        _setf(f)

    # Invoice metadata — always editable
    for f in ('buyers_order_no', 'delivery_note', 'dispatch_doc_no',
              'mode_of_payment', 'other_references', 'dispatched_through',
              'destination', 'terms_of_delivery', 'bill_of_lading',
              'motor_vehicle_no', 'ewaybill_no',
              'irn', 'ack_no'):
        _setf(f)
    for f in ('buyers_order_date', 'delivery_note_date',
              'bill_of_lading_date', 'ack_date'):
        _set_date(f)

    # Line items + amounts — draft-only
    if is_draft:
        new_items_raw = form.getlist('items_json')
        if new_items_raw:
            try:
                parsed = json.loads(new_items_raw[0])
            except Exception:
                parsed = []
            # Drop existing items + append fresh (use clear() for the
            # delete-orphan cascade to fire — avoids the same stale-
            # collection bug we hit on UPVC + Bathqube)
            inv.items.clear()
            db.session.flush()
            for i, row in enumerate(parsed):
                try:
                    qty = float(row.get('quantity') or 0)
                    rate = float(row.get('rate') or 0)
                except (TypeError, ValueError):
                    continue
                if qty <= 0 or rate <= 0:
                    continue
                desc = (row.get('description') or '').strip()[:500]
                if not desc:
                    continue
                inv.items.append(TaxInvoiceItem(
                    sort_order = i,
                    description = desc,
                    hsn_code = (row.get('hsn_code') or '').strip() or None,
                    quantity = qty,
                    unit = (row.get('unit') or 'nos').strip() or 'nos',
                    rate = rate,
                    amount = round(qty * rate, 2),
                    is_extra = bool(row.get('is_extra')),
                ))
        _recompute_tax_invoice_totals(inv)

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Save failed: {e}', 'danger')
        return redirect(url_for('tax_invoice_edit', id=inv.id))

    flash('Invoice saved.' if is_draft else 'Soft fields updated (issued invoice — amounts locked).',
          'success')
    return redirect(url_for('tax_invoice_view', id=inv.id))


@app.route('/tax-invoices/<int:id>/issue', methods=['POST'])
@login_required
def tax_invoice_issue(id):
    """Flip a draft to issued. Locks money + line items going forward."""
    from models import TaxInvoice
    inv = TaxInvoice.query.get_or_404(id)
    if inv.status != 'draft':
        flash(f'Invoice is already {inv.status}.', 'info')
        return redirect(url_for('tax_invoice_view', id=id))
    if not inv.items:
        flash('Cannot issue an empty invoice — add at least one line item first.', 'warning')
        return redirect(url_for('tax_invoice_edit', id=id))
    inv.status = 'issued'
    inv.issued_at = datetime.utcnow()
    _recompute_tax_invoice_totals(inv)
    db.session.commit()
    flash(f'Invoice {inv.invoice_number} issued. Money + line items are now locked.', 'success')
    return redirect(url_for('tax_invoice_view', id=inv.id))


@app.route('/tax-invoices/<int:id>/cancel', methods=['POST'])
@login_required
def tax_invoice_cancel(id):
    """Cancel an invoice. Doesn't delete the row — keeps the audit trail.
    A cancelled invoice frees up the source quote for re-invoicing."""
    from models import TaxInvoice
    inv = TaxInvoice.query.get_or_404(id)
    if inv.status == 'cancelled':
        flash('Invoice already cancelled.', 'info')
        return redirect(url_for('tax_invoice_view', id=id))
    reason = (request.form.get('reason') or '').strip()
    inv.status = 'cancelled'
    inv.cancelled_at = datetime.utcnow()
    inv.cancelled_reason = reason or 'Cancelled by user (no reason given)'
    db.session.commit()
    flash(f'Invoice {inv.invoice_number} cancelled.', 'warning')
    return redirect(url_for('tax_invoice_view', id=inv.id))


@app.route('/tax-invoices/<int:id>/pdf')
@login_required
def tax_invoice_pdf(id):
    """Download the tax invoice as PDF."""
    from models import TaxInvoice
    from utils.tax_invoice_pdf import generate_tax_invoice_pdf
    from flask import send_file
    from io import BytesIO
    inv = TaxInvoice.query.get_or_404(id)
    pdf_bytes = generate_tax_invoice_pdf(inv)
    filename = f"{inv.invoice_number.replace('/', '_')}.pdf"
    return send_file(BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ============================================================================
# GATE PASSES (dispatch / packing slip)
# ============================================================================

def _next_gate_pass_number():
    """Allocate the next sequential per-FY gate-pass number.
    Format: VTS/GP/<FY>/<NNNN> e.g. VTS/GP/2627/0001.
    """
    from models import GatePass
    fy = _current_financial_year()
    prefix = f'VTS/GP/{fy}/'
    last = (GatePass.query
                    .filter(GatePass.gp_number.like(f'{prefix}%'))
                    .order_by(GatePass.id.desc())
                    .first())
    seq = 1
    if last and last.gp_number:
        try:
            seq = int(last.gp_number.rsplit('/', 1)[-1]) + 1
        except (ValueError, IndexError):
            seq = GatePass.query.filter_by(financial_year=fy).count() + 1
    return f'{prefix}{seq:04d}', fy


def _mm_to_inches_display(mm):
    """Render millimetres as a fractional-inch string ("44 11/16") to
    match the Arihant packing slip's ACT(IN) column. Rounded to the
    nearest 16th of an inch — fine enough for fabrication."""
    try:
        mm_f = float(mm or 0)
    except (TypeError, ValueError):
        return ''
    if mm_f <= 0:
        return ''
    total_16ths = round(mm_f / 25.4 * 16)
    whole = total_16ths // 16
    num = total_16ths % 16
    if num == 0:
        return str(whole)
    # Reduce fraction
    from math import gcd
    g = gcd(num, 16)
    num //= g
    den = 16 // g
    return f'{whole} {num}/{den}' if whole else f'{num}/{den}'


def _mm_sqm(width_mm, height_mm, qty):
    """sqm = (w_mm × h_mm × qty) / 1_000_000"""
    try:
        return round(float(width_mm or 0) * float(height_mm or 0) * float(qty or 0) / 1_000_000, 4)
    except (TypeError, ValueError):
        return 0.0


def _mm_sqft(width_mm, height_mm, qty):
    """sqft = sqm × 10.7639104"""
    return round(_mm_sqm(width_mm, height_mm, qty) * 10.7639104, 4)


def _dispatched_qty_so_far(source_kind, source_item_id):
    """Sum qty_this_pass across all ISSUED gate-pass lines that point
    at this source item. Used to compute qty_dispatched_before when
    raising a new gate pass."""
    from models import GatePass, GatePassItem
    if not source_item_id or not source_kind:
        return 0.0
    rows = (db.session.query(db.func.sum(GatePassItem.qty_this_pass))
                       .join(GatePass, GatePassItem.gate_pass_id == GatePass.id)
                       .filter(GatePassItem.source_kind == source_kind,
                               GatePassItem.source_item_id == source_item_id,
                               GatePass.status == 'issued')
                       .scalar())
    return float(rows or 0)


# ──────────────────────────────────────────────────────────────────────────
# Dimension extraction helpers used by gate-pass seeding.
#
# The four sources vary in how (or whether) they carry W × H:
#   - UPVC:          structured (width/height/unit) — handled inline
#   - Regular Quote: structured (actual_width / actual_height / unit)
#   - Bathqube:      NOT on the item row — panels live inside
#                    config_data JSON on the parent quote; also embedded
#                    in the item description text like "46×96 in"
#   - Tax Invoice:   nothing structured; only description text
#
# `_GATE_PASS_DIM_RE` recognises `W × H` (or ×/x/*/X) pairs with an optional
# unit token before/after each number. Matches conservative — we require
# actual digits on both sides of the separator so "Bathroom-1" and phone
# numbers don't get mis-parsed as dims.
#
# `_MM_PER_UNIT` covers the units that show up in Vetrova data: mm/cm/m
# for metric, ft/in for imperial, ' / " (prime marks) for BD-typed text.

_GATE_PASS_DIM_RE = re.compile(
    r'(\d+(?:\.\d+)?)\s*(mm|cm|ft|in|m|["′″\'])?\s*'
    r'[×xX*]'
    r'\s*(\d+(?:\.\d+)?)\s*(mm|cm|ft|in|m|["′″\'])?',
    re.IGNORECASE,
)

_MM_PER_UNIT = {
    'mm': 1.0, 'cm': 10.0, 'm': 1000.0,
    'in': 25.4, '"': 25.4, '″': 25.4,
    'ft': 304.8, "'": 304.8, '′': 304.8,
}


def _to_mm(value, unit, default_unit='mm'):
    """Convert `value` in `unit` (or `default_unit` when unit is blank) to mm.
    Returns None on unparseable input."""
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    u = (unit or default_unit or 'mm').strip().lower()
    factor = _MM_PER_UNIT.get(u)
    if factor is None:
        return None
    return round(v * factor, 2)


def _parse_dims_from_text(text, default_unit='mm'):
    """Best-effort: pull the most-plausible (w_mm, h_mm) pair out of free
    text like:
        '46×96 in', '1200x2400', '48.00" × 91.00"', '4' × 7\\''
        'Panel 1: 46.00" × 96.00" (customer: 46×96 in) [30.66 sqft]'
        '30x40 — Sliding · 3-track · White · 100×100 in'   ← label at start

    When a description contains MULTIPLE candidate pairs, prefer one that
    has an explicit unit token attached — those are typed by BD as real
    dimensions, whereas unit-less pairs at the start are often the line's
    label (e.g. UPVC opening code "30x40"). Falls back to the first match
    when nothing has a unit.

    Returns (None, None) when no plausible match. Uses `default_unit` for
    the mm conversion when the winning match has no unit token.
    """
    if not text:
        return None, None
    best = None       # (w, h, unit) for the first unit-tagged match
    fallback = None   # (w, h, None) for the first plain match
    for m in _GATE_PASS_DIM_RE.finditer(text):
        w_raw, wu, h_raw, hu = m.group(1), m.group(2), m.group(3), m.group(4)
        unit_tok = (hu or wu or '').strip().lower() or None
        if unit_tok and best is None:
            best = (w_raw, h_raw, unit_tok)
        elif fallback is None:
            fallback = (w_raw, h_raw, None)
    chosen = best or fallback
    if not chosen:
        return None, None
    w_raw, h_raw, unit_tok = chosen
    unit = (unit_tok or default_unit or 'mm')
    return _to_mm(w_raw, unit, default_unit), _to_mm(h_raw, unit, default_unit)


def _bathqube_panel_dims_mm(quote):
    """Walk a BathqubeQuote's config_data JSON and yield the flattened panel
    dimensions in mm, in enclosure/panel order — same order they land in
    BathqubeQuoteItem rows on save.

    Yields (idx, w_mm, h_mm) for each panel. Silent no-op when config_data
    is missing / malformed / has no enclosures.
    """
    if not getattr(quote, 'config_data', None):
        return
    try:
        cd = json.loads(quote.config_data)
    except Exception:
        return
    default_unit = (cd.get('dimensionUnit') or 'mm').lower()
    idx = 0
    for enc in (cd.get('enclosures') or []):
        enc_unit = (enc.get('dimensionUnit') or default_unit).lower()
        for p in (enc.get('glassPanels') or []):
            w_mm = _to_mm(p.get('width'), enc_unit, enc_unit)
            h_mm = _to_mm(p.get('height'), enc_unit, enc_unit)
            if w_mm and h_mm:
                yield idx, w_mm, h_mm
            idx += 1


def _seed_gate_pass_from_tax_invoice(inv, gp):
    """Seed gate-pass items from a tax invoice's line items.
    Tax-invoice lines don't carry W×H structurally — best-effort regex-parse
    the description text; BD can override on the form."""
    from models import GatePassItem
    for idx, it in enumerate(inv.items):
        if it.is_extra:
            continue  # skip transport/install/discount rows
        already = _dispatched_qty_so_far('tax_invoice', it.id)
        w_mm, h_mm = _parse_dims_from_text(it.description or '')
        gp.items.append(GatePassItem(
            sort_order   = idx,
            material_spec= (it.description or '')[:200],
            ref_code     = '',
            work_order_no= '',
            width_mm     = w_mm,
            height_mm    = h_mm,
            width_in_display  = _mm_to_inches_display(w_mm) if w_mm else '',
            height_in_display = _mm_to_inches_display(h_mm) if h_mm else '',
            qty_ordered  = float(it.quantity or 0),
            qty_dispatched_before = already,
            qty_this_pass = max(0.0, float(it.quantity or 0) - already),
            source_kind  = 'tax_invoice',
            source_item_id = it.id,
        ))


def _seed_gate_pass_from_bathqube(quote, gp):
    """Seed a Bathqube quote's items into the gate pass.

    W × H per line: try TWO sources in priority order:
      1. `quote.config_data.enclosures[].glassPanels[]` — structured panel
         dimensions from the configurator, converted to mm using each
         enclosure's `dimensionUnit`. This is the source of truth. Panels
         are flattened in enclosure order — same order that item rows
         land in on save.
      2. Regex-parse from the item description text (e.g. "Panel 1:
         46.00\" × 96.00\"") as fallback when the snapshot is missing
         or the item index doesn't line up (item edited manually after
         save, extra panel inserted, etc.).
    """
    from models import GatePassItem, BathqubeQuoteItem
    rows = (BathqubeQuoteItem.query
                             .filter_by(quote_id=quote.id, is_extra=False)
                             .order_by(BathqubeQuoteItem.sort_order)
                             .all())
    # Fix A: pre-index dims by their flattened panel position.
    snapshot_dims = {i: (w, h) for i, w, h in _bathqube_panel_dims_mm(quote)}

    for idx, it in enumerate(rows):
        already = _dispatched_qty_so_far('bathqube', it.id)
        w_mm, h_mm = snapshot_dims.get(idx, (None, None))
        # Fix B: fall back to parsing the description text.
        if not (w_mm and h_mm):
            w_mm, h_mm = _parse_dims_from_text(it.description or '')
        gp.items.append(GatePassItem(
            sort_order   = idx,
            material_spec= (it.description or '')[:200],
            width_mm     = w_mm,
            height_mm    = h_mm,
            width_in_display  = _mm_to_inches_display(w_mm) if w_mm else '',
            height_in_display = _mm_to_inches_display(h_mm) if h_mm else '',
            qty_ordered  = float(it.quantity or 0),
            qty_dispatched_before = already,
            qty_this_pass = max(0.0, float(it.quantity or 0) - already),
            source_kind  = 'bathqube',
            source_item_id = it.id,
        ))


def _seed_gate_pass_from_upvc(quote, gp):
    from models import GatePassItem
    for idx, it in enumerate(sorted(quote.items, key=lambda x: x.sort_order or 0)):
        already = _dispatched_qty_so_far('upvc', it.id)
        # UPVC items already carry W×H — convert into mm so the PDF
        # column structure matches Arihant.
        from utils.bathqube_dimensions import to_inches
        w_in = float(to_inches(float(it.width or 0), it.unit) or 0)
        h_in = float(to_inches(float(it.height or 0), it.unit) or 0)
        w_mm = round(w_in * 25.4, 2) if w_in else None
        h_mm = round(h_in * 25.4, 2) if h_in else None
        desc_parts = []
        if it.label: desc_parts.append(it.label)
        desc_parts.append(f'UPVC {it.track_type}')
        if it.track_system: desc_parts.append(it.track_system)
        gp.items.append(GatePassItem(
            sort_order   = idx,
            material_spec= ' — '.join(desc_parts)[:200],
            width_mm     = w_mm,
            height_mm    = h_mm,
            width_in_display  = _mm_to_inches_display(w_mm) if w_mm else '',
            height_in_display = _mm_to_inches_display(h_mm) if h_mm else '',
            qty_ordered  = float(it.quantity or 0),
            qty_dispatched_before = already,
            qty_this_pass = max(0.0, float(it.quantity or 0) - already),
            sqft         = float(it.sqft or 0) * float(it.quantity or 0),
            source_kind  = 'upvc',
            source_item_id = it.id,
        ))


def _seed_gate_pass_from_lead_quote(quote, gp):
    """Seed from a regular Quote. The QuoteItem schema uses:
      - particular (description text)
      - actual_width / actual_height (mm; what was measured)
      - unit ('MM', 'sqft', …)
      - quantity (int)
    Hierarchical: rows with is_group=True are headers whose children
    carry the real dimensions. We walk the leaves only.
    """
    from models import GatePassItem, QuoteItem
    from utils.bathqube_dimensions import to_inches
    rows = (QuoteItem.query
                     .filter_by(quote_id=quote.id)
                     .order_by(QuoteItem.sort_order, QuoteItem.item_number)
                     .all())
    idx = 0
    for it in rows:
        if it.is_group:
            continue  # group/header — dimensions live on its children
        already = _dispatched_qty_so_far('lead', it.id)

        # actual_width/height are usually millimetres (unit='MM').
        # If unit is something else, run through to_inches first.
        unit_lc = (it.unit or 'mm').lower()
        w_raw = float(it.actual_width or 0)
        h_raw = float(it.actual_height or 0)
        if unit_lc == 'mm':
            w_mm = round(w_raw, 2) if w_raw else None
            h_mm = round(h_raw, 2) if h_raw else None
        else:
            w_in = float(to_inches(w_raw, unit_lc) or 0) if w_raw else 0
            h_in = float(to_inches(h_raw, unit_lc) or 0) if h_raw else 0
            w_mm = round(w_in * 25.4, 2) if w_in else None
            h_mm = round(h_in * 25.4, 2) if h_in else None

        qty = float(it.quantity or 0)
        # Description: parent group's particular + this row's particular
        # (so the gate pass shows e.g. "Bedroom — 8mm Toughened Glass")
        desc_parts = []
        if it.parent and getattr(it.parent, 'particular', None):
            desc_parts.append(it.parent.particular.strip())
        if it.particular:
            desc_parts.append(it.particular.strip())
        material_spec = ' — '.join(desc_parts)[:200] if desc_parts else None

        # Fix B fallback: many older regular quotes were typed without
        # actual_width/height but BD embedded dims in `particular` free
        # text (e.g. "8mm Toughened Glass 1200 x 2400 mm"). Try to
        # rescue those when the structured columns are empty.
        if not (w_mm and h_mm) and material_spec:
            w_mm, h_mm = _parse_dims_from_text(material_spec)

        gp.items.append(GatePassItem(
            sort_order   = idx,
            material_spec= material_spec,
            width_mm     = w_mm,
            height_mm    = h_mm,
            width_in_display  = _mm_to_inches_display(w_mm) if w_mm else '',
            height_in_display = _mm_to_inches_display(h_mm) if h_mm else '',
            qty_ordered  = qty,
            qty_dispatched_before = already,
            qty_this_pass = max(0.0, qty - already),
            sqft         = _mm_sqft(w_mm, h_mm, qty) if (w_mm and h_mm) else 0,
            sqm          = _mm_sqm(w_mm, h_mm, qty) if (w_mm and h_mm) else 0,
            source_kind  = 'lead',
            source_item_id = it.id,
        ))
        idx += 1


def _create_gate_pass_draft(source_kind, source_obj, creator_id):
    """Allocate the next GP number and seed line items.
    Caller commits. Unlike tax invoices, multiple gate passes per
    source are allowed — no duplicate-check guard."""
    from models import GatePass
    gp_number, fy = _next_gate_pass_number()

    # Pull customer + invoice context from whichever source we got
    customer_name = ''
    delivery_address = ''
    customer_gstin = None
    ref_invoice_no = None
    ref_invoice_date = None

    if source_kind == 'tax_invoice':
        customer_name = source_obj.buyer_name or ''
        delivery_address = source_obj.consignee_address or source_obj.buyer_address or ''
        customer_gstin = source_obj.buyer_gstin
        ref_invoice_no = source_obj.invoice_number
        ref_invoice_date = source_obj.invoice_date
    elif source_kind == 'bathqube':
        customer_name = source_obj.customer_name or ''
        delivery_address = getattr(source_obj, 'site_address', '') or ''
        customer_gstin = getattr(source_obj, 'gstin', None)
    elif source_kind == 'upvc':
        customer_name = source_obj.customer_name or ''
        delivery_address = getattr(source_obj, 'site_address', '') or ''
        customer_gstin = getattr(source_obj, 'gstin', None)
    elif source_kind == 'lead':
        customer_name = source_obj.customer_name or ''
        delivery_address = getattr(source_obj, 'customer_address', '') or ''
        customer_gstin = None

    gp = GatePass(
        gp_number = gp_number,
        financial_year = fy,
        gp_date = datetime.utcnow().date(),
        customer_name = customer_name,
        delivery_address = delivery_address,
        customer_gstin = customer_gstin,
        ref_invoice_no = ref_invoice_no,
        ref_invoice_date = ref_invoice_date,
        status = 'draft',
        prepared_by = creator_id,
    )
    if source_kind == 'tax_invoice':
        gp.tax_invoice_id = source_obj.id
        _seed_gate_pass_from_tax_invoice(source_obj, gp)
    elif source_kind == 'bathqube':
        gp.bathqube_quote_id = source_obj.id
        _seed_gate_pass_from_bathqube(source_obj, gp)
    elif source_kind == 'upvc':
        gp.upvc_quote_id = source_obj.id
        _seed_gate_pass_from_upvc(source_obj, gp)
    elif source_kind == 'lead':
        gp.lead_quote_id = source_obj.id
        _seed_gate_pass_from_lead_quote(source_obj, gp)

    db.session.add(gp)
    db.session.flush()
    return gp


def _recompute_gate_pass_item_metrics(item):
    """Refresh sqft + sqm from w_mm × h_mm × qty_this_pass."""
    qty = float(item.qty_this_pass or 0)
    item.sqm  = _mm_sqm(item.width_mm,  item.height_mm,  qty)
    item.sqft = _mm_sqft(item.width_mm, item.height_mm, qty)


# ── Routes: bridge from each source ──────────────────────────────────────

@app.route('/tax-invoices/<int:id>/gate-pass/new', methods=['POST'])
@login_required
def tax_invoice_gate_pass_new(id):
    from models import TaxInvoice
    inv = TaxInvoice.query.get_or_404(id)
    if inv.status == 'cancelled':
        flash('Cannot raise a gate pass against a cancelled invoice.', 'warning')
        return redirect(url_for('tax_invoice_view', id=id))
    gp = _create_gate_pass_draft('tax_invoice', inv, current_user.id)
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} drafted. Add vehicle, driver, qty and issue.', 'success')
    return redirect(url_for('gate_pass_edit', id=gp.id))


@app.route('/quotes/bathqube/<int:id>/gate-pass/new', methods=['POST'])
@login_required
def bathqube_gate_pass_new(id):
    quote = BathqubeQuote.query.get_or_404(id)
    gp = _create_gate_pass_draft('bathqube', quote, current_user.id)
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} drafted.', 'success')
    return redirect(url_for('gate_pass_edit', id=gp.id))


@app.route('/quotes/upvc/<int:id>/gate-pass/new', methods=['POST'])
@login_required
def upvc_gate_pass_new(id):
    from models import UpvcQuote
    quote = UpvcQuote.query.get_or_404(id)
    gp = _create_gate_pass_draft('upvc', quote, current_user.id)
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} drafted.', 'success')
    return redirect(url_for('gate_pass_edit', id=gp.id))


@app.route('/quotes/<int:id>/gate-pass/new', methods=['POST'])
@login_required
def quote_gate_pass_new(id):
    quote = Quote.query.get_or_404(id)
    gp = _create_gate_pass_draft('lead', quote, current_user.id)
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} drafted.', 'success')
    return redirect(url_for('gate_pass_edit', id=gp.id))


# ── Routes: list / view / edit / issue / cancel / pdf ───────────────────

@app.route('/gate-passes')
@login_required
def gate_pass_list():
    from models import GatePass
    search        = (request.args.get('search') or '').strip()
    status_filter = (request.args.get('status') or '').strip()
    fy_filter     = (request.args.get('fy') or '').strip()
    q = GatePass.query
    if status_filter:
        q = q.filter(GatePass.status == status_filter)
    if fy_filter:
        q = q.filter(GatePass.financial_year == fy_filter)
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            GatePass.gp_number.ilike(like),
            GatePass.customer_name.ilike(like),
            GatePass.vehicle_no.ilike(like),
        ))
    gps = q.order_by(GatePass.id.desc()).limit(500).all()
    return render_template('gate_passes/list.html', gate_passes=gps,
                           search=search, status_filter=status_filter,
                           fy_filter=fy_filter)


@app.route('/gate-passes/<int:id>')
@login_required
def gate_pass_view(id):
    from models import GatePass
    gp = GatePass.query.get_or_404(id)
    return render_template('gate_passes/view.html', gp=gp)


@app.route('/gate-passes/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def gate_pass_edit(id):
    from models import GatePass, GatePassItem
    gp = GatePass.query.get_or_404(id)
    if gp.status == 'cancelled':
        flash('Cancelled gate passes cannot be edited.', 'warning')
        return redirect(url_for('gate_pass_view', id=id))

    is_draft = (gp.status == 'draft')

    if request.method == 'GET':
        return render_template('gate_passes/form.html', gp=gp)

    # POST — save form
    # Header fields (logistics always editable; date/customer only on draft)
    if is_draft:
        gp.customer_name    = (request.form.get('customer_name') or '').strip() or gp.customer_name
        gp.delivery_address = (request.form.get('delivery_address') or '').strip()
        gp.customer_gstin   = (request.form.get('customer_gstin') or '').strip() or None
        gp.ref_invoice_no   = (request.form.get('ref_invoice_no') or '').strip() or None
        gp_date_str         = (request.form.get('gp_date') or '').strip()
        if gp_date_str:
            try:
                gp.gp_date = datetime.strptime(gp_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

    gp.vehicle_no       = (request.form.get('vehicle_no') or '').strip() or None
    gp.transporter_name = (request.form.get('transporter_name') or '').strip() or None
    gp.driver_name      = (request.form.get('driver_name') or '').strip() or None
    gp.driver_phone     = (request.form.get('driver_phone') or '').strip() or None
    gp.lr_number        = (request.form.get('lr_number') or '').strip() or None
    gp.eway_bill_no     = (request.form.get('eway_bill_no') or '').strip() or None
    gp.place_of_supply  = (request.form.get('place_of_supply') or '').strip() or None
    gp.remarks          = (request.form.get('remarks') or '').strip() or None

    if is_draft:
        # Line items — accept item IDs (for existing rows) + a parallel
        # set of indexed new-row fields. We clear then re-build.
        existing_by_id = {it.id: it for it in gp.items}
        keep_ids = set()
        new_items = []

        # Parse rows: each row has fields prefixed item_<idx>_<field>
        # Index is whatever ordering the form posts.
        idx = 0
        while True:
            row_marker = request.form.get(f'item_{idx}_present')
            if row_marker is None:
                break
            existing_id = request.form.get(f'item_{idx}_id', '').strip()
            material    = (request.form.get(f'item_{idx}_material') or '').strip()
            ref_code    = (request.form.get(f'item_{idx}_ref_code') or '').strip()
            work_order  = (request.form.get(f'item_{idx}_work_order') or '').strip()

            def _f(name, default=0):
                v = request.form.get(f'item_{idx}_{name}', '').strip()
                if v == '':
                    return default
                try:
                    return float(v)
                except ValueError:
                    return default

            width_mm    = _f('width_mm', 0) or None
            height_mm   = _f('height_mm', 0) or None
            qty_ordered = _f('qty_ordered', 0)
            qty_before  = _f('qty_before', 0)
            qty_now     = _f('qty_now', 0)

            flag_h   = bool(request.form.get(f'item_{idx}_flag_h'))
            flag_c   = bool(request.form.get(f'item_{idx}_flag_c'))
            flag_sp  = bool(request.form.get(f'item_{idx}_flag_sp'))
            flag_bh  = bool(request.form.get(f'item_{idx}_flag_bh'))
            flag_csk = bool(request.form.get(f'item_{idx}_flag_csk'))

            remarks_field = (request.form.get(f'item_{idx}_remarks') or '').strip() or None

            # Validate qty_now doesn't exceed remaining
            remaining = qty_ordered - qty_before
            if qty_now < 0:
                qty_now = 0
            if qty_ordered > 0 and qty_now > remaining:
                qty_now = remaining

            if existing_id and existing_id.isdigit() and int(existing_id) in existing_by_id:
                row = existing_by_id[int(existing_id)]
                keep_ids.add(row.id)
                row.sort_order = idx
                row.material_spec = material[:200] if material else None
                row.ref_code = ref_code[:60] if ref_code else None
                row.work_order_no = work_order[:60] if work_order else None
                row.width_mm = width_mm
                row.height_mm = height_mm
                row.width_in_display = _mm_to_inches_display(width_mm) if width_mm else ''
                row.height_in_display = _mm_to_inches_display(height_mm) if height_mm else ''
                row.qty_ordered = qty_ordered
                row.qty_dispatched_before = qty_before
                row.qty_this_pass = qty_now
                row.flag_h = flag_h
                row.flag_c = flag_c
                row.flag_sp = flag_sp
                row.flag_bh = flag_bh
                row.flag_csk = flag_csk
                row.remarks = remarks_field
                _recompute_gate_pass_item_metrics(row)
            else:
                # Skip totally empty new rows
                if (not material and not ref_code and qty_now <= 0
                        and not width_mm and not height_mm):
                    idx += 1
                    continue
                new_row = GatePassItem(
                    sort_order = idx,
                    material_spec = material[:200] if material else None,
                    ref_code = ref_code[:60] if ref_code else None,
                    work_order_no = work_order[:60] if work_order else None,
                    width_mm = width_mm,
                    height_mm = height_mm,
                    width_in_display = _mm_to_inches_display(width_mm) if width_mm else '',
                    height_in_display = _mm_to_inches_display(height_mm) if height_mm else '',
                    qty_ordered = qty_ordered,
                    qty_dispatched_before = qty_before,
                    qty_this_pass = qty_now,
                    flag_h = flag_h, flag_c = flag_c, flag_sp = flag_sp,
                    flag_bh = flag_bh, flag_csk = flag_csk,
                    remarks = remarks_field,
                    source_kind = 'manual',
                )
                _recompute_gate_pass_item_metrics(new_row)
                new_items.append(new_row)

            idx += 1

        # Drop rows the user removed from the form
        for it in list(gp.items):
            if it.id not in keep_ids:
                db.session.delete(it)
        for row in new_items:
            gp.items.append(row)

    db.session.commit()
    flash('Gate pass saved.' if is_draft else
          'Logistics fields updated (issued gate pass — qty locked).',
          'success')
    return redirect(url_for('gate_pass_view', id=gp.id))


@app.route('/gate-passes/<int:id>/issue', methods=['POST'])
@login_required
def gate_pass_issue(id):
    from models import GatePass
    gp = GatePass.query.get_or_404(id)
    if gp.status != 'draft':
        flash(f'Gate pass is already {gp.status}.', 'info')
        return redirect(url_for('gate_pass_view', id=id))
    if not gp.items or all(float(it.qty_this_pass or 0) <= 0 for it in gp.items):
        flash('Cannot issue — at least one line must have qty > 0.', 'warning')
        return redirect(url_for('gate_pass_edit', id=id))
    gp.status = 'issued'
    gp.issued_at = datetime.utcnow()
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} issued. Qty is now locked.', 'success')
    return redirect(url_for('gate_pass_view', id=gp.id))


@app.route('/gate-passes/<int:id>/cancel', methods=['POST'])
@login_required
def gate_pass_cancel(id):
    from models import GatePass
    gp = GatePass.query.get_or_404(id)
    if gp.status == 'cancelled':
        flash('Already cancelled.', 'info')
        return redirect(url_for('gate_pass_view', id=id))
    reason = (request.form.get('reason') or '').strip()
    gp.status = 'cancelled'
    gp.cancelled_at = datetime.utcnow()
    gp.cancelled_reason = reason or 'Cancelled by user (no reason given)'
    db.session.commit()
    flash(f'Gate Pass {gp.gp_number} cancelled.', 'warning')
    return redirect(url_for('gate_pass_view', id=gp.id))


@app.route('/gate-passes/<int:id>/pdf')
@login_required
def gate_pass_pdf(id):
    from models import GatePass
    from utils.gate_pass_pdf import generate_gate_pass_pdf
    from flask import send_file
    from io import BytesIO
    gp = GatePass.query.get_or_404(id)
    pdf_bytes = generate_gate_pass_pdf(gp)
    filename = f"{gp.gp_number.replace('/', '_')}.pdf"
    return send_file(BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ============================================================================
# PURCHASE INVOICES
# ============================================================================

def _next_pi_serial():
    """Generate next serial number PI-001, PI-002 …"""
    last = PurchaseInvoice.query.order_by(PurchaseInvoice.id.desc()).first()
    if not last:
        return 'PI-001'
    try:
        num = int(last.serial_number.split('-')[1]) + 1
    except Exception:
        num = PurchaseInvoice.query.count() + 1
    return f'PI-{num:03d}'


@app.route('/tally')
@login_required
def tally_index():
    # Manager/admin only — exposes cross-team revenue, PIs, dispatch state
    # and salesman-level totals. Sales users don't need visibility into
    # peers' commercial data.
    if not current_user.is_manager_or_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    from models import (Quote, PurchaseInvoice, Supplier, User as UserModel,
                        BathqubeQuote, TaxInvoice, GatePass, GatePassItem)

    date_from       = request.args.get('date_from', '')
    date_to         = request.args.get('date_to', '')
    salesman_id     = request.args.get('salesman_id', '')
    client_name     = request.args.get('client_name', '').strip()
    supplier_id     = request.args.get('supplier_id', '')
    delivery_status = request.args.get('delivery_status', '')
    # Phase 1 tally filters (manager brief):
    #   payment_status — Pending / Partial / Received  (derived from amounts)
    #   dispatch_state — None / Partial / Complete     (from GatePass sums)
    #   tax_invoice    — not_raised / draft / issued   (from TaxInvoice.status)
    #   overdue        — '1' to show only overdue rows
    # All four are applied post-aggregation in Python (derived fields aren't
    # cheap to push into SQL without extra joins/subqueries; row count is
    # capped by date+source filters so O(n) Python scan is fine).
    payment_status  = request.args.get('payment_status', '')
    dispatch_state  = request.args.get('dispatch_state', '')
    tax_invoice     = request.args.get('tax_invoice', '')
    overdue_only    = request.args.get('overdue', '') == '1'
    # A row is "overdue" when it was accepted more than OVERDUE_DAYS ago
    # and still has pending payment OR pending delivery.
    OVERDUE_DAYS = 15
    # One-click source toggle — '' = all, 'bathqube' = Bathqube only,
    # 'regular' = everything except Bathqube. Used by the segmented button
    # group at the top of tally/index.html.
    source_filter   = request.args.get('source', '')

    # ── Regular Quote rows (status='Accepted') ────────────────────────────────
    q = Quote.query.filter(Quote.status == 'Accepted')
    if date_from:
        q = q.filter(Quote.quote_date >= date_from)
    if date_to:
        q = q.filter(Quote.quote_date <= date_to)
    if salesman_id:
        q = q.filter(Quote.created_by == int(salesman_id))
    if client_name:
        q = q.filter(Quote.customer_name.ilike(f'%{client_name}%'))
    if delivery_status:
        q = q.filter(Quote.delivery_status == delivery_status)

    # Skip the regular-quote query entirely when the user clicks "Bathqube only".
    quotes = [] if source_filter == 'bathqube' else q.order_by(Quote.quote_date.desc()).all()

    def _row_from(quote, pis, source, sort_date):
        """Build a uniform tally row dict for either source. The template only
        reads normalised keys, not the underlying type — so bathqube and
        regular quotes render with the same column structure."""
        sale_value = float(getattr(quote, 'total', 0) or 0)
        # Bathqube quotes use revised_total when set; fall back to total.
        if source == 'bathqube' and getattr(quote, 'revised_total', None) is not None:
            sale_value = float(quote.revised_total)
        pi_amount   = sum(float(pi.invoice_amount or 0) for pi in pis)
        pi_paid     = sum(float(pi.amount_paid    or 0) for pi in pis)
        misc        = float(quote.misc_purchases or 0)
        total_cost  = pi_amount + misc
        profit      = sale_value - total_cost
        cash_recv   = float(quote.cash_received   or 0)

        # ── Bathqube payment receipts → "Tally updates the moment BD saves
        #    a receipt" ────────────────────────────────────────────────────
        # Legacy path: BD typed amount_received directly via the Tally edit
        # form. New path (since the BathqubePaymentReceipt collection
        # landed): BD records one UTR-audited receipt per inflow at
        # /quotes/bathqube/<id>/receipts. The receipts sum is the source of
        # truth; the legacy flat field is a fallback ONLY for quotes that
        # pre-date the receipts system (zero receipts on file).
        # Net effect: as soon as BD saves a receipt, the next /tally render
        # picks up the new total without any cron / hook / denormalisation.
        receipt_count = 0
        receipts_sum  = 0.0
        if source == 'bathqube':
            receipts_sum  = float(getattr(quote, 'paid_via_receipts', 0) or 0)
            receipt_count = len(getattr(quote, 'payment_receipts', []) or [])
            online_recv = receipts_sum if receipts_sum > 0 else float(quote.amount_received or 0)
        else:
            online_recv = float(quote.amount_received or 0)
        total_recv  = cash_recv + online_recv

        # Per-source display fields. Regular Quote has quote_number, quote_date,
        # creator, customer_city, customer_phone, client_payment_status. Bathqube
        # uses different field names — normalise them here.
        if source == 'bathqube':
            display = {
                'display_number':         quote.estimate_number or f'BQ-{quote.id}',
                'display_date':           quote.created_at.date(),
                'creator_name':           '—',                   # bathqube has no salesperson
                'creator_initials':       'BQ',
                'customer_name':          quote.customer_name,
                'customer_city':          quote.pincode or '',
                'customer_phone':         quote.phone,
                'client_payment_status':  ('Paid' if total_recv >= sale_value > 0
                                            else 'Partial' if total_recv > 0
                                            else 'Unpaid'),
                'view_url':               url_for('bathqube_quote_view', id=quote.id),
                'tally_update_url':       url_for('bathqube_quote_tally_update', id=quote.id),
            }
        else:
            display = {
                'display_number':         quote.quote_number,
                'display_date':           quote.quote_date,
                'creator_name':           quote.creator.username if quote.creator else '—',
                'creator_initials':       (quote.creator.username[:2].upper() if quote.creator else '?'),
                'customer_name':          quote.customer_name,
                'customer_city':          getattr(quote, 'customer_city', '') or '',
                'customer_phone':         getattr(quote, 'customer_phone', '') or '',
                'client_payment_status':  quote.client_payment_status,
                'view_url':               url_for('quote_view', id=quote.id),
                'tally_update_url':       url_for('quote_tally_update', id=quote.id),
            }

        return {
            'quote':             quote,
            'source':            source,                       # 'regular' | 'bathqube'
            'sort_date':         sort_date,
            'purchase_invoices': pis,
            'sale_amount':       sale_value,
            'pi_amount':         pi_amount,
            'pi_paid':           pi_paid,
            'pi_balance':        pi_amount - pi_paid,
            'misc_purchases':    misc,
            'total_cost':        total_cost,
            'profit':            profit,
            'margin':            (profit / sale_value * 100) if sale_value > 0 else 0,
            'cash_received':     cash_recv,
            'online_received':   online_recv,
            'amount_received':   total_recv,
            # Receipts metadata (bathqube only — both 0 for regular quotes).
            # Lets the template surface a "N receipts (₹X)" hint under the
            # amount cell so BD can tell at a glance that the figure came
            # from the receipts collection, not from a manual Tally edit.
            'receipt_count':     receipt_count,
            'receipts_sum':      receipts_sum,
            'client_balance':    sale_value - total_recv,
            'delivery_status':   quote.delivery_status,
            **display,
        }

    # ── Batch-load derived state per source (single query each) ──────────
    # dispatch_state depends on issued GatePassItems summed per source item,
    # then compared to source items' ordered qty. Cheap enough at tally
    # granularity to compute per-request without caching.
    def _dispatch_state_for(source_kind, source_id, ordered_qty_total):
        """Return one of: 'none' | 'partial' | 'complete'"""
        if ordered_qty_total <= 0:
            return 'none'
        # Sum qty_this_pass across all issued gate passes for this source.
        # We match through GatePass's polymorphic FK columns (only one set).
        col = {
            'regular':  GatePass.lead_quote_id,
            'bathqube': GatePass.bathqube_quote_id,
        }[source_kind]
        total_disp = (db.session.query(db.func.coalesce(db.func.sum(GatePassItem.qty_this_pass), 0))
                        .join(GatePass, GatePassItem.gate_pass_id == GatePass.id)
                        .filter(col == source_id, GatePass.status == 'issued')
                        .scalar() or 0)
        total_disp = float(total_disp)
        if total_disp <= 0:
            return 'none'
        if total_disp >= ordered_qty_total - 0.001:  # small epsilon for float compare
            return 'complete'
        return 'partial'

    def _tax_invoice_state_for(source_kind, source_id):
        """Return one of: 'not_raised' | 'draft' | 'issued'
        (cancelled invoices don't count — treat source as 'not_raised' again)."""
        col = {
            'regular':  TaxInvoice.lead_quote_id,
            'bathqube': TaxInvoice.bathqube_quote_id,
        }[source_kind]
        # Prefer 'issued' if any; else 'draft' if any; else 'not_raised'.
        found = (TaxInvoice.query
                    .filter(col == source_id, TaxInvoice.status != 'cancelled')
                    .with_entities(TaxInvoice.status)
                    .all())
        if not found:
            return 'not_raised'
        statuses = {s[0] for s in found}
        if 'issued' in statuses:
            return 'issued'
        return 'draft'

    def _ordered_qty_for(source_kind, quote_obj):
        """Total qty across non-extra source lines. Used for dispatch %."""
        if source_kind == 'regular':
            return sum(float(i.quantity or 0) for i in (quote_obj.items or [])
                       if not getattr(i, 'is_group', False))
        # bathqube
        return sum(float(i.quantity or 0) for i in (quote_obj.items or [])
                   if not getattr(i, 'is_extra', False))

    from datetime import date as _date
    today = _date.today()

    def _enrich(row, quote_obj, source_kind):
        """Add phase-1 derived fields onto a row dict in place."""
        ordered = _ordered_qty_for(source_kind, quote_obj)
        row['dispatch_state']    = _dispatch_state_for(source_kind, quote_obj.id, ordered)
        row['tax_invoice_state'] = _tax_invoice_state_for(source_kind, quote_obj.id)
        # Payment status: normalise to lowercase for the filter — regular Quote
        # yields 'Pending'/'Partial'/'Received', Bathqube 'Paid'/'Partial'/'Unpaid'.
        cps = (row.get('client_payment_status') or '').lower()
        if cps in ('received', 'paid'):
            row['payment_state'] = 'received'
        elif cps == 'partial':
            row['payment_state'] = 'partial'
        else:
            row['payment_state'] = 'pending'
        # Overdue: accepted more than N days ago AND (payment still pending
        # OR delivery still pending). Uses sort_date (quote_date or created_at).
        sd = row.get('sort_date')
        days_old = (today - sd).days if sd else 0
        row['days_since_sale'] = days_old
        row['is_overdue'] = (
            days_old > OVERDUE_DAYS and (
                row['payment_state'] in ('pending', 'partial') or
                row['delivery_status'] != 'Delivered'
            )
        )

    rows = []
    for quote in quotes:
        pis = PurchaseInvoice.query.filter_by(quote_id=quote.id).all()
        if supplier_id and not any(str(pi.supplier_id) == supplier_id for pi in pis):
            continue
        row = _row_from(quote, pis, 'regular', quote.quote_date)
        _enrich(row, quote, 'regular')
        rows.append(row)

    # ── Bathqube quote rows (stage='closed_won') ──────────────────────────────
    bq = BathqubeQuote.query.filter(BathqubeQuote.stage == 'closed_won')
    if date_from:
        bq = bq.filter(BathqubeQuote.created_at >= date_from)
    if date_to:
        # add a day so end-of-day is inclusive
        bq = bq.filter(BathqubeQuote.created_at <= date_to + ' 23:59:59')
    if client_name:
        bq = bq.filter(BathqubeQuote.customer_name.ilike(f'%{client_name}%'))
    if delivery_status:
        bq = bq.filter(BathqubeQuote.delivery_status == delivery_status)
    # salesman_id intentionally not applied to bathqube — these come from the
    # public configurator, no created_by salesperson.

    # Skip the bathqube query when the user clicks "Other" (regular) only.
    bquotes = [] if source_filter == 'regular' else bq.order_by(BathqubeQuote.created_at.desc()).all()
    for quote in bquotes:
        pis = PurchaseInvoice.query.filter_by(bathqube_quote_id=quote.id).all()
        if supplier_id and not any(str(pi.supplier_id) == supplier_id for pi in pis):
            continue
        row = _row_from(quote, pis, 'bathqube', quote.created_at.date())
        _enrich(row, quote, 'bathqube')
        rows.append(row)

    # ── Phase-1 post-filters (payment / dispatch / tax invoice / overdue) ──
    if payment_status:
        rows = [r for r in rows if r['payment_state'] == payment_status]
    if dispatch_state:
        rows = [r for r in rows if r['dispatch_state'] == dispatch_state]
    if tax_invoice:
        rows = [r for r in rows if r['tax_invoice_state'] == tax_invoice]
    if overdue_only:
        rows = [r for r in rows if r['is_overdue']]

    # Sort the combined list by date desc so newest are on top regardless of source.
    rows.sort(key=lambda r: r['sort_date'], reverse=True)

    totals = {
        'sales':             sum(r['sale_amount']     for r in rows),
        'pi_amount':         sum(r['pi_amount']       for r in rows),
        'misc':              sum(r['misc_purchases']  for r in rows),
        'total_cost':        sum(r['total_cost']      for r in rows),
        'profit':            sum(r['profit']          for r in rows),
        'pi_paid':           sum(r['pi_paid']         for r in rows),
        'pi_balance':        sum(r['pi_balance']      for r in rows),
        'cash_received':     sum(r['cash_received']   for r in rows),
        'online_received':   sum(r['online_received'] for r in rows),
        'amount_received':   sum(r['amount_received'] for r in rows),
        'client_balance':    sum(r['client_balance']  for r in rows),
        'unprocessed_count': sum(1 for r in rows if not r['purchase_invoices']),
    }

    salesmen  = UserModel.query.order_by(UserModel.username).all()
    suppliers = Supplier.query.order_by(Supplier.name).all()

    filters = {
        'date_from':       date_from,
        'date_to':         date_to,
        'salesman_id':     salesman_id,
        'client_name':     client_name,
        'supplier_id':     supplier_id,
        'delivery_status': delivery_status,
        'payment_status':  payment_status,
        'dispatch_state':  dispatch_state,
        'tax_invoice':     tax_invoice,
        'overdue':         '1' if overdue_only else '',
        'source':          source_filter,
    }

    return render_template('tally/index.html', rows=rows, totals=totals,
                           salesmen=salesmen, suppliers=suppliers, filters=filters)


@app.route('/quotes/<int:id>/tally-update', methods=['POST'])
@login_required
def quote_tally_update(id):
    if not current_user.is_manager_or_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('quote_view', id=id))
    from models import Quote
    quote = Quote.query.get_or_404(id)
    quote.delivery_status = request.form.get('delivery_status', quote.delivery_status)
    online_recv = request.form.get('amount_received', '').strip()
    cash_recv   = request.form.get('cash_received',   '').strip()
    misc        = request.form.get('misc_purchases',  '').strip()
    quote.amount_received = float(online_recv) if online_recv else quote.amount_received
    quote.cash_received   = float(cash_recv)   if cash_recv   else quote.cash_received
    quote.misc_purchases  = float(misc)        if misc        else quote.misc_purchases
    db.session.commit()
    flash('Tally info updated.', 'success')
    return redirect(url_for('quote_view', id=id))


@app.route('/quotes/bathqube/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def bathqube_quote_delete(id):
    """Admin-only delete. Cascade-deletes items/events/revisions automatically
    (cascade on the relationship). Blocks with a clear message if there are
    linked PurchaseInvoices — those are financial records and shouldn't be
    silently destroyed with the quote."""
    quote = BathqubeQuote.query.get_or_404(id)

    # Refuse if PIs are linked — admin must detach/delete those first.
    pi_count = quote.purchase_invoices.count()
    if pi_count:
        flash(f'Cannot delete: {pi_count} purchase invoice(s) still linked to this quote. '
              f'Delete or unlink them first.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))

    label = quote.estimate_number or f'BQ-{quote.id}'
    db.session.delete(quote)
    db.session.commit()
    flash(f'Bathqube quote {label} deleted.', 'success')
    return redirect(url_for('bathqube_quotes_list'))


@app.route('/quotes/bathqube/<int:id>/tally-update', methods=['POST'])
@login_required
def bathqube_quote_tally_update(id):
    """Mirror of quote_tally_update for bathqube quotes — same form fields,
    same semantics, just a different table."""
    if not current_user.is_manager_or_admin():
        flash('Access denied.', 'danger')
        return redirect(url_for('bathqube_quote_view', id=id))
    quote = BathqubeQuote.query.get_or_404(id)
    quote.delivery_status = request.form.get('delivery_status', quote.delivery_status)
    online_recv = request.form.get('amount_received', '').strip()
    cash_recv   = request.form.get('cash_received',   '').strip()
    misc        = request.form.get('misc_purchases',  '').strip()
    if online_recv:
        quote.amount_received = float(online_recv)
    if cash_recv:
        quote.cash_received = float(cash_recv)
    if misc:
        quote.misc_purchases = float(misc)
    db.session.commit()
    flash('Tally info updated.', 'success')
    # Tally update is usually triggered from the Tally page, so return there.
    return redirect(url_for('tally_index'))


@app.route('/purchase-invoices')
@login_required
def purchase_invoices_list():
    supplier_filter  = request.args.get('supplier', '')
    project_filter   = request.args.get('project', '')
    type_filter      = request.args.get('invoice_type', '')
    status_filter    = request.args.get('status', '')

    query = PurchaseInvoice.query

    if supplier_filter:
        query = query.filter(PurchaseInvoice.supplier_id == supplier_filter)
    if project_filter:
        query = query.filter(PurchaseInvoice.project_id == project_filter)
    if type_filter:
        query = query.filter(PurchaseInvoice.invoice_type == type_filter)
    if status_filter:
        query = query.filter(PurchaseInvoice.status == status_filter)

    invoices  = query.order_by(PurchaseInvoice.id.desc()).all()
    suppliers = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    projects  = Project.query.order_by(Project.name).all()

    return render_template('purchase_invoices/list.html',
                           invoices=invoices,
                           suppliers=suppliers,
                           projects=projects,
                           supplier_filter=supplier_filter,
                           project_filter=project_filter,
                           type_filter=type_filter,
                           status_filter=status_filter)


@app.route('/purchase-invoices/new', methods=['GET', 'POST'])
@login_required
@manager_or_admin_required
def purchase_invoice_new():
    from models import Quote
    suppliers      = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    accepted_quotes = Quote.query.filter_by(status='Accepted').order_by(Quote.quote_date.desc()).all()
    closed_won_bathqube = BathqubeQuote.query.filter_by(stage='closed_won').order_by(BathqubeQuote.created_at.desc()).all()

    if request.method == 'POST':
        supplier_id    = request.form.get('supplier_id', '').strip()
        # linked_quote is "regular:<id>" or "bathqube:<id>". Old "quote_id"
        # field is kept for backwards-compat with any external callers.
        linked_quote   = request.form.get('linked_quote', '').strip()
        legacy_quote_id = request.form.get('quote_id', '').strip()
        bill_number    = request.form.get('bill_number', '').strip()
        invoice_type   = request.form.get('invoice_type', 'GST')
        invoice_amount = request.form.get('invoice_amount', '').strip()
        amount_paid    = request.form.get('amount_paid', '0').strip()
        notes          = request.form.get('notes', '').strip()
        bill_image     = request.files.get('bill_image')

        # Resolve linked quote into (quote_id, bathqube_quote_id)
        quote_id_val = None
        bathqube_quote_id_val = None
        if linked_quote:
            if linked_quote.startswith('bathqube:'):
                bathqube_quote_id_val = int(linked_quote.split(':', 1)[1])
            elif linked_quote.startswith('regular:'):
                quote_id_val = int(linked_quote.split(':', 1)[1])
        elif legacy_quote_id:
            quote_id_val = int(legacy_quote_id)

        errors = []
        if not supplier_id:
            errors.append('Vendor is required.')
        if quote_id_val is None and bathqube_quote_id_val is None:
            errors.append('Linked Quotation is required.')
        if not bill_number:
            errors.append('Bill Number is required.')
        if not bill_image or bill_image.filename == '':
            errors.append('Bill Image is required.')

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('purchase_invoices/form.html',
                                   suppliers=suppliers, accepted_quotes=accepted_quotes,
                                   closed_won_bathqube=closed_won_bathqube, invoice=None)

        bill_image_url = None
        try:
            from utils.s3_upload import S3Uploader
            from werkzeug.utils import secure_filename
            uploader  = S3Uploader()
            filename  = secure_filename(bill_image.filename)
            ext       = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'jpg'
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            s3_key    = f"purchase-invoices/{timestamp}_{filename}"
            content_types = {'pdf': 'application/pdf', 'jpg': 'image/jpeg',
                             'jpeg': 'image/jpeg', 'png': 'image/png'}
            content_type = content_types.get(ext, 'application/octet-stream')
            uploader.s3_client.upload_fileobj(
                bill_image, uploader.bucket_name, s3_key,
                ExtraArgs={'ContentType': content_type}
            )
            bill_image_url = f"https://{uploader.bucket_name}.s3.{uploader.region}.amazonaws.com/{s3_key}"
        except Exception as e:
            flash(f'Image upload failed: {e}', 'warning')

        invoice = PurchaseInvoice(
            serial_number     = _next_pi_serial(),
            supplier_id       = int(supplier_id),
            quote_id          = quote_id_val,
            bathqube_quote_id = bathqube_quote_id_val,
            bill_number       = bill_number,
            bill_image_url    = bill_image_url,
            invoice_type      = invoice_type,
            invoice_amount    = float(invoice_amount) if invoice_amount else None,
            amount_paid       = float(amount_paid) if amount_paid else 0.0,
            notes             = notes or None,
            created_by        = current_user.id,
        )
        db.session.add(invoice)
        db.session.commit()
        flash(f'Purchase Invoice {invoice.serial_number} created successfully.', 'success')
        return redirect(url_for('purchase_invoices_list'))

    return render_template('purchase_invoices/form.html',
                           suppliers=suppliers, accepted_quotes=accepted_quotes,
                           closed_won_bathqube=closed_won_bathqube, invoice=None)


@app.route('/purchase-invoices/<int:id>')
@login_required
def purchase_invoice_view(id):
    invoice = PurchaseInvoice.query.get_or_404(id)
    bill_image_presigned = None
    if invoice.bill_image_url:
        try:
            from utils.s3_upload import S3Uploader
            uploader = S3Uploader()
            s3_key = invoice.bill_image_url.split(f"{uploader.bucket_name}.s3.{uploader.region}.amazonaws.com/")[1]
            bill_image_presigned = uploader.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': uploader.bucket_name, 'Key': s3_key},
                ExpiresIn=3600
            )
        except Exception:
            bill_image_presigned = invoice.bill_image_url
    return render_template('purchase_invoices/view.html', invoice=invoice, bill_image_presigned=bill_image_presigned)


@app.route('/purchase-invoices/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@manager_or_admin_required
def purchase_invoice_edit(id):
    from models import Quote
    invoice         = PurchaseInvoice.query.get_or_404(id)
    suppliers       = Supplier.query.filter_by(is_active=True).order_by(Supplier.name).all()
    accepted_quotes = Quote.query.filter_by(status='Accepted').order_by(Quote.quote_date.desc()).all()

    if request.method == 'POST':
        invoice.supplier_id  = int(request.form.get('supplier_id'))
        qid = request.form.get('quote_id', '').strip()
        invoice.quote_id     = int(qid) if qid else None
        invoice.bill_number  = request.form.get('bill_number', '').strip()
        invoice.invoice_type = request.form.get('invoice_type', 'GST')
        amt  = request.form.get('invoice_amount', '').strip()
        paid = request.form.get('amount_paid', '0').strip()
        invoice.invoice_amount = float(amt)  if amt  else None
        invoice.amount_paid    = float(paid) if paid else 0.0
        invoice.notes          = request.form.get('notes', '').strip() or None

        new_image = request.files.get('bill_image')
        if new_image and new_image.filename != '':
            try:
                from utils.s3_upload import S3Uploader
                from werkzeug.utils import secure_filename
                uploader  = S3Uploader()
                filename  = secure_filename(new_image.filename)
                ext       = filename.rsplit('.', 1)[-1].lower() if '.' in filename else 'jpg'
                timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
                s3_key    = f"purchase-invoices/{timestamp}_{filename}"
                content_types = {'pdf': 'application/pdf', 'jpg': 'image/jpeg',
                                 'jpeg': 'image/jpeg', 'png': 'image/png'}
                content_type = content_types.get(ext, 'application/octet-stream')
                uploader.s3_client.upload_fileobj(
                    new_image, uploader.bucket_name, s3_key,
                    ExtraArgs={'ContentType': content_type}
                )
                invoice.bill_image_url = f"https://{uploader.bucket_name}.s3.{uploader.region}.amazonaws.com/{s3_key}"
            except Exception as e:
                flash(f'Image upload failed: {e}', 'warning')

        db.session.commit()
        flash('Purchase Invoice updated successfully.', 'success')
        return redirect(url_for('purchase_invoice_view', id=invoice.id))

    bill_image_presigned = None
    if invoice.bill_image_url:
        try:
            from utils.s3_upload import S3Uploader
            uploader = S3Uploader()
            s3_key = invoice.bill_image_url.split(f"{uploader.bucket_name}.s3.{uploader.region}.amazonaws.com/")[1]
            bill_image_presigned = uploader.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': uploader.bucket_name, 'Key': s3_key},
                ExpiresIn=3600
            )
        except Exception:
            bill_image_presigned = invoice.bill_image_url
    return render_template('purchase_invoices/form.html',
                           suppliers=suppliers, accepted_quotes=accepted_quotes, invoice=invoice,
                           bill_image_presigned=bill_image_presigned)


@app.route('/purchase-invoices/<int:id>/delete', methods=['POST'])
@login_required
@admin_required
def purchase_invoice_delete(id):
    invoice = PurchaseInvoice.query.get_or_404(id)
    db.session.delete(invoice)
    db.session.commit()
    flash('Purchase Invoice deleted.', 'success')
    return redirect(url_for('purchase_invoices_list'))


# Quick-add vendor (inline modal)
@app.route('/api/suppliers/quick-add', methods=['POST'])
@login_required
@manager_or_admin_required
def supplier_quick_add():
    name  = request.form.get('name',  '').strip()
    phone = request.form.get('phone', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Vendor name is required.'}), 400
    if not phone:
        return jsonify({'success': False, 'error': 'Phone number is required.'}), 400
    if Supplier.query.filter_by(name=name).first():
        return jsonify({'success': False, 'error': 'Vendor already exists.'}), 400
    supplier = Supplier(
        name=name,
        phone=phone,
        gstin=request.form.get('gstin', '').strip() or None,
        address=request.form.get('address', '').strip() or None,
        is_active=True,
    )
    db.session.add(supplier)
    db.session.commit()
    return jsonify({'success': True, 'id': supplier.id, 'name': supplier.name})


# Quick-add project (inline modal)
@app.route('/api/projects/quick-add', methods=['POST'])
@login_required
@manager_or_admin_required
def project_quick_add():
    name = request.form.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'error': 'Project name is required.'}), 400
    project = Project(
        name=name,
        status='Not Started',
        owner_id=current_user.id,
    )
    db.session.add(project)
    db.session.commit()
    return jsonify({'success': True, 'id': project.id, 'name': project.name})


# ============================================================================
# MEETINGS MODULE
# ============================================================================

MEETING_TYPES = ['Lead Visit', 'Site Survey', 'Installation', 'Follow-up', 'General']
MEETING_STATUSES = ['Scheduled', 'Checked In', 'Completed', 'Cancelled']


@app.route('/meetings')
@login_required
def meetings_list():
    """List all meetings — manager/admin see all, promotor sees own"""
    from models import Meeting, User

    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    status_filter = request.args.get('status', '')
    type_filter = request.args.get('meeting_type', '')
    user_filter = request.args.get('user_id', '')

    query = Meeting.query
    if not current_user.is_manager_or_admin():
        query = query.filter(Meeting.user_id == current_user.id)
    elif user_filter:
        try:
            query = query.filter(Meeting.user_id == int(user_filter))
        except ValueError:
            pass

    if status_filter:
        query = query.filter(Meeting.status == status_filter)
    if type_filter:
        query = query.filter(Meeting.meeting_type == type_filter)
    if date_from:
        try:
            query = query.filter(Meeting.scheduled_at >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Meeting.scheduled_at <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            pass

    # Portable NULL-last ordering: IS NULL sorts 0 (non-null) before 1 (null)
    meetings = query.order_by(Meeting.scheduled_at.is_(None), Meeting.scheduled_at.desc(), Meeting.created_at.desc()).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all() if current_user.is_manager_or_admin() else []

    return render_template('meetings/list.html',
                           meetings=meetings,
                           users=users,
                           meeting_types=MEETING_TYPES,
                           meeting_statuses=MEETING_STATUSES)


@app.route('/meetings/new', methods=['GET', 'POST'])
@login_required
def meeting_new():
    """Create a new meeting"""
    from models import Meeting, Lead, User

    if request.method == 'POST':
        try:
            data = request.form
            scheduled_str = data.get('scheduled_at', '').strip()
            scheduled_at = datetime.strptime(scheduled_str, '%Y-%m-%dT%H:%M') if scheduled_str else None

            lead_id = int(data.get('lead_id')) if data.get('lead_id') else None
            project_id = int(data.get('project_id')) if data.get('project_id') else None
            user_id = int(data.get('user_id')) if data.get('user_id') else current_user.id

            meeting = Meeting(
                title=data.get('title', '').strip(),
                user_id=user_id,
                created_by=current_user.id,
                lead_id=lead_id,
                project_id=project_id,
                client_name=data.get('client_name', '').strip() or None,
                client_phone=data.get('client_phone', '').strip() or None,
                client_address=data.get('client_address', '').strip() or None,
                meeting_type=data.get('meeting_type', 'Lead Visit'),
                scheduled_at=scheduled_at,
                notes=data.get('notes', '').strip() or None,
                status='Scheduled',
            )
            db.session.add(meeting)
            db.session.commit()
            flash(f'Meeting "{meeting.title}" created successfully!', 'success')
            return redirect(url_for('meeting_view', id=meeting.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating meeting: {str(e)}', 'danger')

    leads = Lead.query.order_by(Lead.name).all()
    projects = Project.query.order_by(Project.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all() if current_user.is_manager_or_admin() else [current_user]
    return render_template('meetings/form.html', title='New Meeting',
                           meeting=None, leads=leads, projects=projects,
                           users=users, meeting_types=MEETING_TYPES)


@app.route('/meetings/<int:id>')
@login_required
def meeting_view(id):
    """View meeting details"""
    from models import Meeting
    meeting = Meeting.query.get_or_404(id)
    if not current_user.is_manager_or_admin() and meeting.user_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('meetings_list'))
    return render_template('meetings/view.html', meeting=meeting)


@app.route('/meetings/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def meeting_edit(id):
    """Edit a meeting"""
    from models import Meeting, Lead, User
    meeting = Meeting.query.get_or_404(id)

    if not current_user.is_manager_or_admin() and meeting.user_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('meetings_list'))

    if request.method == 'POST':
        try:
            data = request.form
            scheduled_str = data.get('scheduled_at', '').strip()
            meeting.scheduled_at = datetime.strptime(scheduled_str, '%Y-%m-%dT%H:%M') if scheduled_str else None
            meeting.title = data.get('title', '').strip()
            meeting.meeting_type = data.get('meeting_type', 'Lead Visit')
            meeting.lead_id = int(data.get('lead_id')) if data.get('lead_id') else None
            meeting.project_id = int(data.get('project_id')) if data.get('project_id') else None
            if current_user.is_manager_or_admin():
                meeting.user_id = int(data.get('user_id')) if data.get('user_id') else meeting.user_id
            meeting.client_name = data.get('client_name', '').strip() or None
            meeting.client_phone = data.get('client_phone', '').strip() or None
            meeting.client_address = data.get('client_address', '').strip() or None
            meeting.notes = data.get('notes', '').strip() or None
            meeting.outcome = data.get('outcome', '').strip() or None
            meeting.status = data.get('status', meeting.status)
            meeting.updated_at = datetime.utcnow()
            db.session.commit()
            flash('Meeting updated successfully!', 'success')
            return redirect(url_for('meeting_view', id=meeting.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating meeting: {str(e)}', 'danger')

    leads = Lead.query.order_by(Lead.name).all()
    projects = Project.query.order_by(Project.name).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all() if current_user.is_manager_or_admin() else [current_user]
    return render_template('meetings/form.html', title='Edit Meeting',
                           meeting=meeting, leads=leads, projects=projects,
                           users=users, meeting_types=MEETING_TYPES,
                           meeting_statuses=MEETING_STATUSES)


@app.route('/meetings/<int:id>/delete', methods=['POST'])
@login_required
def meeting_delete(id):
    """Delete a meeting"""
    from models import Meeting
    meeting = Meeting.query.get_or_404(id)
    if not current_user.is_manager_or_admin() and meeting.user_id != current_user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('meetings_list'))
    try:
        db.session.delete(meeting)
        db.session.commit()
        flash('Meeting deleted.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('meetings_list'))


@app.route('/meetings/<int:id>/checkin', methods=['POST'])
@login_required
def meeting_checkin(id):
    """AJAX: GPS check-in for a meeting"""
    from models import Meeting
    meeting = Meeting.query.get_or_404(id)
    if meeting.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Only the assigned person can check in'}), 403
    try:
        meeting.check_in_lat = float(request.form.get('lat'))
        meeting.check_in_lng = float(request.form.get('lng'))
        meeting.check_in_accuracy = float(request.form.get('accuracy', 0))
        meeting.check_in_address = request.form.get('address', '')[:500]
        meeting.check_in_time = datetime.utcnow()
        meeting.status = 'Checked In'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/meetings/<int:id>/checkout', methods=['POST'])
@login_required
def meeting_checkout(id):
    """AJAX: GPS check-out for a meeting"""
    from models import Meeting
    meeting = Meeting.query.get_or_404(id)
    if meeting.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Only the assigned person can check out'}), 403
    try:
        meeting.check_out_lat = float(request.form.get('lat', 0)) or None
        meeting.check_out_lng = float(request.form.get('lng', 0)) or None
        meeting.check_out_time = datetime.utcnow()
        meeting.outcome = request.form.get('outcome', '').strip() or meeting.outcome
        meeting.status = 'Completed'
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/meetings/<int:id>/upload-photo', methods=['POST'])
@login_required
def meeting_upload_photo(id):
    """Upload a photo for a meeting"""
    from models import Meeting, MeetingPhoto
    meeting = Meeting.query.get_or_404(id)
    if not current_user.is_manager_or_admin() and meeting.user_id != current_user.id:
        return jsonify({'success': False, 'error': 'Access denied'}), 403

    file = request.files.get('photo')
    if not file or not file.filename:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    try:
        import boto3
        from werkzeug.utils import secure_filename
        bucket = os.getenv('AWS_BUCKET_NAME', 'glassyimages')
        region = os.getenv('AWS_REGION', 'ap-south-1')
        s3 = boto3.client('s3', region_name=region)
        filename = f"meeting_photos/{meeting.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}_{secure_filename(file.filename)}"
        s3.upload_fileobj(file, bucket, filename, ExtraArgs={'ContentType': file.content_type or 'image/jpeg'})
        url = f"https://{bucket}.s3.{region}.amazonaws.com/{filename}"

        photo = MeetingPhoto(
            meeting_id=meeting.id,
            photo_url=url,
            caption=request.form.get('caption', '').strip() or None,
            uploaded_by=current_user.id,
        )
        db.session.add(photo)
        db.session.commit()
        return jsonify({'success': True, 'url': url, 'photo_id': photo.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/meetings/<int:id>/delete-photo/<int:photo_id>', methods=['POST'])
@login_required
def meeting_delete_photo(id, photo_id):
    """Delete a meeting photo"""
    from models import MeetingPhoto
    photo = MeetingPhoto.query.get_or_404(photo_id)
    if photo.meeting_id != id:
        return jsonify({'success': False, 'error': 'Not found'}), 404
    try:
        db.session.delete(photo)
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/meetings/map')
@login_required
@manager_or_admin_required
def meetings_map():
    """Map view of all field visits (manager/admin only)"""
    from models import Meeting, User

    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    user_filter = request.args.get('user_id', '')

    query = Meeting.query.filter(Meeting.check_in_lat.isnot(None))
    if user_filter:
        try:
            query = query.filter(Meeting.user_id == int(user_filter))
        except ValueError:
            pass
    if date_from:
        try:
            query = query.filter(Meeting.check_in_time >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            query = query.filter(Meeting.check_in_time <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            pass

    meetings = query.order_by(Meeting.check_in_time.desc()).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    meetings_json = []
    for m in meetings:
        ist_time = (m.check_in_time + timedelta(hours=5, minutes=30)).strftime('%d %b %Y, %I:%M %p') if m.check_in_time else ''
        meetings_json.append({
            'id': m.id,
            'title': m.title,
            'user': m.user.username if m.user else '',
            'type': m.meeting_type,
            'client': m.client_name or (m.lead.name if m.lead else ''),
            'lat': m.check_in_lat,
            'lng': m.check_in_lng,
            'address': m.check_in_address or '',
            'time': ist_time,
            'status': m.status,
        })

    return render_template('meetings/map.html',
                           meetings_json=meetings_json,
                           users=users,
                           total=len(meetings))


# ============================================================================
# BATHQUBE QUOTATIONS
# ----------------------------------------------------------------------------
# Mirrored from glassyplatform's bathspace-quotes collection. Glassy POSTs to
# /api/bathqube/quotes/ingest after the configurator submits. Ops manages the
# 5-stage lifecycle from /quotes/bathqube/*.
# ============================================================================

import hmac as _hmac
import hashlib as _hashlib

from models import (
    BathqubeQuote, BathqubeStatusEvent, BathqubeQuoteItem, BathqubeQuoteRevision,
    BathqubeWorkOrder, BathqubeStageAttachment,
    BATHQUBE_STAGES, BATHQUBE_OPS_STAGES, BATHQUBE_OPS_ACTIVE_STAGES,
    UpvcQuote, UpvcQuoteItem, UpvcQuoteRevision, UpvcStatusEvent,
    UPVC_STAGES, UPVC_ACTIVE_STAGES,
    VetrovaQuote, VetrovaQuoteItem, VetrovaStatusEvent, VetrovaQuoteRevision,
    VETROVA_STAGES, VETROVA_ACTIVE_STAGES, VETROVA_STAGE_LABELS,
    VETROVA_CATEGORIES, VETROVA_CATEGORY_LABELS,
)
from utils.bathqube_messages import render_stage_message, STAGE_LABELS


def _bathqube_seed_items_from_config(quote):
    """First-time revise: build starting line items from the configurator snapshot.

    Supports two configData shapes:
      - NEW (schemaVersion=2): cfg['enclosures'] = [{ name, typeLabel, glassPanels[], pricePerSqft, quantity, ... }]
        → one BathqubeQuoteItem per panel per enclosure.
      - LEGACY (flat fields at top-level): cfg['glassPanels'] = [...] with single type/material/etc.
        → wrap as one synthetic enclosure, then same per-panel expansion.

    Panel sizes render in INCHES for the staff (vcore is always inches),
    with the original-unit value in parens so the staff can cross-check
    against what the customer typed. Pre-feature quotes (no dimensionUnit
    on configData) keep the historical "Wft × Hft" rendering — those
    quotes were typed in feet, no conversion needed.
    """
    from utils.bathqube_dimensions import to_inches, get_dimension_unit
    cfg = quote.config or {}
    enclosures = _bathqube_enclosures_from_cfg(cfg)
    quote_dim_unit = get_dimension_unit(cfg)  # 'mm'|'cm'|'in'|'m' or None for legacy

    items = []
    sort = 0
    for enc in enclosures:
        # PROD bug fix (2026-06-27): a customer-facing configurator quote
        # came through with values in mm but `dimensionUnit` missing from
        # the top-level cfg. The seeder fell back to the legacy
        # "values-are-feet" branch and produced a ₹103-crore line item.
        # The revise UI now persists a per-enclosure `dimensionUnit` so
        # the BD's correction sticks; honour that here in priority order
        # (per-enclosure > top-level > legacy/None).
        raw_enc_unit = enc.get('dimensionUnit')
        if raw_enc_unit in ('mm', 'cm', 'in', 'm', 'ft'):
            enc_dim_unit = None if raw_enc_unit == 'ft' else raw_enc_unit
        else:
            enc_dim_unit = quote_dim_unit
        enc_name = enc.get('name') or 'Enclosure'
        type_label = enc.get('typeLabel') or 'Shower Enclosure'
        spec_bits = [b for b in (
            enc.get('materialLabel'),
            enc.get('thicknessLabel'),
            enc.get('fittingLabel'),
        ) if b]
        spec = ' · '.join(spec_bits)
        try:
            price_per_sqft = float(enc.get('pricePerSqft') or 0)
        except (TypeError, ValueError):
            price_per_sqft = 0.0
        try:
            qty = float(enc.get('quantity') or 1)
        except (TypeError, ValueError):
            qty = 1.0
        panels = enc.get('glassPanels') or []
        if not panels:
            # Enclosure with no panels — skip rather than emitting an empty row
            continue
        for pidx, p in enumerate(panels, start=1):
            try:
                width = float(p.get('width') or 0)
                height = float(p.get('height') or 0)
            except (TypeError, ValueError):
                width = height = 0.0
            # Compute sqft authoritatively SERVER-side using the resolved
            # unit (per-enc > top-level > legacy). The JS-supplied sqft
            # on the panel was buggy on quotes ingested before the
            # 2026-06-27 fix (it would equal `width × height` even when
            # the values were in mm), so re-deriving here is the only
            # way to stop a stale 1,877,920-sqft figure from leaking
            # into BathqubeQuoteItem.amount on the next revise save.
            if enc_dim_unit:
                w_in = to_inches(width, enc_dim_unit)
                h_in = to_inches(height, enc_dim_unit)
                sqft = (w_in * h_in) / 144
                size_str = (
                    f'{w_in:.2f}" × {h_in:.2f}" '
                    f"(customer: {width:g}×{height:g} {enc_dim_unit})"
                )
            else:
                # Legacy pre-feature quote — values typed in feet, no
                # conversion needed (matches the historical math).
                sqft = width * height
                size_str = f"{width:g}×{height:g}ft"
            # KAN-45: include the per-sqft rate so the revised quote
            # surfaces the unit price (same info the initial Bathqube
            # PDF shows). Embedded inside the bracketed sqft block so
            # the PDF renderer can pull the whole "[…]" out and style
            # it as a muted sub-line under the main description.
            rate_str = f" @ ₹{price_per_sqft:,.0f}/sq.ft" if price_per_sqft > 0 else ""
            desc = (
                f"{enc_name} — {type_label}"
                + (f" ({spec})" if spec else '')
                + f" · Panel {pidx}: {size_str} [{sqft:.2f} sq.ft{rate_str}]"
            )
            rate = round(sqft * price_per_sqft, 2)
            amount = round(rate * qty, 2)
            items.append(BathqubeQuoteItem(
                sort_order=sort, description=desc[:500],
                quantity=qty, rate=rate, amount=amount,
            ))
            sort += 1

    # Fallback: if for any reason we built zero items but a subtotal exists,
    # emit a single placeholder so the revise UI isn't empty.
    if not items and float(quote.subtotal or 0) > 0:
        items.append(BathqubeQuoteItem(
            sort_order=0,
            description='Shower Enclosure (auto-imported)',
            quantity=1.0, rate=float(quote.subtotal), amount=float(quote.subtotal),
        ))
    return items


def _bathqube_enclosures_from_cfg(cfg):
    """Return list of enclosure dicts in the NEW shape, normalising legacy data.

    Each returned dict has at least: name, typeLabel, materialLabel,
    thicknessLabel, fittingLabel, glassPanels (list), pricePerSqft, quantity.
    """
    # PROD bug fix (BSP-000190, 2026-06-30): dimensionUnit was silently
    # dropped by this normaliser. The seeder's per-enclosure unit priority
    # (added 2026-06-27) then had nothing to prefer, and fell back to the
    # top-level unit — treating 840mm as 840ft → ₹18 crore subtotals.
    # Preserve dimensionUnit end-to-end so the BD's per-enclosure
    # correction actually reaches the seeder on Save.
    if isinstance(cfg.get('enclosures'), list) and cfg['enclosures']:
        out = []
        for idx, raw in enumerate(cfg['enclosures'], start=1):
            if not isinstance(raw, dict):
                continue
            out.append({
                'name': raw.get('name') or f'Enclosure {idx}',
                'typeLabel': raw.get('typeLabel') or 'Shower Enclosure',
                'materialLabel': raw.get('materialLabel') or '',
                'thicknessLabel': raw.get('thicknessLabel') or '',
                'fittingLabel': raw.get('fittingLabel') or '',
                'hardwareTypeLabel': raw.get('hardwareTypeLabel') or '',
                'dimensionUnit': raw.get('dimensionUnit'),  # ← preserve
                'glassPanels': raw.get('glassPanels') or [],
                'pricePerSqft': raw.get('pricePerSqft') or 0,
                'quantity': raw.get('quantity') or 1,
                'sqft': raw.get('sqft'),
                'subtotal': raw.get('subtotal'),
            })
        return out
    # Legacy flat shape — wrap as a single enclosure
    return [{
        'name': 'Enclosure 1',
        'typeLabel': cfg.get('typeLabel') or 'Shower Enclosure',
        'materialLabel': cfg.get('materialLabel') or '',
        'thicknessLabel': cfg.get('thicknessLabel') or '',
        'fittingLabel': cfg.get('fittingLabel') or '',
        'hardwareTypeLabel': cfg.get('hardwareTypeLabel') or '',
        'dimensionUnit': cfg.get('dimensionUnit'),  # ← preserve
        'glassPanels': cfg.get('glassPanels') or [],
        'pricePerSqft': cfg.get('pricePerSqft') or 0,
        'quantity': cfg.get('quantity') or 1,
        'sqft': cfg.get('sqft'),
        'subtotal': cfg.get('subtotal'),
    }]


def _bathqube_recompute_totals(quote):
    """Sum items → subtotal; apply discount %; apply GST % → CGST + SGST; sum → revised_total.

    Calculation chain (discount applies BEFORE GST, industry standard):
        subtotal        = Σ items.amount
        discount_amount = subtotal × (discount_percent / 100)
        taxable         = subtotal − discount_amount
        cgst = sgst     = taxable × (gst_percent / 2 / 100)
        revised_total   = taxable + cgst + sgst
    """
    subtotal = sum(float(it.amount or 0) for it in quote.items)
    discount_pct = float(quote.discount_percent or 0)
    discount_amt = round(subtotal * discount_pct / 100, 2)
    taxable = max(0.0, subtotal - discount_amt)
    gst_pct = float(quote.gst_percentage or 0)
    cgst = round(taxable * gst_pct / 2 / 100, 2)
    sgst = round(taxable * gst_pct / 2 / 100, 2)

    quote.subtotal = subtotal
    quote.discount_amount = discount_amt
    quote.cgst = cgst
    quote.sgst = sgst
    quote.revised_total = round(taxable + cgst + sgst, 2)

# Email-sending stages — go through bathqube_quote_action (subject/body editor + send).
# Revise has its own dedicated /revise editor (richer UI with item editing).
BATHQUBE_ACTIONS = ('awaiting_payment',)

# No-email stage transitions — single-click move via bathqube_quote_set_stage.
BATHQUBE_STAGE_TRANSITIONS = ('in_pipeline', 'closed_won', 'junk', 'rejected')

# ─── Vetrova stage machine helpers (Phase 2 2026-07-29) ──────────────────────
# Whitelist of stages a BD may set from the view page's stage-move form.
# Everything else (revision, quote_generated) is set as a side-effect of
# other actions (Revise save → revision, ingest → quote_generated).
VETROVA_STAGE_TRANSITIONS = ('in_pipeline', 'awaiting_payment', 'closed_won', 'junk', 'rejected')


def _vetrova_edit_lock_reason(quote, user):
    """Returns a string reason if the quote is locked from mutation,
    else None. Admins always bypass.

    Rule: once stage=closed_won, non-admin BDs can neither revise the
    line items nor move the stage. Prevents accidental clobber of a
    deal already booked into ops / invoicing.

    Applies to: /revise POST, /stage POST, /delete (delete already
    admin-only). Read paths are never locked.
    """
    try:
        if user.is_admin():
            return None
    except Exception:
        pass
    if getattr(quote, 'stage', None) == 'closed_won':
        return ('This quote is Closed Won — only admins can edit or move it. '
                'Ask an admin to unlock if you truly need to make a change.')
    return None


# ─── BD-side Bathqube quote creator helpers (KAN-67 follow-up) ────────────────
# Vcore-native form at /quotes/bathqube/new lets BD build a Bathqube quote
# from scratch (instead of waiting for a customer to submit the public
# configurator). Pricing data flows from a glassyplatform read-only options
# API so vcore never drifts from Payload's source of truth.

# Module-level cache for the options JSON. The configurator-options endpoint
# already edge-caches for 60s; this in-process cache stops us hitting the
# Next.js app on every keystroke when BD opens / refreshes the form. The
# expiry is intentionally short (90s) so a price change in Payload reaches
# vcore inside two minutes.
_BD_OPTIONS_CACHE = {'data': None, 'expires_at': 0}


def _glassyplatform_base_url():
    """Where the configurator API lives. Defaulted so a missing env doesn't
    break local dev — prod sets it explicitly."""
    return os.getenv('GLASSYPLATFORM_URL', 'https://platform.glassy.in').rstrip('/')


def _fetch_configurator_options():
    """Return the same options/prices the public configurator uses, cached
    in-process for 90 seconds.

    Shape:
      {
        basePricePerSqft: <num>,
        types | materials | thicknesses | fittings | hardwareTypes:
          [{value, label, surcharge}, ...]
      }

    Raises RuntimeError if the upstream is unreachable or returns invalid
    data — the form GET handler catches this and renders an error state
    instead of a broken page.
    """
    import time, requests as req_lib
    now = time.time()
    cached = _BD_OPTIONS_CACHE.get('data')
    if cached is not None and _BD_OPTIONS_CACHE.get('expires_at', 0) > now:
        return cached
    url = f'{_glassyplatform_base_url()}/api/configurator-options'
    try:
        resp = req_lib.get(url, timeout=10)
    except Exception as e:
        raise RuntimeError(f'Could not reach glassyplatform: {e}') from e
    if resp.status_code != 200:
        raise RuntimeError(
            f'glassyplatform returned HTTP {resp.status_code} for {url}'
        )
    try:
        data = resp.json()
    except ValueError as e:
        raise RuntimeError(f'glassyplatform returned non-JSON: {e}') from e
    # Validate the shape at least loosely so a broken upstream doesn't
    # silently render an empty form.
    if not isinstance(data, dict) or 'basePricePerSqft' not in data:
        raise RuntimeError('glassyplatform returned malformed options payload')
    _BD_OPTIONS_CACHE['data'] = data
    _BD_OPTIONS_CACHE['expires_at'] = now + 90
    return data


def _resolve_price_per_sqft(options, type_label, material_label, thickness_label,
                            fitting_label, hardware_label):
    """Sum base + surcharges by matching the chosen LABEL against each option
    list. Mirrors the JS configurator's pricing math so server + client agree.

    None / unknown labels contribute 0 (treated as "Other"). Returns a float
    rounded to 2dp.
    """
    def _surcharge(opt_list, label):
        if not label:
            return 0.0
        for o in (opt_list or []):
            if o.get('label') == label:
                try:
                    return float(o.get('surcharge') or 0)
                except (TypeError, ValueError):
                    return 0.0
        return 0.0  # "Other (specify)" — no surcharge

    base = float(options.get('basePricePerSqft') or 0)
    base += _surcharge(options.get('types'),         type_label)
    base += _surcharge(options.get('materials'),     material_label)
    base += _surcharge(options.get('thicknesses'),   thickness_label)
    base += _surcharge(options.get('fittings'),      fitting_label)
    base += _surcharge(options.get('hardwareTypes'), hardware_label)
    return round(max(0.0, base), 2)


def _bd_panel_sqft(width, height, unit):
    """Same formula as the JS panelSqft(): convert each dim to inches
    using the unified UNIT_TO_INCHES table, multiply, divide by 144.
    Default to 'ft' for unknown / missing units (legacy)."""
    from utils.bathqube_dimensions import to_inches, _SUPPORTED_UNITS
    u = unit if unit in _SUPPORTED_UNITS else 'ft'
    try:
        w_in = to_inches(float(width or 0), u)
        h_in = to_inches(float(height or 0), u)
    except (TypeError, ValueError):
        return 0.0
    return (w_in * h_in) / 144.0


def _compute_bd_quote_totals(enclosures, gst_percent):
    """Server-side recompute matching the JS preview formula:

        encSqft       = Σ panelSqft(w, h, unit)
        encSubtotal   = encSqft × pricePerSqft × max(1, quantity)
        subtotal      = Σ encSubtotal
        cgst = sgst   = subtotal × (gst/2) / 100
        total         = subtotal + cgst + sgst

    No discount (BD adds discounts later via Revise — keeps the create
    form clean per the agreed scope).
    """
    subtotal = 0.0
    for enc in enclosures:
        unit = enc.get('dimensionUnit') or 'ft'
        ppsft = float(enc.get('pricePerSqft') or 0)
        try:
            qty = max(1, int(float(enc.get('quantity') or 1)))
        except (TypeError, ValueError):
            qty = 1
        enc_sqft = 0.0
        for p in (enc.get('glassPanels') or []):
            enc_sqft += _bd_panel_sqft(p.get('width'), p.get('height'), unit)
        subtotal += enc_sqft * ppsft * qty
    gst_pct = float(gst_percent or 0)
    cgst = round(subtotal * gst_pct / 2 / 100, 2)
    sgst = round(subtotal * gst_pct / 2 / 100, 2)
    total = round(subtotal + cgst + sgst, 2)
    return {'subtotal': round(subtotal, 2), 'cgst': cgst, 'sgst': sgst, 'total': total}


def _post_bd_quote_to_glassyplatform(payload):
    """HMAC-sign + POST a BD-created quote to glassyplatform so the
    bathspace-quotes row exists in Payload with a shared BSP-NNNNNN
    number. Returns (id, estimateNumber) on success; raises on failure.

    Body shape expected by glassyplatform's /api/bathspace/bd-quote:
        { name, phone, email?, pincode?, siteAddress?, configData }
    """
    import hmac as _hmac
    import hashlib as _hashlib
    import requests as req_lib

    secret = os.getenv('VCORE_INGEST_SECRET', '')
    if not secret:
        raise RuntimeError(
            'VCORE_INGEST_SECRET not configured — cannot sign BD-quote request'
        )
    body = json.dumps(payload, separators=(',', ':')).encode('utf-8')
    sig = _hmac.new(secret.encode('utf-8'), body, _hashlib.sha256).hexdigest()
    url = f'{_glassyplatform_base_url()}/api/bathspace/bd-quote'
    try:
        resp = req_lib.post(
            url, data=body,
            headers={
                'Content-Type': 'application/json',
                'X-Vcore-Auth': f'sha256={sig}',
            },
            timeout=20,
        )
    except Exception as e:
        raise RuntimeError(f'POST failed: {e}') from e
    try:
        data = resp.json()
    except ValueError:
        data = {}
    if resp.status_code != 200 or not data.get('ok'):
        raise RuntimeError(
            f'glassyplatform rejected BD quote: HTTP {resp.status_code}: '
            f'{(data.get("error") if isinstance(data, dict) else None) or resp.text[:200]}'
        )
    return str(data.get('id') or ''), str(data.get('estimateNumber') or '')


def _bathqube_verify_signature(raw_body: bytes, header_sig: str) -> bool:
    secret = os.getenv('VCORE_INGEST_SECRET', '')
    if not secret or not header_sig:
        return False
    expected = _hmac.new(secret.encode('utf-8'), raw_body, _hashlib.sha256).hexdigest()
    # accept either "sha256=<hex>" or plain "<hex>"
    candidate = header_sig.split('=', 1)[1] if header_sig.startswith('sha256=') else header_sig
    return _hmac.compare_digest(expected, candidate.strip())


@app.route('/api/bathqube/quotes/ingest', methods=['POST'])
@limiter.limit("60 per minute")
def bathqube_ingest():
    """Webhook from glassyplatform: create/upsert a BathqubeQuote."""
    raw = request.get_data(cache=True)
    sig = request.headers.get('X-Bathqube-Signature', '')
    if not _bathqube_verify_signature(raw, sig):
        return jsonify({'error': 'invalid signature'}), 401

    try:
        data = json.loads(raw.decode('utf-8'))
    except Exception:
        return jsonify({'error': 'invalid json'}), 400

    external_id = str(data.get('externalId') or data.get('id') or '').strip()
    if not external_id:
        return jsonify({'error': 'externalId required'}), 400

    name = (data.get('name') or '').strip()
    phone = (data.get('phone') or '').strip()
    if not name or not phone:
        return jsonify({'error': 'name and phone required'}), 400

    cfg = data.get('configData') or {}
    if not isinstance(cfg, dict):
        cfg = {}

    quote = BathqubeQuote.query.filter_by(external_id=external_id).first()
    is_new = quote is None
    if is_new:
        # New ingest from the bathspace configurator webhook lands at the
        # top of the sales pipeline. Sales then moves it through In Pipeline
        # → Revision → Awaiting Payment → Closed Won (or Junk / Rejected).
        quote = BathqubeQuote(external_id=external_id, stage='quote_generated')
        db.session.add(quote)

    quote.estimate_number = data.get('estimateNumber') or quote.estimate_number
    quote.customer_name = name
    quote.phone = phone
    quote.email = data.get('email') or None
    quote.pincode = data.get('pincode') or None
    quote.site_address = cfg.get('siteAddress') or data.get('siteAddress') or None
    quote.source_path = data.get('sourcePath') or None
    quote.variant_size = data.get('variantSize') or None
    quote.variant_material = data.get('variantMaterial') or None
    quote.config_data = json.dumps(cfg) if cfg else None
    quote.subtotal = cfg.get('subtotal') or 0
    quote.cgst = cfg.get('cgst') or 0
    quote.sgst = cfg.get('sgst') or 0
    quote.total = cfg.get('total') or 0

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'db error', 'detail': str(e)}), 500

    return jsonify({'ok': True, 'id': quote.id, 'created': is_new}), 200


@app.route('/quotes/bathqube')
@login_required
def bathqube_quotes_list():
    from models import BATHQUBE_ACTIVE_STAGES
    search = (request.args.get('search') or '').strip()
    stage = request.args.get('stage') or ''
    include_archived = request.args.get('archived') == '1'
    q = BathqubeQuote.query
    if search:
        like = f'%{search}%'
        q = q.filter(
            (BathqubeQuote.customer_name.ilike(like))
            | (BathqubeQuote.phone.ilike(like))
            | (BathqubeQuote.estimate_number.ilike(like))
            | (BathqubeQuote.email.ilike(like))
        )
    if stage:
        # Explicit stage filter wins over the archived/active default.
        q = q.filter_by(stage=stage)
    elif not include_archived:
        # Default view: hide junk + rejected.
        q = q.filter(BathqubeQuote.stage.in_(BATHQUBE_ACTIVE_STAGES))
    quotes = q.order_by(BathqubeQuote.created_at.desc()).all()
    return render_template('quotes/bathqube_list.html',
                           quotes=quotes, search=search, stage=stage,
                           include_archived=include_archived,
                           stage_labels=STAGE_LABELS, stages=BATHQUBE_STAGES)


# ── Bulk WhatsApp template send from Bathqube Quotations ─────────────────────
# BD picks Bathqube quotes to blast with an approved template (default:
# 'welcome'). The brand is IMPLICIT — always Bathqube, since every row on
# this list is a Bathqube customer. Same 25-per-batch cap as any bulk send.

MAX_BULK_WA_QUOTES = 25


@app.route('/quotes/bathqube/bulk-send-whatsapp', methods=['POST'])
@login_required
def bathqube_bulk_send_whatsapp():
    """Bulk-send an approved WhatsApp template to selected Bathqube quotes.

    Input (POST form):
        quote_ids[]     — the bathqube_quote IDs to send to
        template_name   (optional, default 'welcome')
        language        (optional, default 'en')

    Behaviour:
        - Sender is the Bathqube Brand row (looked up by slug='bathqube').
        - Loops sequentially, capped at MAX_BULK_WA_QUOTES per request.
        - Skips quotes with no phone.
        - Passes the customer's first_name as {{1}} with a param-error
          fallback (matches lead_send_welcome for parity).
        - Logs each attempt to whatsapp_messages with brand_id + the
          new bathqube_quote_id so the per-quote 'sent' count works.
    """
    from models import Brand, BathqubeQuote, WhatsAppMessage
    from utils.whatsapp import send_template, normalize_phone

    brand = Brand.query.filter_by(slug='bathqube', is_active=True).first()
    if not brand:
        return jsonify({'success': False, 'error': 'Bathqube brand is not configured. Add a row to brands with slug=bathqube.'}), 500

    template_name = (request.form.get('template_name') or 'welcome').strip()
    language = (request.form.get('language') or 'en').strip()

    quote_ids = request.form.getlist('quote_ids')
    if not quote_ids:
        return jsonify({'success': False, 'error': 'No quotes selected.'}), 400
    try:
        quote_ids = [int(x) for x in quote_ids]
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid quote ids.'}), 400

    quotes = (BathqubeQuote.query
              .filter(BathqubeQuote.id.in_(quote_ids))
              .limit(MAX_BULK_WA_QUOTES + 1)
              .all())
    truncated = len(quotes) > MAX_BULK_WA_QUOTES
    quotes = quotes[:MAX_BULK_WA_QUOTES]

    sent = 0
    failed = 0
    skipped = 0
    failures = []

    for quote in quotes:
        if not quote.phone:
            skipped += 1
            failures.append({'quote_id': quote.id, 'name': quote.customer_name, 'error': 'no phone number'})
            continue

        first_name = (quote.customer_name or 'there').strip().split()[0]
        variables = [first_name]

        msg = WhatsAppMessage(
            bathqube_quote_id=quote.id,
            brand_id=brand.id,
            to_number=normalize_phone(quote.phone) or quote.phone,
            template_name=template_name,
            language=language,
            variables_json=json.dumps(variables),
            sent_by=current_user.id,
            status='queued',
        )
        db.session.add(msg)
        db.session.flush()

        result = send_template(
            to=quote.phone,
            template_name=template_name,
            language=language,
            variables=variables,
            brand=brand,
        )
        if not result.get('success') and 'parameter' in (result.get('error') or '').lower():
            result = send_template(
                to=quote.phone,
                template_name=template_name,
                language=language,
                variables=None,
                brand=brand,
            )
            if result.get('success'):
                msg.variables_json = json.dumps([])

        if result.get('success'):
            msg.status = 'sent'
            msg.wamid = result.get('wamid')
            sent += 1
        else:
            msg.status = 'failed'
            msg.error_message = result.get('error', 'Unknown error')
            failed += 1
            failures.append({'quote_id': quote.id, 'name': quote.customer_name, 'error': msg.error_message})

    db.session.commit()

    return jsonify({
        'success': True,
        'sent': sent,
        'failed': failed,
        'skipped': skipped,
        'attempted': len(quotes),
        'truncated': truncated,
        'cap': MAX_BULK_WA_QUOTES,
        'brand': brand.slug,
        'template': template_name,
        'failures': failures[:20],
    })


# ─── Vetrova Quotes — configurator quotes from vetrova.in ────────────────────
# Ingested from glassyplatform's /api/vetrova/quote-request via the same
# HMAC-signed webhook pattern as Bathqube. BD works the funnel here.

# ─── BD-side Vetrova quote creator helpers (2026-08-01) ──────────────────────
# Vcore-native form at /quotes/vetrova/new lets BD build a Vetrova quote
# from scratch. Reads the same Payload configurator globals that the
# storefront reads, via three read-only glassyplatform endpoints:
#   /api/vetrova/categories               → dropdown list
#   /api/vetrova/configurator-schema/<slug> → steps + options for one category
#   /api/vetrova/price                    → live subtotal (single source of truth)
#
# In-process caches keep the categories + per-slug schema for 90s so a BD
# refreshing the form doesn't hammer glassyplatform. Pricing is never cached
# — every change fires a fresh POST so BD sees the latest per-category rates.

_VETROVA_CATEGORIES_CACHE = {'data': None, 'expires_at': 0}
_VETROVA_SCHEMA_CACHE = {}  # slug → {'data': …, 'expires_at': …}


def _fetch_vetrova_categories():
    """Read-only category list for the BD form's dropdown. Cached 90s.
    Raises RuntimeError on upstream failure so the form GET handler
    can flash a proper error instead of rendering a blank dropdown."""
    import time, requests as req_lib
    now = time.time()
    if _VETROVA_CATEGORIES_CACHE.get('data') is not None and \
       _VETROVA_CATEGORIES_CACHE.get('expires_at', 0) > now:
        return _VETROVA_CATEGORIES_CACHE['data']
    url = f'{_glassyplatform_base_url()}/api/vetrova/categories'
    try:
        resp = req_lib.get(url, timeout=10)
    except Exception as e:
        raise RuntimeError(f'Could not reach glassyplatform: {e}') from e
    if resp.status_code != 200:
        raise RuntimeError(f'glassyplatform returned HTTP {resp.status_code} for {url}')
    try:
        data = resp.json()
    except ValueError as e:
        raise RuntimeError(f'glassyplatform returned non-JSON: {e}') from e
    if not isinstance(data, dict) or not isinstance(data.get('categories'), list):
        raise RuntimeError('glassyplatform returned malformed categories payload')
    _VETROVA_CATEGORIES_CACHE['data'] = data
    _VETROVA_CATEGORIES_CACHE['expires_at'] = now + 90
    return data


def _fetch_vetrova_schema(slug):
    """Per-category configurator schema. Cached per-slug for 90s so
    switching between categories in the same form session doesn't
    re-hit glassyplatform every click."""
    import time, requests as req_lib
    now = time.time()
    entry = _VETROVA_SCHEMA_CACHE.get(slug)
    if entry and entry.get('data') is not None and entry.get('expires_at', 0) > now:
        return entry['data']
    url = f'{_glassyplatform_base_url()}/api/vetrova/configurator-schema/{slug}'
    try:
        resp = req_lib.get(url, timeout=10)
    except Exception as e:
        raise RuntimeError(f'Could not reach glassyplatform: {e}') from e
    if resp.status_code == 404:
        raise RuntimeError(f'Unknown Vetrova category: {slug}')
    if resp.status_code != 200:
        raise RuntimeError(f'glassyplatform returned HTTP {resp.status_code} for {url}')
    try:
        data = resp.json()
    except ValueError as e:
        raise RuntimeError(f'glassyplatform returned non-JSON: {e}') from e
    if not isinstance(data, dict) or 'steps' not in data:
        raise RuntimeError('glassyplatform returned malformed schema payload')
    _VETROVA_SCHEMA_CACHE[slug] = {'data': data, 'expires_at': now + 90}
    return data


def _mint_bd_vetrova_quote_ref():
    """Mint a unique VQ-BD-NNNNN reference for a BD-created quote.
    Retries on collision (should be exceedingly rare given the 5-digit
    random tail and vcore's row count, but the DB unique index will
    catch any repeat and we resample rather than crash the request)."""
    import random
    for _ in range(10):
        candidate = f'VQ-BD-{random.randint(10000, 99999)}'
        exists = VetrovaQuote.query.filter_by(quote_ref=candidate).first()
        if exists is None:
            return candidate
    # Astronomically unlikely; escalate rather than hand back a duplicate.
    raise RuntimeError('Could not mint a unique VQ-BD-NNNNN ref after 10 tries')

@app.route('/api/vetrova/quotes/ingest', methods=['POST'])
@limiter.limit("60 per minute")
def vetrova_ingest():
    """Webhook from glassyplatform: create/upsert a VetrovaQuote + items.

    Accepts BOTH payload shapes glassyplatform sends:
      - Multi-run: `runs: [{ categorySlug, selections, runningFt, quantity,
                            subtotal, panels[], fabricCode,
                            uploadedImageDataUrl, dimensionKind,
                            dimensionUnit }, …]`
      - Legacy single-run: top-level `categorySlug/selections/runningFt/
        quantity/total` (synthesized into a single run before persist).

    Upserts by `quote_ref`. On re-ingest of the same ref we WIPE the
    existing items and re-create — the vetrova.in wire payload is the
    source of truth for a "fresh" submission; BD edits happen after via
    the revise route, and those never re-trigger ingest.
    """
    raw = request.get_data(cache=True)
    sig = request.headers.get('X-Bathqube-Signature', '')
    if not _bathqube_verify_signature(raw, sig):
        return jsonify({'error': 'invalid signature'}), 401

    try:
        data = json.loads(raw.decode('utf-8'))
    except Exception:
        return jsonify({'error': 'invalid json'}), 400

    quote_ref = (data.get('quoteRef') or '').strip()
    if not quote_ref:
        return jsonify({'error': 'quoteRef required'}), 400

    contact = data.get('contact') or {}
    name = (contact.get('name') or '').strip()
    phone = (contact.get('phone') or '').strip()
    if not name or not phone:
        return jsonify({'error': 'name and phone required'}), 400

    external_id = str(data.get('externalId') or quote_ref)

    # Normalise to runs[] regardless of payload shape.
    runs_raw = data.get('runs')
    if not isinstance(runs_raw, list) or not runs_raw:
        # Legacy single-run fallback — synthesize from top-level fields.
        legacy_slug = (data.get('categorySlug') or '').strip()
        if not legacy_slug:
            return jsonify({'error': 'runs[] or categorySlug required'}), 400
        runs_raw = [{
            'categorySlug':  legacy_slug,
            'categoryLabel': data.get('categoryLabel')
                             or VETROVA_CATEGORY_LABELS.get(legacy_slug, legacy_slug.title()),
            'selections':    data.get('selections') or {},
            'runningFt':     data.get('runningFt') or 0,
            'quantity':      data.get('quantity') or 1,
            'subtotal':      data.get('total') or 0,
        }]

    quote = VetrovaQuote.query.filter_by(quote_ref=quote_ref).first()
    is_new = quote is None
    if is_new:
        quote = VetrovaQuote(quote_ref=quote_ref, stage='quote_generated')
        db.session.add(quote)

    quote.external_id = external_id
    quote.customer_name = name
    quote.phone = phone
    quote.email = contact.get('email') or None
    quote.pincode = contact.get('pincode') or None
    quote.site_address = contact.get('siteAddress') or None
    quote.notes = contact.get('notes') or None

    # Header category — first run's slug (matches glassyplatform's own
    # convention that VQ-CAT- ref prefix uses runs[0]'s categorySlug).
    first_slug = (runs_raw[0].get('categorySlug') or '').strip() or 'unknown'
    quote.category_slug = first_slug
    # Multi-run label — "Multi-configuration" when runs span categories.
    uniq_labels = {(r.get('categoryLabel') or VETROVA_CATEGORY_LABELS.get(
        (r.get('categorySlug') or '').strip(),
        (r.get('categorySlug') or '').strip().title())) for r in runs_raw}
    quote.category_label = next(iter(uniq_labels)) if len(uniq_labels) == 1 else 'Multi-configuration'

    # Wipe + re-create items on every ingest (fresh submission).
    if not is_new:
        VetrovaQuoteItem.query.filter_by(quote_id=quote.id).delete()
    else:
        # Flush so quote.id is populated BEFORE the item loop below
        # composes VetrovaQuoteItem(quote_id=quote.id, …). Without this,
        # every brand-new quote from vetrova.in dies with a NOT NULL
        # violation on vetrova_quote_items.quote_id — re-ingests of an
        # existing ref survive because .first() returns a row with id.
        # Placed here so all NOT NULL parent columns (category_slug,
        # customer_name, phone) are already set on the quote row.
        # Same pattern used by upvc_quotes_ingest (line 1538) etc.
        db.session.flush()

    for idx, r in enumerate(runs_raw):
        slug = (r.get('categorySlug') or '').strip() or first_slug
        label = r.get('categoryLabel') or VETROVA_CATEGORY_LABELS.get(slug, slug.title())
        selections = r.get('selections') if isinstance(r.get('selections'), dict) else {}
        panels = r.get('panels') if isinstance(r.get('panels'), list) else None
        running_ft = float(r.get('runningFt') or 0)
        qty = float(r.get('quantity') or 1)
        item_subtotal = float(r.get('subtotal') or 0)
        # Rate derivation mirrors glassyplatform's PDF math: panelsMode
        # runs pre-bake qty into runningFt so divide by runningFt alone;
        # single-row runs multiply externally so divide by both.
        if panels:
            denom = running_ft
        else:
            denom = running_ft * qty
        rate = round(item_subtotal / denom, 2) if denom > 0 else 0

        img_data_url = r.get('uploadedImageDataUrl')
        # Guard rail: only accept data:image/* URLs so we don't accidentally
        # persist an attacker-supplied http:// URL.
        if not (isinstance(img_data_url, str) and img_data_url.startswith('data:image/')):
            img_data_url = None

        item = VetrovaQuoteItem(
            quote_id=quote.id,
            sort_order=idx + 1,
            category_slug=slug,
            category_label=label,
            selections=json.dumps(selections) if selections else None,
            dimension_kind=r.get('dimensionKind') or 'square_feet',
            dimension_unit=r.get('dimensionUnit') or 'ft',
            running_ft=running_ft,
            quantity=qty,
            panels=json.dumps(panels) if panels else None,
            fabric_code=(r.get('fabricCode') or None),
            uploaded_image_data_url=img_data_url,
            rate_per_unit=rate,
            subtotal=item_subtotal,
        )
        db.session.add(item)

    # Legacy per-quote fields — populated from run[0] so the list view keeps
    # working. Multi-run quotes still show first-run summary in the header.
    r0 = runs_raw[0]
    quote.selections = json.dumps(r0.get('selections') or {}) if r0.get('selections') else None
    quote.running_ft = r0.get('runningFt') or 0
    quote.quantity = int(r0.get('quantity') or 1)
    quote.total = data.get('total') or sum(float(r.get('subtotal') or 0) for r in runs_raw)

    try:
        db.session.flush()  # get item PKs so recompute_totals sees them
        # Reload items via relationship so recompute_totals picks them up.
        db.session.refresh(quote)
        quote.recompute_totals()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': 'db error', 'detail': str(e)}), 500

    return jsonify({'ok': True, 'id': quote.id, 'created': is_new,
                    'itemCount': len(runs_raw)}), 200


@app.route('/quotes/vetrova/new', methods=['GET', 'POST'])
@login_required
def vetrova_quote_create():
    """BD-driven Vetrova quote create. GET renders the form; POST persists
    a VetrovaQuote + items with source='bd' and redirects to the view page.

    The form is schema-driven — the category dropdown fetches
    /api/vetrova/categories on load; picking a category fetches
    /api/vetrova/configurator-schema/<slug> and renders the correct
    step selects + dimensions grid via /quotes/vetrova/new/api/schema/<slug>
    and /quotes/vetrova/new/api/price (JS-fronted proxies that keep this
    session-authenticated and CORS-free).
    """
    if request.method == 'GET':
        try:
            cats = _fetch_vetrova_categories()
        except RuntimeError as e:
            app.logger.error(f'[vetrova/new GET] categories fetch failed: {e}')
            flash(
                'Could not load Vetrova categories from glassyplatform. '
                f'Reason: {e}. Try again in a minute, or check that '
                'platform.glassy.in is reachable.',
                'danger',
            )
            return redirect(url_for('vetrova_quotes_list'))
        return render_template(
            'quotes/vetrova_new.html',
            categories=cats.get('categories', []),
            meta=cats.get('meta', {}),
        )

    # ── POST — persist the BD-built quote ──
    form = request.form
    customer_name = (form.get('customer_name') or '').strip()
    phone = (form.get('phone') or '').strip()
    if not customer_name or not phone:
        flash('Customer name and phone are required.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    raw_state = (form.get('state_json') or '').strip()
    if not raw_state:
        flash('Configuration missing — pick a category and enter dimensions first.', 'danger')
        return redirect(url_for('vetrova_quote_create'))
    try:
        state = json.loads(raw_state)
    except Exception:
        flash('Form payload was malformed — refresh and try again.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    slug = (state.get('slug') or '').strip().lower()
    if not slug:
        flash('Category missing from configuration.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    # ── Server-side re-price via the same /api/vetrova/price endpoint the
    # JS uses, so the stored subtotal is canonical even if BD's browser
    # was tampered / stale. The client sends its own snapshot in
    # state.breakdown; we ignore it entirely and use what the server
    # returns here as truth.
    import requests as req_lib
    price_body = {
        'categorySlug':  slug,
        'dimensionMode': state.get('dimensionMode') or 'sqft',
        'panels':        state.get('panels') or [],
        'totalSqft':     state.get('totalSqft') or 0,
        'runningFt':     state.get('runningFt') or 0,
        'qty':           state.get('qty') or 1,
        'selections': [
            {
                'stepId':      k,
                'stepLabel':   k,
                'optionId':    v,
                'optionLabel': v,
            } for k, v in (state.get('selections') or {}).items() if v
        ],
    }
    try:
        pr = req_lib.post(
            f'{_glassyplatform_base_url()}/api/vetrova/price',
            json=price_body, timeout=15,
        )
        pr.raise_for_status()
        priced = pr.json()
    except Exception as e:
        app.logger.error(f'[vetrova/new POST] price fetch failed: {e}')
        flash(f'Could not re-verify pricing with vetrova.in: {e}. Quote NOT saved.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    breakdown = priced.get('breakdown') or {}
    subtotal = float(breakdown.get('subtotal') or 0)
    if subtotal <= 0:
        flash('Priced subtotal is ₹0 — dimensions may be empty. Fix and re-save.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    # ── Build DB rows ──
    try:
        quote_ref = _mint_bd_vetrova_quote_ref()
    except RuntimeError as e:
        flash(f'Could not mint quote ref: {e}. Try again.', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    category_label = (state.get('categoryName') or slug).strip()
    dimension_mode = state.get('dimensionMode') or 'sqft'
    dimension_kind = ('square_feet' if dimension_mode != 'runningFt' else 'running_feet')
    total_units = float(breakdown.get('sqft') or 0)
    per_unit = float(breakdown.get('perSqft') or 0)

    quote = VetrovaQuote(
        quote_ref=quote_ref,
        external_id=quote_ref,   # BD-created → no external system, mirror ref
        source='bd',
        stage='in_pipeline',     # BD is already working it; skip quote_generated
        category_slug=slug,
        category_label=category_label,
        customer_name=customer_name,
        phone=phone,
        email=(form.get('email') or '').strip() or None,
        pincode=(form.get('pincode') or '').strip() or None,
        site_address=(form.get('site_address') or '').strip() or None,
        notes=(form.get('notes') or '').strip() or None,
        selections=json.dumps(state.get('selections') or {}) if state.get('selections') else None,
        running_ft=total_units,
        quantity=int(state.get('qty') or 1),
        total=subtotal,
    )
    db.session.add(quote)
    db.session.flush()  # get quote.id for item

    item = VetrovaQuoteItem(
        quote_id=quote.id,
        sort_order=1,
        category_slug=slug,
        category_label=category_label,
        selections=json.dumps(state.get('selections') or {}) if state.get('selections') else None,
        dimension_kind=dimension_kind,
        dimension_unit='ft',
        running_ft=total_units,
        quantity=int(state.get('qty') or 1),
        panels=json.dumps(state.get('panels') or []) if state.get('panels') else None,
        fabric_code=(state.get('fabricCode') or None),
        rate_per_unit=per_unit,
        subtotal=subtotal,
        notes=(state.get('lineNotes') or None),
    )
    db.session.add(item)

    # Audit event so the view page's timeline shows who created the quote.
    db.session.add(VetrovaStatusEvent(
        quote_id=quote.id,
        from_stage=None,
        to_stage='in_pipeline',
        note=f'BD-created via /quotes/vetrova/new (₹{subtotal:,.2f} subtotal).',
        channel='none',
        send_status='skipped',
        triggered_by=current_user.id,
    ))

    try:
        quote.recompute_totals()
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[vetrova/new POST] commit failed: {e}')
        flash(f'Save failed: {e}', 'danger')
        return redirect(url_for('vetrova_quote_create'))

    flash(f'Created {quote_ref} for {customer_name}.', 'success')
    return redirect(url_for('vetrova_quote_view', id=quote.id))


# JS-fronted proxies so the browser talks only to vcore (same origin,
# session-authenticated). vcore's cached wrappers pass through to
# glassyplatform. Keeps CORS out of the equation and gives us one
# place to add rate-limiting / auth checks if the endpoints ever
# stop being public on the glassyplatform side.

@app.route('/quotes/vetrova/new/api/schema/<slug>', methods=['GET'])
@login_required
def vetrova_new_api_schema(slug):
    try:
        return jsonify(_fetch_vetrova_schema(slug))
    except RuntimeError as e:
        code = 404 if 'Unknown Vetrova category' in str(e) else 502
        return jsonify({'error': str(e)}), code


@app.route('/quotes/vetrova/new/api/price', methods=['POST'])
@login_required
def vetrova_new_api_price():
    import requests as req_lib
    try:
        resp = req_lib.post(
            f'{_glassyplatform_base_url()}/api/vetrova/price',
            json=request.get_json(silent=True) or {},
            timeout=10,
        )
    except Exception as e:
        return jsonify({'error': f'upstream unreachable: {e}'}), 502
    return (resp.text, resp.status_code, {'Content-Type': 'application/json'})


@app.route('/quotes/vetrova')
@login_required
def vetrova_quotes_list():
    search = (request.args.get('search') or '').strip()
    stage = request.args.get('stage') or ''
    category = request.args.get('category') or ''
    include_archived = request.args.get('archived') == '1'
    q = VetrovaQuote.query
    if search:
        like = f'%{search}%'
        q = q.filter(
            (VetrovaQuote.customer_name.ilike(like))
            | (VetrovaQuote.phone.ilike(like))
            | (VetrovaQuote.quote_ref.ilike(like))
            | (VetrovaQuote.email.ilike(like))
        )
    if category:
        q = q.filter_by(category_slug=category)
    if stage:
        q = q.filter_by(stage=stage)
    elif not include_archived:
        q = q.filter(VetrovaQuote.stage.in_(VETROVA_ACTIVE_STAGES))
    quotes = q.order_by(VetrovaQuote.created_at.desc()).all()
    return render_template('quotes/vetrova_list.html',
                           quotes=quotes, search=search, stage=stage,
                           category=category, include_archived=include_archived,
                           stage_labels=VETROVA_STAGE_LABELS, stages=VETROVA_STAGES,
                           categories=VETROVA_CATEGORIES)


@app.route('/quotes/vetrova/<int:id>')
@login_required
def vetrova_quote_view(id):
    quote = VetrovaQuote.query.get_or_404(id)
    return render_template('quotes/vetrova_view.html',
                           quote=quote,
                           stage_labels=VETROVA_STAGE_LABELS,
                           stages=VETROVA_STAGES,
                           category_label=VETROVA_CATEGORY_LABELS.get(quote.category_slug, quote.category_label))


@app.route('/quotes/vetrova/<int:id>/stage', methods=['POST'])
@login_required
def vetrova_quote_stage(id):
    quote = VetrovaQuote.query.get_or_404(id)
    target = (request.form.get('stage') or '').strip()
    if target not in VETROVA_STAGES:
        flash('Invalid stage.', 'danger')
        return redirect(url_for('vetrova_quote_view', id=id))
    # Phase 2: restrict manual transitions to the whitelist. `revision`
    # and `quote_generated` are set implicitly by other actions (Revise
    # save / ingest) and should not appear as click-to-move options.
    if target not in VETROVA_STAGE_TRANSITIONS:
        flash(f'Stage "{VETROVA_STAGE_LABELS.get(target, target)}" is not '
              'directly settable — it moves as a side-effect of other actions.',
              'warning')
        return redirect(url_for('vetrova_quote_view', id=id))
    if target == quote.stage:
        flash(f'Already in {VETROVA_STAGE_LABELS.get(target, target)}.', 'info')
        return redirect(url_for('vetrova_quote_view', id=id))
    # Phase 2 closed-won lock. Non-admin cannot move a closed_won quote.
    lock = _vetrova_edit_lock_reason(quote, current_user)
    if lock:
        flash(lock, 'warning')
        return redirect(url_for('vetrova_quote_view', id=id))
    event = VetrovaStatusEvent(
        quote_id=quote.id,
        from_stage=quote.stage,
        to_stage=target,
        note=(request.form.get('note') or '').strip() or None,
        channel='none',
        send_status='skipped',
        triggered_by=current_user.id,
    )
    quote.stage = target
    db.session.add(event)
    db.session.commit()
    flash(f'Moved to {VETROVA_STAGE_LABELS.get(target, target)}.', 'success')
    return redirect(url_for('vetrova_quote_view', id=id))


@app.route('/quotes/vetrova/<int:id>/delete', methods=['POST'])
@login_required
def vetrova_quote_delete(id):
    if not current_user.is_admin():
        flash('Only admins can delete quotes.', 'danger')
        return redirect(url_for('vetrova_quote_view', id=id))
    quote = VetrovaQuote.query.get_or_404(id)
    ref = quote.quote_ref
    db.session.delete(quote)
    db.session.commit()
    flash(f'Deleted {ref}.', 'success')
    return redirect(url_for('vetrova_quotes_list'))


@app.route('/quotes/vetrova/<int:id>/revise', methods=['GET', 'POST'])
@login_required
def vetrova_quote_revise(id):
    """Full editor for a Vetrova quote.

    GET  — renders the form pre-populated with current quote + items.
    POST — accepts application/json:
      {
        customer: {name, phone, email, pincode, siteAddress, notes},
        pricing:  {transportCharges, gstPercentage, revisedTotal, amountReceived},
        items:    [
          { id?: int,           # existing item id; omit for new
            sortOrder: int,
            categorySlug: str, categoryLabel: str,
            label?: str,
            selections: { key: value, … },
            dimensionKind: 'square_feet' | 'running_feet',
            dimensionUnit: 'ft' | 'in',
            runningFt: number, quantity: number,
            panels: [{ widthFt, heightFt, qty, kind: 'panel'|'door' }, …] | null,
            fabricCode?: str,
            ratePerUnit: number, subtotal: number,
            notes?: str },
          …
        ]
      }

    The server WIPES existing items and re-creates from the payload
    (items are cheap; keeping this simple avoids per-row diff bugs).
    Bumps `revision_count` on every save; Phase-2 snapshots the pre-edit
    state into `vetrova_quote_revisions` for the audit trail + history
    panel — see VetrovaQuoteRevision (:2760).

    Phase 2 additions on the POST body (all optional, back-compat):
      customer.pan            → quote.customer_pan
      pricing.discountPercent → quote.discount_percent
      items[i].isExtra        → items[i].is_extra
    """
    quote = VetrovaQuote.query.get_or_404(id)

    if request.method == 'POST':
        # Phase 2 closed-won lock. Non-admin cannot revise a closed_won
        # quote. Admins are silently allowed (audit trail captures WHO).
        lock = _vetrova_edit_lock_reason(quote, current_user)
        if lock:
            return jsonify({'error': lock, 'locked': True}), 423  # 423 Locked

        try:
            data = request.get_json(force=True, silent=False) or {}
        except Exception:
            return jsonify({'error': 'invalid json body'}), 400

        cust = data.get('customer') or {}
        pricing = data.get('pricing') or {}
        items_in = data.get('items') or []
        if not isinstance(items_in, list) or not items_in:
            return jsonify({'error': 'items[] required (need at least one line)'}), 400

        # ── Phase 2 audit: snapshot the pre-edit state BEFORE mutating.
        # If the save later blows up (bad numeric etc.), the snapshot is
        # only written on the successful commit path below — pending
        # snapshot object stays in the session.
        pre_subtotal = float(quote.subtotal or 0)
        pre_total = float(quote.grand_total or 0)
        pre_discount_pct = float(quote.discount_percent or 0)
        pre_snapshot = json.dumps({
            'customer': {
                'name': quote.customer_name,
                'phone': quote.phone,
                'email': quote.email,
                'pincode': quote.pincode,
                'siteAddress': quote.site_address,
                'notes': quote.notes,
                'pan': getattr(quote, 'customer_pan', None),
            },
            'pricing': {
                'transportCharges': float(quote.transport_charges or 0),
                'deliveryCharge': float(quote.delivery_charge or 0),
                'gstPercentage': float(quote.gst_percentage or 18),
                'discountPercent': pre_discount_pct,
                'discountAmount': float(quote.discount_amount or 0),
                'revisedTotal': (float(quote.revised_total) if quote.revised_total is not None else None),
                'amountReceived': float(quote.amount_received or 0),
                'subtotal': pre_subtotal,
                'cgst': float(quote.cgst or 0),
                'sgst': float(quote.sgst or 0),
                'grandTotal': pre_total,
            },
            'items': [
                {
                    'id': it.id,
                    'sortOrder': it.sort_order,
                    'categorySlug': it.category_slug,
                    'categoryLabel': it.category_label,
                    'label': it.label,
                    'selections': it.selections_parsed,
                    'dimensionKind': it.dimension_kind,
                    'dimensionUnit': it.dimension_unit,
                    'runningFt': float(it.running_ft or 0),
                    'quantity': float(it.quantity or 1),
                    'panels': it.panels_parsed,
                    'fabricCode': it.fabric_code,
                    'uploadedImageDataUrl': it.uploaded_image_data_url,
                    'ratePerUnit': float(it.rate_per_unit or 0),
                    'subtotal': float(it.subtotal or 0),
                    'isExtra': bool(getattr(it, 'is_extra', False)),
                    'notes': it.notes,
                } for it in quote.items
            ],
        })

        # Customer fields
        name = (cust.get('name') or '').strip()
        phone = (cust.get('phone') or '').strip()
        if not name or not phone:
            return jsonify({'error': 'customer name + phone required'}), 400
        quote.customer_name = name
        quote.phone = phone
        quote.email = (cust.get('email') or '').strip() or None
        quote.pincode = (cust.get('pincode') or '').strip() or None
        quote.site_address = (cust.get('siteAddress') or '').strip() or None
        quote.notes = (cust.get('notes') or '').strip() or None
        # Phase 2: optional PAN for B2B invoice.
        quote.customer_pan = (cust.get('pan') or '').strip() or None

        # Pricing overrides
        try:
            quote.transport_charges = float(pricing.get('transportCharges') or 0)
            quote.gst_percentage = float(pricing.get('gstPercentage') or 18)
            rev = pricing.get('revisedTotal')
            quote.revised_total = float(rev) if rev not in (None, '', 0, '0') else None
            quote.amount_received = float(pricing.get('amountReceived') or 0)
            # Phase 2: bill-level percent discount (0–100). Snapshotted
            # into quote.discount_amount by recompute_totals below.
            dpct = float(pricing.get('discountPercent') or 0)
            if dpct < 0 or dpct > 100:
                return jsonify({'error': 'discountPercent must be between 0 and 100'}), 400
            quote.discount_percent = dpct
        except (TypeError, ValueError):
            return jsonify({'error': 'invalid pricing numeric field'}), 400

        # Wipe existing items + insert fresh from payload.
        VetrovaQuoteItem.query.filter_by(quote_id=quote.id).delete()
        db.session.flush()

        for idx, r in enumerate(items_in):
            slug = (r.get('categorySlug') or '').strip() or 'unknown'
            selections = r.get('selections') if isinstance(r.get('selections'), dict) else {}
            panels = r.get('panels') if isinstance(r.get('panels'), list) else None
            try:
                running_ft = float(r.get('runningFt') or 0)
                qty = float(r.get('quantity') or 1)
                rate = float(r.get('ratePerUnit') or 0)
                subtotal = float(r.get('subtotal') or 0)
            except (TypeError, ValueError):
                return jsonify({'error': f'invalid numeric on item {idx + 1}'}), 400

            item = VetrovaQuoteItem(
                quote_id=quote.id,
                sort_order=int(r.get('sortOrder') or (idx + 1)),
                category_slug=slug,
                category_label=(r.get('categoryLabel') or VETROVA_CATEGORY_LABELS.get(slug, slug.title())),
                label=(r.get('label') or '').strip() or None,
                selections=json.dumps(selections) if selections else None,
                dimension_kind=(r.get('dimensionKind') or 'square_feet'),
                dimension_unit=(r.get('dimensionUnit') or 'ft'),
                running_ft=running_ft,
                quantity=qty,
                panels=json.dumps(panels) if panels else None,
                fabric_code=(r.get('fabricCode') or '').strip() or None,
                # Uploaded image data URL is preserved through revise —
                # BD isn't expected to re-upload; the client just echoes back
                # what was already there (form injects it hidden).
                uploaded_image_data_url=(r.get('uploadedImageDataUrl') or None),
                rate_per_unit=rate,
                subtotal=subtotal,
                notes=(r.get('notes') or '').strip() or None,
                # Phase 2: extras/discount marker.
                is_extra=bool(r.get('isExtra')),
            )
            db.session.add(item)

        # Bump revision count so BD can see quote has been edited.
        quote.revision_count = (quote.revision_count or 0) + 1

        # Auto-move to Revision stage on first save (stays put on subsequent
        # revisions so BD doesn't lose their manual stage moves).
        if quote.stage == 'quote_generated':
            evt = VetrovaStatusEvent(
                quote_id=quote.id,
                from_stage=quote.stage,
                to_stage='revision',
                note=f'Auto: revised by {current_user.email}',
                channel='none',
                send_status='skipped',
                triggered_by=current_user.id,
            )
            quote.stage = 'revision'
            db.session.add(evt)

        # Phase 2: snapshot the PRE-edit state into vetrova_quote_revisions.
        # Added to session BEFORE flush/commit so it commits atomically
        # with the item/quote mutations — either everything sticks or
        # nothing does.
        try:
            db.session.flush()
            db.session.refresh(quote)
            quote.recompute_totals()

            revision = VetrovaQuoteRevision(
                quote_id=quote.id,
                revision_number=quote.revision_count,
                prev_subtotal=pre_subtotal,
                new_subtotal=float(quote.subtotal or 0),
                prev_total=pre_total,
                new_total=float(quote.grand_total or 0),
                discount_percent=pre_discount_pct,
                snapshot=pre_snapshot,
                triggered_by=current_user.id,
                note=(data.get('revisionNote') or '').strip() or None,
            )
            db.session.add(revision)

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': 'db error', 'detail': str(e)}), 500

        return jsonify({'ok': True, 'id': quote.id,
                        'revisionCount': quote.revision_count,
                        'redirect': url_for('vetrova_quote_view', id=quote.id)}), 200

    # GET — render the editor form.
    # Serialize the items into a JSON structure the client-side JS reads.
    items_json = []
    for it in quote.items:
        items_json.append({
            'id':             it.id,
            'sortOrder':      it.sort_order,
            'categorySlug':   it.category_slug,
            'categoryLabel':  it.category_label,
            'label':          it.label,
            'selections':     it.selections_parsed,
            'dimensionKind':  it.dimension_kind or 'square_feet',
            'dimensionUnit':  it.dimension_unit or 'ft',
            'runningFt':      float(it.running_ft or 0),
            'quantity':       float(it.quantity or 1),
            'panels':         it.panels_parsed,
            'fabricCode':     it.fabric_code,
            'uploadedImageDataUrl': it.uploaded_image_data_url,
            'ratePerUnit':    float(it.rate_per_unit or 0),
            'subtotal':       float(it.subtotal or 0),
            'notes':          it.notes,
        })
    return render_template('quotes/vetrova_revise.html',
                           quote=quote,
                           items_json=items_json,
                           categories=VETROVA_CATEGORIES,
                           stage_labels=VETROVA_STAGE_LABELS)


@app.route('/quotes/vetrova/<int:id>/revisions/<int:rid>/snapshot', methods=['GET'])
@login_required
def vetrova_revision_snapshot(id, rid):
    """View-only JSON dump of a single revision's snapshot.

    Phase 2 audit-trail read path. The snapshot column is stored as
    JSON text; this route parses + returns the object with `Content-Type:
    application/json` so BD can eyeball what the quote looked like at
    that point in history. Deliberately no restore endpoint — history
    is view-only in Phase 2 (matches Bathqube's model).
    """
    revision = VetrovaQuoteRevision.query.filter_by(
        id=rid, quote_id=id,
    ).first_or_404()
    return jsonify({
        'revisionNumber': revision.revision_number,
        'quoteId': revision.quote_id,
        'createdAt': revision.created_at.isoformat() if revision.created_at else None,
        'triggeredByEmail': (revision.user.email if revision.user else None),
        'prevSubtotal': float(revision.prev_subtotal or 0),
        'newSubtotal': float(revision.new_subtotal or 0),
        'prevTotal': float(revision.prev_total or 0),
        'newTotal': float(revision.new_total or 0),
        'discountPercent': float(revision.discount_percent or 0),
        'note': revision.note,
        'snapshot': revision.snapshot_parsed,
    })


@app.route('/quotes/vetrova/<int:id>/send-first-quote', methods=['POST'])
@login_required
def vetrova_send_first_quote(id):
    """WhatsApp the customer the current Vetrova estimate PDF using the
    approved utility template 'first_quote_generated'. Phase 2 addition
    2026-07-29 — mirrors bathqube_send_first_quote (:9834), reusing the
    same WABA (Bathqube WhatsApp setup [[bathqube-whatsapp-setup]]) +
    the same 'first_quote_generated' template (already approved for
    two body variables: {{1}} first name, {{2}} quote ref).

    Flow:
      1. Render current customer-facing Vetrova estimate PDF
      2. Upload to Meta media store → media_id
      3. Send template with media_id populating the DOCUMENT header
      4. Log a WhatsAppMessage row (webhook updates status → delivered → read)
      5. Log a VetrovaStatusEvent with channel='whatsapp' so the
         Revisions/Stage history panel shows the send outcome
    """
    from models import WhatsAppMessage
    from utils.whatsapp import (upload_media, send_template_with_document,
                                normalize_phone)
    from utils.vetrova_quote_pdf import generate_vetrova_quote_pdf

    quote = VetrovaQuote.query.get_or_404(id)
    if not quote.phone:
        return jsonify({'success': False, 'error': 'Quote has no customer phone'}), 400

    # 1) Render the customer estimate PDF
    try:
        pdf_bytes = generate_vetrova_quote_pdf(quote)
    except Exception as e:
        return jsonify({'success': False, 'error': f'PDF render failed: {e}'}), 500

    filename = f'{quote.quote_ref or ("Vetrova-Q" + str(quote.id))}.pdf'

    # 2) Upload to Meta media store (media_id valid ~30 days)
    upload = upload_media(pdf_bytes, filename=filename, mime_type='application/pdf')
    if not upload.get('success'):
        # Persist the failure so BD can see it in Stage history.
        db.session.add(VetrovaStatusEvent(
            quote_id=quote.id,
            from_stage=quote.stage,
            to_stage=quote.stage,
            channel='whatsapp',
            subject='first_quote_generated (media upload failed)',
            send_status='failed',
            send_error=str(upload.get('error'))[:2000],
            triggered_by=current_user.id,
        ))
        db.session.commit()
        return jsonify({'success': False, 'error': f'Media upload failed: {upload.get("error")}'}), 502
    media_id = upload['media_id']

    # 3) Template body has two placeholders (approved on Meta):
    #    {{1}} = customer first name  {{2}} = quote ref (VQ-CAT-NNNN)
    first_name = (quote.customer_name or 'there').strip().split()[0]
    quote_ref  = quote.quote_ref or f'VQ-{quote.id}'
    body_variables = [first_name, quote_ref]

    to_number = normalize_phone(quote.phone) or quote.phone

    msg = WhatsAppMessage(
        lead_id=None,
        to_number=to_number,
        template_name='first_quote_generated',
        language='en',
        variables_json=json.dumps({'body': body_variables,
                                   'filename': filename,
                                   'media_id': media_id}),
        sent_by=current_user.id,
        status='queued',
    )
    db.session.add(msg)
    db.session.flush()

    # 4) Single send — template body + PDF header in one API call
    result = send_template_with_document(
        to=quote.phone,
        template_name='first_quote_generated',
        media_id=media_id,
        filename=filename,
        language='en',
        body_variables=body_variables,
    )

    # 5) Persist WhatsAppMessage + VetrovaStatusEvent audit trail
    if not result.get('success'):
        msg.status = 'failed'
        msg.error_message = result.get('error', 'Unknown error')
        db.session.add(VetrovaStatusEvent(
            quote_id=quote.id,
            from_stage=quote.stage,
            to_stage=quote.stage,
            channel='whatsapp',
            subject=f'first_quote_generated → {to_number}',
            message=json.dumps(body_variables),
            send_status='failed',
            send_error=(msg.error_message or '')[:2000],
            triggered_by=current_user.id,
        ))
        db.session.commit()
        return jsonify({'success': False, 'error': msg.error_message}), 502

    msg.status = 'sent'
    msg.wamid = result.get('wamid')
    db.session.add(VetrovaStatusEvent(
        quote_id=quote.id,
        from_stage=quote.stage,
        to_stage=quote.stage,
        channel='whatsapp',
        subject=f'first_quote_generated → {to_number}',
        message=json.dumps(body_variables),
        send_status='sent',
        provider_message_id=msg.wamid,
        triggered_by=current_user.id,
    ))
    db.session.commit()

    return jsonify({
        'success': True,
        'wamid': msg.wamid,
        'to': to_number,
    })


@app.route('/quotes/vetrova/<int:id>/pdf', methods=['GET'])
@login_required
def vetrova_quote_pdf(id):
    """Regenerate the customer-facing Vetrova quote PDF from current DB
    state. Uses vcore-side ReportLab (not glassyplatform's @react-pdf)
    so it reflects whatever BD just edited via revise.
    """
    from flask import send_file
    from utils.vetrova_quote_pdf import generate_vetrova_quote_pdf
    quote = VetrovaQuote.query.get_or_404(id)
    pdf_bytes = generate_vetrova_quote_pdf(quote)
    filename = f'{quote.quote_ref}.pdf'
    return send_file(io.BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=False, download_name=filename)


# ─── BD-side Bathqube quote creator routes (KAN-67 follow-up) ────────────────
# Companion to /quotes/upvc/new, but for Bathqube. BD picks options + types
# panels in vcore (no visuals, no leaving the app), the form server-fetches
# pricing from glassyplatform so it never drifts, and on Save a row is
# created in BOTH glassyplatform's bathspace-quotes (for BSP-NNNNNN
# numbering + admin parity) AND vcore's bathqube_quotes (with stage='draft'
# so BD reviews before sending).

UPVC_BATHQUBE_DIM_UNITS = ('mm', 'cm', 'm', 'ft', 'in')  # match the revise form selector


def _build_bd_config_data(form_enclosures, totals, dimension_unit_hint=None):
    """Compose the `configData` JSON shape that glassyplatform stores +
    that vcore's _bathqube_seed_items_from_config expects.

    Identical shape to the customer configurator's webhook payload so the
    seeder + ingest paths can't tell a BD quote from a customer one. Each
    enclosure carries its own `dimensionUnit` (post-2026-06-27 fix), so
    the top-level dimensionUnit is just a convenience hint for downstream
    UI rendering — typically the unit of the first enclosure.
    """
    cfg = {
        'schemaVersion': 2,
        'enclosures': form_enclosures,
        'subtotal': totals['subtotal'],
        'cgst': totals['cgst'],
        'sgst': totals['sgst'],
        'total': totals['total'],
    }
    if dimension_unit_hint:
        cfg['dimensionUnit'] = dimension_unit_hint
    elif form_enclosures and form_enclosures[0].get('dimensionUnit'):
        cfg['dimensionUnit'] = form_enclosures[0]['dimensionUnit']
    return cfg


@app.route('/quotes/bathqube/new', methods=['GET', 'POST'])
@login_required
def bathqube_quote_create():
    """BD-driven Bathqube quote create. GET renders the form (with options
    fetched live from glassyplatform), POST creates the row in
    glassyplatform + locally as stage='draft'."""
    if request.method == 'GET':
        try:
            options = _fetch_configurator_options()
        except RuntimeError as e:
            app.logger.error(f'[bathqube/new GET] options fetch failed: {e}')
            flash(
                'Could not load configurator options from glassyplatform. '
                f'Reason: {e}. Try again in a minute, or check that '
                'platform.glassy.in is reachable.',
                'danger',
            )
            return redirect(url_for('bathqube_quotes_list'))
        return render_template(
            'quotes/bathqube_new.html',
            options=options,
            dim_units=UPVC_BATHQUBE_DIM_UNITS,
            default_gst=18,
        )

    # POST — create the quote
    form = request.form
    customer_name = (form.get('customer_name') or '').strip()
    phone = (form.get('phone') or '').strip()
    if not customer_name or not phone:
        flash('Customer name and phone are required.', 'danger')
        return redirect(url_for('bathqube_quote_create'))

    raw_enclosures = (form.get('enclosures_json') or '').strip()
    if not raw_enclosures:
        flash('Add at least one enclosure with a panel before saving.', 'danger')
        return redirect(url_for('bathqube_quote_create'))
    try:
        parsed = json.loads(raw_enclosures)
    except Exception:
        flash('Form payload was malformed — refresh and try again.', 'danger')
        return redirect(url_for('bathqube_quote_create'))
    if not isinstance(parsed, list) or not parsed:
        flash('Add at least one enclosure before saving.', 'danger')
        return redirect(url_for('bathqube_quote_create'))

    try:
        gst_pct = float(form.get('gst_percentage') or 18)
    except ValueError:
        gst_pct = 18.0

    # Server-truth totals (don't trust the JS preview — server recomputes
    # using the same formula so the BD-stored subtotal is canonical even
    # if a stray copy-paste mucked with the JS-computed `subtotal` field).
    totals = _compute_bd_quote_totals(parsed, gst_pct)

    config_data = _build_bd_config_data(parsed, totals)

    # ── Push to glassyplatform first (gets the canonical BSP-NNNNNN) ──
    push_payload = {
        'name': customer_name,
        'phone': phone,
        'email': (form.get('email') or '').strip() or None,
        'pincode': (form.get('pincode') or '').strip() or None,
        'siteAddress': (form.get('site_address') or '').strip() or None,
        'configData': config_data,
    }
    try:
        external_id, estimate_number = _post_bd_quote_to_glassyplatform(push_payload)
    except RuntimeError as e:
        app.logger.error(f'[bathqube/new POST] glassyplatform push failed: {e}')
        flash(
            f'Could not save quote to glassyplatform: {e}. '
            'No vcore record was created — try again.',
            'danger',
        )
        return redirect(url_for('bathqube_quote_create'))

    # ── Mirror locally in vcore (stage=draft, BD reviews before send) ──
    quote = BathqubeQuote(
        external_id=external_id,
        estimate_number=estimate_number,
        customer_name=customer_name,
        phone=phone,
        email=push_payload['email'],
        pincode=push_payload['pincode'],
        site_address=push_payload['siteAddress'],
        config_data=json.dumps(config_data),
        original_config_data=json.dumps(config_data),  # capture the BD's original
        subtotal=totals['subtotal'],
        cgst=totals['cgst'],
        sgst=totals['sgst'],
        total=totals['total'],
        revised_total=totals['total'],  # no revision yet — current state = total
        gst_percentage=gst_pct,
        stage='draft',
        has_revision=False,
    )
    db.session.add(quote)
    db.session.flush()
    # Seed line items via the existing helper so a BD-created quote
    # renders identically on the view page to a customer-created one.
    for item in _bathqube_seed_items_from_config(quote):
        quote.items.append(item)

    db.session.add(BathqubeStatusEvent(
        quote_id=quote.id, from_stage=None, to_stage='draft',
        channel='none', send_status='skipped',
        subject='Created in vcore',
        message=f'BD-created from /quotes/bathqube/new ({len(parsed)} enclosure'
                f'{"s" if len(parsed) != 1 else ""}, total ₹{totals["total"]:,.2f}).',
        triggered_by=current_user.id,
    ))

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f'[bathqube/new POST] local commit failed: {e}')
        # The glassyplatform-side row exists but vcore-side doesn't — log
        # the BSP number so an admin can reconcile manually.
        flash(
            f'Quote {estimate_number} saved on glassyplatform but vcore '
            f'commit failed: {e}. Reach out to engineering with this '
            f'estimate number for reconciliation.',
            'danger',
        )
        return redirect(url_for('bathqube_quotes_list'))

    flash(
        f'Draft created ({estimate_number}). Review the items + totals, '
        'then click "Send to customer" to email + flip to Quote Generated.',
        'success',
    )
    return redirect(url_for('bathqube_quote_view', id=quote.id))


@app.route('/quotes/bathqube/<int:id>/send-first-quote', methods=['POST'])
@login_required
def bathqube_send_first_quote(id):
    """WhatsApp the customer the initial estimate PDF using the approved
    utility template 'first_quote_generated'. BD hits this the moment a
    fresh Bathqube quote lands from the configurator.

    Flow:
      1. Render the customer-facing estimate PDF
      2. Upload it to Meta's media store, get a media_id
      3. Send the template with the media_id populating its DOCUMENT
         header — PDF ships with the template body in a single message,
         no 24-hour window dependency (2026-07-03 rewrite; previous
         `first_qoute` template had a TEXT header so we needed a
         follow-up document that failed for cold leads).
      4. Log to WhatsAppMessage (webhook updates status from sent →
         delivered → read as it moves through Meta)
    """
    from models import BathqubeQuote, WhatsAppMessage
    from utils.whatsapp import (upload_media, send_template_with_document,
                                normalize_phone)
    from utils.bathqube_pdf import generate_bathqube_pdf

    quote = BathqubeQuote.query.get_or_404(id)
    if not quote.phone:
        return jsonify({'success': False, 'error': 'Quote has no customer phone'}), 400

    # 1) Render the customer estimate PDF
    try:
        pdf_bytes = generate_bathqube_pdf(quote)
    except Exception as e:
        return jsonify({'success': False, 'error': f'PDF render failed: {e}'}), 500

    filename = f'{quote.estimate_number or ("Bathqube-Q" + str(quote.id))}.pdf'

    # 2) Upload to Meta media store (media_id valid ~30 days)
    upload = upload_media(pdf_bytes, filename=filename, mime_type='application/pdf')
    if not upload.get('success'):
        return jsonify({'success': False, 'error': f'Media upload failed: {upload.get("error")}'}), 502
    media_id = upload['media_id']

    # 3) Template body has TWO placeholders (approved on Meta):
    #    {{1}} = customer first name
    #    {{2}} = quote / estimate number (falls back to internal id when
    #            the configurator hasn't assigned a BSP- number yet)
    # Header is a DOCUMENT header — media_id ships in the same request.
    first_name = (quote.customer_name or 'there').strip().split()[0]
    quote_ref  = quote.estimate_number or f'BQ-{quote.id}'
    body_variables = [first_name, quote_ref]

    to_number = normalize_phone(quote.phone) or quote.phone

    msg = WhatsAppMessage(
        lead_id=None,
        to_number=to_number,
        template_name='first_quote_generated',
        language='en',
        variables_json=json.dumps({'body': body_variables,
                                   'filename': filename,
                                   'media_id': media_id}),
        sent_by=current_user.id,
        status='queued',
    )
    db.session.add(msg)
    db.session.flush()

    # 4) Single send — template body + PDF header in one API call
    result = send_template_with_document(
        to=quote.phone,
        template_name='first_quote_generated',
        media_id=media_id,
        filename=filename,
        language='en',
        body_variables=body_variables,
    )

    if not result.get('success'):
        msg.status = 'failed'
        msg.error_message = result.get('error', 'Unknown error')
        db.session.commit()
        return jsonify({'success': False, 'error': msg.error_message}), 502

    msg.status = 'sent'
    msg.wamid = result.get('wamid')
    db.session.commit()

    return jsonify({
        'success': True,
        'wamid': msg.wamid,
        'to': to_number,
    })


@app.route('/quotes/bathqube/<int:id>/send-draft', methods=['POST'])
@login_required
def bathqube_send_draft(id):
    """Promote a draft → quote_generated + email the customer the PDF.

    Only valid when stage='draft'. The same /action/awaiting_payment flow
    already exists for later-stage email sends; this is the dedicated
    draft → quote_generated transition that BD hits after reviewing a
    self-created quote.
    """
    from utils.bathqube_pdf import generate_bathqube_pdf
    quote = BathqubeQuote.query.get_or_404(id)
    if quote.stage != 'draft':
        flash(
            f'Quote is in {STAGE_LABELS.get(quote.stage, quote.stage)} — '
            '"Send to customer" only applies to drafts.',
            'info',
        )
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    estimate = quote.estimate_number or f'BQ-{quote.id}'
    subject = f'Your Bathqube Estimate — {estimate}'
    message = (
        f'Dear {quote.customer_name},\n\n'
        f'Please find your Bathqube estimate attached.\n\n'
        f'Estimate #: {estimate}\n'
        f'Total: INR {float(quote.total or 0):,.2f}\n\n'
        f'Estimate validity: 15 days from issue. For any questions reply '
        f'to this email or WhatsApp +91 85500 11196.\n\n'
        f'Bathqube'
    )

    try:
        pdf_bytes = generate_bathqube_pdf(quote)
    except Exception as e:
        flash(f'PDF generation failed: {e}', 'danger')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    event = _bathqube_send_and_log(
        quote, action='quote_generated', subject=subject, message=message,
        attachments=[{
            'filename': f'{estimate}.pdf',
            'content': pdf_bytes,
            'content_type': 'application/pdf',
        }],
    )
    db.session.commit()

    if event.send_status == 'sent':
        flash(f'Estimate emailed to {quote.email}. Stage moved to Quote Generated.', 'success')
    elif event.send_status == 'skipped':
        flash('No email on file — stage moved to Quote Generated but no email sent.', 'info')
    else:
        flash(f'Email failed: {event.send_error}. Stage moved anyway.', 'warning')
    return redirect(url_for('bathqube_quote_view', id=quote.id))


def _bathqube_send_and_log(quote, *, action, subject, message, attachments=None):
    """Send the customer email (if email present), log a status event, update stage."""
    from utils.email_service import EmailService
    from_stage = quote.stage
    quote.stage = action
    if action == 'order_confirmation' and quote.purchased_at is None:
        quote.purchased_at = datetime.utcnow()

    event = BathqubeStatusEvent(
        quote_id=quote.id, from_stage=from_stage, to_stage=action,
        channel='email', subject=subject, message=message,
        triggered_by=current_user.id, send_status='pending',
    )
    db.session.add(event)

    if not quote.email:
        event.send_status = 'skipped'
        event.send_error = 'no email on file'
        return event

    try:
        svc = EmailService()
        res = svc.send_email(
            to=quote.email, subject=subject, body=message,
            attachments=attachments,
            from_email=os.getenv('BATHQUBE_FROM_EMAIL', 'Bathqube <support@bathqube.com>'),
        )
        if res.get('success'):
            event.send_status = 'sent'
            event.provider_message_id = res.get('message_id')
        else:
            event.send_status = 'failed'
            event.send_error = res.get('error', 'unknown error')
    except Exception as e:
        event.send_status = 'failed'
        event.send_error = str(e)
    return event


@app.route('/quotes/bathqube/<int:id>')
@login_required
def bathqube_quote_view(id):
    quote = BathqubeQuote.query.get_or_404(id)
    return render_template('quotes/bathqube_view.html',
                           quote=quote, stage_labels=STAGE_LABELS,
                           actions=BATHQUBE_ACTIONS,
                           stage_transitions=BATHQUBE_STAGE_TRANSITIONS)


@app.route('/quotes/bathqube/<int:id>/set-stage/<stage>', methods=['POST'])
@login_required
def bathqube_quote_set_stage(id, stage):
    """No-email stage transition. For pipeline markers (in_pipeline, awaiting_payment,
    closed_won) and dispositions (junk, rejected). Logs an audit event but doesn't
    touch email."""
    if stage not in BATHQUBE_STAGE_TRANSITIONS:
        flash('Unknown stage.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))

    quote = BathqubeQuote.query.get_or_404(id)
    from_stage = quote.stage
    if from_stage == stage:
        flash(f'Already in {STAGE_LABELS.get(stage, stage)}.', 'info')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    quote.stage = stage
    db.session.add(BathqubeStatusEvent(
        quote_id=quote.id, from_stage=from_stage, to_stage=stage,
        channel='internal', subject=None, message=None,
        triggered_by=current_user.id, send_status='skipped',
        send_error='no email — internal pipeline transition',
    ))
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save: {e}', 'danger')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    flash(f'Moved to {STAGE_LABELS.get(stage, stage)}.', 'success')
    return redirect(url_for('bathqube_quote_view', id=quote.id))


@app.route('/quotes/bathqube/<int:id>/action/<action>', methods=['GET', 'POST'])
@login_required
def bathqube_quote_action(id, action):
    """Generic stage-email action — send Order Confirmation, Processing, Order Ready, Thank You."""
    if action not in BATHQUBE_ACTIONS:
        flash('Unknown action.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))
    quote = BathqubeQuote.query.get_or_404(id)

    if request.method == 'POST':
        subject = (request.form.get('subject') or '').strip()
        message = (request.form.get('message') or '').strip()
        event = _bathqube_send_and_log(quote, action=action, subject=subject, message=message)
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Could not save: {e}', 'danger')
            return redirect(url_for('bathqube_quote_view', id=quote.id))

        if event.send_status == 'sent':
            flash(f'{STAGE_LABELS[action]} email sent.', 'success')
        elif event.send_status == 'skipped':
            flash(f'Stage updated to {STAGE_LABELS[action]} — no email (customer has no email on file).', 'info')
        else:
            flash(f'Stage updated but email failed: {event.send_error}', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    subject, body = render_stage_message(quote, action)
    return render_template('quotes/bathqube_action.html',
                           quote=quote, action=action,
                           action_label=STAGE_LABELS[action],
                           subject=subject, body=body)


@app.route('/quotes/bathqube/<int:id>/pdf', methods=['GET'])
@login_required
def bathqube_quote_pdf(id):
    """Download the current Bathqube quote as a PDF (for sales person to keep a local copy)."""
    from utils.bathqube_pdf import generate_bathqube_pdf
    from flask import send_file
    from io import BytesIO

    quote = BathqubeQuote.query.get_or_404(id)
    pdf_bytes = generate_bathqube_pdf(quote)
    filename = f"{quote.estimate_number or ('BQ-' + str(quote.id))}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename,
    )


def _ensure_work_order(quote):
    """Return the BathqubeWorkOrder row for this quote, creating an empty
    one the first time it's needed. Keeps the workshop docs working
    even on quotes that pre-date the work_order auto-create logic."""
    from models import BathqubeWorkOrder
    wo = quote.work_order  # backref, uselist=False
    if wo is None:
        wo = BathqubeWorkOrder(quote_id=quote.id)
        db.session.add(wo)
        db.session.flush()
        # Re-read so the relationship sees it
        quote.work_order = wo
    return wo


@app.route('/quotes/bathqube/<int:id>/work-order.pdf', methods=['GET'])
@login_required
def bathqube_work_order_pdf(id):
    """Generate a compact glass-workshop Work Order PDF.

    Manual button — appears in the quote view once stage = closed_won (the
    quote view template gates the button; we permit the route itself for
    any stage so BD can pre-print before confirming the close, but emit a
    Flash warning when stage hasn't reached closed_won yet)."""
    from utils.bathqube_pdf import generate_bathqube_work_order_pdf
    from flask import send_file
    from io import BytesIO

    quote = BathqubeQuote.query.get_or_404(id)
    # Lazily create the work_order row if missing so cutting_notes /
    # ops_notes can attach to something stable.
    _ensure_work_order(quote)
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    pdf_bytes = generate_bathqube_work_order_pdf(quote)
    filename = f"WO-{quote.estimate_number or ('BQ-' + str(quote.id))}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename,
    )


@app.route('/quotes/bathqube/<int:id>/cutting-notes', methods=['POST'])
@login_required
def bathqube_cutting_notes_save(id):
    """Save workshop fields on the quote's work_order: cutting_notes,
    delivery_eta, priority. All three are optional and editable
    independently — submit the whole form to update any of them. The
    Work Order PDF re-renders fresh on every download, so these edits
    are visible the next time BD clicks Generate Work Order PDF."""
    from datetime import date as _date

    quote = BathqubeQuote.query.get_or_404(id)
    wo = _ensure_work_order(quote)

    notes = (request.form.get('cutting_notes') or '').strip()
    if len(notes) > 4000:
        flash('Workshop notes are too long (max 4,000 chars).', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))
    wo.cutting_notes = notes or None

    eta_raw = (request.form.get('delivery_eta') or '').strip()
    if eta_raw:
        try:
            wo.delivery_eta = _date.fromisoformat(eta_raw)
        except ValueError:
            flash('Invalid delivery ETA date.', 'warning')
            return redirect(url_for('bathqube_quote_view', id=quote.id))
    else:
        wo.delivery_eta = None

    priority = (request.form.get('priority') or 'normal').strip().lower()
    if priority not in ('low', 'normal', 'urgent'):
        priority = 'normal'
    wo.priority = priority

    try:
        db.session.commit()
        flash('Workshop details saved.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save: {e}', 'danger')
    return redirect(url_for('bathqube_quote_view', id=quote.id))


@app.route('/quotes/bathqube/<int:id>/receipts', methods=['POST'])
@login_required
def bathqube_receipt_create(id):
    """Record a customer payment against this quote. Creates one
    BathqubePaymentReceipt row + immediately generates its PDF for
    download. Multiple receipts per quote — one per inflow (10% advance,
    install milestone, balance, etc.)."""
    from models import BathqubePaymentReceipt
    from utils.bathqube_pdf import next_receipt_number
    from datetime import date as _date

    quote = BathqubeQuote.query.get_or_404(id)

    # Validate the form
    raw_amount = (request.form.get('amount') or '').strip()
    try:
        amount = float(raw_amount)
    except ValueError:
        flash('Amount must be a number.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))
    if amount <= 0:
        flash('Amount must be greater than 0.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    method = (request.form.get('payment_method') or 'bank_transfer').strip()
    if method not in ('bank_transfer', 'upi', 'cash', 'cheque'):
        flash('Invalid payment method.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    utr = (request.form.get('utr_number') or '').strip() or None
    cheque = (request.form.get('cheque_number') or '').strip() or None
    if method in ('bank_transfer', 'upi') and not utr:
        flash('UTR / reference number is required for bank transfer / UPI payments.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))
    if method == 'cheque' and not cheque:
        flash('Cheque number is required for cheque payments.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    notes = (request.form.get('notes') or '').strip() or None

    received_raw = (request.form.get('received_at') or '').strip()
    if received_raw:
        try:
            received_at = _date.fromisoformat(received_raw)
        except ValueError:
            flash('Invalid received-at date.', 'warning')
            return redirect(url_for('bathqube_quote_view', id=quote.id))
    else:
        received_at = _date.today()

    # Mint the receipt number + insert
    receipt_number = next_receipt_number(db.session)
    rcpt = BathqubePaymentReceipt(
        quote_id=quote.id,
        receipt_number=receipt_number,
        received_at=received_at,
        amount=amount,
        payment_method=method,
        utr_number=utr,
        cheque_number=cheque,
        notes=notes,
        created_by=current_user.id,
    )
    db.session.add(rcpt)
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save receipt: {e}', 'danger')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    flash(f'Payment ₹{amount:,.2f} recorded · receipt {receipt_number}', 'success')
    # Open the receipt PDF in a new tab so BD can immediately share it.
    return redirect(url_for('bathqube_receipt_pdf', id=quote.id, receipt_id=rcpt.id))


@app.route('/quotes/bathqube/<int:id>/receipts/<int:receipt_id>.pdf', methods=['GET'])
@login_required
def bathqube_receipt_pdf(id, receipt_id):
    """Download a single payment receipt PDF. Re-generated fresh from
    the row each time — the cumulative running total reflects all
    receipts dated up to and including this one."""
    from models import BathqubePaymentReceipt
    from utils.bathqube_pdf import generate_bathqube_receipt_pdf
    from flask import send_file
    from io import BytesIO

    receipt = BathqubePaymentReceipt.query.get_or_404(receipt_id)
    if receipt.quote_id != id:
        # Defensive — shouldn't happen via the UI but guard against
        # URL-tampering where someone references a receipt that belongs
        # to a different quote.
        flash('Receipt does not belong to this quote.', 'warning')
        return redirect(url_for('bathqube_quote_view', id=id))

    pdf_bytes = generate_bathqube_receipt_pdf(receipt)
    filename = f"{receipt.receipt_number}.pdf"
    return send_file(
        BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=False,  # open in new tab; BD can save from browser
        download_name=filename,
    )


@app.route('/quotes/bathqube/<int:id>/revise', methods=['GET', 'POST'])
@login_required
def bathqube_quote_revise(id):
    """Sales-person bill revision UI.

    Full control over the bill:
      - Edit enclosures (type, material, panels, qty, price/sqft) like the web configurator
      - Add free-form 'extras' line items (installation, trim, etc. — can be negative for manual discounts)
      - Apply percentage discount on subtotal (BEFORE GST)
      - GST % and amount received
      - Auto-recompute totals; preview is live in the browser via JS
    """
    quote = BathqubeQuote.query.get_or_404(id)

    # First-time revise: snapshot customer's ORIGINAL submission, then seed
    # editable line items from the configurator data so the sales person
    # starts with the same line items the customer saw.
    if not quote.has_revision and not quote.items:
        if quote.original_config_data is None and quote.config_data:
            quote.original_config_data = quote.config_data
        for item in _bathqube_seed_items_from_config(quote):
            quote.items.append(item)
        db.session.flush()

    if request.method == 'POST':
        # ─── Capture BEFORE state for the revision audit log ─────────────
        # Done BEFORE any field mutation so we have a clean snapshot of what
        # the bill looked like prior to this save.
        prev_subtotal = float(quote.subtotal or 0)
        prev_total = float(quote.revised_total if quote.revised_total is not None else quote.total or 0)

        # Customer fields
        quote.customer_name = request.form.get('customer_name', quote.customer_name).strip()
        quote.phone = request.form.get('phone', quote.phone).strip()
        quote.email = (request.form.get('email') or '').strip() or None
        quote.pincode = (request.form.get('pincode') or '').strip() or None
        quote.customer_pan = (request.form.get('customer_pan') or '').strip().upper() or None
        quote.site_address = (request.form.get('site_address') or '').strip() or None
        quote.notes = (request.form.get('notes') or '').strip() or None

        # GST %, discount %, amount received
        try:
            quote.gst_percentage = float(request.form.get('gst_percentage') or 18)
        except ValueError:
            flash('Invalid GST %', 'warning')
        try:
            quote.discount_percent = max(0.0, min(100.0, float(request.form.get('discount_percent') or 0)))
        except ValueError:
            quote.discount_percent = 0
        try:
            quote.amount_received = float(request.form.get('amount_received') or 0)
        except ValueError:
            pass

        # ─── Enclosures: the structured part of the bill ─────────────────
        # Frontend serializes the enclosures array into a hidden JSON field.
        # We persist it to config_data (so the view template + revise page
        # both render from the same source of truth), then auto-generate
        # BathqubeQuoteItem rows from it (one item per panel per enclosure).
        enclosures_json = (request.form.get('enclosures_json') or '').strip()
        enclosures_list = []
        if enclosures_json:
            try:
                parsed = json.loads(enclosures_json)
                if isinstance(parsed, list):
                    enclosures_list = parsed
            except Exception as e:
                flash(f'Could not parse enclosures: {e}', 'warning')

        if enclosures_list:
            # Update config_data with edited enclosures (preserve other keys)
            current_cfg = quote.config if isinstance(quote.config, dict) else {}
            current_cfg['enclosures'] = enclosures_list
            current_cfg['schemaVersion'] = 2
            quote.config_data = json.dumps(current_cfg)

        # ─── Drop existing items, regenerate from enclosures + extras ────
        # Use collection.clear() (cascade='all, delete-orphan' handles DB delete)
        # instead of db.session.delete() in a loop — the latter removes from DB but
        # leaves stale objects in quote.items until commit, so the appends below
        # see old+new and double-count subtotal/snapshot/PDF.
        quote.items.clear()
        db.session.flush()

        sort_order = 0

        # Enclosure-derived items (one per panel) — flagged is_extra=False
        for enc_item in _bathqube_seed_items_from_config(quote):
            enc_item.sort_order = sort_order
            enc_item.is_extra = False
            quote.items.append(enc_item)
            sort_order += 1

        # ─── Extra free-form line items (additional charges / manual discounts) ─
        descs = request.form.getlist('extra_description')
        qtys = request.form.getlist('extra_quantity')
        rates = request.form.getlist('extra_rate')
        for d, q_, r in zip(descs, qtys, rates):
            desc = (d or '').strip()
            if not desc:
                continue
            try:
                qty = float(q_ or 0)
                rate = float(r or 0)
            except ValueError:
                qty, rate = 0, 0
            quote.items.append(BathqubeQuoteItem(
                sort_order=sort_order, description=desc[:500],
                quantity=qty, rate=rate, amount=round(qty * rate, 2),
                is_extra=True,
            ))
            sort_order += 1

        # Auto-recompute totals (subtotal → discount → GST → revised_total)
        _bathqube_recompute_totals(quote)
        quote.has_revision = True

        # ─── Insert audit row: one BathqubeQuoteRevision per save ────────
        # Full snapshot of what the bill looks like AFTER this save, plus
        # the before/after totals so the view page can show "Rev N: ₹X → ₹Y".
        quote.revision_count = (quote.revision_count or 0) + 1
        new_subtotal = float(quote.subtotal or 0)
        new_total = float(quote.revised_total or 0)
        snapshot = {
            'customer': {
                'name': quote.customer_name, 'phone': quote.phone, 'email': quote.email,
                'pincode': quote.pincode, 'site_address': quote.site_address,
            },
            'enclosures': enclosures_list,  # the JSON the sales person just submitted
            'items': [
                {'description': it.description, 'quantity': float(it.quantity or 0),
                 'rate': float(it.rate or 0), 'amount': float(it.amount or 0),
                 'is_extra': bool(it.is_extra)}
                for it in quote.items
            ],
            'gst_percentage': float(quote.gst_percentage or 0),
            'discount_percent': float(quote.discount_percent or 0),
            'discount_amount': float(quote.discount_amount or 0),
            'amount_received': float(quote.amount_received or 0),
        }
        revision = BathqubeQuoteRevision(
            quote_id=quote.id,
            revision_number=quote.revision_count,
            prev_subtotal=prev_subtotal, new_subtotal=new_subtotal,
            prev_total=prev_total, new_total=new_total,
            discount_percent=float(quote.discount_percent or 0),
            snapshot=json.dumps(snapshot),
            triggered_by=current_user.id if current_user.is_authenticated else None,
        )
        db.session.add(revision)

        # Email send?
        send = request.form.get('send_email') == 'on'
        subject = (request.form.get('subject') or '').strip()
        message = (request.form.get('message') or '').strip()

        try:
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            flash(f'Could not save: {e}', 'danger')
            return redirect(url_for('bathqube_quote_view', id=quote.id))

        event = None
        if send:
            try:
                from utils.bathqube_pdf import generate_bathqube_pdf
                pdf_bytes = generate_bathqube_pdf(quote)
                attachments = [{
                    'filename': f"{quote.estimate_number or ('BQ-' + str(quote.id))}-revised.pdf",
                    'content': pdf_bytes,
                }]
            except Exception as e:
                attachments = None
                flash(f'PDF generation failed, sending without attachment: {e}', 'warning')
            event = _bathqube_send_and_log(
                quote, action='revision',
                subject=subject, message=message, attachments=attachments,
            )

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f'Could not save: {e}', 'danger')
            return redirect(url_for('bathqube_quote_view', id=quote.id))

        if event is None:
            flash('Bill revised. No email sent.', 'success')
        elif event.send_status == 'sent':
            flash('Bill revised and revised estimate emailed (PDF attached).', 'success')
        elif event.send_status == 'skipped':
            flash('Bill revised — no email (customer has no email on file).', 'info')
        else:
            flash(f'Bill revised but email failed: {event.send_error}', 'warning')
        return redirect(url_for('bathqube_quote_view', id=quote.id))

    # GET — render editor
    subject, body = render_stage_message(quote, 'revision')

    # Enclosures: normalised from current config_data (handles legacy flat shape)
    enclosures = _bathqube_enclosures_from_cfg(quote.config or {})
    # Extras: line items added by the sales person on top of enclosure-derived rows
    extras = [it for it in quote.items if it.is_extra]

    return render_template('quotes/bathqube_revise.html',
                           quote=quote, subject=subject, body=body,
                           enclosures=enclosures, extras=extras,
                           option_lists=BATHQUBE_REVISE_OPTIONS)


# Standard configurator options exposed to the revise UI — sales person sees
# these as dropdown options, plus an "Other (specify)" choice that reveals a
# free-text input. Keep labels + surcharges in sync with glassyplatform's
# DEFAULT_* lists in ShowerConfigurator.tsx. Surcharge = ₹/sqft added to base
# rate when the option is picked. "Other" surcharges default to 0.
BATHQUBE_REVISE_OPTIONS = {
    'types': [
        ('Straight Shower Enclosure', 0),
        ('Curved Shower Enclosure', 80),
        ('L-Shaped Glass Enclosure', 80),
        ('Half Wall Shower Enclosure', 30),
        ('Walk-In Shower Enclosure', -30),
        ('Frameless Shower Enclosure', 120),
        ('Quadrant Shower Enclosure', 60),
        ('Sliding Door Shower Enclosure', 40),
    ],
    'materials': [
        ('Clear', 0), ('Frosted', 80), ('Fluted', 150), ('Tinted', 100),
    ],
    'fittings': [
        ('Chrome', 0), ('Matte Black', 80), ('Brushed Gold', 150), ('Rose Gold', 200),
    ],
    'hardwareTypes': [
        ('Glossy', 0), ('Matte', 100),
    ],
    'thicknesses': [
        ('8mm', 0), ('10mm', 120), ('12mm', 240),
    ],
}


# ============================================================================
# BATHQUBE OPS / FULFILLMENT
# ----------------------------------------------------------------------------
# Post-sale flow. After a quote hits closed_won the ops team picks it up here
# and drives it through site measurement → fabrication → installation →
# handover. All routes mounted under /quotes/bathqube/ops/* — separate from
# the sales-facing /quotes/bathqube/* views so each team sees only their work.
#
# Reuses BathqubeStatusEvent for the audit log and the same single `stage`
# column on bathqube_quotes (just extended with BATHQUBE_OPS_STAGES). Ops-
# specific data (assignee, scheduling, notes) lives on bathqube_work_orders;
# files/photos on bathqube_stage_attachments.
# ============================================================================

def _get_or_create_work_order(quote):
    """Return the BathqubeWorkOrder for a quote, creating an empty one on
    first access. Caller is responsible for db.session.commit()."""
    if quote.work_order:
        return quote.work_order
    wo = BathqubeWorkOrder(quote_id=quote.id)
    db.session.add(wo)
    db.session.flush()
    return wo


def _parse_form_datetime(s):
    """HTML <input type=datetime-local> submits "YYYY-MM-DDTHH:MM". Be lenient."""
    s = (s or '').strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%d %H:%M', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_form_date(s):
    s = (s or '').strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except ValueError:
        return None


@app.route('/quotes/bathqube/ops')
@login_required
def bathqube_ops_list():
    """Ops fulfillment list. Default view shows orders that need ops action:
    closed_won (just handed over by sales) through installed. handover_complete
    is filtered out unless the user opts in via ?handed_over=1."""
    search = (request.args.get('search') or '').strip()
    stage = (request.args.get('stage') or '').strip()
    assignee_id = (request.args.get('assignee') or '').strip()
    include_handed_over = request.args.get('handed_over') == '1'

    q = BathqubeQuote.query
    if search:
        like = f'%{search}%'
        q = q.filter(
            (BathqubeQuote.customer_name.ilike(like))
            | (BathqubeQuote.phone.ilike(like))
            | (BathqubeQuote.estimate_number.ilike(like))
        )

    if stage and stage in (BATHQUBE_OPS_STAGES + ('closed_won',)):
        q = q.filter_by(stage=stage)
    elif include_handed_over:
        q = q.filter(BathqubeQuote.stage.in_(('closed_won',) + BATHQUBE_OPS_STAGES))
    else:
        q = q.filter(BathqubeQuote.stage.in_(BATHQUBE_OPS_ACTIVE_STAGES))

    if assignee_id:
        try:
            aid = int(assignee_id)
            q = q.join(BathqubeWorkOrder, BathqubeQuote.id == BathqubeWorkOrder.quote_id) \
                 .filter(BathqubeWorkOrder.ops_assignee_id == aid)
        except ValueError:
            pass

    quotes = q.order_by(BathqubeQuote.updated_at.desc()).all()
    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    return render_template('quotes/bathqube_ops_list.html',
                           quotes=quotes, search=search, stage=stage,
                           assignee_id=assignee_id,
                           include_handed_over=include_handed_over,
                           users=users,
                           stage_labels=STAGE_LABELS,
                           ops_stages=BATHQUBE_OPS_STAGES,
                           ops_active_stages=BATHQUBE_OPS_ACTIVE_STAGES)


@app.route('/quotes/bathqube/ops/<int:id>')
@login_required
def bathqube_ops_view(id):
    """Work-order detail page — stage timeline, assignee + scheduling form,
    attachment gallery grouped by stage."""
    quote = BathqubeQuote.query.get_or_404(id)
    wo = _get_or_create_work_order(quote)
    # Persist lazy creation so subsequent loads don't keep flushing.
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()

    attachments_by_stage = {}
    for att in quote.stage_attachments:
        attachments_by_stage.setdefault(att.stage, []).append(att)

    users = User.query.filter_by(is_active=True).order_by(User.username).all()

    return render_template('quotes/bathqube_ops_view.html',
                           quote=quote, work_order=wo,
                           attachments_by_stage=attachments_by_stage,
                           users=users,
                           stage_labels=STAGE_LABELS,
                           ops_stages=BATHQUBE_OPS_STAGES,
                           events=quote.events)


@app.route('/quotes/bathqube/ops/<int:id>/stage/<stage>', methods=['POST'])
@login_required
def bathqube_ops_set_stage(id, stage):
    """Move a quote through the ops pipeline. Logs a BathqubeStatusEvent
    audit row for every transition. Customer messaging is wired up in
    Stage 4 (utils/bathqube_messages.py)."""
    if stage not in BATHQUBE_OPS_STAGES:
        flash('Unknown ops stage.', 'warning')
        return redirect(url_for('bathqube_ops_view', id=id))

    quote = BathqubeQuote.query.get_or_404(id)
    from_stage = quote.stage
    if from_stage == stage:
        flash(f'Already in {STAGE_LABELS.get(stage, stage)}.', 'info')
        return redirect(url_for('bathqube_ops_view', id=quote.id))

    # First ops transition — make sure the work order exists.
    _get_or_create_work_order(quote)

    quote.stage = stage
    db.session.add(BathqubeStatusEvent(
        quote_id=quote.id, from_stage=from_stage, to_stage=stage,
        channel='internal', subject=None, message=None,
        triggered_by=current_user.id, send_status='skipped',
        send_error='ops stage transition — customer messaging wired up in Stage 4',
    ))
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save: {e}', 'danger')
        return redirect(url_for('bathqube_ops_view', id=quote.id))

    flash(f'Moved to {STAGE_LABELS.get(stage, stage)}.', 'success')
    return redirect(url_for('bathqube_ops_view', id=quote.id))


@app.route('/quotes/bathqube/ops/<int:id>/details', methods=['POST'])
@login_required
def bathqube_ops_update_details(id):
    """Save the work-order detail form: assignee, scheduling dates, ops notes."""
    quote = BathqubeQuote.query.get_or_404(id)
    wo = _get_or_create_work_order(quote)

    assignee_raw = (request.form.get('ops_assignee_id') or '').strip()
    if assignee_raw == '':
        wo.ops_assignee_id = None
    else:
        try:
            wo.ops_assignee_id = int(assignee_raw)
        except ValueError:
            flash('Invalid assignee.', 'warning')
            return redirect(url_for('bathqube_ops_view', id=id))

    wo.measurement_scheduled_at  = _parse_form_datetime(request.form.get('measurement_scheduled_at'))
    wo.installation_scheduled_at = _parse_form_datetime(request.form.get('installation_scheduled_at'))
    wo.delivery_eta              = _parse_form_date(request.form.get('delivery_eta'))
    notes = (request.form.get('ops_notes') or '').strip()
    wo.ops_notes = notes or None

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save: {e}', 'danger')
        return redirect(url_for('bathqube_ops_view', id=id))

    flash('Ops details updated.', 'success')
    return redirect(url_for('bathqube_ops_view', id=id))


@app.route('/quotes/bathqube/ops/<int:id>/upload', methods=['POST'])
@login_required
def bathqube_ops_upload(id):
    """Attach a photo/document to a specific ops stage of the quote.
    Uploads to S3 (no watermark — these are internal ops files)."""
    quote = BathqubeQuote.query.get_or_404(id)

    file = request.files.get('file')
    if not file or not file.filename:
        flash('No file selected.', 'warning')
        return redirect(url_for('bathqube_ops_view', id=id))

    stage = (request.form.get('stage') or quote.stage).strip()
    if stage not in (BATHQUBE_OPS_STAGES + ('closed_won',)):
        flash(f'Cannot attach to non-ops stage "{stage}".', 'warning')
        return redirect(url_for('bathqube_ops_view', id=id))

    kind = (request.form.get('kind') or 'photo').strip()
    if kind not in ('photo', 'document', 'signature'):
        kind = 'photo'
    caption = (request.form.get('caption') or '').strip() or None

    try:
        url = S3Uploader().upload_bathqube_attachment(file, quote_id=quote.id, stage=stage)
    except Exception as e:
        flash(f'Upload failed: {e}', 'danger')
        return redirect(url_for('bathqube_ops_view', id=id))

    if not url:
        flash('Upload failed (S3 returned no URL).', 'danger')
        return redirect(url_for('bathqube_ops_view', id=id))

    db.session.add(BathqubeStageAttachment(
        quote_id=quote.id, stage=stage, kind=kind,
        file_url=url, caption=caption, uploaded_by=current_user.id,
    ))
    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Could not save attachment record: {e}', 'danger')
        return redirect(url_for('bathqube_ops_view', id=id))

    flash('File attached.', 'success')
    return redirect(url_for('bathqube_ops_view', id=id))


# ============================================================================
# RUN APPLICATION
# ============================================================================

def _run_migrations():
    """Safe ALTER TABLE migrations for columns added after initial schema creation."""
    migrations = [
        "ALTER TABLE leads ADD COLUMN facebook_lead_id VARCHAR(50) NULL UNIQUE",
    ]
    with db.engine.connect() as conn:
        for sql in migrations:
            try:
                conn.execute(db.text(sql))
                conn.commit()
            except Exception:
                pass  # Column already exists — ignore


# ============================================================================
# VETROVA INTERNI · UPVC QUOTES (KAN-67)
# ============================================================================
# BD-driven UPVC quotation flow. No public configurator — BD types each
# opening's spec + price by hand. Price written by BD is the TAXABLE
# amount; GST adds on top per the directive. PDF carries Vetrova Interni
# branding + Vetrova Tech Services Pvt Ltd legal entity + 20-year
# warranty highlight strip.

UPVC_STAGE_LABELS = {
    'draft':             'Draft',
    'sent':              'Sent',
    'revision':          'Revision',
    'awaiting_payment':  'Awaiting Payment',
    'closed_won':        'Closed Won',
    'rejected':          'Rejected',
    'junk':              'Junk',
}

# Per-stage Bootstrap badge colours — mirrors the Leadfy LEAD_STAGE_BADGE_CLASSES
# pattern so a UPVC quote's status reads at a glance in the list view.
# Colours chosen to match each stage's emotional weight:
#   draft     → secondary (gray — not yet live)
#   sent      → primary   (blue — out with customer)
#   revision  → warning   (yellow — in-flight changes, BD action pending)
#   awaiting  → warning   (yellow — money pending)
#   closed_won→ success   (green — paid)
#   rejected  → danger    (red — lost)
#   junk      → dark      (black — dead)
UPVC_STAGE_BADGE_CLASSES = {
    'draft':            'secondary',
    'sent':             'primary',
    'revision':         'warning',
    'awaiting_payment': 'warning',
    'closed_won':       'success',
    'rejected':         'danger',
    'junk':             'dark',
}


def upvc_stage_badge_class(stage):
    """Return the bootstrap badge colour suffix for the given UPVC stage.
    Falls back to 'info' for unknown values so the badge still renders."""
    return UPVC_STAGE_BADGE_CLASSES.get(stage, 'info')

UPVC_TRACK_TYPES   = ('swing', 'sliding', 'louvers')
# Sub-type options per track type. The form template + parser look up
# the valid set based on the chosen track_type — never show 2-track for
# Louvers or Fixed for Sliding.
UPVC_TRACK_SYSTEMS = ('2-track', '2.5-track', '3-track')   # Sliding only
UPVC_LOUVER_SUBTYPES = ('fixed', 'movable')                # Louvers only
UPVC_COLOURS       = ('white', 'black', 'wooden')
UPVC_UNITS         = ('mm', 'cm', 'm', 'ft', 'in')
# Stage transitions BD can flip directly from the view page without an
# email (mirrors Bathqube's BATHQUBE_STAGE_TRANSITIONS).
UPVC_STAGE_TRANSITIONS = ('revision', 'awaiting_payment', 'closed_won', 'rejected', 'junk')


def _upvc_line_sqft(width, height, unit):
    """Compute square feet from width × height + unit, using the same
    unit-aware to_inches() table as the Bathqube flow.

    Formula:
        w_in = width × UNIT_TO_INCHES[unit]
        h_in = height × UNIT_TO_INCHES[unit]
        sqft = (w_in × h_in) / 144

    Where UNIT_TO_INCHES = { mm: 1/25.4, cm: 1/2.54, in: 1,
                              m: 39.37007874015748, ft: 12 }

    The same panel typed in mm vs ft (e.g. 1219.2mm vs 4ft for width)
    must produce identical sqft to 4dp — that's the precision contract.

    Returns 0.0 on missing/invalid input — the calling parser already
    rejects rows without width/height, so a 0 here is a guard against
    later schema changes, not a silent degrade.
    """
    from utils.bathqube_dimensions import to_inches, _SUPPORTED_UNITS
    u = unit if unit in _SUPPORTED_UNITS else 'ft'
    try:
        w_in = to_inches(float(width or 0), u)
        h_in = to_inches(float(height or 0), u)
    except (TypeError, ValueError):
        return 0.0
    return round((w_in * h_in) / 144.0, 4)


def _next_upvc_estimate_number():
    """VI-UPVC-YYYY-NNNN. NNNN counts within the calendar year so the
    number stays short forever (resets each Jan). Uses MAX(estimate_number)
    pattern matching for atomic-ish next-id without a dedicated sequence
    table — collisions only possible under heavy concurrent create which
    UPVC quotes never see (BD-driven, manual)."""
    year = datetime.utcnow().year
    prefix = f'VI-UPVC-{year}-'
    last = (UpvcQuote.query
                     .filter(UpvcQuote.estimate_number.like(f'{prefix}%'))
                     .order_by(UpvcQuote.id.desc())
                     .first())
    seq = 1
    if last and last.estimate_number:
        try:
            seq = int(last.estimate_number.rsplit('-', 1)[-1]) + 1
        except Exception:
            pass
    return f'{prefix}{seq:04d}'


def _upvc_recompute_totals(quote):
    """Sum item amounts → subtotal; add transport; apply GST → total.

    Per KAN-67 answer #1: BD's per-line price IS the taxable amount, GST
    calculated on top. Transportation charge added 2026-07-21 — folded
    into the taxable base BEFORE GST so the PDF shows:
        Subtotal
        Transportation
        CGST/SGST (on subtotal + transport)
        Grand Total

    Math:
        subtotal      = Σ items.amount                        (unchanged)
        taxable       = subtotal + transport_charges          (new)
        cgst = sgst   = taxable × (gst_percent / 2 / 100)
        total         = taxable + cgst + sgst
    """
    subtotal  = sum(float(it.amount or 0) for it in quote.items)
    transport = float(quote.transport_charges or 0)
    taxable   = subtotal + transport
    gst_pct   = float(quote.gst_percentage or 0)
    cgst      = round(taxable * gst_pct / 2 / 100, 2)
    sgst      = round(taxable * gst_pct / 2 / 100, 2)
    quote.subtotal = subtotal
    quote.cgst = cgst
    quote.sgst = sgst
    quote.total = round(taxable + cgst + sgst, 2)


def _upvc_parse_items_from_form(form):
    """Parse the create/edit form's repeated-row encoding into clean dicts.

    Form encoding (BD's create form repeats per row, index i):
        items[i][label]        — optional opening label
        items[i][track_type]   — 'swing' | 'sliding' | 'louvers'   (required)
        items[i][track_system] — '2-track' | '2.5-track' | '3-track'  (sliding only)
                                  OR 'fixed' | 'movable'              (louvers only)
                                  OR NULL                              (swing)
        items[i][width]        — number > 0  (in `unit`)              (required)
        items[i][height]       — number > 0  (in `unit`)              (required)
        items[i][unit]         — 'mm' | 'cm' | 'm' | 'ft' | 'in' (KAN-34)
        items[i][colour]       — 'white' | 'black' | 'wooden'
        items[i][quantity]     — int / float, defaults to 1 if blank
        items[i][rate]         — BD-typed PER-SQFT price (₹). Amount = qty×sqft×rate.

    Empty rows (no track_type / rate <= 0) are skipped silently — keeps
    the form forgiving when BD drafts and removes a row by clearing it.
    """
    import re as _re
    from collections import defaultdict
    rows = defaultdict(dict)
    pat = _re.compile(r'^items\[(\d+)\]\[(\w+)\]$')
    # Werkzeug MultiDict — iterate every key (form.items() returns one
    # value per key by default; for our shape that's exactly what we want).
    for k in form.keys():
        m = pat.match(k)
        if not m:
            continue
        rows[int(m.group(1))][m.group(2)] = (form.get(k) or '').strip()
    out = []
    for idx in sorted(rows.keys()):
        r = rows[idx]
        track_type = (r.get('track_type') or '').strip().lower()
        if track_type not in UPVC_TRACK_TYPES:
            continue
        try:
            rate = float(r.get('rate') or 0)
        except ValueError:
            continue
        if rate <= 0:
            continue
        # Quantity defaults to 1 — BD can leave it blank for a single
        # opening. Treat 0 / negative as 1 so a misclick doesn't silently
        # erase the line.
        try:
            qty = float(r.get('quantity') or 1)
        except ValueError:
            qty = 1
        if qty <= 0:
            qty = 1
        colour = (r.get('colour') or '').strip().lower()
        if colour not in UPVC_COLOURS:
            continue
        unit = (r.get('unit') or 'ft').strip().lower()
        if unit not in UPVC_UNITS:
            unit = 'ft'
        # Resolve track_system based on track_type. Sliding accepts
        # 2-track/2.5-track/3-track, Louvers accepts fixed/movable,
        # Swing has none. Invalid combos collapse to NULL rather than
        # erroring — the form's conditional dropdown shouldn't let BD
        # submit a wrong combo, but be defensive against direct POSTs.
        track_system_raw = (r.get('track_system') or '').strip().lower()
        if track_type == 'sliding' and track_system_raw in UPVC_TRACK_SYSTEMS:
            track_system = track_system_raw
        elif track_type == 'louvers' and track_system_raw in UPVC_LOUVER_SUBTYPES:
            track_system = track_system_raw
        else:
            track_system = None
        # Width/height are REQUIRED now — without them sqft can't be
        # computed and the line amount would be 0. Reject silently
        # (matches how missing rate/colour rows are filtered).
        def _num_or_none(v):
            try:
                return float(v) if v not in ('', None) else None
            except ValueError:
                return None
        width = _num_or_none(r.get('width'))
        height = _num_or_none(r.get('height'))
        if width is None or width <= 0 or height is None or height <= 0:
            continue
        out.append({
            'sort_order':   idx,
            'label':        (r.get('label') or '').strip()[:200] or None,
            'track_type':   track_type,
            'track_system': track_system,
            'width':        width,
            'height':       height,
            'unit':         unit,
            'colour':       colour,
            'quantity':     qty,
            'rate':         rate,
        })
    return out


def _upvc_apply_items(quote, item_dicts):
    """Wipe existing items + recreate from parsed dicts. Called by create
    + edit + revise to keep one path that writes items.

    Uses the SQLAlchemy `clear()` pattern on the delete-orphan relationship
    rather than per-row `session.delete(it)`. The latter marks rows for
    deletion in the DB but leaves their references in `quote.items`'s
    in-memory cache — so the very next iteration (e.g. `_upvc_recompute_totals`)
    would still see them and sum BOTH old + new amounts, producing
    inflated subtotals on every edit. The `clear()` path drops them
    from the collection AND triggers the delete cascade.
    """
    quote.items.clear()
    db.session.flush()
    for d in item_dicts:
        rate = float(d['rate'])           # ₹/sqft
        qty  = float(d.get('quantity') or 1)
        sqft = _upvc_line_sqft(d.get('width'), d.get('height'), d.get('unit'))
        quote.items.append(UpvcQuoteItem(
            sort_order   = d['sort_order'],
            label        = d.get('label'),
            track_type   = d['track_type'],
            track_system = d.get('track_system'),
            width        = d.get('width'),
            height       = d.get('height'),
            unit         = d['unit'],
            colour       = d['colour'],
            quantity     = qty,
            sqft         = sqft,
            rate         = rate,
            amount       = round(qty * sqft * rate, 2),
        ))


def _upvc_send_estimate_email(quote):
    """Send the customer the estimate PDF + log a UpvcStatusEvent.

    Reuses the same EmailService + Resend setup as Bathqube but with
    Vetrova Interni branding. Per KAN-46/47 we send from support@glassy.in
    once Resend verifies the glassy.in domain — until then this falls
    back to whatever RESEND_SENDER_EMAIL is configured.

    Returns the UpvcStatusEvent row (caller decides whether to flash
    success / failure to the user)."""
    from utils.email_service import EmailService
    from utils.vetrova_upvc_pdf import generate_upvc_quote_pdf

    from_stage = quote.stage
    if from_stage == 'draft':
        # draft → sent on first email; subsequent re-sends stay on whatever
        # stage they're on (revision, awaiting_payment) and just log the email.
        quote.stage = 'sent'

    subject = f'Vetrova Interni Estimate — {quote.estimate_number or quote.id}'
    body = (
        f'Dear {quote.customer_name},\n\n'
        f'Please find your Vetrova Interni UPVC estimate attached.\n\n'
        f'Estimate #: {quote.estimate_number or quote.id}\n'
        f'Total: INR {float(quote.total or 0):,.2f}\n'
        f'Valid for: {quote.validity_days} days from today.\n\n'
        f'Every Vetrova Interni UPVC installation is backed by our 20-year warranty.\n\n'
        f'For any questions, reply to this email or call us.\n\n'
        f'Vetrova Interni\n'
        f'(Vetrova Tech Services Private Limited)'
    )

    event = UpvcStatusEvent(
        quote_id=quote.id, from_stage=from_stage, to_stage=quote.stage,
        channel='email', subject=subject, message=body,
        triggered_by=current_user.id, send_status='pending',
    )
    db.session.add(event)

    if not quote.email:
        event.send_status = 'skipped'
        event.send_error  = 'no email on file'
        return event

    try:
        pdf_bytes = generate_upvc_quote_pdf(quote)
        svc = EmailService()
        res = svc.send_email(
            to=quote.email,
            subject=subject,
            body=body,
            attachments=[{
                'filename': f'{quote.estimate_number or ("VI-UPVC-" + str(quote.id))}.pdf',
                'content':  pdf_bytes,
                'content_type': 'application/pdf',
            }],
            from_email=os.getenv('RESEND_SENDER_EMAIL', 'Vetrova Interni <support@glassy.in>'),
        )
        if res.get('success'):
            event.send_status = 'sent'
            event.provider_message_id = res.get('message_id')
        else:
            event.send_status = 'failed'
            event.send_error  = res.get('error', 'unknown error')
    except Exception as e:
        event.send_status = 'failed'
        event.send_error  = str(e)
    return event


@app.route('/quotes/upvc', methods=['GET'])
@login_required
def upvc_quotes_list():
    """BD's list view. Defaults to active stages (hides junk + rejected)."""
    search = (request.args.get('search') or '').strip()
    stage_filter = (request.args.get('stage') or '').strip()
    include_archived = request.args.get('archived') == '1'

    q = UpvcQuote.query
    if search:
        like = f'%{search}%'
        q = q.filter(db.or_(
            UpvcQuote.customer_name.ilike(like),
            UpvcQuote.phone.ilike(like),
            UpvcQuote.email.ilike(like),
            UpvcQuote.estimate_number.ilike(like),
        ))
    if stage_filter:
        q = q.filter(UpvcQuote.stage == stage_filter)
    elif not include_archived:
        q = q.filter(UpvcQuote.stage.in_(UPVC_ACTIVE_STAGES))

    quotes = q.order_by(UpvcQuote.created_at.desc()).limit(500).all()
    return render_template('quotes/upvc_list.html',
                           quotes=quotes,
                           stages=UPVC_STAGES,
                           stage_labels=UPVC_STAGE_LABELS,
                           stage_badges=UPVC_STAGE_BADGE_CLASSES,
                           search=search,
                           stage=stage_filter,
                           include_archived=include_archived)


@app.route('/quotes/upvc/new', methods=['GET', 'POST'])
@login_required
def upvc_quote_create():
    if request.method == 'GET':
        return render_template('quotes/upvc_form.html',
                               quote=None, items=[],
                               track_types=UPVC_TRACK_TYPES,
                               track_systems=UPVC_TRACK_SYSTEMS,
                               louver_subtypes=UPVC_LOUVER_SUBTYPES,
                               colours=UPVC_COLOURS,
                               units=UPVC_UNITS,
                               default_gst=18,
                               default_transport_charges=0,
                               default_validity_days=10)

    # POST — create
    form = request.form
    customer_name = (form.get('customer_name') or '').strip()
    phone = (form.get('phone') or '').strip()
    if not customer_name or not phone:
        flash('Name and phone are required.', 'danger')
        return redirect(url_for('upvc_quote_create'))

    item_dicts = _upvc_parse_items_from_form(form)
    if not item_dicts:
        flash('Add at least one opening with a valid price before saving.', 'danger')
        return redirect(url_for('upvc_quote_create'))

    try:
        gst_pct = float(form.get('gst_percentage') or 18)
    except ValueError:
        gst_pct = 18
    try:
        validity = int(form.get('validity_days') or 10)
    except ValueError:
        validity = 10

    try:
        transport_charges = float(form.get('transport_charges') or 0)
        if transport_charges < 0:
            transport_charges = 0
    except ValueError:
        transport_charges = 0

    quote = UpvcQuote(
        estimate_number=_next_upvc_estimate_number(),
        customer_name=customer_name,
        phone=phone,
        email=(form.get('email') or '').strip() or None,
        pincode=(form.get('pincode') or '').strip() or None,
        customer_pan=(form.get('customer_pan') or '').strip().upper() or None,
        site_address=(form.get('site_address') or '').strip() or None,
        gst_percentage=gst_pct,
        transport_charges=transport_charges,
        validity_days=validity,
        notes=(form.get('notes') or '').strip() or None,
        stage='draft',
        created_by=current_user.id,
    )
    db.session.add(quote)
    db.session.flush()  # need quote.id for items + estimate_number visibility

    _upvc_apply_items(quote, item_dicts)
    _upvc_recompute_totals(quote)
    db.session.commit()

    flash(f'Draft created — {quote.estimate_number}. Review then send to the customer.', 'success')
    return redirect(url_for('upvc_quote_view', id=quote.id))


@app.route('/quotes/upvc/<int:id>', methods=['GET'])
@login_required
def upvc_quote_view(id):
    quote = UpvcQuote.query.get_or_404(id)
    return render_template('quotes/upvc_view.html',
                           quote=quote,
                           stage_labels=UPVC_STAGE_LABELS,
                           stage_badges=UPVC_STAGE_BADGE_CLASSES,
                           stage_transitions=UPVC_STAGE_TRANSITIONS,
                           track_systems=UPVC_TRACK_SYSTEMS)


def _upvc_save_form(quote, form, *, mode):
    """Shared save path used by Edit and Revise.

    mode='edit'   — silent fix: no revision row, no counter bump, no
                    stage flip. For typo correction by BD. Stage stays
                    where it was (draft / sent / revision / ...).
    mode='revise' — proper revision: writes UpvcQuoteRevision audit row,
                    bumps revision_count, flips a sent quote to 'revision'
                    stage. For "customer asked for changes" workflow.

    Returns (flash_message, flash_level) for the caller to flash.
    """
    customer_name = (form.get('customer_name') or '').strip()
    phone = (form.get('phone') or '').strip()
    if not customer_name or not phone:
        return ('Name and phone are required.', 'danger')

    item_dicts = _upvc_parse_items_from_form(form)
    if not item_dicts:
        return ('Add at least one opening with a valid price before saving.', 'danger')

    # Snapshot pre-save state for the revision audit log (only used in
    # 'revise' mode but cheap to compute either way).
    prev_subtotal = float(quote.subtotal or 0)
    prev_total    = float(quote.total or 0)
    snapshot = {
        'customer': {
            'name': quote.customer_name, 'phone': quote.phone,
            'email': quote.email, 'pincode': quote.pincode,
            'site_address': quote.site_address,
        },
        'items': [{
            'label': it.label, 'track_type': it.track_type,
            'track_system': it.track_system,
            'width': float(it.width) if it.width else None,
            'height': float(it.height) if it.height else None,
            'unit': it.unit, 'colour': it.colour,
            'quantity': float(it.quantity or 1),
            'sqft': float(it.sqft or 0),
            'rate': float(it.rate or 0), 'amount': float(it.amount or 0),
        } for it in quote.items],
        'gst_percentage': float(quote.gst_percentage or 0),
        'transport_charges': float(quote.transport_charges or 0),
    }

    quote.customer_name = customer_name
    quote.phone = phone
    quote.email = (form.get('email') or '').strip() or None
    quote.pincode = (form.get('pincode') or '').strip() or None
    quote.customer_pan = (form.get('customer_pan') or '').strip().upper() or None
    quote.site_address = (form.get('site_address') or '').strip() or None
    quote.notes = (form.get('notes') or '').strip() or None
    try:
        quote.gst_percentage = float(form.get('gst_percentage') or quote.gst_percentage or 18)
    except ValueError:
        pass
    try:
        quote.validity_days = int(form.get('validity_days') or quote.validity_days or 10)
    except ValueError:
        pass
    try:
        transport_charges = float(form.get('transport_charges') or 0)
        quote.transport_charges = max(0, transport_charges)
    except ValueError:
        # Leave the existing value untouched on bad input rather than
        # silently zeroing out a previously-typed transport charge.
        pass

    _upvc_apply_items(quote, item_dicts)
    _upvc_recompute_totals(quote)

    if mode == 'revise':
        quote.revision_count = int(quote.revision_count or 0) + 1
        db.session.add(UpvcQuoteRevision(
            quote_id=quote.id,
            revision_number=quote.revision_count,
            prev_subtotal=prev_subtotal,
            new_subtotal=float(quote.subtotal or 0),
            prev_total=prev_total,
            new_total=float(quote.total or 0),
            snapshot=json.dumps(snapshot),
            triggered_by=current_user.id,
        ))
        # First revise after sending flips stage to 'revision' so the
        # list view shows at a glance which sent quotes have been edited.
        if quote.stage == 'sent':
            db.session.add(UpvcStatusEvent(
                quote_id=quote.id, from_stage='sent', to_stage='revision',
                channel='none', send_status='skipped',
                subject='Revised in vcore',
                message=f'Quote revised (revision #{quote.revision_count}).',
                triggered_by=current_user.id,
            ))
            quote.stage = 'revision'
        return (f'Saved revision #{quote.revision_count}.', 'success')

    # mode == 'edit' — silent fix, no audit row, no counter bump
    return ('Changes saved.', 'success')


@app.route('/quotes/upvc/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def upvc_quote_edit(id):
    """Silent edit — fix a typo / wrong price / wrong dimension. Does
    NOT bump revision_count or write a UpvcQuoteRevision row. Use this
    when BD made a data-entry mistake. For real customer-driven changes
    after the quote has been sent, use /revise instead."""
    quote = UpvcQuote.query.get_or_404(id)

    if request.method == 'GET':
        return render_template('quotes/upvc_form.html',
                               quote=quote, items=quote.items, mode='edit',
                               track_types=UPVC_TRACK_TYPES,
                               track_systems=UPVC_TRACK_SYSTEMS,
                               louver_subtypes=UPVC_LOUVER_SUBTYPES,
                               colours=UPVC_COLOURS,
                               units=UPVC_UNITS,
                               default_gst=float(quote.gst_percentage or 18),
                               default_transport_charges=float(quote.transport_charges or 0),
                               default_validity_days=int(quote.validity_days or 10))

    msg, level = _upvc_save_form(quote, request.form, mode='edit')
    if level != 'success':
        flash(msg, level)
        return redirect(url_for('upvc_quote_edit', id=quote.id))
    db.session.commit()
    flash(msg, 'success')
    return redirect(url_for('upvc_quote_view', id=quote.id))


@app.route('/quotes/upvc/<int:id>/revise', methods=['GET', 'POST'])
@login_required
def upvc_quote_revise(id):
    """Proper revision — opens the same editor as /edit, but saving
    writes a UpvcQuoteRevision row + bumps revision_count. Use this
    when the customer asked for changes after seeing the original
    quote. Saved as Revision #1, #2, ...; sent quotes automatically
    move to the 'revision' stage on first save here."""
    quote = UpvcQuote.query.get_or_404(id)

    if request.method == 'GET':
        return render_template('quotes/upvc_form.html',
                               quote=quote, items=quote.items, mode='revise',
                               track_types=UPVC_TRACK_TYPES,
                               track_systems=UPVC_TRACK_SYSTEMS,
                               louver_subtypes=UPVC_LOUVER_SUBTYPES,
                               colours=UPVC_COLOURS,
                               units=UPVC_UNITS,
                               default_gst=float(quote.gst_percentage or 18),
                               default_transport_charges=float(quote.transport_charges or 0),
                               default_validity_days=int(quote.validity_days or 10))

    msg, level = _upvc_save_form(quote, request.form, mode='revise')
    if level != 'success':
        flash(msg, level)
        return redirect(url_for('upvc_quote_revise', id=quote.id))
    db.session.commit()
    flash(msg, 'success')
    return redirect(url_for('upvc_quote_view', id=quote.id))


@app.route('/quotes/upvc/<int:id>/set-stage/<stage>', methods=['POST'])
@login_required
def upvc_quote_set_stage(id, stage):
    """Single-click stage flip from the view page. No email — just a log row."""
    if stage not in UPVC_STAGES:
        flash('Invalid stage.', 'danger')
        return redirect(url_for('upvc_quote_view', id=id))
    quote = UpvcQuote.query.get_or_404(id)
    if quote.stage == stage:
        flash(f'Already in {UPVC_STAGE_LABELS.get(stage, stage)}.', 'info')
        return redirect(url_for('upvc_quote_view', id=id))

    db.session.add(UpvcStatusEvent(
        quote_id=quote.id, from_stage=quote.stage, to_stage=stage,
        channel='none', send_status='skipped',
        message=f'Stage changed manually to {UPVC_STAGE_LABELS.get(stage, stage)}.',
        triggered_by=current_user.id,
    ))
    quote.stage = stage
    if stage == 'closed_won' and quote.purchased_at is None:
        quote.purchased_at = datetime.utcnow()
    db.session.commit()
    flash(f'Moved to {UPVC_STAGE_LABELS.get(stage, stage)}.', 'success')
    return redirect(url_for('upvc_quote_view', id=id))


@app.route('/quotes/upvc/<int:id>/send', methods=['POST'])
@login_required
def upvc_quote_send(id):
    """Email the customer the current estimate PDF + log the send."""
    quote = UpvcQuote.query.get_or_404(id)
    if not quote.items:
        flash('Cannot send an empty quote — add at least one opening first.', 'danger')
        return redirect(url_for('upvc_quote_view', id=id))
    event = _upvc_send_estimate_email(quote)
    db.session.commit()
    if event.send_status == 'sent':
        flash(f'Estimate emailed to {quote.email}.', 'success')
    elif event.send_status == 'skipped':
        flash('No email on file — stage updated but no email sent.', 'info')
    else:
        flash(f'Email failed: {event.send_error}', 'danger')
    return redirect(url_for('upvc_quote_view', id=id))


@app.route('/quotes/upvc/<int:id>/pdf', methods=['GET'])
@login_required
def upvc_quote_pdf(id):
    """Download the current estimate as a PDF (BD-facing)."""
    from utils.vetrova_upvc_pdf import generate_upvc_quote_pdf
    from flask import send_file
    from io import BytesIO

    quote = UpvcQuote.query.get_or_404(id)
    pdf_bytes = generate_upvc_quote_pdf(quote)
    filename = f"{quote.estimate_number or ('VI-UPVC-' + str(quote.id))}.pdf"
    return send_file(BytesIO(pdf_bytes), mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


@app.route('/quotes/upvc/<int:id>/delete', methods=['POST'])
@login_required
def upvc_quote_delete(id):
    """Admin-only hard delete (mirrors bathqube_quote_delete)."""
    if not current_user.is_admin():
        flash('Only admins can delete quotes.', 'danger')
        return redirect(url_for('upvc_quote_view', id=id))
    quote = UpvcQuote.query.get_or_404(id)
    estimate = quote.estimate_number or f'VI-UPVC-{quote.id}'
    db.session.delete(quote)
    db.session.commit()
    flash(f'Deleted quote {estimate}.', 'success')
    return redirect(url_for('upvc_quotes_list'))


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        _run_migrations()
    app.run(debug=True, host='0.0.0.0', port=8080)
