"""
Import products from rohit_products_final.xlsx into the database
"""
import openpyxl
import json
from app import app
from models import db, Product

# Fields that have dedicated columns (not stored in JSON)
DEDICATED_FIELDS = {
    'Category', 'Product Name', 'Product URL', 'Product Price',
    'Image_1_URL', 'Image_2_URL', 'Image_3_URL', 'Image_4_URL',
    'Availability', 'Description', 'Material', 'Brand',
    'Usage/Application', 'Thickness', 'Shape', 'Pattern'
}

def import_products_from_excel(excel_file='rohit_products_final.xlsx'):
    """Import products from Excel file"""
    with app.app_context():
        print(f"Reading Excel file: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"Found {len(headers)} columns")
        
        # Track statistics
        imported_count = 0
        skipped_count = 0
        
        # Process each row
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            try:
                # Extract values
                row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
                
                # Skip if no product name
                if not row_data.get('Product Name'):
                    skipped_count += 1
                    continue
                
                # Create product instance
                product = Product(
                    category=row_data.get('Category', 'Uncategorized'),
                    product_name=row_data.get('Product Name'),
                    product_url=row_data.get('Product URL'),
                    price=row_data.get('Product Price'),
                    image_1_url=row_data.get('Image_1_URL'),
                    image_2_url=row_data.get('Image_2_URL'),
                    image_3_url=row_data.get('Image_3_URL'),
                    image_4_url=row_data.get('Image_4_URL'),
                    availability=row_data.get('Availability'),
                    description=row_data.get('Description'),
                    material=row_data.get('Material'),
                    brand=row_data.get('Brand'),
                    usage_application=row_data.get('Usage/Application'),
                    thickness=row_data.get('Thickness'),
                    shape=row_data.get('Shape'),
                    pattern=row_data.get('Pattern')
                )
                
                # Collect additional specifications (sparse fields)
                specifications = {}
                for header, value in row_data.items():
                    # Skip dedicated fields and empty values
                    if header not in DEDICATED_FIELDS and value is not None and str(value).strip():
                        specifications[header] = str(value).strip()
                
                # Store specifications as JSON
                if specifications:
                    product.set_specifications(specifications)
                
                # Add to database
                db.session.add(product)
                imported_count += 1
                
                if imported_count % 10 == 0:
                    print(f"  Processed {imported_count} products...")
                
            except Exception as e:
                print(f"✗ Error processing row {row_idx}: {e}")
                skipped_count += 1
                continue
        
        # Commit all changes
        try:
            db.session.commit()
            print(f"\n✓ Successfully imported {imported_count} products!")
            if skipped_count > 0:
                print(f"  Skipped {skipped_count} rows")
            
            # Show summary
            print("\nImport Summary:")
            print(f"  Total products: {Product.query.count()}")
            print(f"  Categories: {len(Product.get_categories())}")
            print(f"  Brands: {len(Product.get_brands())}")
            
        except Exception as e:
            db.session.rollback()
            print(f"\n✗ Error committing to database: {e}")
            return False
        
        return True

if __name__ == '__main__':
    import_products_from_excel()
